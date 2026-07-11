from __future__ import annotations

import io
import gc
import json
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import tarfile
import tempfile
import time
import unittest
import zipfile
from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path
from urllib.parse import unquote
from unittest.mock import patch


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def load_web_module():
    spec = spec_from_file_location("bld_matcher_test_web", PROJECT_ROOT / "app.py")
    module = module_from_spec(spec)
    assert spec and spec.loader
    sys.modules["bld_matcher_test_web"] = module
    spec.loader.exec_module(module)
    return module


def pollute_xlsx_tail(path: Path, *, declared_rows: int = 2000, after_row: int = 251) -> None:
    temporary_path = path.with_suffix(".polluted.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(temporary_path, "w") as target:
        for entry in source.infolist():
            data = source.read(entry.filename)
            if entry.filename == "xl/worksheets/sheet1.xml":
                data = re.sub(rb'<dimension ref="[^"]+"', f'<dimension ref="A1:A{declared_rows}"'.encode(), data, count=1)
                empty_rows = b"".join(
                    f'<row r="{row_index}" s="1" customFormat="1"/>'.encode()
                    for row_index in range(after_row, declared_rows + 1)
                )
                data = data.replace(b"</sheetData>", empty_rows + b"</sheetData>")
            target.writestr(entry, data)
    temporary_path.replace(path)


def strip_xlsx_dimension(path: Path) -> None:
    temporary_path = path.with_suffix(".no-dimension.xlsx")
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(temporary_path, "w") as target:
        for entry in source.infolist():
            data = source.read(entry.filename)
            if entry.filename == "xl/worksheets/sheet1.xml":
                data = re.sub(rb"<dimension\b[^>]*/>", b"", data, count=1)
            target.writestr(entry, data)
    temporary_path.replace(path)


class WebAppTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp = tempfile.TemporaryDirectory()
        root = Path(cls.tmp.name)
        cls.root = root
        os.environ["SECRET_KEY"] = "test-secret"
        os.environ["MAX_UPLOAD_MB"] = "20"
        os.environ["PRODUCT_SYNC_MAX_UPLOAD_MB"] = "512"
        os.environ["BLD_DATA_DIR"] = str(root / "data")
        os.environ["BLD_UPLOAD_DIR"] = str(root / "uploads")
        os.environ["BLD_OUTPUT_DIR"] = str(root / "outputs")
        os.environ["DEFAULT_ADMIN_PASSWORD"] = "test-admin-pw"
        os.environ["INTERNAL_API_TOKEN"] = ""
        for module_name in [name for name in sys.modules if name == "app" or name.startswith("app.")]:
            sys.modules.pop(module_name, None)
        cls.web = load_web_module()
        if not cls.web.DB_PATH.resolve().is_relative_to(root.resolve()):
            raise RuntimeError(f"Tests must use the isolated database under {root}, got {cls.web.DB_PATH}")
        cls.web.app.config["TESTING"] = True
        cls.client = cls.web.app.test_client()

    @classmethod
    def tearDownClass(cls):
        cls.client = None
        cls.web = None
        gc.collect()
        cls.tmp.cleanup()

    def login(self):
        return self.client.post(
            "/login",
            data={"username": "007", "password": "test-admin-pw", "next": "/"},
            follow_redirects=False,
        )

    def create_internal_api_token(self, *, scopes=None, name="OpenClaw Test"):
        from app.platform.api_keys import create_internal_api_key

        with self.web.connect(self.web.DB_PATH) as conn:
            return create_internal_api_key(conn, actor="tester", name=name, scopes=scopes)

    def test_login_and_homepage(self):
        response = self.client.get("/login")
        self.assertEqual(response.status_code, 200)

        response = self.login()
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/"))

        response = self.client.get("/")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("BLD", html)
        self.assertLess(html.index('class="messages'), html.index('class="inquiry-landing"'))
        self.assertIn('class="embedded-submit" type="submit">开始匹配', html)
        self.assertIn('class="embedded-input-control"', html)
        self.assertIn('class="embedded-submit" type="submit">搜索', html)
        nav_order = ["询价处理", "报价记录", "合同管理", "产品目录", "生产料单", "货物识别"]
        nav_positions = [html.index(label) for label in nav_order]
        self.assertEqual(nav_positions, sorted(nav_positions))

    def test_quick_inquiry_results_can_filter_by_match_source(self):
        from app.database import upsert_product

        self.login()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "QF6010B",
                    "series": "TEST",
                    "item": "BLD FILTER HIT",
                    "oe_no_1": "OE-BLD-FILTER",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "QF-OE-HIT",
                    "series": "TEST",
                    "item": "OE FILTER HIT",
                    "oe_no_1": "QF6010-OE",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "QF-BRAND-HIT",
                    "series": "TEST",
                    "item": "BRAND FILTER HIT",
                    "oe_no_2": "QF6010-BRAND",
                    "active": "1",
                },
                actor="tester",
            )
            conn.commit()

        response = self.client.get("/?quick_oe=QF6010")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("只看BLD号", html)
        self.assertIn("只看OE号", html)
        self.assertIn("只看品牌号", html)
        self.assertIn("QF6010B", html)
        self.assertIn("QF-OE-HIT", html)
        self.assertIn("QF-BRAND-HIT", html)
        self.assertIn("命中BLD号：", html)
        self.assertIn("命中OE号：", html)
        self.assertIn("命中品牌号：", html)
        self.assertIn("QF6010-OE", html)
        self.assertIn("QF6010-BRAND", html)
        self.assertIn('data-quick-results data-initial-filter=""', html)
        self.assertIn('data-match-type="bld"', html)
        self.assertIn('data-match-type="oe"', html)
        self.assertIn('data-match-type="brand"', html)

        response = self.client.get("/?quick_oe=QF6010&quick_filter=bld")
        html = response.get_data(as_text=True)
        self.assertIn('data-quick-results data-initial-filter="bld"', html)
        self.assertIn("QF6010B", html)
        self.assertIn("QF-OE-HIT", html)
        self.assertIn("QF-BRAND-HIT", html)

        response = self.client.get("/?quick_oe=QF6010&quick_filter=oe")
        html = response.get_data(as_text=True)
        self.assertIn('data-quick-results data-initial-filter="oe"', html)
        self.assertIn("QF6010B", html)
        self.assertIn("QF-OE-HIT", html)
        self.assertIn("QF-BRAND-HIT", html)

        response = self.client.get("/?quick_oe=QF6010&quick_filter=brand")
        html = response.get_data(as_text=True)
        self.assertIn('data-quick-results data-initial-filter="brand"', html)
        self.assertIn("QF6010B", html)
        self.assertIn("QF-BRAND-HIT", html)
        self.assertIn("QF-OE-HIT", html)

    def test_login_next_rejects_external_url(self):
        response = self.client.post(
            "/login",
            data={"username": "007", "password": "test-admin-pw", "next": "https://example.com/phish"},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertNotIn("example.com", response.headers["Location"])

    def test_download_does_not_send_directories(self):
        self.login()
        output_dir = self.root / "outputs" / "u1-007"
        output_dir.mkdir(parents=True, exist_ok=True)
        response = self.client.get("/download/u1-007", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/"))

    def test_core_admin_pages_load(self):
        self.login()
        for path in ["/quotes", "/contracts", "/contracts/sales", "/products", "/materials", "/material-drawings", "/shipping-notices", "/shipment-recognition", "/purchase-contracts", "/users", "/internal-api-key", "/logs", "/system-updates", "/product-data-sync"]:
            with self.subTest(path=path):
                response = self.client.get(path)
                self.assertEqual(response.status_code, 200)

    def test_material_drawings_page_lists_codes_and_previews_pdf(self):
        self.login()
        drawing_dir = self.root / "data" / "material_drawings"
        drawing_dir.mkdir(parents=True, exist_ok=True)
        (drawing_dir / "QD1000.pdf").write_bytes(b"%PDF-1.4\n% test drawing\n")
        (drawing_dir / "QD999.pdf").write_bytes(b"%PDF-1.4\n% test drawing\n")

        response = self.client.get("/material-drawings?q=1000&category=%E7%90%83%E9%94%80")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("当前 1 个 / 共 2 个", html)
        self.assertIn("<strong>QD1000</strong>", html)
        self.assertIn("球销", html)
        self.assertIn('data-material-drawing-select', html)
        self.assertIn('data-material-drawing-frame', html)
        self.assertIn("/material-drawings/preview/QD1000.pdf", html)
        self.assertIn("/material-drawings/preview/QD1000.pdf#page=1&zoom=100", html)
        self.assertNotIn("<strong>QD999</strong>", html)

        selected_page = self.client.get("/material-drawings?selected=QD999.pdf")
        selected_html = selected_page.get_data(as_text=True)
        self.assertEqual(selected_page.status_code, 200)
        self.assertIn('data-material-drawing-current-code>QD999</h2>', selected_html)
        self.assertIn('data-material-drawing-current-download href="/material-drawings/QD999.pdf"', selected_html)

        preview = self.client.get("/material-drawings/preview/QD1000.pdf")
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.mimetype, "application/pdf")
        self.assertIn("inline", preview.headers.get("Content-Disposition", ""))
        preview.close()

    def _build_product_sync_package(self, rows: list[dict], *, media: bool = False) -> Path:
        package_path = self.root / "incoming-product-sync.tar.gz"
        work_dir = self.root / "incoming-product-sync"
        if work_dir.exists():
            shutil.rmtree(work_dir)
        if package_path.exists():
            package_path.unlink()
        data_dir = work_dir / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        product_db = data_dir / "products.sqlite3"
        target = sqlite3.connect(product_db)
        try:
            with self.web.connect(self.web.DB_PATH) as source:
                target.row_factory = sqlite3.Row
                schema = source.execute(
                    "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'products'"
                ).fetchone()["sql"]
                target.execute(schema)
                columns = [row["name"] for row in source.execute("PRAGMA table_info(products)")]
                insert_columns = ", ".join(columns)
                placeholders = ", ".join("?" for _ in columns)
                for index, row in enumerate(rows, start=1):
                    values = []
                    for column in columns:
                        if column == "id":
                            values.append(index)
                        else:
                            values.append(row.get(column, "" if column != "price_cny" else None))
                    target.execute(f"INSERT INTO products ({insert_columns}) VALUES ({placeholders})", values)
                target.commit()
        finally:
            target.close()
        (work_dir / "manifest.json").write_text(
            '{"package_type":"bld_product_data","version":1}',
            encoding="utf-8",
        )
        if media:
            image_dir = data_dir / "product_images"
            image_dir.mkdir(parents=True, exist_ok=True)
            (image_dir / "SYNC001.png").write_bytes(b"fake-image")
        with tarfile.open(package_path, "w:gz") as archive:
            archive.add(product_db, arcname="data/products.sqlite3")
            archive.add(work_dir / "manifest.json", arcname="manifest.json")
            if media:
                archive.add(data_dir / "product_images" / "SYNC001.png", arcname="data/product_images/SYNC001.png")
        return package_path

    def test_product_data_sync_exports_products_without_api_keys(self):
        from app.database import upsert_product

        self.login()
        self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "SYNC-EXPORT",
                    "series": "SYNC",
                    "item": "Export Test",
                    "oe_no_1": "SYNC-EXPORT-OE",
                    "active": "1",
                },
                actor="tester",
            )

        response = self.client.post("/product-data-sync/export")
        self.assertEqual(response.status_code, 200)
        package_path = self.root / "exported-product-data.tar.gz"
        package_path.write_bytes(response.data)
        response.close()
        with tarfile.open(package_path, "r:gz") as archive:
            names = set(archive.getnames())
            self.assertIn("data/products.sqlite3", names)
            self.assertIn("manifest.json", names)
            archive.extract("data/products.sqlite3", self.root / "export-check", filter="data")

        exported_db = self.root / "export-check" / "data" / "products.sqlite3"
        conn = sqlite3.connect(exported_db)
        try:
            tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'")}
            count = conn.execute("SELECT COUNT(*) FROM products WHERE bld_no = 'SYNC-EXPORT'").fetchone()[0]
        finally:
            conn.close()
        self.assertEqual(tables, {"products", "sqlite_sequence"})
        self.assertEqual(count, 1)

    def test_product_data_sync_imports_incrementally_and_preserves_api_key(self):
        from app.database import upsert_product
        from app.platform.api_keys import internal_api_key_status

        self.login()
        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "SYNC001",
                    "series": "LOCAL",
                    "item": "Local Item",
                    "oe_no_1": "OLD-OE",
                    "price_cny": "10",
                    "active": "1",
                },
                actor="tester",
            )

        package_path = self._build_product_sync_package(
            [
                {
                    "bld_no": "SYNC001",
                    "series": "NAS",
                    "item": "NAS Item",
                    "oe_no_1": "NEW-OE",
                    "price_cny": 12.5,
                    "image_path": "data_product_images/SYNC001.png",
                    "active": 1,
                    "source": "nas",
                    "created_at": "2026-05-01 00:00:00",
                    "updated_at": "2099-05-27 10:00:00",
                },
                {
                    "bld_no": "SYNC002",
                    "series": "NAS",
                    "item": "New Product",
                    "oe_no_1": "NEW-ONLY",
                    "price_cny": 20,
                    "active": 1,
                    "source": "nas",
                    "created_at": "2026-05-27 10:00:00",
                    "updated_at": "2099-05-27 10:00:00",
                },
            ],
            media=True,
        )
        response = self.client.post(
            "/product-data-sync/import/preview",
            data={
                "include_images": "1",
                "package": (package_path.open("rb"), package_path.name),
            },
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("新增产品", html)
        self.assertIn("SYNC001", html)
        self.assertIn("SYNC002", html)
        path_match = re.search(r'name="package_path" value="([^"]+)"', html)
        self.assertIsNotNone(path_match)

        response = self.client.post(
            "/product-data-sync/import/apply",
            data={"include_images": "1", "package_path": path_match.group(1)},
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("产品数据导入完成：新增 1 条，更新 1 条", html)
        with self.web.connect(self.web.DB_PATH) as conn:
            row = conn.execute("SELECT * FROM products WHERE bld_no = 'SYNC001'").fetchone()
            new_row = conn.execute("SELECT * FROM products WHERE bld_no = 'SYNC002'").fetchone()
            status = internal_api_key_status(conn)
        self.assertEqual(row["series"], "NAS")
        self.assertEqual(row["oe_no_1"], "NEW-OE")
        self.assertEqual(new_row["item"], "New Product")
        self.assertTrue(status["enabled"])
        self.assertTrue(token.endswith(status["preview"][-6:]))
        self.assertTrue((self.root / "data" / "product_images" / "SYNC001.png").exists())

    def test_product_data_sync_rolls_back_media_when_database_apply_fails(self):
        self.login()
        target = self.root / "data" / "product_images" / "SYNC001.png"
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(b"original-image")
        package_path = self._build_product_sync_package(
            [
                {
                    "bld_no": "SYNC-ROLLBACK",
                    "series": "SYNC",
                    "item": "Rollback Test",
                    "image_path": "data_product_images/SYNC001.png",
                    "active": 1,
                    "source": "test",
                    "created_at": "2026-07-11 00:00:00",
                    "updated_at": "2099-07-11 00:00:00",
                }
            ],
            media=True,
        )
        preview = self.client.post(
            "/product-data-sync/import/preview",
            data={"include_images": "1", "package": (package_path.open("rb"), package_path.name)},
            content_type="multipart/form-data",
        )
        path_match = re.search(r'name="package_path" value="([^"]+)"', preview.get_data(as_text=True))
        self.assertIsNotNone(path_match)

        with patch("app.routes.product_sync._apply_products", side_effect=RuntimeError("forced apply failure")):
            response = self.client.post(
                "/product-data-sync/import/apply",
                data={"include_images": "1", "package_path": path_match.group(1)},
                follow_redirects=True,
            )
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("forced apply failure", html)
        self.assertIn("已恢复本次媒体文件变更", html)
        self.assertEqual(target.read_bytes(), b"original-image")
        backups = sorted((self.root / "data" / "local-backups").glob("*/products.sqlite3"))
        self.assertTrue(backups)
        with sqlite3.connect(backups[-1]) as backup:
            self.assertEqual(backup.execute("PRAGMA integrity_check").fetchone()[0], "ok")

    def test_product_data_sync_skips_older_package_rows(self):
        from app.database import upsert_product

        self.login()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "SYNC-STALE",
                    "series": "LOCAL",
                    "item": "Local Newer",
                    "oe_no_1": "LOCAL-OE",
                    "price_cny": "30",
                    "active": "1",
                },
                actor="tester",
            )
            conn.execute(
                "UPDATE products SET updated_at = ? WHERE bld_no = ?",
                ("2099-01-01 00:00:00", "SYNC-STALE"),
            )
            conn.commit()

        package_path = self._build_product_sync_package(
            [
                {
                    "bld_no": "SYNC-STALE",
                    "series": "OLD-PACKAGE",
                    "item": "Should Not Overwrite",
                    "oe_no_1": "OLD-OE",
                    "price_cny": 1,
                    "active": 1,
                    "source": "old-package",
                    "created_at": "2020-01-01 00:00:00",
                    "updated_at": "2020-01-01 00:00:00",
                }
            ],
        )
        preview = self.client.post(
            "/product-data-sync/import/preview",
            data={"package": (package_path.open("rb"), package_path.name)},
            content_type="multipart/form-data",
        )
        preview_html = preview.get_data(as_text=True)
        self.assertEqual(preview.status_code, 200)
        self.assertIn("包内旧数据", preview_html)
        path_match = re.search(r'name="package_path" value="([^"]+)"', preview_html)
        self.assertIsNotNone(path_match)

        applied = self.client.post(
            "/product-data-sync/import/apply",
            data={"package_path": path_match.group(1)},
            follow_redirects=True,
        )
        applied_html = applied.get_data(as_text=True)
        self.assertEqual(applied.status_code, 200)
        self.assertIn("跳过包内旧数据 1 条", applied_html)
        with self.web.connect(self.web.DB_PATH) as conn:
            row = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("SYNC-STALE",)).fetchone()
        self.assertEqual(row["series"], "LOCAL")
        self.assertEqual(row["oe_no_1"], "LOCAL-OE")

    def test_product_data_sync_can_deactivate_local_only_rows_after_preview(self):
        from app.database import upsert_product

        self.login()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "SYNC-LOCAL-ONLY",
                    "series": "LOCAL",
                    "item": "Only On This Machine",
                    "oe_no_1": "LOCAL-ONLY-OE",
                    "price_cny": "30",
                    "active": "1",
                },
                actor="tester",
            )

        package_path = self._build_product_sync_package(
            [
                {
                    "bld_no": "SYNC-PACKAGE-ONLY",
                    "series": "NAS",
                    "item": "Package Product",
                    "oe_no_1": "PACKAGE-OE",
                    "price_cny": 20,
                    "active": 1,
                    "source": "nas",
                    "created_at": "2026-06-01 00:00:00",
                    "updated_at": "2099-06-01 00:00:00",
                }
            ],
        )
        preview = self.client.post(
            "/product-data-sync/import/preview",
            data={"package": (package_path.open("rb"), package_path.name)},
            content_type="multipart/form-data",
        )
        html = preview.get_data(as_text=True)
        self.assertEqual(preview.status_code, 200)
        self.assertIn("当前系统独有", html)
        self.assertIn("SYNC-LOCAL-ONLY", html)
        self.assertIn("deactivate_local_only", html)
        path_match = re.search(r'name="package_path" value="([^"]+)"', html)
        self.assertIsNotNone(path_match)

        response = self.client.post(
            "/product-data-sync/import/apply",
            data={"package_path": path_match.group(1)},
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("停用本机独有 0 条", html)
        with self.web.connect(self.web.DB_PATH) as conn:
            row = conn.execute("SELECT active FROM products WHERE bld_no = ?", ("SYNC-LOCAL-ONLY",)).fetchone()
        self.assertEqual(row["active"], 1)

        response = self.client.post(
            "/product-data-sync/import/apply",
            data={"package_path": path_match.group(1), "deactivate_local_only": "1"},
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("停用本机独有", html)
        with self.web.connect(self.web.DB_PATH) as conn:
            row = conn.execute("SELECT active FROM products WHERE bld_no = ?", ("SYNC-LOCAL-ONLY",)).fetchone()
        self.assertEqual(row["active"], 0)

    def test_shipment_recognition_requires_photos(self):
        self.login()
        response = self.client.post(
            "/shipment-recognition/run",
            data={"provider": "tesseract"},
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("请选择 jpg、png、webp、bmp、tif、heic 或 heif 照片", html)

    def test_shipping_notice_template_preview_and_generate(self):
        from openpyxl import Workbook, load_workbook

        template_book = Workbook()
        template_sheet = template_book.active
        template_sheet.title = "通知模板"
        template_sheet.append(["客户", "商品编码", "数量", "备注"])
        template_sheet.append(["ABC", "", "", "固定"])
        template_buffer = io.BytesIO()
        template_book.save(template_buffer)
        template_book.close()
        template_buffer.seek(0)

        self.login()
        upload = self.client.post(
            "/shipping-notices/templates/upload",
            data={
                "customer": "ABC客户",
                "template_name": "标准模板",
                "template": (template_buffer, "abc-template.xlsx"),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        self.assertEqual(upload.status_code, 302)
        template_id = re.search(r"template_id=([^&]+)", upload.headers["Location"]).group(1)

        page = self.client.get(f"/shipping-notices?template_id={template_id}")
        html = page.get_data(as_text=True)
        self.assertEqual(page.status_code, 200)
        self.assertIn("发货通知", html)
        self.assertIn("选择客户模板", html)
        self.assertIn("模板管理", html)
        self.assertIn("data-open-shipping-template-action-modal", html)
        self.assertIn("ABC客户", html)
        self.assertIn("data-file-drop-zone", html)
        self.assertIn("shipping-history-drawer", html)

        data_book = Workbook()
        data_sheet = data_book.active
        data_sheet.title = "发货数据"
        data_sheet.append(["商品编码", "数量"])
        data_sheet.append(["K8001LA", 12])
        data_sheet.append(["K8001RA", 8])
        data_buffer = io.BytesIO()
        data_book.save(data_buffer)
        data_book.close()
        data_buffer.seek(0)

        preview = self.client.post(
            "/shipping-notices/preview",
            data={
                "template_id": template_id,
                "shipment_data": (data_buffer, "shipment.xlsx"),
            },
            content_type="multipart/form-data",
        )
        preview_html = preview.get_data(as_text=True)
        self.assertEqual(preview.status_code, 200)
        self.assertIn("发货通知预览", preview_html)
        self.assertIn("生成前预览", preview_html)
        self.assertIn("K8001LA", preview_html)
        upload_match = re.search(r'name="upload_path" value="([^"]+)"', preview_html)
        self.assertIsNotNone(upload_match)

        generated = self.client.post(
            "/shipping-notices/generate",
            data={"template_id": template_id, "upload_path": upload_match.group(1)},
            follow_redirects=False,
        )
        self.assertEqual(generated.status_code, 302)
        output_name = unquote(re.search(r"generated=([^&]+)", generated.headers["Location"]).group(1))
        output_path = self.root / "outputs" / output_name
        self.assertTrue(output_path.exists())
        workbook = load_workbook(output_path, data_only=True)
        sheet = workbook.active
        self.assertEqual(sheet.cell(2, 2).value, "K8001LA")
        self.assertEqual(sheet.cell(2, 3).value, 12)
        self.assertEqual(sheet.cell(3, 2).value, "K8001RA")
        self.assertEqual(sheet.cell(3, 3).value, 8)
        workbook.close()

    def test_shipping_notice_batch_template_upload(self):
        from openpyxl import Workbook

        self.login()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["商品编码", "数量"])
        sheet.append(["", ""])
        template_buffer = io.BytesIO()
        workbook.save(template_buffer)
        workbook.close()

        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w") as archive:
            archive.writestr("批量客户-发货模板.xlsx", template_buffer.getvalue())
        zip_buffer.seek(0)
        batch = self.client.post(
            "/shipping-notices/templates/batch",
            data={"template_zip": (zip_buffer, "templates.zip")},
            content_type="multipart/form-data",
            follow_redirects=True,
        )
        batch_html = batch.get_data(as_text=True)
        self.assertEqual(batch.status_code, 200)
        self.assertIn("已导入 1 个模板", batch_html)
        self.assertIn("批量客户", batch_html)

    def test_shipment_recognition_async_job_completes(self):
        self.login()
        received_args = {}

        def fake_recognize_photo(job, args):
            received_args.update({"model": args.model, "base_url": args.base_url})
            return {
                "relative_name": job.relative_name,
                "path": str(job.path),
                "status": "ok",
                "seconds": 0.01,
                "model": "fake-qwen-vl",
                "usage": {"prompt_tokens": 11, "completion_tokens": 7, "total_tokens": 18},
                "result": {
                    "photo_summary": "测试照片",
                    "labels": [
                        {
                            "label_index": 1,
                            "visible": True,
                            "label_type": "part",
                            "numbers": ["54501-8Y50B"],
                            "part_no": "54501-8Y50B",
                            "bld_no": "",
                            "oe_no": "",
                            "customer_code": "",
                            "product_name": "CONTROL ARM",
                            "models": "TEST CAR",
                            "quantity": 10,
                            "carton_size": "",
                            "barcode": "",
                            "confidence": 0.95,
                            "notes": "",
                        }
                    ],
                },
                "error": "",
            }

        png_bytes = (
            b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
            b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\xff\xff?"
            b"\x00\x05\xfe\x02\xfeA\x89\xa3\x95\x00\x00\x00\x00IEND\xaeB`\x82"
        )
        with patch("app.routes.shipment_recognition.recognizer.recognize_photo", side_effect=fake_recognize_photo):
            response = self.client.post(
                "/shipment-recognition/run",
                data={
                    "provider": "openai-compatible",
                    "model": "attacker-model",
                    "base_url": "https://attacker.invalid/v1",
                    "shipment_photos": (io.BytesIO(png_bytes), "IMG_0001.png"),
                },
                headers={"Accept": "application/json", "X-Requested-With": "fetch"},
                content_type="multipart/form-data",
            )
            payload = response.get_json()
            self.assertEqual(response.status_code, 202)
            self.assertTrue(payload["ok"])

            status_payload = None
            for _ in range(30):
                status = self.client.get(payload["status_url"], headers={"Accept": "application/json"})
                status_payload = status.get_json()
                if status_payload["job"]["status"] == "completed":
                    break
                time.sleep(0.1)

        self.assertIsNotNone(status_payload)
        job = status_payload["job"]
        self.assertEqual(job["status"], "completed")
        self.assertEqual(job["result"]["photos"], 1)
        self.assertEqual(job["result"]["labels"], 1)
        self.assertEqual(job["result"]["total_tokens"], 18)
        self.assertTrue(job["result"]["excel_url"].startswith("/download/"))
        self.assertEqual(received_args, {"model": None, "base_url": None})

    def test_shipment_recognition_prepares_heic_images(self):
        from PIL import Image
        from tools.shipment_photo_recognition import _prepare_image, find_photos

        image_dir = self.root / "uploads" / "heic-test"
        image_dir.mkdir(parents=True, exist_ok=True)
        image_path = image_dir / "label.heic"
        Image.new("RGB", (12, 8), "white").save(image_path, format="HEIF")

        jobs = find_photos(image_dir)
        raw, mime = _prepare_image(image_path, max_side=2200)

        self.assertEqual([job.relative_name for job in jobs], ["label.heic"])
        self.assertEqual(mime, "image/jpeg")
        self.assertTrue(raw.startswith(b"\xff\xd8\xff"))

    def test_internal_api_numbers_generate_openclaw_workbook(self):
        from app.database import upsert_product
        from openpyxl import load_workbook

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-001",
                    "series": "HYUNDAI",
                    "item": "API CONTROL ARM",
                    "oe_no_1": "API-001",
                    "models": "ApiTester",
                    "price_cny": "88.8",
                    "active": "1",
                },
                actor="tester",
            )

        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={
                "numbers": ["API-001", "NO-MATCH-001"],
                "source_name": "机器人询价结果",
                "price_mode": "net",
                "rows_limit": 10,
                "export": True,
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["mode"], "new-workbook")
        self.assertEqual(payload["matched_count"], 1)
        self.assertEqual(payload["unmatched_count"], 1)
        self.assertEqual(payload["rows"][0]["bld_no"], "K-API-001")
        self.assertEqual(payload["rows"][0]["export_price"], 81)
        self.assertIn("NO-MATCH-001", payload["unmatched_list"])
        self.assertTrue(payload["output_path"].endswith(payload["output_name"]))
        self.assertRegex(payload["output_name"], r"^re\d{6}_机器人询价结果\.xlsx$")
        output_path = Path(payload["output_path"])
        self.assertEqual(output_path.parent.resolve(), (self.root / "outputs" / "openclaw").resolve())
        self.assertTrue(output_path.exists())

        workbook = load_workbook(output_path)
        sheet = workbook.active
        self.assertEqual(sheet.cell(1, 1).value, "OE号")
        self.assertEqual(sheet.cell(1, 2).value, "BLD NO.")
        self.assertEqual(sheet.cell(1, 3).value, "不含税单价")
        self.assertEqual(sheet.cell(2, 2).value, "K-API-001")
        self.assertEqual(sheet.cell(2, 3).value, 81)
        workbook.close()

        rejected_export = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["API-001"], "export": True},
            headers={"Authorization": f"Bearer {token}"},
        )
        rejected_payload = rejected_export.get_json()
        self.assertEqual(rejected_export.status_code, 400)
        self.assertFalse(rejected_payload["ok"])
        self.assertIn("必须传 source_name", rejected_payload["error"])

    def test_internal_api_numbers_use_oe_suffix_variant_matching(self):
        from app.database import upsert_product

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8041LB",
                    "series": "VW",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "561407151A\n561407151C",
                    "active": "1",
                },
                actor="tester",
            )

        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["561407151D"], "rows_limit": 10},
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["matched_count"], 1)
        self.assertEqual(payload["unmatched_count"], 0)
        self.assertEqual(payload["rows"][0]["bld_no"], "K8041LB")
        self.assertEqual(payload["rows"][0]["match_reason"], "OE 尾字母容错命中")

    def test_internal_api_numbers_use_bld_fragment_lookup(self):
        from app.database import upsert_product
        from openpyxl import load_workbook

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8072LA",
                    "series": "NISSAN",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "54501-TEST-LA",
                    "oe_no_2": "Moog: TEST-LA",
                    "models": "VERSA TEST",
                    "price_cny": "43",
                    "image_path": "data_product_images/K8072LA.png",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K8072RA",
                    "series": "NISSAN",
                    "item": "Front Right Lower Control Arm",
                    "price_cny": "43",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-OE-ONLY",
                    "series": "TEST",
                    "item": "OE SHOULD NOT WIN BLD SHORTHAND",
                    "oe_no_1": "8072",
                    "active": "1",
                },
                actor="tester",
            )

        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["8072"], "rows_limit": 10},
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["matched_count"], 2)
        self.assertEqual(payload["unmatched_count"], 0)
        self.assertEqual([row["bld_no"] for row in payload["rows"]], ["K8072LA", "K8072RA"])
        self.assertEqual(payload["rows"][0]["original_number"], "8072")
        self.assertEqual(payload["rows"][0]["match_reason"], "BLD NO. 片段命中")
        self.assertEqual(payload["rows"][0]["price_cny"], 43.0)
        self.assertEqual(payload["rows"][0]["product"]["item"], "Front Left Lower Control Arm")
        self.assertEqual(payload["rows"][0]["product"]["oe_no_1"], "54501-TEST-LA")
        self.assertEqual(payload["rows"][0]["product"]["oe_no_2"], "Moog: TEST-LA")
        self.assertEqual(payload["rows"][0]["product"]["models"], "VERSA TEST")
        self.assertEqual(payload["rows"][0]["product"]["image_paths"], ["data_product_images/K8072LA.png"])

        k_response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["K8072"], "rows_limit": 10},
            headers={"Authorization": f"Bearer {token}"},
        )
        k_payload = k_response.get_json()
        self.assertEqual(k_response.status_code, 200)
        self.assertEqual(k_payload["matched_count"], 2)
        self.assertEqual([row["bld_no"] for row in k_payload["rows"]], ["K8072LA", "K8072RA"])

        export_response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["8072"], "source_name": "片段查询", "export": True, "rows_limit": 10},
            headers={"Authorization": f"Bearer {token}"},
        )
        export_payload = export_response.get_json()
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(export_payload["matched_count"], 2)
        self.assertEqual([row["bld_no"] for row in export_payload["rows"]], ["K8072LA", "K8072RA"])
        generated = load_workbook(Path(export_payload["output_path"]), read_only=True, data_only=True)
        try:
            sheet = generated.active
            self.assertEqual(sheet.cell(2, 1).value, "8072")
            self.assertEqual(sheet.cell(2, 2).value, "K8072LA")
            self.assertEqual(sheet.cell(3, 1).value, "8072")
            self.assertEqual(sheet.cell(3, 2).value, "K8072RA")
        finally:
            generated.close()

    def test_internal_api_numbers_use_psa_352x_dot_rule(self):
        from app.database import upsert_product

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-PSA-352126",
                    "series": "PEUGEOT\nCITROEN",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "3521.26",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-GM-352126",
                    "series": "GM\nOPEL",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "352126",
                    "active": "1",
                },
                actor="tester",
            )

        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["3521.26"], "rows_limit": 10},
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["matched_count"], 1)
        self.assertEqual(payload["rows"][0]["bld_no"], "K-API-PSA-352126")
        self.assertEqual(payload["rows"][0]["match_reason"], "PSA 号码点号容错命中")

    def test_internal_api_file_augment_and_analyze(self):
        from app.database import upsert_product
        from openpyxl import Workbook, load_workbook

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-FILE",
                    "series": "HYUNDAI",
                    "item": "API FILE ARM",
                    "oe_no_1": "API-FILE-OE",
                    "models": "ApiFileTester",
                    "price_cny": "79.2",
                    "active": "1",
                },
                actor="tester",
            )

        source_path = self.root / "uploads" / "api-file-source.xlsx"
        source_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["编号", "数量"])
        sheet.append(["API-FILE-OE", 2])
        workbook.save(source_path)
        workbook.close()

        analyze = self.client.post(
            "/api/internal/inquiry/analyze",
            json={"file_path": str(source_path), "match_column": "A", "price_mode": "tax"},
            headers={"Authorization": f"Bearer {token}"},
        )
        analyze_payload = analyze.get_json()
        self.assertEqual(analyze.status_code, 200)
        self.assertTrue(analyze_payload["ok"])
        self.assertEqual(analyze_payload["mode"], "augment-source-workbook")
        self.assertEqual(analyze_payload["summary"]["output_generated"], False)
        self.assertIsNone(analyze_payload["output_path"])
        self.assertEqual(analyze_payload["matched_count"], 1)

        response = self.client.post(
            "/api/internal/inquiry/file",
            json={
                "file_path": str(source_path),
                "match_column": "A",
                "price_mode": "usd",
                "exchange_rate": "7.2",
                "export": True,
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["mode"], "augment-source-workbook")
        self.assertEqual(payload["rows"][0]["export_price"], 10)
        output_path = Path(payload["output_path"])
        self.assertEqual(output_path.parent.resolve(), (self.root / "outputs" / "openclaw").resolve())
        self.assertRegex(payload["output_name"], r"^re\d{6}_api-file-source\.xlsx$")
        self.assertTrue(output_path.exists())

        generated = load_workbook(output_path)
        sheet = generated.active
        self.assertEqual(sheet.cell(1, 3).value, "BLD NO.")
        self.assertEqual(sheet.cell(1, 4).value, "美金价")
        self.assertEqual(sheet.cell(1, 5).value, "Product Status")
        self.assertEqual(sheet.cell(1, 6).value, "匹配说明")
        self.assertEqual(sheet.cell(2, 3).value, "K-API-FILE")
        self.assertEqual(sheet.cell(2, 4).value, 10)
        generated.close()

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-FILE-MULTI",
                    "series": "HYUNDAI",
                    "item": "API MULTI ARM",
                    "oe_no_1": "API-FILE-REF",
                    "active": "1",
                },
                actor="tester",
            )
        multi_source = self.root / "uploads" / "api-file-multi-source.xlsx"
        multi_workbook = Workbook()
        multi_sheet = multi_workbook.active
        multi_sheet.append(["客户号码", "参考号"])
        multi_sheet.append(["NO-HIT-API", "API-FILE-REF"])
        multi_workbook.save(multi_source)
        multi_workbook.close()

        multi_response = self.client.post(
            "/api/internal/inquiry/analyze",
            json={"file_path": str(multi_source), "match_columns": ["A", "B"], "rows_limit": 10},
            headers={"Authorization": f"Bearer {token}"},
        )
        multi_payload = multi_response.get_json()
        self.assertEqual(multi_response.status_code, 200)
        self.assertEqual(multi_payload["matched_count"], 1)
        self.assertEqual(multi_payload["rows"][0]["bld_no"], "K-API-FILE-MULTI")
        self.assertIn("B列：API-FILE-REF", multi_payload["rows"][0]["match_note"])

    def test_internal_api_defaults_to_analysis_and_restricts_file_path(self):
        from app.database import upsert_product
        from openpyxl import Workbook

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-DEFAULT",
                    "oe_no_1": "API-DEFAULT-OE",
                    "active": "1",
                },
                actor="tester",
            )

        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["API-DEFAULT-OE"]},
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["matched_count"], 1)
        self.assertFalse(payload["summary"]["output_generated"])
        self.assertIsNone(payload["output_path"])
        openclaw_upload_dir = self.root / "uploads" / "openclaw"
        before_uploads = set(openclaw_upload_dir.glob("*")) if openclaw_upload_dir.exists() else set()

        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["API-DEFAULT-OE"], "export": False},
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["matched_count"], 1)
        self.assertFalse(payload["summary"]["output_generated"])
        self.assertIsNone(payload["output_path"])
        after_uploads = set(openclaw_upload_dir.glob("*")) if openclaw_upload_dir.exists() else set()
        self.assertEqual(after_uploads, before_uploads)

        outside_path = self.root / "outside-api-source.xlsx"
        workbook = Workbook()
        workbook.active.append(["OE号"])
        workbook.active.append(["API-DEFAULT-OE"])
        workbook.save(outside_path)
        workbook.close()

        rejected = self.client.post(
            "/api/internal/inquiry/file",
            json={"file_path": str(outside_path), "export": True},
            headers={"Authorization": f"Bearer {token}"},
        )
        rejected_payload = rejected.get_json()
        self.assertEqual(rejected.status_code, 400)
        self.assertFalse(rejected_payload["ok"])
        self.assertIn("file_path 不在允许读取范围内", rejected_payload["error"])

    def test_internal_api_requires_api_key(self):
        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["API-001"]},
        )
        self.assertEqual(response.status_code, 401)
        self.assertFalse(response.get_json()["ok"])

        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-AUTH",
                    "oe_no_1": "API-AUTH-OE",
                    "active": "1",
                },
                actor="tester",
            )
        token = self.create_internal_api_token()
        response = self.client.post(
            "/api/internal/inquiry/analyze",
            json={"numbers": ["API-AUTH-OE"]},
            headers={"Authorization": f"Bearer {token}"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["ok"])

    def test_v1_api_index_and_openapi_use_stable_contract(self):
        unauthorized = self.client.get("/api/v1", headers={"X-Request-ID": "v1-unauthorized-1"})
        unauthorized_payload = unauthorized.get_json()
        self.assertEqual(unauthorized.status_code, 401)
        self.assertEqual(unauthorized_payload["api_version"], "1")
        self.assertEqual(unauthorized_payload["request_id"], "v1-unauthorized-1")
        self.assertEqual(unauthorized_payload["error"]["code"], "auth.unauthorized")

        token = self.create_internal_api_token()
        headers = {"Authorization": f"Bearer {token}", "X-Request-ID": "v1-index-1"}
        response = self.client.get("/api/v1", headers=headers)
        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["X-Request-ID"], "v1-index-1")
        self.assertEqual(response.headers["Cache-Control"], "no-store")
        self.assertEqual(payload["api_version"], "1")
        self.assertEqual(payload["request_id"], "v1-index-1")
        self.assertEqual(payload["data"]["name"], "bld-matcher")
        self.assertIn("openapi", payload["data"]["capabilities"])

        document_response = self.client.get("/api/v1/openapi.json", headers=headers)
        document = document_response.get_json()
        self.assertEqual(document_response.status_code, 200)
        self.assertEqual(document["openapi"], "3.1.0")
        self.assertIn("/api/v1", document["paths"])
        self.assertIn("/api/v1/openapi.json", document["paths"])
        self.assertEqual(
            document["paths"]["/api/v1"]["get"]["x-required-scopes"],
            ["api:read"],
        )
        self.assertIn("PlatformInfoEnvelope", document["components"]["schemas"])

    def test_admin_can_generate_and_disable_internal_api_key(self):
        self.login()
        page = self.client.get("/internal-api-key")
        html = page.get_data(as_text=True)
        self.assertEqual(page.status_code, 200)
        self.assertIn("内部 API Key", html)
        self.assertIn("生成 API Key", html)

        no_scope = self.client.post(
            "/internal-api-key/generate",
            data={"name": "No Scope", "scope_selection_present": "1"},
            follow_redirects=True,
        )
        self.assertEqual(no_scope.status_code, 200)
        self.assertIn("API Key 至少需要一个 Scope", no_scope.get_data(as_text=True))

        generated = self.client.post(
            "/internal-api-key/generate",
            data={"name": "OpenClaw Visual"},
        )
        html = generated.get_data(as_text=True)
        self.assertEqual(generated.status_code, 200)
        self.assertEqual(generated.headers.get("Cache-Control"), "no-store")
        token_match = re.search(r'id="generated-api-key">(bld_sk_[^<]+)</code>', html)
        self.assertIsNotNone(token_match)
        token = token_match.group(1)
        self.assertIn("OpenClaw Visual", html)
        self.assertIn("quotes:read", html)
        self.assertIn(token, html)

        generated_second = self.client.post(
            "/internal-api-key/generate",
            data={"name": "OpenClaw Backup"},
        )
        html = generated_second.get_data(as_text=True)
        second_match = re.search(r'id="generated-api-key">(bld_sk_[^<]+)</code>', html)
        self.assertIsNotNone(second_match)
        second_token = second_match.group(1)
        self.assertNotIn(token, html)
        self.assertIn(second_token, html)

        with self.web.connect(self.web.DB_PATH) as conn:
            first_key = conn.execute(
                "SELECT id FROM internal_api_keys WHERE name = ?",
                ("OpenClaw Visual",),
            ).fetchone()
            key_columns = {
                row["name"] for row in conn.execute("PRAGMA table_info(internal_api_keys)").fetchall()
            }
        self.assertIsNotNone(first_key)
        self.assertNotIn("token_plain", key_columns)
        self.assertIn("scopes", key_columns)
        self.assertIn("expires_at", key_columns)

        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-VISUAL",
                    "oe_no_1": "API-VISUAL-OE",
                    "active": "1",
                },
                actor="tester",
            )
        api_response = self.client.post(
            "/api/internal/inquiry/analyze",
            json={"numbers": ["API-VISUAL-OE"]},
            headers={"Authorization": f"Bearer {token}"},
        )
        self.assertEqual(api_response.status_code, 200)
        second_api_response = self.client.post(
            "/api/internal/inquiry/analyze",
            json={"numbers": ["API-VISUAL-OE"]},
            headers={"Authorization": f"Bearer {second_token}"},
        )
        self.assertEqual(second_api_response.status_code, 200)

        disabled = self.client.post("/internal-api-key/disable", data={"key_id": str(first_key["id"])})
        self.assertEqual(disabled.status_code, 302)
        rejected = self.client.post(
            "/api/internal/inquiry/analyze",
            json={"numbers": ["NO-MATCH-VISUAL"]},
            headers={"Authorization": f"Bearer {token}"},
        )
        self.assertEqual(rejected.status_code, 401)
        still_accepted = self.client.post(
            "/api/internal/inquiry/analyze",
            json={"numbers": ["API-VISUAL-OE"]},
            headers={"Authorization": f"Bearer {second_token}"},
        )
        self.assertEqual(still_accepted.status_code, 200)

    def test_purchase_contract_can_generate_pdf(self):
        from PIL import Image
        from app.database import upsert_product

        image_dir = self.root / "data" / "product_images"
        image_dir.mkdir(parents=True, exist_ok=True)
        Image.new("RGB", (96, 64), "white").save(image_dir / "K-OUT-001.png")
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-OUT-001",
                    "oe_no_1": "OE-CATALOG-001",
                    "item": "目录外购支架",
                    "models": "目录车型",
                    "price_cny": "45.5",
                    "image_path": "data_product_images/K-OUT-001.png",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/contracts")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("合同管理", html)
        self.assertIn("采购合同", html)
        self.assertIn("销售合同", html)
        self.assertIn('class="contract-history-drawer" id="contract-history">', html)
        self.assertIn("历史合同记录", html)
        self.assertNotIn("销售合同模板后续接入", html)
        self.assertIn("玉环博莱德机械有限公司", html)
        self.assertIn("浙江省玉环市金汇路11号", html)
        self.assertIn("月结 30 天", html)
        self.assertIn('name="product_code[]"', html)
        self.assertIn('name="oe_no[]"', html)
        self.assertIn('name="models[]"', html)
        self.assertIn('data-add-purchase-row', html)
        self.assertIn('data-supplier-sign-name', html)
        self.assertIn('data-purchase-confirm-modal', html)
        self.assertIn("确认生成 PDF", html)
        self.assertNotIn("统一社会信用代码", html)
        self.assertIn('name="buyer_signature_address"', html)
        self.assertIn('name="supplier_signature_address"', html)
        self.assertIn('name="buyer_bank"', html)
        self.assertIn('name="supplier_bank_account"', html)
        self.assertIn('name="buyer_signature_date"', html)
        self.assertIn("supplier-detail-line", html)

        lookup = self.client.get("/purchase-contracts/product-lookup", query_string={"bld": "K-OUT-001"})
        self.assertEqual(lookup.status_code, 200)
        payload = lookup.get_json()
        self.assertTrue(payload["found"])
        self.assertEqual(payload["oe_no"], "OE-CATALOG-001")
        self.assertEqual(payload["product_name"], "目录外购支架")
        self.assertEqual(payload["models"], "目录车型")
        self.assertEqual(payload["price_cny"], 45.5)
        self.assertIn("product-image-thumbs", payload["thumb_url"])

        response = self.client.post(
            "/purchase-contracts/generate",
            data={
                "contract_no": "CG-TEST-001",
                "contract_date": "2026-05-05",
                "buyer_name": "玉环博莱德机械有限公司",
                "buyer_contact": "李四",
                "buyer_phone": "13900000000",
                "supplier_name": "外购供应商",
                "supplier_contact": "张三",
                "supplier_phone": "13800000000",
                "buyer_signature_address": "甲方签章地址",
                "supplier_signature_address": "乙方签章地址",
                "buyer_signature_phone": "0576-11111111",
                "supplier_signature_phone": "0576-22222222",
                "buyer_bank": "甲方开户行",
                "supplier_bank": "乙方开户行",
                "buyer_bank_account": "11112222",
                "supplier_bank_account": "33334444",
                "buyer_signature_date": "2026-05-06",
                "supplier_signature_date": "2026-05-07",
                "delivery_address": "浙江省玉环市",
                "price_note": "以上价格为含税价（增值税税率13%），含包装费及运费，送达甲方指定地点。",
                "payment_terms": "月结",
                "quality_terms": "按图纸执行",
                "product_code[]": ["K-OUT-001", ""],
                "oe_no[]": ["手填OE", ""],
                "product_name[]": ["手填名称", ""],
                "models[]": ["手填车型", ""],
                "quantity[]": ["10", ""],
                "unit_price[]": ["25.5", ""],
                "delivery_date[]": ["2026-05-20", ""],
                "item_note[]": ["加急", ""],
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/pdf")
        self.assertIn("CG-TEST-001", response.headers["Content-Disposition"])
        self.assertTrue(response.get_data().startswith(b"%PDF-"))
        response.close()
        files = list((self.root / "outputs").glob("u*-007/采购合同/外购供应商/CG-TEST-001外购供应商.pdf"))
        self.assertEqual(len(files), 1)
        response = self.client.get("/contracts", query_string={"contract_q": "CG-TEST-001"})
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("已生成合同", html)
        self.assertIn('class="contract-history-drawer" id="contract-history" open>', html)
        self.assertIn("外购供应商", html)
        self.assertIn("CG-TEST-001外购供应商.pdf", html)
        self.assertIn("下载", html)

    def test_sales_contract_can_generate_pdf_with_customer_code(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-SALE-001",
                    "oe_no_1": "OE-SALE-001",
                    "item": "销售控制臂",
                    "models": "销售车型",
                    "price_cny": "88.8",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/contracts/sales")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("产 品 销 售 合 同", html)
        self.assertIn('class="contract-history-drawer" id="contract-history">', html)
        self.assertIn("供方（甲方）", html)
        self.assertIn("需方（乙方）", html)
        self.assertIn("客户编码", html)
        self.assertIn('name="customer_code[]"', html)
        self.assertIn('action="/sales-contracts/generate"', html)
        self.assertIn("甲方按行业通用标准及乙方要求进行包装", html)
        self.assertIn("增值税专用发票（税率 13%）的开具时间由双方另行约定", html)

        response = self.client.post(
            "/sales-contracts/generate",
            data={
                "contract_no": "XS-TEST-001",
                "contract_date": "2026-05-06",
                "buyer_name": "玉环博莱德机械有限公司",
                "buyer_contact": "李四",
                "buyer_phone": "13900000000",
                "supplier_name": "销售客户",
                "supplier_contact": "王五",
                "supplier_phone": "13700000000",
                "delivery_address": "客户仓库",
                "price_note": "以上价格为含税价。",
                "payment_terms": "月结 30 天",
                "quality_terms": "按封样执行",
                "product_code[]": ["K-SALE-001"],
                "customer_code[]": ["CUST-001"],
                "oe_no[]": ["手填销售OE"],
                "product_name[]": ["手填销售名称"],
                "models[]": ["手填销售车型"],
                "quantity[]": ["3"],
                "unit_price[]": ["88.8"],
                "delivery_date[]": ["2026-05-25"],
                "item_note[]": ["销售备注"],
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/pdf")
        self.assertIn("XS-TEST-001", response.headers["Content-Disposition"])
        self.assertTrue(response.get_data().startswith(b"%PDF-"))
        response.close()
        files = list((self.root / "outputs").glob("u*-007/销售合同/销售客户/XS-TEST-001销售客户.pdf"))
        self.assertEqual(len(files), 1)
        response = self.client.get("/contracts", query_string={"contract_type": "sales", "contract_q": "销售客户"})
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("已生成合同", html)
        self.assertIn("销售合同", html)
        self.assertIn("销售客户", html)
        self.assertIn("XS-TEST-001销售客户.pdf", html)

    def test_purchase_contract_signature_fields_are_optional(self):
        from app.purchase_contract import purchase_contract_from_form

        class FormData(dict):
            def getlist(self, key):
                value = self.get(key, [])
                return value if isinstance(value, list) else [value]

        contract = purchase_contract_from_form(
            FormData(
                {
                    "contract_no": "CG-OPTIONAL-SIGN",
                    "contract_date": "2026-05-06",
                    "buyer_name": "甲方公司",
                    "supplier_name": "乙方公司",
                    "product_code[]": ["K-OPTIONAL-001"],
                    "quantity[]": ["1"],
                    "unit_price[]": ["2.50"],
                }
            )
        )

        self.assertEqual(contract["buyer_signature_address"], "")
        self.assertEqual(contract["supplier_signature_phone"], "")
        self.assertEqual(contract["buyer_bank"], "")
        self.assertEqual(contract["supplier_bank_account"], "")
        self.assertEqual(contract["buyer_signature_date"], "")

    def test_purchase_contracts_are_admin_only(self):
        from app.database import save_user

        with self.web.connect(self.web.DB_PATH) as conn:
            save_user(
                conn,
                {
                    "username": "editor-contracts",
                    "display_name": "Editor Contracts",
                    "password": "editor-pw",
                    "role": "editor",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        admin_page = self.client.get("/").get_data(as_text=True)
        self.assertIn("合同管理", admin_page)
        self.client.post("/logout")

        login = self.client.post(
            "/login",
            data={"username": "editor-contracts", "password": "editor-pw", "next": "/"},
            follow_redirects=False,
        )
        self.assertEqual(login.status_code, 302)

        editor_page = self.client.get("/").get_data(as_text=True)
        self.assertNotIn("合同管理", editor_page)
        for path in ["/contracts", "/contracts/sales", "/purchase-contracts"]:
            with self.subTest(path=path):
                response = self.client.get(path, follow_redirects=False)
                self.assertEqual(response.status_code, 302)
                self.assertTrue(response.headers["Location"].endswith("/"))

        response = self.client.get("/purchase-contracts/product-lookup", query_string={"bld": "K-OUT-001"}, follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/"))

        for path in ["/purchase-contracts/generate", "/sales-contracts/generate"]:
            with self.subTest(path=path):
                response = self.client.post(path, follow_redirects=False)
                self.assertEqual(response.status_code, 302)
                self.assertTrue(response.headers["Location"].endswith("/"))
        self.client.post("/logout")

    def test_quote_records_page_can_create_search_and_edit(self):
        self.login()
        response = self.client.get("/quotes")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("报价记录", html)
        self.assertIn('action="/quotes#quote-results"', html)
        self.assertIn('name="customer_name"', html)
        self.assertIn('name="bld_no"', html)
        self.assertIn('name="customer_product_code"', html)
        self.assertIn('name="tax_price"', html)
        self.assertIn('name="net_price"', html)
        self.assertIn('name="date_from"', html)
        self.assertIn('name="currency"', html)
        self.assertIn("导入报价记录", html)
        self.assertIn('action="/quotes/import/preview"', html)
        self.assertIn('name="customer_name"', html)
        self.assertIn('name="currency"', html)
        self.assertIn('name="quote_file"', html)
        self.assertNotIn("MOQ", html)
        self.assertNotIn("删除", html)

        response = self.client.post(
            "/quotes/save",
            data={
                "customer_name": "博世",
                "bld_no": "K48620",
                "customer_product_code": "48620-0K040",
                "tax_price": "5.35",
                "net_price": "4.73",
                "currency": "USD",
                "quote_date": "2026-06-10",
                "quoted_by": "007",
                "source_type": "manual",
                "source_text": "博世 K48620 USD 5.35",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

        response = self.client.post(
            "/quotes/save",
            data={
                "customer_name": "博世",
                "bld_no": "K48620",
                "customer_product_code": "48620-0K040",
                "tax_price": "5.55",
                "net_price": "4.91",
                "currency": "USD",
                "quote_date": "2026-06-11",
                "quoted_by": "sales",
                "source_type": "wechat",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

        response = self.client.get("/quotes", query_string={"customer_name": "博世", "bld_no": "K48620"})
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("最近一次报价", html)
        self.assertIn("含税 USD 5.5500", html)
        self.assertIn("不含税 USD 4.9100", html)
        self.assertIn("博世 K48620 USD 5.35", html)
        self.assertIn("修正", html)
        self.assertNotIn("data-open-customer-price-delete", html)

        with self.web.connect(self.web.DB_PATH) as conn:
            quote_id = conn.execute(
                "SELECT id FROM quote_records WHERE customer_name = ? AND bld_no = ? ORDER BY quote_date DESC, id DESC",
                ("博世", "K48620"),
            ).fetchone()["id"]

        response = self.client.post(
            f"/quotes/{quote_id}/edit",
            data={
                "customer_name": "博世",
                "bld_no": "K48620",
                "customer_product_code": "BOSCH-K48620",
                "tax_price": "5.65",
                "net_price": "5.00",
                "currency": "USD",
                "quote_date": "2026-06-11",
                "quoted_by": "sales",
                "source_type": "wechat",
                "remark": "人工复核",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        with self.web.connect(self.web.DB_PATH) as conn:
            revised = conn.execute("SELECT * FROM quote_records WHERE id = ?", (quote_id,)).fetchone()
            revisions = conn.execute("SELECT COUNT(*) FROM quote_record_revisions WHERE quote_id = ?", (quote_id,)).fetchone()[0]
        self.assertEqual(revised["bld_no"], "K48620")
        self.assertEqual(revised["customer_product_code"], "BOSCH-K48620")
        self.assertEqual(revised["tax_price"], 5.65)
        self.assertEqual(revised["net_price"], 5.00)
        self.assertEqual(revisions, 1)

        response = self.client.post(
            f"/quotes/{quote_id}/edit",
            data={
                "customer_name": "博世",
                "bld_no": "K48620",
                "tax_price": "5.65",
                "net_price": "",
                "currency": "USD",
                "quote_date": "2026-06-11",
                "quoted_by": "sales",
                "source_type": "wechat",
                "source_text": "",
                "attachment_path": "",
                "remark": "",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        with self.web.connect(self.web.DB_PATH) as conn:
            cleared = conn.execute("SELECT * FROM quote_records WHERE id = ?", (quote_id,)).fetchone()
        self.assertIsNone(cleared["net_price"])
        self.assertEqual(cleared["remark"], "")

        old_path = self.client.get("/customer-prices", follow_redirects=False)
        self.assertEqual(old_path.status_code, 302)
        self.assertTrue(old_path.headers["Location"].endswith("/quotes"))

    def test_quote_api_requires_key_validates_and_keeps_revision_log(self):
        response = self.client.post("/api/quotes", json={"customer_name": "ACME"})
        self.assertEqual(response.status_code, 401)

        token = self.create_internal_api_token()
        headers = {"Authorization": f"Bearer {token}", "X-Quote-Actor": "spoofed-client"}
        response = self.client.post(
            "/api/quotes",
            json={
                "customer_name": "HermesBosch",
                "bld_no": "HERMES-48620",
                "tax_price": "bad",
                "currency": "USD",
            },
            headers=headers,
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("tax_price", response.get_json()["error"])

        response = self.client.post(
            "/api/quotes",
            json={
                "customer_name": "HermesBosch",
                "bld_no": "HERMES-48620",
                "customer_product_code": "CUST-48620",
                "tax_price": "5.35",
                "net_price": "4.73",
                "currency": "USD",
                "quote_date": "2026-06-10",
                "quoted_by": "hermes",
                "source_type": "wechat",
                "source_text": "HermesBosch HERMES-48620 USD 5.35",
            },
            headers=headers,
        )
        self.assertEqual(response.status_code, 201)
        quote = response.get_json()["quote"]
        self.assertEqual(quote["customer_name"], "HermesBosch")
        self.assertEqual(quote["bld_no"], "HERMES-48620")
        self.assertEqual(quote["customer_product_code"], "CUST-48620")
        self.assertEqual(quote["tax_price"], 5.35)
        self.assertEqual(quote["net_price"], 4.73)

        response = self.client.get(
            "/api/quotes/latest",
            query_string={"customer_name": "HermesBosch", "bld_no": "HERMES-48620"},
            headers=headers,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()["quote"]["id"], quote["id"])

        response = self.client.put(
            f"/api/quotes/{quote['id']}",
            json={"tax_price": "5.45", "net_price": "", "currency": "USD", "remark": ""},
            headers=headers,
        )
        self.assertEqual(response.status_code, 200)
        updated_quote = response.get_json()["quote"]
        self.assertEqual(updated_quote["tax_price"], 5.45)
        self.assertIsNone(updated_quote["net_price"])
        self.assertEqual(updated_quote["remark"], "")
        with self.web.connect(self.web.DB_PATH) as conn:
            revisions = conn.execute("SELECT * FROM quote_record_revisions WHERE quote_id = ?", (quote["id"],)).fetchall()
        self.assertEqual(len(revisions), 1)
        self.assertEqual(revisions[0]["changed_by"], "OpenClaw Test")
        self.assertNotEqual(revisions[0]["changed_by"], "spoofed-client")
        self.assertIn('"tax_price": 5.35', revisions[0]["before_json"])
        self.assertIn('"tax_price": 5.45', revisions[0]["after_json"])

        response = self.client.get(
            "/api/quotes",
            query_string={"customer_name": "Hermes", "currency": "USD"},
            headers=headers,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.get_json()["quotes"]), 1)

    def test_product_inquiry_v1_and_artifact_consumer_contract(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "V1-INQUIRY-001",
                    "series": "CONTRACT",
                    "item": "Consumer Contract Arm",
                    "oe_no_1": "V1-OE-001",
                    "models": "Contract Model",
                    "price_cny": "110",
                    "active": "1",
                },
                actor="tester",
            )

        token = self.create_internal_api_token(
            scopes=["products:read", "inquiries:run", "artifacts:read"],
            name="WorkBuddy Inquiry V1",
        )
        authorization = {"Authorization": f"Bearer {token}"}

        product_response = self.client.get(
            "/api/v1/products/search",
            query_string={"oe": "V1-OE-001", "limit": 10},
            headers=authorization,
        )
        product_payload = product_response.get_json()
        self.assertEqual(product_response.status_code, 200)
        self.assertEqual(product_payload["data"]["total"], 1)
        self.assertEqual(product_payload["data"]["products"][0]["bld_no"], "V1-INQUIRY-001")

        missing_idempotency = self.client.post(
            "/api/v1/inquiries/analyze",
            json={"numbers": ["V1-OE-001"]},
            headers=authorization,
        )
        self.assertEqual(missing_idempotency.status_code, 400)
        self.assertEqual(missing_idempotency.get_json()["error"]["code"], "idempotency.required")

        analyze_response = self.client.post(
            "/api/v1/inquiries/analyze",
            json={"numbers": ["V1-OE-001", "V1-MISSING"], "price_mode": "net"},
            headers={**authorization, "Idempotency-Key": "inquiry-analyze-v1-001"},
        )
        analyze_payload = analyze_response.get_json()
        self.assertEqual(analyze_response.status_code, 200)
        self.assertEqual(analyze_payload["data"]["summary"]["matched_count"], 1)
        self.assertEqual(analyze_payload["data"]["rows"][0]["bld_no"], "V1-INQUIRY-001")
        self.assertEqual(analyze_payload["data"]["rows"][0]["export_price"], 100)
        self.assertIsNone(analyze_payload["data"]["artifact"])

        legacy_response = self.client.post(
            "/api/internal/inquiry/analyze",
            json={"numbers": ["V1-OE-001"]},
            headers=authorization,
        )
        self.assertEqual(legacy_response.status_code, 200)
        self.assertEqual(legacy_response.get_json()["rows"][0]["bld_no"], "V1-INQUIRY-001")

        export_body = {
            "numbers": ["V1-OE-001"],
            "source_name": "consumer-contract",
            "price_mode": "tax",
        }
        export_headers = {**authorization, "Idempotency-Key": "inquiry-export-v1-001"}
        export_response = self.client.post(
            "/api/v1/inquiries/export",
            json=export_body,
            headers=export_headers,
        )
        export_payload = export_response.get_json()
        self.assertEqual(export_response.status_code, 201)
        artifact = export_payload["data"]["artifact"]
        self.assertTrue(artifact["id"].startswith("art_"))
        self.assertRegex(artifact["filename"], r"^re\d{6}_consumer-contract\.xlsx$")
        self.assertNotIn("output_path", json.dumps(export_payload, ensure_ascii=False))
        self.assertNotIn(str(self.root), json.dumps(export_payload, ensure_ascii=False))

        replay = self.client.post(
            "/api/v1/inquiries/export",
            json=export_body,
            headers=export_headers,
        )
        self.assertEqual(replay.status_code, 201)
        self.assertEqual(replay.headers["Idempotency-Replayed"], "true")
        self.assertEqual(replay.get_json()["data"]["artifact"]["id"], artifact["id"])

        download = self.client.get(artifact["download_url"], headers=authorization)
        self.assertEqual(download.status_code, 200)
        self.assertTrue(download.data.startswith(b"PK"))
        self.assertIn("attachment", download.headers["Content-Disposition"])
        self.assertEqual(download.headers["Cache-Control"], "private, no-store")
        download.close()

        other_token = self.create_internal_api_token(
            scopes=["artifacts:read"],
            name="Other Artifact Consumer",
        )
        denied = self.client.get(
            artifact["download_url"],
            headers={"Authorization": f"Bearer {other_token}"},
        )
        self.assertEqual(denied.status_code, 404)
        self.assertEqual(denied.get_json()["error"]["code"], "artifact.not_found")

        with self.web.connect(self.web.DB_PATH) as conn:
            stored = conn.execute(
                "SELECT owner_id, storage_path, sha256 FROM api_artifacts WHERE id = ?",
                (artifact["id"],),
            ).fetchone()
            audit = conn.execute(
                "SELECT actor FROM audit_logs WHERE action = ? ORDER BY id DESC LIMIT 1",
                ("内部 API 生成号码结果",),
            ).fetchone()
        self.assertIsNotNone(stored)
        self.assertTrue(
            Path(stored["storage_path"]).resolve().is_relative_to((self.root / "outputs").resolve())
        )
        self.assertEqual(len(stored["sha256"]), 64)
        self.assertEqual(audit["actor"], "WorkBuddy Inquiry V1")

    def test_quote_v1_contract_idempotency_and_optimistic_concurrency(self):
        token = self.create_internal_api_token(
            scopes=["api:read", "quotes:read", "quotes:write"],
            name="WorkBuddy Quote V1",
        )
        authorization = {"Authorization": f"Bearer {token}"}
        create_headers = {**authorization, "Idempotency-Key": "quote-create-v1-001"}
        invalid = self.client.post(
            "/api/v1/quotes",
            json={
                "customer_name": "V1 Contract Customer",
                "bld_no": "V1-BLD-001",
                "tax_price": "12.34",
                "currency": "USD",
                "attachment_path": "/tmp/private-quote.pdf",
            },
            headers=create_headers,
        )
        self.assertEqual(invalid.status_code, 422)
        self.assertEqual(invalid.get_json()["error"]["code"], "request.invalid")
        self.assertNotIn("private-quote.pdf", invalid.get_data(as_text=True))

        create_payload = {
            "customer_name": "V1 Contract Customer",
            "bld_no": "V1-BLD-001",
            "customer_product_code": "V1-CUSTOMER-001",
            "tax_price": "12.34",
            "net_price": "11.22",
            "currency": "USD",
            "quote_date": "2026-07-11",
            "source_type": "wechat",
            "source_text": "V1 quote source",
            "on_behalf_of": "sales-operator",
        }
        created = self.client.post(
            "/api/v1/quotes",
            json=create_payload,
            headers=create_headers,
        )
        self.assertEqual(created.status_code, 201)
        created_body = created.get_json()
        quote = created_body["data"]["quote"]
        self.assertEqual(quote["version"], 1)
        self.assertEqual(created.headers["ETag"], '"1"')
        self.assertNotIn("attachment_path", quote)
        quote_id = quote["id"]

        replayed = self.client.post(
            "/api/v1/quotes",
            json=create_payload,
            headers=create_headers,
        )
        self.assertEqual(replayed.status_code, 201)
        self.assertEqual(replayed.headers["Idempotency-Replayed"], "true")
        self.assertEqual(replayed.headers["ETag"], '"1"')
        self.assertEqual(replayed.get_json(), created_body)

        fetched = self.client.get(f"/api/v1/quotes/{quote_id}", headers=authorization)
        self.assertEqual(fetched.status_code, 200)
        self.assertEqual(fetched.headers["ETag"], '"1"')
        self.assertEqual(fetched.get_json()["data"]["quote"]["id"], quote_id)

        listed = self.client.get(
            "/api/v1/quotes?customer_name=V1%20Contract%20Customer&limit=10",
            headers=authorization,
        )
        listed_data = listed.get_json()["data"]
        self.assertEqual(listed.status_code, 200)
        self.assertEqual(listed_data["total"], 1)
        self.assertEqual([item["id"] for item in listed_data["quotes"]], [quote_id])

        missing_precondition = self.client.patch(
            f"/api/v1/quotes/{quote_id}",
            json={"remark": "first revision"},
            headers={**authorization, "Idempotency-Key": "quote-update-v1-missing"},
        )
        self.assertEqual(missing_precondition.status_code, 428)
        self.assertEqual(missing_precondition.get_json()["error"]["code"], "precondition.required")

        updated = self.client.patch(
            f"/api/v1/quotes/{quote_id}",
            json={"remark": "first revision", "on_behalf_of": "sales-operator"},
            headers={
                **authorization,
                "Idempotency-Key": "quote-update-v1-001",
                "If-Match": '"1"',
            },
        )
        updated_quote = updated.get_json()["data"]["quote"]
        self.assertEqual(updated.status_code, 200)
        self.assertEqual(updated_quote["version"], 2)
        self.assertEqual(updated_quote["remark"], "first revision")
        self.assertEqual(updated.headers["ETag"], '"2"')

        stale = self.client.patch(
            f"/api/v1/quotes/{quote_id}",
            json={"remark": "stale revision"},
            headers={
                **authorization,
                "Idempotency-Key": "quote-update-v1-stale",
                "If-Match": '"1"',
            },
        )
        stale_error = stale.get_json()["error"]
        self.assertEqual(stale.status_code, 412)
        self.assertEqual(stale_error["code"], "quote.version_conflict")
        self.assertEqual(stale_error["details"]["current_version"], 2)

        with self.web.connect(self.web.DB_PATH) as conn:
            quote_count = conn.execute(
                "SELECT COUNT(*) FROM quote_records WHERE customer_name = ?",
                ("V1 Contract Customer",),
            ).fetchone()[0]
            revisions = conn.execute(
                "SELECT changed_by, before_json, after_json FROM quote_record_revisions WHERE quote_id = ?",
                (quote_id,),
            ).fetchall()
            audit = conn.execute(
                "SELECT actor, detail FROM audit_logs WHERE action = 'API mutation' AND target_key = ? ORDER BY id",
                ("quote_v1_api.update_quote_v1",),
            ).fetchall()
        self.assertEqual(quote_count, 1)
        self.assertEqual(len(revisions), 1)
        self.assertEqual(revisions[0]["changed_by"], "WorkBuddy Quote V1")
        self.assertIn('"version": 1', revisions[0]["before_json"])
        self.assertIn('"version": 2', revisions[0]["after_json"])
        self.assertTrue(any(row["actor"] == "WorkBuddy Quote V1" for row in audit))

        document = self.client.get("/api/v1/openapi.json", headers=authorization).get_json()
        self.assertIn("/api/v1/quotes", document["paths"])
        self.assertIn("/api/v1/quotes/{quote_id}", document["paths"])
        create_operation = document["paths"]["/api/v1/quotes"]["post"]
        patch_operation = document["paths"]["/api/v1/quotes/{quote_id}"]["patch"]
        self.assertEqual(create_operation["x-required-scopes"], ["quotes:write"])
        self.assertIn("requestBody", create_operation)
        self.assertTrue(
            any(parameter["name"] == "Idempotency-Key" for parameter in create_operation["parameters"])
        )
        self.assertTrue(
            any(parameter["name"] == "If-Match" for parameter in patch_operation["parameters"])
        )
        self.assertIn("ETag", patch_operation["responses"]["200"]["headers"])

    def test_quote_records_can_import_excel_into_quote_table(self):
        from openpyxl import Workbook

        self.login()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["BLD号", "客户产品编码", "含税单价", "不含税单价", "报价日期", "报价人", "来源", "原文", "备注"])
        sheet.append(["K-IMPORT-QUOTE", "CUST-IMPORT", 12.34, 10.92, "2026-07-01", "importer", "excel", "导入客户 K-IMPORT-QUOTE USD 12.34", "批量导入"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        buffer.seek(0)

        preview = self.client.post(
            "/quotes/import/preview",
            data={"customer_name": "导入客户", "currency": "USD", "quote_file": (buffer, "quotes.xlsx")},
            content_type="multipart/form-data",
        )
        html = preview.get_data(as_text=True)
        self.assertEqual(preview.status_code, 200)
        self.assertIn("导入报价记录", html)
        self.assertIn("K-IMPORT-QUOTE", html)
        payload_match = re.search(r'name="payload" value="([^"]+)"', html)
        self.assertIsNotNone(payload_match)

        payload = __import__("html").unescape(payload_match.group(1))
        apply = self.client.post(
            "/quotes/import/apply",
            data={"payload": payload},
            follow_redirects=False,
        )
        self.assertEqual(apply.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            row = conn.execute("SELECT * FROM quote_records WHERE bld_no = ?", ("K-IMPORT-QUOTE",)).fetchone()
        self.assertIsNotNone(row)
        self.assertEqual(row["customer_name"], "导入客户")
        self.assertEqual(row["customer_product_code"], "CUST-IMPORT")
        self.assertEqual(row["currency"], "USD")
        self.assertEqual(row["tax_price"], 12.34)
        self.assertEqual(row["net_price"], 10.92)

    def test_quote_api_oversized_request_returns_json(self):
        response = self.client.post(
            "/api/quotes",
            data=b"x" * (21 * 1024 * 1024),
            content_type="application/json",
            follow_redirects=False,
        )
        payload = response.get_json()
        self.assertEqual(response.status_code, 413)
        self.assertEqual(payload["ok"], False)
        self.assertIn("上传文件不能超过", payload["error"])

    def test_quotes_are_admin_only(self):
        from app.database import save_user

        with self.web.connect(self.web.DB_PATH) as conn:
            save_user(
                conn,
                {
                    "username": "editor-prices",
                    "display_name": "Editor Prices",
                    "password": "editor-pw",
                    "role": "editor",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        admin_page = self.client.get("/").get_data(as_text=True)
        self.assertIn("报价记录", admin_page)
        self.client.post("/logout")

        login = self.client.post(
            "/login",
            data={"username": "editor-prices", "password": "editor-pw", "next": "/"},
            follow_redirects=False,
        )
        self.assertEqual(login.status_code, 302)

        editor_page = self.client.get("/").get_data(as_text=True)
        self.assertNotIn("报价记录", editor_page)
        response = self.client.get("/quotes", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/"))
        self.client.post("/logout")

    def test_products_search_uses_results_anchor(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-FILTER-HYUNDAI",
                    "series": "HYUNDAI",
                    "item": "CONTROL ARM",
                    "oe_no_1": "FILTER-001",
                    "models": "Sportage",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-FILTER-HONDA",
                    "series": "HONDA",
                    "item": "CONTROL ARM",
                    "oe_no_1": "FILTER-002",
                    "models": "Civic",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-FILTER-DOT-OE",
                    "series": "PEUGEOT",
                    "item": "CONTROL ARM",
                    "oe_no_1": "3521.R1",
                    "models": "C-CROSSER",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/products")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('id="products-results"', html)
        self.assertIn('action="/products#products-results"', html)
        self.assertIn("按 BLD / 品牌 / 车型搜索", html)
        self.assertIn('<button class="linear-button primary" type="submit">搜索</button>', html)
        self.assertIn('class="embedded-submit" type="submit">上传预览', html)
        self.assertIn('class="embedded-submit" type="submit">确认导入', html)
        self.assertIn('id="product-modal"', html)
        self.assertIn('id="product-edit-modal"', html)
        self.assertIn("data-draggable-modal-panel", html)

        for query in ["HYUNDAI", "Sportage"]:
            with self.subTest(query=query):
                response = self.client.get("/products", query_string={"bld": query})
                html = response.get_data(as_text=True)
                self.assertIn("K-FILTER-HYUNDAI", html)
                self.assertNotIn("K-FILTER-HONDA", html)

        for query in ["3521.r1", "3521R1", "3521-R1"]:
            with self.subTest(query=query):
                response = self.client.get("/products", query_string={"oe": query})
                html = response.get_data(as_text=True)
                self.assertIn("K-FILTER-DOT-OE", html)

    def test_products_oe_search_psa_352x_dot_does_not_show_gm_exact(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-PRODUCT-PSA-352125",
                    "series": "PEUGEOT\nCITROEN",
                    "item": "CONTROL ARM",
                    "oe_no_1": "3521.25",
                    "models": "C5",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-PRODUCT-GM-352125",
                    "series": "GM\nOPEL",
                    "item": "CONTROL ARM",
                    "oe_no_1": "352125",
                    "models": "OPEL",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        dotted = self.client.get("/products", query_string={"oe": "3521.25"})
        dotted_html = dotted.get_data(as_text=True)
        self.assertEqual(dotted.status_code, 200)
        self.assertIn("K-PRODUCT-PSA-352125", dotted_html)
        self.assertNotIn("K-PRODUCT-GM-352125", dotted_html)

        undotted = self.client.get("/products", query_string={"oe": "352125"})
        undotted_html = undotted.get_data(as_text=True)
        self.assertEqual(undotted.status_code, 200)
        self.assertIn("K-PRODUCT-PSA-352125", undotted_html)
        self.assertIn("K-PRODUCT-GM-352125", undotted_html)

    def test_products_use_bld_natural_order(self):
        from app.bld_sort import bld_sort_key
        from app.database import upsert_product

        self.assertEqual(
            sorted(["K8274LA", "K8274RA", "K8274LB", "K8274RB"], key=bld_sort_key),
            ["K8274LA", "K8274RA", "K8274LB", "K8274RB"],
        )
        self.assertEqual(
            sorted(["K8058LA-1", "K8058RA-1", "K8058LB", "K8058RB"], key=bld_sort_key),
            ["K8058LA-1", "K8058RA-1", "K8058LB", "K8058RB"],
        )

        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no in ["K8274LA", "K8274RA", "K8274LB", "K8274RB", "K8058LA-1", "K8058RA-1", "K8058LB", "K8058RB"]:
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "SORT",
                        "item": "SORT TEST",
                        "oe_no_1": f"OE-{bld_no}",
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.get("/products", query_string={"bld": "K8274"})
        html = response.get_data(as_text=True)
        self.assertLess(html.index("K8274LA"), html.index("K8274RA"))
        self.assertLess(html.index("K8274RA"), html.index("K8274LB"))
        self.assertLess(html.index("K8274LB"), html.index("K8274RB"))

        response = self.client.get("/products", query_string={"bld": "K8058"})
        html = response.get_data(as_text=True)
        self.assertLess(html.index("K8058LA-1"), html.index("K8058RA-1"))
        self.assertLess(html.index("K8058RA-1"), html.index("K8058LB"))
        self.assertLess(html.index("K8058LB"), html.index("K8058RB"))

    def test_products_are_paginated(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            for index in range(121):
                upsert_product(
                    conn,
                    {
                        "bld_no": f"K-PAGE-{index:03d}",
                        "series": "PAGED",
                        "item": "PAGED PART",
                        "oe_no_1": f"PAGE-{index:03d}",
                        "models": "Batch Tester",
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.get("/products", query_string={"bld": "K-PAGE"})
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("第 1-50 条 / 共 121 条", html)
        self.assertIn("第 1 / 3 页，每页 50 条", html)
        self.assertEqual(html.count('aria-label="产品分页"'), 1)
        self.assertIn("K-PAGE-000", html)
        self.assertIn("K-PAGE-049", html)
        self.assertNotIn("K-PAGE-050", html)
        self.assertNotIn("/products/rows", html)

        third_page = self.client.get("/products", query_string={"bld": "K-PAGE", "page": "3"})
        third_html = third_page.get_data(as_text=True)
        self.assertEqual(third_page.status_code, 200)
        self.assertNotIn("第 101-121 条 / 共 121 条", third_html)
        self.assertIn("第 3 / 3 页，每页 50 条", third_html)
        self.assertIn("K-PAGE-100", third_html)
        self.assertIn("K-PAGE-120", third_html)
        self.assertNotIn("K-PAGE-099", third_html)

    def test_product_drawing_upload_preview_and_batch_entry(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-DRAW-001",
                    "series": "TEST",
                    "item": "DRAWING PART",
                    "oe_no_1": "DRAW-001",
                    "models": "Tester",
                    "active": "1",
                },
                actor="tester",
            )
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-DRAW-001",)).fetchone()

        self.login()
        response = self.client.get("/products", query_string={"bld": "K-DRAW-001"})
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("data-open-edit-product-modal", html)
        self.assertNotIn("PDF图纸", html)
        self.assertIn("批量上传图纸", html)
        self.assertNotIn('name="drawing"', html)
        self.assertNotIn(f'href="/products/{product["id"]}/drawing"', html)

        edit = self.client.get(f"/products/{product['id']}/edit")
        edit_html = edit.get_data(as_text=True)
        self.assertEqual(edit.status_code, 200)
        for slot in range(1, 6):
            self.assertIn(f'name="product_image_{slot}"', edit_html)
        self.assertIn("file-picker-clear", edit_html)
        self.assertIn("/static/app.js", edit_html)
        self.assertIn('name="drawing"', edit_html)

        embedded = self.client.get(f"/products/{product['id']}/edit", query_string={"embedded": "1"})
        embedded_html = embedded.get_data(as_text=True)
        self.assertEqual(embedded.status_code, 200)
        self.assertIn("embedded-product-form-page", embedded_html)
        self.assertIn('name="embedded" value="1"', embedded_html)
        self.assertNotIn("返回目录", embedded_html)

        embedded_save = self.client.post(
            "/products/save",
            data={
                "embedded": "1",
                "bld_no": "K-DRAW-001",
                "series": "TEST",
                "item": "DRAWING PART",
                "oe_no_1": "DRAW-001",
                "oe_no_2": "",
                "models": "Tester",
                "price_cny": "",
                "active": "1",
            },
            follow_redirects=False,
        )
        self.assertEqual(embedded_save.status_code, 200)
        self.assertIn("window.parent.location.reload()", embedded_save.get_data(as_text=True))

        upload = self.client.post(
            "/products/save",
            data={
                "bld_no": "K-DRAW-001",
                "series": "TEST",
                "item": "DRAWING PART",
                "oe_no_1": "DRAW-001",
                "oe_no_2": "",
                "models": "Tester",
                "price_cny": "",
                "active": "1",
                "product_image_1": (io.BytesIO(b"\x89PNG\r\n\x1a\nproduct image 1"), "K-DRAW-001.png"),
                "product_image_2": (io.BytesIO(b"\x89PNG\r\n\x1a\nproduct image 2"), "K-DRAW-001-2.png"),
                "drawing": (io.BytesIO(b"%PDF-1.4\nfirst drawing\n%%EOF"), "K-DRAW-001.pdf"),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        self.assertEqual(upload.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            updated = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-DRAW-001",)).fetchone()
        drawing_path = self.root / "data" / updated["drawing_path"]
        image_path = self.root / "data" / "product_images" / "K-DRAW-001.png"
        image_path_2 = self.root / "data" / "product_images" / "K-DRAW-001-2.png"
        self.assertTrue(drawing_path.exists())
        self.assertTrue(image_path.exists())
        self.assertTrue(image_path_2.exists())
        self.assertEqual(updated["drawing_original_name"], "K-DRAW-001.pdf")
        self.assertEqual(updated["image_path"], "data_product_images/K-DRAW-001.png")
        self.assertEqual(updated["image_path_2"], "data_product_images/K-DRAW-001-2.png")

        response = self.client.get("/products", query_string={"bld": "K-DRAW-001"})
        html = response.get_data(as_text=True)
        self.assertIn(f'href="/products/{product["id"]}/drawing"', html)
        self.assertNotIn("替换图纸", html)
        self.assertIn("/product-image-thumbs/K-DRAW-001.png", html)
        self.assertIn("/product-images/K-DRAW-001.png", html)
        self.assertIn("/product-images/K-DRAW-001-2.png", html)

        image = self.client.get("/product-images/K-DRAW-001.png")
        self.assertEqual(image.status_code, 200)
        self.assertTrue(image.get_data().startswith(b"\x89PNG"))
        image.close()

        preview = self.client.get(f"/products/{product['id']}/drawing")
        self.assertEqual(preview.status_code, 200)
        self.assertTrue(preview.get_data().startswith(b"%PDF-1.4"))
        preview.close()

        replace = self.client.post(
            "/products/save",
            data={
                "bld_no": "K-DRAW-001",
                "series": "TEST",
                "item": "DRAWING PART",
                "oe_no_1": "DRAW-001",
                "oe_no_2": "",
                "models": "Tester",
                "price_cny": "",
                "active": "1",
                "drawing": (io.BytesIO(b"%PDF-1.4\nsecond drawing\n%%EOF"), "K-DRAW-001-v2.pdf"),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        self.assertEqual(replace.status_code, 302)
        archive_dir = self.root / "data" / "drawings" / "archive" / "K-DRAW-001"
        self.assertTrue(list(archive_dir.glob("*.pdf")))

        batch = self.client.get("/products/drawings/batch")
        self.assertEqual(batch.status_code, 200)
        self.assertIn("暂未开放", batch.get_data(as_text=True))

    def test_product_save_can_clear_price_and_reject_invalid_price(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-PRICE-CLEAR",
                    "series": "OLD",
                    "item": "PRICE TEST",
                    "oe_no_1": "PRICE-CLEAR-OE",
                    "models": "Tester",
                    "price_cny": "88.5",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        clear = self.client.post(
            "/products/save",
            data={
                "bld_no": "K-PRICE-CLEAR",
                "series": "CLEARED",
                "item": "PRICE TEST",
                "oe_no_1": "PRICE-CLEAR-OE",
                "oe_no_2": "",
                "models": "Tester",
                "price_cny": "",
                "active": "1",
            },
            follow_redirects=False,
        )
        self.assertEqual(clear.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-PRICE-CLEAR",)).fetchone()
        self.assertEqual(product["series"], "CLEARED")
        self.assertIsNone(product["price_cny"])

        invalid = self.client.post(
            "/products/save",
            data={
                "bld_no": "K-PRICE-CLEAR",
                "series": "BAD",
                "item": "PRICE TEST",
                "oe_no_1": "PRICE-CLEAR-OE",
                "oe_no_2": "",
                "models": "Tester",
                "price_cny": "abc",
                "active": "1",
            },
            follow_redirects=False,
        )
        self.assertEqual(invalid.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-PRICE-CLEAR",)).fetchone()
        self.assertEqual(product["series"], "CLEARED")
        self.assertIsNone(product["price_cny"])

    def test_product_status_can_be_edited_and_shown_in_catalog(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-STATUS-001",
                    "series": "TEST",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "STATUS-OE-001",
                    "models": "Tester",
                    "price_cny": "45",
                    "product_status": "1 个球头 2 个衬套",
                    "active": "1",
                },
                actor="tester",
            )
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-STATUS-001",)).fetchone()

        self.login()
        page = self.client.get("/products", query_string={"bld": "K-STATUS-001"})
        html = page.get_data(as_text=True)
        self.assertEqual(page.status_code, 200)
        self.assertIn("<th>产品状态</th>", html)
        self.assertIn("1 ball joint 2 bushings", html)

        edit = self.client.get(f"/products/{product['id']}/edit")
        edit_html = edit.get_data(as_text=True)
        self.assertEqual(edit.status_code, 200)
        self.assertIn('name="product_status"', edit_html)
        self.assertIn("1 个球头 2 个衬套", edit_html)

        save = self.client.post(
            "/products/save",
            data={
                "bld_no": "K-STATUS-001",
                "series": "TEST",
                "item": "Front Left Lower Control Arm",
                "oe_no_1": "STATUS-OE-001",
                "oe_no_2": "",
                "models": "Tester",
                "price_cny": "45",
                "product_status": "0 个球头 1 个衬套",
                "active": "1",
            },
            follow_redirects=False,
        )
        self.assertEqual(save.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            updated = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-STATUS-001",)).fetchone()
        self.assertEqual(updated["product_status"], "0 个球头 1 个衬套")

    def test_product_save_rejects_invalid_image_before_updating_fields(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-IMAGE-FAIL",
                    "series": "OLD",
                    "item": "IMAGE TEST",
                    "oe_no_1": "IMAGE-FAIL-OE",
                    "models": "Tester",
                    "price_cny": "55",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.post(
            "/products/save",
            data={
                "bld_no": "K-IMAGE-FAIL",
                "series": "NEW",
                "item": "IMAGE TEST UPDATED",
                "oe_no_1": "IMAGE-FAIL-OE",
                "oe_no_2": "",
                "models": "Tester",
                "price_cny": "66",
                "active": "1",
                "product_image_1": (io.BytesIO(b"not really a png"), "K-IMAGE-FAIL.png"),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-IMAGE-FAIL",)).fetchone()
        self.assertEqual(product["series"], "OLD")
        self.assertEqual(product["item"], "IMAGE TEST")
        self.assertEqual(product["price_cny"], 55)
        self.assertEqual(product["image_path"], "")

    def test_product_image_table_uses_generated_thumbnail(self):
        from PIL import Image

        from app.database import upsert_product

        image_dir = self.root / "data" / "product_images"
        image_dir.mkdir(parents=True, exist_ok=True)
        image_path = image_dir / "K-THUMB-001.png"
        Image.new("RGB", (640, 360), "white").save(image_path)

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-THUMB-001",
                    "series": "TEST",
                    "item": "THUMB PART",
                    "oe_no_1": "THUMB-001",
                    "models": "Tester",
                    "image_path": "data_product_images/K-THUMB-001.png",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/products", query_string={"bld": "K-THUMB-001"})
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("/product-image-thumbs/K-THUMB-001.png", html)
        self.assertIn("/product-images/K-THUMB-001.png", html)

        thumb = self.client.get("/product-image-thumbs/K-THUMB-001.png")
        self.assertEqual(thumb.status_code, 200)
        with Image.open(io.BytesIO(thumb.get_data())) as generated:
            self.assertLessEqual(generated.width, 160)
            self.assertLessEqual(generated.height, 120)
        thumb.close()
        self.assertTrue((image_dir / "thumbs" / "K-THUMB-001.png").exists())

    def test_product_edit_can_delete_product(self):
        from app.database import save_alias, upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-DELETE-001",
                    "series": "TEST",
                    "item": "DELETE TARGET",
                    "oe_no_1": "DELETE-001",
                    "models": "Tester",
                    "active": "1",
                },
                actor="tester",
            )
            save_alias(conn, "DELETE-ALIAS-001", "K-DELETE-001", actor="tester")
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-DELETE-001",)).fetchone()

        self.login()
        edit = self.client.get(f"/products/{product['id']}/edit")
        edit_html = edit.get_data(as_text=True)
        self.assertEqual(edit.status_code, 200)
        self.assertIn("删除产品", edit_html)
        self.assertIn(f'formaction="/products/{product["id"]}/delete"', edit_html)
        self.assertIn('data-confirm="确认删除 K-DELETE-001', edit_html)

        delete = self.client.post(f"/products/{product['id']}/delete", follow_redirects=False)
        self.assertEqual(delete.status_code, 302)
        self.assertTrue(delete.headers["Location"].endswith("/products"))

        with self.web.connect(self.web.DB_PATH) as conn:
            deleted = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-DELETE-001",)).fetchone()
            alias = conn.execute("SELECT * FROM aliases WHERE source_code = ?", ("DELETEALIAS001",)).fetchone()
            log = conn.execute(
                "SELECT * FROM audit_logs WHERE action = ? AND target_key = ? ORDER BY id DESC LIMIT 1",
                ("删除产品", "K-DELETE-001"),
            ).fetchone()

        self.assertIsNone(deleted)
        self.assertEqual(alias["active"], 0)
        self.assertIsNotNone(log)
        self.assertIn("DELETE TARGET", log["detail"])

    def test_system_updates_page_reads_handoff_notes(self):
        self.login()
        response = self.client.get("/system-updates")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("系统更新", html)
        self.assertIn('data-page="admin.system_updates"', html)
        self.assertIn('data-page-type="system-admin"', html)
        self.assertIn("当前最近重要变更", html)
        self.assertIn("项目交接说明.md", html)
        self.assertIn("建立长期项目治理基线并加固安全边界", html)
        self.assertIn("2026-07-10", html)
        self.assertIn("补齐系统更新记录并设为强制提交要求", html)
        self.assertIn("d1ab621", html)
        self.assertIn("改进 OpenClaw 询价命令行启动和调用体验", html)
        self.assertIn("ac3aa1a", html)
        self.assertIn("新增系统更新页面", html)

    def test_new_material_item_uses_modal(self):
        self.login()
        response = self.client.get("/materials")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('href="/materials/items"', html)
        self.assertNotIn('id="materials-results"', html)
        self.assertIn('class="embedded-submit" type="submit">生成并下载', html)

        response = self.client.get("/materials/items")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("data-open-material-modal", html)
        self.assertIn('id="material-modal"', html)
        self.assertIn('action="/materials/items/save"', html)
        self.assertIn('id="materials-results"', html)
        self.assertIn('action="/materials/items#materials-results"', html)
        self.assertIn("data-enter-navigation", html)
        self.assertIn('name="spec_text"', html)
        self.assertIn('<button class="linear-button" type="submit">搜索</button>', html)
        self.assertIn('class="embedded-submit" type="submit">确认导入', html)
        self.assertIn("母件编码", html)
        self.assertIn("零件编码", html)
        self.assertRegex(html, r'<input name="code"[^>]*required')
        self.assertRegex(html, r'<input name="part"[^>]*required')
        self.assertIn("<th>母件编码</th>", html)
        self.assertIn("<th>零件编码</th>", html)
        self.assertIn("<th>单件重量kg</th>", html)
        self.assertNotIn("<th>型号</th>", html)
        self.assertNotIn("<th>编码</th>", html)
        self.assertNotIn('name="thickness"', html)
        self.assertNotIn('name="width"', html)
        self.assertNotIn('name="length"', html)
        self.assertNotIn('href="/materials/items/new"', html)
        self.assertLess(html.index('name="part"'), html.index('name="pieces"'))
        self.assertLess(html.index('name="spec_text"'), html.index('name="category"'))
        self.assertLess(html.index('name="category"'), html.index('name="car"'))

    def test_admin_materials_page_shows_all_recent_material_files_with_operator(self):
        output_root = self.root / "outputs"
        current_user_dir = output_root / "u1-007"
        other_user_dir = output_root / "u88-other-material"
        current_user_dir.mkdir(parents=True, exist_ok=True)
        other_user_dir.mkdir(parents=True, exist_ok=True)
        current_file = current_user_dir / "current-user-260701料单.xlsx"
        other_file = other_user_dir / "other-user-260701料单.xlsx"
        unrelated_file = other_user_dir / "other-user-quote.xlsx"
        current_file.write_bytes(b"current")
        other_file.write_bytes(b"other")
        unrelated_file.write_bytes(b"quote")

        self.login()
        response = self.client.get("/materials")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("最近料单", html)
        self.assertIn("操作用户", html)
        self.assertIn("current-user-260701料单.xlsx", html)
        self.assertIn("other-user-260701料单.xlsx", html)
        self.assertIn("other-material", html)
        self.assertNotIn("other-user-quote.xlsx", html)

        response = self.client.get("/materials", query_string={"material_history_q": "other-material"})
        html = response.get_data(as_text=True)
        self.assertIn('class="materials-history-drawer" open', html)
        self.assertIn("other-user-260701料单.xlsx", html)
        self.assertNotIn("current-user-260701料单.xlsx", html)

    def test_material_item_save_calculates_spec_text(self):
        self.login()
        examples = [
            ("T-SPEC-WEB-SPACE", "2.5 357 1260", "2.5×357×1260"),
            ("T-SPEC-WEB-STAR", "2.5*357*1260", "2.5×357×1260"),
            ("T-SPEC-WEB-DASH", "2.5-357-1260", "2.5×357×1260"),
            ("T-SPEC-WEB-SLASH", "2.5/357/1260", "2.5×357×1260"),
        ]
        for model, spec_text, expected in examples:
            with self.subTest(spec_text=spec_text):
                response = self.client.post(
                    "/materials/items/save",
                    data={
                        "model": model,
                        "code": "KA-TEST",
                        "category": "测试类别",
                        "car": "测试车型",
                        "part": "测试零件",
                        "pieces": "2",
                        "spec_text": spec_text,
                        "active": "1",
                    },
                    follow_redirects=False,
                )
                self.assertEqual(response.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            rows = conn.execute(
                "SELECT model, spec_text, thickness, width, length FROM material_items WHERE model LIKE 'T-SPEC-WEB-%'"
            ).fetchall()
        saved = {row["model"]: row for row in rows}
        for model, _, expected in examples:
            self.assertIn(model, saved)
            self.assertEqual(saved[model]["spec_text"], expected)
            self.assertEqual(saved[model]["thickness"], 2.5)
            self.assertEqual(saved[model]["width"], 357)
            self.assertEqual(saved[model]["length"], 1260)

        response = self.client.get("/materials/items?q=T-SPEC-WEB-SPACE")
        html = response.get_data(as_text=True)
        self.assertIn("单件重量kg", html)
        self.assertIn("4.41", html)
        for query in ["357", "2.5 357", "357/1260", "2.5-1260", "2.5*357*1260"]:
            with self.subTest(query=query):
                response = self.client.get("/materials/items", query_string={"q": query})
                html = response.get_data(as_text=True)
                self.assertIn("T-SPEC-WEB-SPACE", html)
        response = self.client.get("/materials/items", query_string={"q": "2.5 999"})
        html = response.get_data(as_text=True)
        self.assertNotIn("T-SPEC-WEB-SPACE", html)

    def test_material_item_requires_code_and_part(self):
        self.login()
        for field, data in [
            (
                "code",
                {
                    "model": "T-SPEC-REQUIRED-CODE",
                    "part": "测试零件",
                    "pieces": "2",
                    "spec_text": "2.5 357 1260",
                    "active": "1",
                },
            ),
            (
                "part",
                {
                    "model": "T-SPEC-REQUIRED-PART",
                    "code": "KA-TEST",
                    "pieces": "2",
                    "spec_text": "2.5 357 1260",
                    "active": "1",
                },
            ),
        ]:
            with self.subTest(field=field):
                response = self.client.post("/materials/items/save", data=data, follow_redirects=False)
                self.assertEqual(response.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            count = conn.execute(
                "SELECT COUNT(*) FROM material_items WHERE model IN (?, ?)",
                ("T-SPEC-REQUIRED-CODE", "T-SPEC-REQUIRED-PART"),
            ).fetchone()[0]
        self.assertEqual(count, 0)

    def test_material_import_calculates_spec_text_from_dimensions(self):
        from app.database import import_materials_from_excel
        from openpyxl import Workbook

        path = self.root / "stale-material-spec.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "材料数据"
        sheet.append(["型号", "型号", "类别", "车型", "零件名称", "规格尺寸", "下料只数", "单重", "规格1", "规格2", "规格3"])
        sheet.append(["T-SPEC-IMPORT", "KA-IMPORT", "测试类别", "测试车型", "测试零件", "旧规格", 3, "", 4, 92.5, 1260])
        workbook.save(path)
        workbook.close()

        with self.web.connect(self.web.DB_PATH) as conn:
            imported = import_materials_from_excel(conn, path, replace=False, actor="tester")
            row = conn.execute("SELECT spec_text FROM material_items WHERE model = ?", ("T-SPEC-IMPORT",)).fetchone()

        self.assertEqual(imported, 1)
        self.assertIsNotNone(row)
        self.assertEqual(row["spec_text"], "4.0×92.5×1260")

    def test_material_source_sync_rewrites_spec_text_column(self):
        from app.material_sheet import sync_material_specs_from_dimensions
        from openpyxl import Workbook, load_workbook

        path = self.root / "sync-material-spec.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "材料数据"
        sheet.append(["型号", "型号", "类别", "车型", "零件名称", "规格尺寸", "下料只数", "单重", "规格1", "规格2", "规格3"])
        sheet.append(["T-SPEC-SOURCE", "KA-SOURCE", "测试类别", "测试车型", "测试零件", "旧规格", 3, "", 2.5, 312, 1260])
        workbook.save(path)
        workbook.close()

        changed = sync_material_specs_from_dimensions(path)
        synced = load_workbook(path, read_only=True, data_only=True)
        try:
            self.assertEqual(changed, 1)
            self.assertEqual(synced["材料数据"].cell(2, 6).value, "2.5×312×1260")
        finally:
            synced.close()

    def test_upload_limits_keep_product_sync_headroom(self):
        self.assertEqual(self.web.MAX_UPLOAD_MB, 20)
        self.assertEqual(self.web.PRODUCT_SYNC_MAX_UPLOAD_MB, 512)
        self.assertEqual(self.web.app.config["MAX_CONTENT_LENGTH"], 512 * 1024 * 1024)

    def test_oversized_upload_redirects(self):
        self.login()
        original_limit = self.web.app.config["MAX_CONTENT_LENGTH"]
        self.web.app.config["MAX_CONTENT_LENGTH"] = 10
        try:
            big_file = io.BytesIO(b"x" * 11)
            response = self.client.post(
                "/catalog",
                data={"catalog": (big_file, "big.xlsx")},
                content_type="multipart/form-data",
                follow_redirects=False,
            )
            self.assertEqual(response.status_code, 302)
            self.assertTrue(response.headers["Location"].endswith("/products"))
            response.close()
        finally:
            self.web.app.config["MAX_CONTENT_LENGTH"] = original_limit

    def test_async_shipment_oversized_upload_returns_json(self):
        self.login()
        original_limit = self.web.app.config["MAX_CONTENT_LENGTH"]
        self.web.app.config["MAX_CONTENT_LENGTH"] = 100
        try:
            response = self.client.post(
                "/shipment-recognition/run",
                data={
                    "provider": "tesseract",
                    "shipment_photos": (io.BytesIO(b"x" * 200), "photo.jpg"),
                },
                headers={"Accept": "application/json", "X-Requested-With": "fetch"},
                content_type="multipart/form-data",
                follow_redirects=False,
            )
            payload = response.get_json()
            self.assertEqual(response.status_code, 413)
            self.assertEqual(payload["ok"], False)
            self.assertIn("上传文件不能超过", payload["error"])
            response.close()
        finally:
            self.web.app.config["MAX_CONTENT_LENGTH"] = original_limit

    def test_migrations_are_recorded(self):
        with self.web.connect(self.web.DB_PATH) as conn:
            rows = conn.execute("SELECT id FROM schema_migrations ORDER BY id").fetchall()
        self.assertEqual(
            [row["id"] for row in rows],
            [
                "001_audit_log_actor",
                "002_product_price_and_image",
                "003_product_drawings",
                "004_product_image_slots",
                "005_internal_api_keys",
                "006_shipment_recognition_jobs",
                "007_product_status",
                "008_internal_api_key_plaintext",
                "009_quote_records",
                "010_quote_record_bld_prices",
                "011_customer_price_bld_index",
                "012_scrub_internal_api_key_plaintext",
                "013_api_principal_scopes_and_idempotency",
                "014_quote_record_version",
                "015_idempotency_response_headers",
                "016_api_artifacts",
            ],
        )

    def test_migration_scrubs_historical_api_key_plaintext(self):
        from app.migrations import run_migrations
        from app.platform.api_keys import verify_internal_api_token
        from app.platform.api_principal import LEGACY_COMPATIBILITY_SCOPES

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            conn.execute("ALTER TABLE internal_api_keys ADD COLUMN token_plain TEXT DEFAULT ''")
            conn.execute(
                "UPDATE internal_api_keys SET token_plain = ? WHERE id = (SELECT MIN(id) FROM internal_api_keys)",
                ("historical-plaintext",),
            )
            conn.execute(
                "DELETE FROM schema_migrations WHERE id = '012_scrub_internal_api_key_plaintext'"
            )
            conn.execute("UPDATE internal_api_keys SET scopes = '[]'")
            conn.execute(
                "DELETE FROM schema_migrations WHERE id = '013_api_principal_scopes_and_idempotency'"
            )
            conn.commit()
            run_migrations(conn)
            columns = {
                row["name"] for row in conn.execute("PRAGMA table_info(internal_api_keys)").fetchall()
            }
            principal = verify_internal_api_token(conn, token)
        self.assertNotIn("token_plain", columns)
        self.assertEqual(principal.integration_name, "OpenClaw Test")
        self.assertEqual(principal.scopes, LEGACY_COMPATIBILITY_SCOPES)

    def test_concurrent_database_initialization_is_process_safe(self):
        database_path = self.root / "concurrent-init.sqlite3"
        gate_path = self.root / "concurrent-init.go"
        script = """
import sys
import time
from pathlib import Path
from app.database import connect

database_path = Path(sys.argv[1])
gate_path = Path(sys.argv[2])
while not gate_path.exists():
    time.sleep(0.005)
with connect(database_path) as conn:
    conn.execute("SELECT COUNT(*) FROM schema_migrations").fetchone()
"""
        processes = [
            subprocess.Popen(
                [sys.executable, "-c", script, str(database_path), str(gate_path)],
                cwd=PROJECT_ROOT,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
            )
            for _ in range(6)
        ]
        gate_path.touch()
        results = [process.communicate(timeout=20) for process in processes]
        failures = [stderr for process, (_stdout, stderr) in zip(processes, results) if process.returncode != 0]
        self.assertEqual(failures, [])

    def test_default_admin_requires_explicit_password_and_never_rewrites_existing_hash(self):
        from app.config import DEFAULT_ADMIN_PASSWORD_PLACEHOLDER
        from app.database import ensure_default_admin, now_text

        with self.web.connect(self.web.DB_PATH) as conn:
            with self.assertRaisesRegex(RuntimeError, "DEFAULT_ADMIN_PASSWORD"):
                ensure_default_admin(
                    conn,
                    username="placeholder-admin",
                    password=DEFAULT_ADMIN_PASSWORD_PLACEHOLDER,
                )
            legacy_hash = "scrypt:32768:8:1$legacy$unsupported-hash"
            stamp = now_text()
            conn.execute(
                """
                INSERT INTO users (username, display_name, password_hash, role, active, created_at, updated_at)
                VALUES (?, '', ?, 'admin', 1, ?, ?)
                """,
                ("legacy-admin", legacy_hash, stamp, stamp),
            )
            conn.commit()
            ensure_default_admin(conn, username="legacy-admin", password="replacement-password")
            stored = conn.execute(
                "SELECT password_hash FROM users WHERE username = ?",
                ("legacy-admin",),
            ).fetchone()["password_hash"]
        self.assertEqual(stored, legacy_hash)

    def test_generated_files_are_scoped_to_user(self):
        self.login()
        user_files = set((self.root / "outputs").glob("u*-007/catalog-export-bld-007-*.xlsx"))
        response = self.client.post("/products/export", data={"status": "active", "export_format": "bld"})
        self.assertEqual(response.status_code, 200)
        response.close()

        files = set((self.root / "outputs").glob("u*-007/catalog-export-bld-007-*.xlsx")) - user_files
        self.assertEqual(len(files), 1)
        self.assertFalse(list((self.root / "outputs").glob("catalog-export-bld-007-*.xlsx")))

    def test_catalog_export_is_admin_only(self):
        from app.database import save_user

        with self.web.connect(self.web.DB_PATH) as conn:
            save_user(
                conn,
                {
                    "username": "editor-export",
                    "display_name": "Editor Export",
                    "password": "editor-pw",
                    "role": "editor",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        admin_page = self.client.get("/products").get_data(as_text=True)
        self.assertIn("导出目录", admin_page)

        self.client.post("/logout")
        login = self.client.post(
            "/login",
            data={"username": "editor-export", "password": "editor-pw", "next": "/"},
            follow_redirects=False,
        )
        self.assertEqual(login.status_code, 302)

        editor_page = self.client.get("/products").get_data(as_text=True)
        self.assertNotIn("导出目录", editor_page)
        self.assertNotIn('action="/products/export"', editor_page)

        response = self.client.post("/products/export", data={"status": "active", "export_format": "bld"})
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/"))
        self.assertFalse(list((self.root / "outputs").glob("**/catalog-export-bld-editor-export-*.xlsx")))
        self.client.post("/logout")

    def test_product_export_embeds_main_image(self):
        from openpyxl import load_workbook
        from PIL import Image

        from app.database import upsert_product

        image_dir = self.root / "data" / "product_images"
        image_dir.mkdir(parents=True, exist_ok=True)
        image_path = image_dir / "K-EXPORT-IMG.png"
        Image.new("RGB", (80, 40), "white").save(image_path)

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-EXPORT-IMG",
                    "series": "TEST",
                    "item": "EXPORT IMAGE",
                    "oe_no_1": "EXPORT-IMAGE-001",
                    "models": "Tester",
                    "image_path": "data_product_images/K-EXPORT-IMG.png",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.post("/products/export", data={"status": "active", "export_format": "bld"})
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(io.BytesIO(response.data))
        sheet = workbook["产品目录"]
        row_index = next(row[0].row for row in sheet.iter_rows(min_row=2) if row[0].value == "K-EXPORT-IMG")

        self.assertIsNone(sheet.cell(row_index, 7).value)
        self.assertGreaterEqual(len(sheet._images), 1)
        self.assertGreaterEqual(sheet.row_dimensions[row_index].height, 62)
        workbook.close()
        response.close()

    def test_catalog_export_uses_bld_natural_order(self):
        from openpyxl import load_workbook

        from app.database import upsert_product

        expected = ["K8274LA", "K8274RA", "K8274LB", "K8274RB"]
        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no in expected:
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "SORT EXPORT",
                        "item": "SORT EXPORT TEST",
                        "oe_no_1": f"OE-{bld_no}",
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.post("/products/export", data={"status": "active", "export_format": "bld"})
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(io.BytesIO(response.data), read_only=True)
        sheet = workbook["产品目录"]
        exported = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] in expected]

        self.assertEqual(exported, expected)
        workbook.close()
        response.close()

    def test_admin_homepage_shows_all_recent_outputs(self):
        output_root = self.root / "outputs"
        other_user_dir = output_root / "u99-other"
        other_user_dir.mkdir(parents=True, exist_ok=True)
        root_file = output_root / "old-root-result.xlsx"
        other_user_file = other_user_dir / "other-user-result.xlsx"
        catalog_file = output_root / "catalog-export-bld-history-sample.xlsx"
        material_file = output_root / "26年4月冲压生产计划260423料单.xlsx"
        root_file.write_bytes(b"legacy")
        other_user_file.write_bytes(b"other")
        catalog_file.write_bytes(b"catalog")
        material_file.write_bytes(b"materials")

        self.login()
        response = self.client.get("/")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("操作用户", html)
        self.assertIn("data-history-loader", html)
        self.assertIn("展开后加载", html)
        self.assertIn("data-file-drop-zone", html)
        self.assertIn("可拖入询价文件", html)
        self.assertIn("输入 OE或 BLD 号", html)
        self.assertIn("file-picker-clear", html)
        self.assertNotIn("old-root-result.xlsx", html)
        self.assertNotIn("other-user-result.xlsx", html)

        response = self.client.get("/history-files")
        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertIsNotNone(payload)
        names = [item["name"] for item in payload["rows"]]
        self.assertIn("old-root-result.xlsx", names)
        self.assertIn("other-user-result.xlsx", names)
        operators = {item["name"]: item["operator"] for item in payload["rows"]}
        self.assertEqual(operators["other-user-result.xlsx"], "other")
        self.assertNotIn("catalog-export-bld-history-sample.xlsx", names)
        self.assertNotIn("26年4月冲压生产计划260423料单.xlsx", names)

        response = self.client.get("/?history_q=other-user")
        html = response.get_data(as_text=True)
        self.assertIn("other-user-result.xlsx", html)
        self.assertNotIn("old-root-result.xlsx", html)

        response = self.client.get("/?history_q=other")
        html = response.get_data(as_text=True)
        self.assertIn("other-user-result.xlsx", html)
        self.assertNotIn("old-root-result.xlsx", html)

        response = self.client.get("/history-files?history_q=other")
        payload = response.get_json()
        names = [item["name"] for item in payload["rows"]]
        self.assertIn("other-user-result.xlsx", names)
        self.assertNotIn("old-root-result.xlsx", names)

    def test_quick_oe_lookup_on_homepage(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K6004LB",
                    "series": "HYUNDAI",
                    "item": "CONTROL ARM",
                    "oe_no_1": "55270-2Z000",
                    "models": "Sportage",
                    "image_path": "product_images/K6004LB.jpg",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/?quick_oe=55270-2Z000")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("快速号码查询", html)
        self.assertIn("K6004LB", html)
        self.assertIn("OE 精准命中", html)
        self.assertIn("data-quick-oe-image", html)
        self.assertIn('id="quick-oe-image-modal"', html)

    def test_quick_brand_code_lookup_on_homepage(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K6004BR",
                    "series": "HYUNDAI",
                    "item": "CONTROL ARM",
                    "oe_no_1": "55270-2Z010",
                    "oe_no_2": "MOOG：K623123",
                    "models": "Sportage",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        for query in ["623123", "K623123", "MOOG：K623123"]:
            with self.subTest(query=query):
                response = self.client.get("/", query_string={"quick_oe": query})
                html = response.get_data(as_text=True)

                self.assertEqual(response.status_code, 200)
                self.assertIn("快速号码查询", html)
                self.assertIn("K6004BR", html)
                self.assertIn("品牌号码精准命中", html)

    def test_quick_bld_lookup_on_homepage(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-BLD-LOOKUP",
                    "series": "HYUNDAI",
                    "item": "CONTROL ARM",
                    "oe_no_1": "BLDLOOKUP-OE",
                    "models": "Sportage",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/", query_string={"quick_oe": "K-BLD-LOOKUP"})
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("快速号码查询", html)
        self.assertIn("K-BLD-LOOKUP", html)
        self.assertIn("BLD NO. 精准命中", html)

    def test_quick_bld_fragment_lookup_on_homepage(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no in ["K6004LB", "K6004RB", "K6015B"]:
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "HYUNDAI",
                        "item": "CONTROL ARM",
                        "oe_no_1": f"OE-{bld_no}",
                        "models": "Sportage",
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.get("/", query_string={"quick_oe": "6004"})
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("快速号码查询", html)
        self.assertIn("K6004LB", html)
        self.assertIn("K6004RB", html)
        self.assertNotIn("K6015B", html)
        self.assertIn("BLD NO. 片段命中", html)

    def test_quick_partial_number_lookup_checks_bld_oe_and_brand_codes(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no, oe_no, brand_no in [
                ("K-DV613-L", "DV613A424AF", "X15CJ6600"),
                ("K-DV613-R", "DV613A423AF", "X15CJ6601"),
                ("K-NUM-5450", "54500-2D000", "BRAND-54500"),
            ]:
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "FORD",
                        "item": "CONTROL ARM",
                        "oe_no_1": oe_no,
                        "oe_no_2": brand_no,
                        "models": "Transit",
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.get("/", query_string={"quick_oe": "dv613"})
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("K-DV613-L", html)
        self.assertIn("K-DV613-R", html)
        self.assertIn("OE 前缀命中", html)

        response = self.client.get("/", query_string={"quick_oe": "5450"})
        html = response.get_data(as_text=True)
        self.assertIn("K-NUM-5450", html)
        self.assertIn("OE 前缀命中", html)

        response = self.client.get("/", query_string={"quick_oe": "15CJ"})
        html = response.get_data(as_text=True)
        self.assertIn("K-DV613-L", html)
        self.assertIn("K-DV613-R", html)
        self.assertIn("品牌号码片段命中", html)

    def test_quick_lookup_requires_at_least_four_normalized_chars(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-SHORT-001",
                    "series": "FORD",
                    "item": "CONTROL ARM",
                    "oe_no_1": "ABC12345",
                    "models": "Transit",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/", query_string={"quick_oe": "abc"})
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("请输入至少 4 位号码", html)
        self.assertNotIn("K-SHORT-001", html)

    def test_quick_lookup_uses_unique_oe_suffix_variant(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8041LB",
                    "series": "VW",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "561407151A\n561407151C",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/", query_string={"quick_oe": "561407151D"})
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("K8041LB", html)
        self.assertIn("OE 尾字母容错命中", html)

    def test_quick_lookup_psa_352x_dot_prefers_psa_over_gm_exact(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-PSA-352088-QUICK",
                    "series": "PEUGEOT\nCITROEN",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "3520.88",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-GM-352088-QUICK",
                    "series": "GM\nOPEL",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "352088",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/", query_string={"quick_oe": "3520.88"})
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("K-PSA-352088-QUICK", html)
        self.assertIn("PSA 号码点号容错命中", html)
        self.assertNotIn("K-GM-352088-QUICK", html)

    def test_single_pasted_code_keeps_quick_lookup(self):
        self.login()
        response = self.client.post("/match", data={"quick_oe": "55270-2Z000"})
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/?quick_oe=55270-2Z000"))

    def test_pasted_inquiry_has_character_limit(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-LIMIT-001",
                    "oe_no_1": "LIMIT-001",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.post(
            "/match",
            data={"quick_oe": "A" * 5001},
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("粘贴号码最多支持 5000 个字符", html)

    def test_pasted_multiple_codes_generates_match_excel(self):
        from app.database import upsert_product
        from openpyxl import load_workbook

        products = [
            ("K54500L", "54500-2D000", "79.2"),
            ("K54501L", "54501-2D000", "39.6"),
            ("K54501A", "54501-A0000", "118.8"),
        ]
        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no, oe_no, price_cny in products:
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "HYUNDAI",
                        "item": "CONTROL ARM",
                        "oe_no_1": oe_no,
                        "models": "Elantra",
                        "price_cny": price_cny,
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.post(
            "/match",
            data={"quick_oe": "54500-2d000 54501-2d000 54501-a0000"},
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("匹配结果", html)
        self.assertIn("下载 Excel", html)
        self.assertIn("K54500L", html)
        self.assertIn("K54501L", html)
        self.assertIn("K54501A", html)
        self.assertIn("粘贴号码询价.xlsx", html)
        self.assertIn("含税单价", html)
        self.assertIn("¥79.20", html)
        self.assertIn("<td>1</td>", html)
        self.assertIn("<td>2</td>", html)
        self.assertIn("<td>3</td>", html)
        self.assertNotIn("<td>4</td>", html)
        self.assertIn('id="download-excel-modal"', html)
        self.assertIn('action="/match/download"', html)
        self.assertNotIn("返回上一步", html)

        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)
        upload_path = upload_match.group(1)
        output_name = output_match.group(1)
        output_path = self.root / "outputs" / "u1-007" / output_name

        download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_path,
                "original_filename": "粘贴号码询价.xlsx",
                "output_name": output_name,
                "match_column": "",
                "price_mode": "usd",
                "exchange_rate": "7.2",
            },
        )
        self.assertEqual(download.status_code, 200)
        download.close()
        self.assertTrue(output_path.exists())

        generated = load_workbook(output_path)
        sheet = generated.active
        self.assertEqual(sheet.cell(1, 1).value, "OE号")
        self.assertEqual(sheet.cell(1, 2).value, "BLD NO.")
        self.assertEqual(sheet.cell(1, 3).value, "美金价")
        self.assertEqual(sheet.cell(1, 4).value, "Product Status")
        self.assertEqual(sheet.cell(1, 5).value, "匹配说明")
        self.assertEqual(sheet.cell(2, 1).value, "54500-2d000")
        self.assertEqual(sheet.cell(2, 2).value, "K54500L")
        self.assertEqual(sheet.cell(2, 3).value, 10)
        self.assertEqual(sheet.cell(3, 2).value, "K54501L")
        self.assertEqual(sheet.cell(3, 3).value, 5)
        self.assertEqual(sheet.cell(4, 2).value, "K54501A")
        self.assertEqual(sheet.cell(4, 3).value, 15)
        generated.close()

    def test_pasted_combined_oe_prefix_stays_one_query(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8282RA",
                    "series": "FORD",
                    "item": "Front Right Lower Control Arm",
                    "oe_no_1": "F1F1-3A423-AAA\nF1F1-3A423-AAB",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K8235RA",
                    "series": "FORD",
                    "item": "Front Right Lower Control Arm",
                    "oe_no_1": "JX61\n3A423\nAPB",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.post("/match", data={"quick_oe": "F1F1 3A423"})
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("匹配结果", html)
        self.assertIn("K8282RA", html)
        self.assertIn("OE 组合前缀命中", html)
        self.assertNotIn("K8235RA", html)
        self.assertIn("<td>1</td>", html)
        self.assertNotIn("<td>2</td>", html)

    def test_uploaded_inquiry_combined_oe_prefix_matches_before_fragments(self):
        from app.database import upsert_product
        from app.excel_io import generate_excel_with_bld
        from app.helpers import load_catalog
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8282RA",
                    "series": "FORD",
                    "item": "Front Right Lower Control Arm",
                    "oe_no_1": "F1F1-3A423-AAA\nF1F1-3A423-AAB",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K8235RA",
                    "series": "FORD",
                    "item": "Front Right Lower Control Arm",
                    "oe_no_1": "JX61\n3A423\nAPB",
                    "active": "1",
                },
                actor="tester",
            )

        inquiry_path = self.root / "uploads" / "combined-prefix.xlsx"
        inquiry_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["F1F1 3A423"])
        workbook.save(inquiry_path)
        workbook.close()

        summary = generate_excel_with_bld(
            inquiry_path,
            self.root / "outputs" / "combined-prefix-result.xlsx",
            load_catalog(),
            write_output=False,
        )

        self.assertEqual(summary["total"], 1)
        self.assertEqual(summary["matched"], 1)
        self.assertEqual(summary["rows"][0]["bld_no"], "K8282RA")
        self.assertEqual(summary["rows"][0]["reason"], "OE 组合前缀命中")

    def test_uploaded_inquiry_integer_decimal_text_matches_prefix(self):
        from app.database import upsert_product
        from app.excel_io import generate_excel_with_bld
        from app.helpers import load_catalog
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8041LB",
                    "series": "VW",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "561407151A\n561407151C",
                    "active": "1",
                },
                actor="tester",
            )

        inquiry_path = self.root / "uploads" / "integer-decimal-text.xlsx"
        inquiry_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["561407151.0"])
        workbook.save(inquiry_path)
        workbook.close()

        summary = generate_excel_with_bld(
            inquiry_path,
            self.root / "outputs" / "integer-decimal-text-result.xlsx",
            load_catalog(),
            write_output=False,
        )

        self.assertEqual(summary["total"], 1)
        self.assertEqual(summary["matched"], 1)
        self.assertEqual(summary["rows"][0]["bld_no"], "K8041LB")
        self.assertEqual(summary["rows"][0]["reason"], "OE 组合前缀命中")

    def test_uploaded_inquiry_oe_suffix_variant_matches_unique_base(self):
        from app.database import upsert_product
        from app.excel_io import generate_excel_with_bld
        from app.helpers import load_catalog
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8041LB",
                    "series": "VW",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "561407151A\n561407151C",
                    "active": "1",
                },
                actor="tester",
            )

        inquiry_path = self.root / "uploads" / "oe-suffix-variant.xlsx"
        inquiry_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["561407151D"])
        workbook.save(inquiry_path)
        workbook.close()

        summary = generate_excel_with_bld(
            inquiry_path,
            self.root / "outputs" / "oe-suffix-variant-result.xlsx",
            load_catalog(),
            write_output=False,
        )

        self.assertEqual(summary["total"], 1)
        self.assertEqual(summary["matched"], 1)
        self.assertEqual(summary["rows"][0]["bld_no"], "K8041LB")
        self.assertEqual(summary["rows"][0]["reason"], "OE 尾字母容错命中")

    def test_uploaded_inquiry_split_oe_suffix_variants_match_same_product(self):
        from app.database import upsert_product
        from app.excel_io import generate_excel_with_bld
        from app.helpers import load_catalog
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8321LB",
                    "series": "VW",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "2QD407151",
                    "active": "1",
                },
                actor="tester",
            )

        inquiry_path = self.root / "uploads" / "split-oe-suffix-variants.xlsx"
        inquiry_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["2QD407151A;2QD407151C"])
        workbook.save(inquiry_path)
        workbook.close()

        summary = generate_excel_with_bld(
            inquiry_path,
            self.root / "outputs" / "split-oe-suffix-variants-result.xlsx",
            load_catalog(),
            write_output=False,
        )

        self.assertEqual(summary["total"], 1)
        self.assertEqual(summary["matched"], 1)
        self.assertEqual(summary["rows"][0]["bld_no"], "K8321LB")
        self.assertEqual(summary["rows"][0]["reason"], "OE 尾字母容错命中")
        self.assertEqual(summary["rows"][0]["matched_oe_codes"], ["2QD407151A", "2QD407151C"])
        self.assertIn("命中号码：2QD407151A, 2QD407151C", summary["rows"][0]["match_note"])

    def test_psa_352x_dot_matches_psa_before_gm_exact(self):
        from app.matcher import ProductCatalog

        catalog = ProductCatalog(
            [
                {
                    "BLD NO.": "K-PSA-352123",
                    "SERIES": "PEUGEOT\nCITROEN",
                    "ITEM": "Front Left Lower Control Arm",
                    "OE NO.1": "3521.23",
                },
                {
                    "BLD NO.": "K-GM-352123",
                    "SERIES": "GM\nOPEL",
                    "ITEM": "Front Left Lower Control Arm",
                    "OE NO.1": "352123",
                },
            ]
        )

        match = catalog.match("", "3521.23")

        self.assertIsNotNone(match)
        self.assertEqual(match.bld_no, "K-PSA-352123")
        self.assertEqual(match.reason, "PSA 号码点号容错命中")

    def test_psa_352x_without_dot_is_ambiguous_when_gm_exact_also_exists(self):
        from app.matcher import ProductCatalog

        catalog = ProductCatalog(
            [
                {
                    "BLD NO.": "K-PSA-352123",
                    "SERIES": "PEUGEOT\nCITROEN",
                    "ITEM": "Front Left Lower Control Arm",
                    "OE NO.1": "3521.23",
                },
                {
                    "BLD NO.": "K-GM-352123",
                    "SERIES": "GM\nOPEL",
                    "ITEM": "Front Left Lower Control Arm",
                    "OE NO.1": "352123",
                },
            ]
        )

        match = catalog.match("", "352123")

        self.assertIsNotNone(match)
        self.assertIn("K-PSA-352123", match.bld_no)
        self.assertIn("K-GM-352123", match.bld_no)
        self.assertEqual(match.reason, "3520/3521 号码同时命中 PSA 与其他品牌，请人工确认")

    def test_psa_352x_without_dot_still_matches_gm_when_no_psa_exists(self):
        from app.matcher import ProductCatalog

        catalog = ProductCatalog(
            [
                {
                    "BLD NO.": "K-GM-352023",
                    "SERIES": "GM\nOPEL",
                    "ITEM": "Front Left Lower Control Arm",
                    "OE NO.1": "352023",
                }
            ]
        )

        match = catalog.match("", "352023")

        self.assertIsNotNone(match)
        self.assertEqual(match.bld_no, "K-GM-352023")
        self.assertEqual(match.reason, "OE 精准命中")

    def test_uploaded_inquiry_psa_352x_dot_does_not_match_gm_exact(self):
        from app.database import upsert_product
        from app.excel_io import generate_excel_with_bld
        from app.helpers import load_catalog
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-PSA-352124-FILE",
                    "series": "PEUGEOT\nCITROEN",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "3521.24",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-GM-352124-FILE",
                    "series": "GM\nOPEL",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "352124",
                    "active": "1",
                },
                actor="tester",
            )

        inquiry_path = self.root / "uploads" / "psa-352x-dot.xlsx"
        inquiry_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["3521.24"])
        workbook.save(inquiry_path)
        workbook.close()

        summary = generate_excel_with_bld(
            inquiry_path,
            self.root / "outputs" / "psa-352x-dot-result.xlsx",
            load_catalog(),
            write_output=False,
        )

        self.assertEqual(summary["total"], 1)
        self.assertEqual(summary["matched"], 1)
        self.assertEqual(summary["rows"][0]["bld_no"], "K-PSA-352124-FILE")
        self.assertEqual(summary["rows"][0]["reason"], "PSA 号码点号容错命中")

    def test_pasted_multiple_bld_codes_generates_match_excel(self):
        from app.database import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no in ["K-BLD-BATCH-1", "K-BLD-BATCH-2"]:
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "HYUNDAI",
                        "item": "CONTROL ARM",
                        "oe_no_1": f"{bld_no}-OE",
                        "models": "Elantra",
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.post("/match", data={"quick_oe": "K-BLD-BATCH-1 K-BLD-BATCH-2"})
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("匹配结果", html)
        self.assertIn("K-BLD-BATCH-1", html)
        self.assertIn("K-BLD-BATCH-2", html)
        self.assertIn("BLD NO. 精准命中", html)

    def test_uploaded_inquiry_can_export_tax_price(self):
        from app.database import upsert_product
        from openpyxl import Workbook, load_workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "KPRICE01",
                    "series": "HYUNDAI",
                    "item": "CONTROL ARM",
                    "oe_no_1": "PRICE-001",
                    "models": "Elantra",
                    "price_cny": "88.8",
                    "active": "1",
                },
                actor="tester",
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["PRICE-001"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        self.login()
        response = self.client.post(
            "/match",
            data={"inquiry": (buffer, "price-export.xlsx")},
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("选择匹配列", html)
        self.assertIn("没有识别到明确的 OE 号码表头", html)
        self.assertNotIn("KPRICE01", html)
        self.assertNotIn('id="download-excel-modal"', html)

        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)
        upload_path = upload_match.group(1)
        output_name = output_match.group(1)
        output_path = self.root / "outputs" / "u1-007" / output_name

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_path,
                "original_filename": "price-export.xlsx",
                "output_name": output_name,
                "match_column": "0",
            },
        )
        result_html = result.get_data(as_text=True)

        self.assertEqual(result.status_code, 200)
        self.assertIn("KPRICE01", result_html)
        self.assertIn("¥88.80", result_html)
        self.assertIn('id="download-excel-modal"', result_html)
        self.assertIn('value="net">带不含税单价', result_html)
        self.assertIn("返回上一步", result_html)

        download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_path,
                "original_filename": "price-export.xlsx",
                "output_name": output_name,
                "match_column": "0",
                "price_mode": "tax",
            },
        )
        self.assertEqual(download.status_code, 200)
        download.close()
        self.assertTrue(output_path.exists())

        generated = load_workbook(output_path)
        sheet = generated.active
        self.assertEqual(sheet.cell(1, 2).value, "BLD NO.")
        self.assertEqual(sheet.cell(1, 3).value, "含税单价")
        self.assertEqual(sheet.cell(1, 4).value, "产品状态")
        self.assertEqual(sheet.cell(1, 5).value, "匹配说明")
        self.assertEqual(sheet.cell(2, 2).value, "KPRICE01")
        self.assertEqual(sheet.cell(2, 3).value, 88.8)
        generated.close()

        net_download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_path,
                "original_filename": "price-export.xlsx",
                "output_name": output_name,
                "match_column": "0",
                "price_mode": "net",
            },
        )
        self.assertEqual(net_download.status_code, 200)
        net_download.close()

        generated = load_workbook(output_path)
        sheet = generated.active
        self.assertEqual(sheet.cell(1, 3).value, "不含税单价")
        self.assertEqual(sheet.cell(2, 3).value, 81)
        self.assertEqual(sheet.cell(2, 3).number_format, "0")
        generated.close()

    def test_uploaded_polluted_xlsx_uses_cleaned_copy_without_skipping_late_rows(self):
        from app.database import upsert_product
        from openpyxl import Workbook, load_workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no, oe_no in [("KCLEAN01", "CLEAN-002"), ("KCLEAN02", "CLEAN-250")]:
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "HYUNDAI",
                        "item": "CLEAN TEST ARM",
                        "oe_no_1": oe_no,
                        "active": "1",
                    },
                    actor="tester",
                )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["CLEAN-002"])
        sheet.cell(250, 1).value = "CLEAN-250"
        polluted_path = self.root / "uploads" / "polluted-inquiry.xlsx"
        polluted_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(polluted_path)
        workbook.close()
        pollute_xlsx_tail(polluted_path, declared_rows=2000, after_row=251)

        self.login()
        with polluted_path.open("rb") as handle:
            response = self.client.post(
                "/match",
                data={"inquiry": (io.BytesIO(handle.read()), "polluted-inquiry.xlsx")},
                content_type="multipart/form-data",
            )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("已自动清理 Excel 尾部空白格式", html)
        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)
        self.assertIn("inquiry-cleaned", upload_match.group(1))
        cleaned_workbook = load_workbook(Path(upload_match.group(1)), read_only=True, data_only=True)
        self.assertEqual(cleaned_workbook.active.cell(2, 1).value, "CLEAN-002")
        cleaned_workbook.close()

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "polluted-inquiry.xlsx",
                "output_name": output_match.group(1),
                "match_column": "0",
            },
        )
        result_html = result.get_data(as_text=True)

        self.assertEqual(result.status_code, 200)
        self.assertIn("共 2 行，命中 2 行，未找到 0 行", result_html)
        self.assertIn("KCLEAN01", result_html)
        self.assertIn("KCLEAN02", result_html)
        self.assertIn("<td>250</td>", result_html)

    def test_uploaded_inquiry_can_match_multiple_selected_columns(self):
        from app.database import upsert_product
        from openpyxl import Workbook, load_workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "KMULTI02",
                    "series": "HYUNDAI",
                    "item": "MULTI COLUMN ARM",
                    "oe_no_1": "REF-MULTI-002",
                    "price_cny": "77",
                    "active": "1",
                },
                actor="tester",
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["客户OE", "参考号"])
        sheet.append(["NO-HIT-001", "REF-MULTI-002"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        self.login()
        response = self.client.post(
            "/match",
            data={"inquiry": (buffer, "multi-column.xlsx")},
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('name="match_columns" value="0"', html)
        self.assertIn('name="match_columns" value="1"', html)
        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "multi-column.xlsx",
                "output_name": output_match.group(1),
                "match_columns": ["0", "1"],
            },
        )
        result_html = result.get_data(as_text=True)

        self.assertEqual(result.status_code, 200)
        self.assertIn("KMULTI02", result_html)
        self.assertIn("命中列：B列：REF-MULTI-002", result_html)

        output_path = self.root / "outputs" / "u1-007" / output_match.group(1)
        download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "multi-column.xlsx",
                "output_name": output_match.group(1),
                "match_columns": ["0", "1"],
                "price_mode": "tax",
            },
        )
        self.assertEqual(download.status_code, 200)
        download.close()

        generated = load_workbook(output_path)
        generated_sheet = generated.active
        self.assertEqual(generated_sheet.cell(1, 3).value, "BLD NO.")
        self.assertEqual(generated_sheet.cell(2, 3).value, "KMULTI02")
        self.assertIn("命中列：B列：REF-MULTI-002", generated_sheet.cell(2, 6).value)
        generated.close()

    def test_item_header_with_code_values_prompts_for_match_column(self):
        from app.database import upsert_product
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "KPIKA01",
                    "series": "HYUNDAI",
                    "item": "LOWER ARM",
                    "oe_no_1": "TST545012B000",
                    "models": "SANTA FE 2006",
                    "price_cny": "66",
                    "active": "1",
                },
                actor="tester",
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["PIKA BOOKING ORDER-200426"])
        sheet.append([])
        sheet.append(["SN", "ITEM", "DESCRIPTION", "QTY", "PRICE", "PICTURE", "BRAND"])
        sheet.append(["1", "TST545012B000 ", "LOWER ARM-HYUNDAI SANTA FE 2006", 60, None, None, "L-TGL"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        self.login()
        response = self.client.post(
            "/match",
            data={"inquiry": (buffer, "pika-order.xlsx")},
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("选择匹配列", html)
        self.assertIn("没有识别到明确的 OE 号码表头", html)
        self.assertIn("match-preview-table", html)
        self.assertIn("match-preview-cell", html)
        self.assertIn("<th>源行</th>", html)

        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "pika-order.xlsx",
                "output_name": output_match.group(1),
                "match_column": "1",
            },
        )
        result_html = result.get_data(as_text=True)

        self.assertEqual(result.status_code, 200)
        self.assertIn("共 1 行，命中 1 行，未找到 0 行", result_html)
        self.assertIn("KPIKA01", result_html)
        self.assertIn("¥66.00", result_html)
        self.assertIn("<td>4</td>", result_html)
        self.assertNotIn("<td>3</td>", result_html)

    def test_xlsx_without_dimension_can_preview_match_columns(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["فحمات", None])
        sheet.append(["رقم", "كمية"])
        sheet.append([446633220, 300])
        sheet.append(["58101F2A00", 300])
        inquiry_path = self.root / "uploads" / "arabic-no-dimension.xlsx"
        inquiry_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(inquiry_path)
        workbook.close()
        strip_xlsx_dimension(inquiry_path)

        self.login()
        with inquiry_path.open("rb") as handle:
            response = self.client.post(
                "/match",
                data={"inquiry": (handle, "هونداي و تويوتا.xlsx")},
                content_type="multipart/form-data",
            )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("选择匹配列", html)
        self.assertIn("فحمات", html)
        self.assertIn("رقم", html)
        self.assertIn('name="match_columns" value="0"', html)
        self.assertIn('name="match_columns" value="1"', html)
        self.assertNotIn("生成失败", html)

    def test_segmented_merged_headers_do_not_count_as_inquiry_rows(self):
        from app.database import upsert_product
        from openpyxl import Workbook, load_workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            for index in range(1, 5):
                upsert_product(
                    conn,
                    {
                        "bld_no": f"KSEG{index:02d}",
                        "series": "TEST",
                        "item": "SEGMENTED ARM",
                        "oe_no_1": f"SEG-OE-{index:03d}",
                        "active": "1",
                    },
                    actor="tester",
                )

        workbook = Workbook()
        sheet = workbook.active
        row = 1
        for section in range(2):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            sheet.cell(row, 1).value = f"分段 {section + 1}"
            row += 1
            sheet.append(["序号", "OE号", "数量"])
            row += 1
            for item in range(2):
                number = section * 2 + item + 1
                sheet.append([number, f"SEG-OE-{number:03d}", 10])
                row += 1
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        buffer.seek(0)

        self.login()
        response = self.client.post(
            "/match",
            data={"inquiry": (buffer, "segmented-merged.xlsx")},
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "segmented-merged.xlsx",
                "output_name": output_match.group(1),
                "match_column": "1",
            },
        )
        result_html = result.get_data(as_text=True)

        self.assertEqual(result.status_code, 200)
        self.assertIn("共 4 行，命中 4 行，未找到 0 行", result_html)
        self.assertNotIn("<td>2</td>", result_html)
        self.assertNotIn("<td>6</td>", result_html)
        for index in range(1, 5):
            self.assertIn(f"KSEG{index:02d}", result_html)

        download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "segmented-merged.xlsx",
                "output_name": output_match.group(1),
                "match_column": "1",
            },
        )
        self.assertEqual(download.status_code, 200)
        download.close()
        generated = load_workbook(self.root / "outputs" / "u1-007" / output_match.group(1), data_only=True)
        generated_sheet = generated.active
        self.assertEqual(generated_sheet.cell(2, 4).value, "BLD NO.")
        self.assertEqual(generated_sheet.cell(3, 4).value, "KSEG01")
        self.assertEqual(generated_sheet.cell(4, 4).value, "KSEG02")
        self.assertIsNone(generated_sheet.cell(6, 4).value)
        self.assertEqual(generated_sheet.cell(7, 4).value, "KSEG03")
        self.assertEqual(generated_sheet.cell(8, 4).value, "KSEG04")
        generated.close()

    def test_catalog_import_recognizes_chinese_brand_number_header(self):
        from app.matcher import ProductCatalog
        from openpyxl import Workbook

        for header in ["品牌号码", "Other Reference"]:
            with self.subTest(header=header):
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["BLD NO.", "品牌", "产品名称", "OE Reference", header, "车型"])
                sheet.append(["K6004CN", "HYUNDAI", "CONTROL ARM", "55270-2Z020", "BRAND-CN-55270", "Sportage"])
                catalog_path = self.root / f"catalog-brand-number-{header.replace(' ', '-').lower()}.xlsx"
                workbook.save(catalog_path)

                catalog = ProductCatalog.from_excel(catalog_path)
                match = catalog.match("", "BRAND-CN-55270")

                self.assertIsNotNone(match)
                self.assertEqual(match.bld_no, "K6004CN")
                self.assertEqual(match.reason, "品牌号码精准命中")

    def test_manual_column_result_defers_excel_until_download(self):
        from app.database import upsert_product
        from openpyxl import Workbook, load_workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K6004LC",
                    "series": "HYUNDAI",
                    "item": "CONTROL ARM",
                    "oe_no_1": "55270-2Z001",
                    "models": "Sportage",
                    "price_cny": "55",
                    "active": "1",
                },
                actor="tester",
            )
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K6004LC",)).fetchone()

        self.login()
        drawing_upload = self.client.post(
            f"/products/{product['id']}/drawing",
            data={"drawing": (io.BytesIO(b"%PDF-1.4\nK6004LC drawing\n%%EOF"), "K6004LC.pdf")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        self.assertEqual(drawing_upload.status_code, 302)

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["客户号码", "数量"])
        sheet.append(["55270-2Z001", 1])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = self.client.post(
            "/match",
            data={"inquiry": (buffer, "manual-column.xlsx")},
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("选择匹配列", html)
        self.assertIn("返回上一步", html)
        self.assertNotIn("返回首页", html)
        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)
        upload_path = upload_match.group(1)
        output_name = output_match.group(1)
        output_path = self.root / "outputs" / "u1-007" / output_name

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_path,
                "original_filename": "manual-column.xlsx",
                "output_name": output_name,
                "match_column": "0",
            },
        )
        result_html = result.get_data(as_text=True)

        self.assertEqual(result.status_code, 200)
        self.assertIn("Excel 文件将在点击下载时生成", result_html)
        self.assertIn("下载 Excel", result_html)
        self.assertIn("下载图纸包", result_html)
        self.assertIn("返回上一步", result_html)
        self.assertNotIn("返回首页", result_html)
        self.assertIn("K6004LC", result_html)
        self.assertIn("¥55.00", result_html)
        self.assertIn('id="download-excel-modal"', result_html)
        self.assertIn('name="price_mode"', result_html)
        self.assertFalse(output_path.exists())

        drawing_zip = self.client.post(
            "/match/drawings/download",
            data={
                "upload_path": upload_path,
                "original_filename": "manual-column.xlsx",
                "match_column": "0",
            },
        )
        self.assertEqual(drawing_zip.status_code, 200)
        with zipfile.ZipFile(io.BytesIO(drawing_zip.get_data())) as archive:
            self.assertIn("K6004LC_55270-2Z001.pdf", archive.namelist())
        drawing_zip.close()

        back = self.client.post(
            "/match/column/back",
            data={
                "upload_path": upload_path,
                "original_filename": "manual-column.xlsx",
                "output_name": output_name,
                "match_column": "0",
            },
        )
        back_html = back.get_data(as_text=True)
        self.assertEqual(back.status_code, 200)
        self.assertIn("选择匹配列", back_html)
        self.assertIn("返回上一步", back_html)
        self.assertNotIn("返回首页", back_html)
        self.assertRegex(back_html, r'name="match_columns" value="0"[^>]*checked')

        download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_path,
                "original_filename": "manual-column.xlsx",
                "output_name": output_name,
                "match_column": "0",
                "price_mode": "tax",
            },
        )
        self.assertEqual(download.status_code, 200)
        download.close()
        self.assertTrue(output_path.exists())

        generated = load_workbook(output_path)
        generated_sheet = generated.active
        self.assertEqual(generated_sheet.cell(1, 3).value, "BLD NO.")
        self.assertEqual(generated_sheet.cell(1, 4).value, "含税单价")
        self.assertEqual(generated_sheet.cell(1, 5).value, "产品状态")
        self.assertEqual(generated_sheet.cell(1, 6).value, "匹配说明")
        self.assertEqual(generated_sheet.cell(2, 3).value, "K6004LC")
        self.assertEqual(generated_sheet.cell(2, 4).value, 55)
        generated.close()

    def test_uploaded_files_are_scoped_to_user(self):
        self.login()
        response = self.client.post(
            "/prices/import/preview",
            data={"price_file": (io.BytesIO(b"not a real workbook"), "same-name.xlsx")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        uploads = list((self.root / "uploads" / "u1-007").glob("price-*-007-same-name.xlsx"))
        self.assertEqual(len(uploads), 1)

    def test_import_lock_blocks_parallel_imports(self):
        from app.locks import ImportLockError, import_lock

        with import_lock("tester", "测试导入"):
            with self.assertRaises(ImportLockError):
                with import_lock("other", "第二个导入"):
                    pass


if __name__ == "__main__":
    unittest.main()

from __future__ import annotations

from tests.web_app_test_base import (
    WebAppTestBase,
    Path,
    io,
    json,
    patch,
    re,
    shutil,
    sqlite3,
    tarfile,
    unittest,
)


class TestWebProductImportAndSync(WebAppTestBase):
    def test_business_data_sync_exports_material_drawings_with_materials(self):
        self.login()
        drawing_dir = self.root / "data" / "material_drawings"
        drawing_dir.mkdir(parents=True, exist_ok=True)
        drawing = drawing_dir / "WEB-SYNC-MATERIAL.pdf"
        drawing.write_bytes(b"%PDF-1.4\nweb material drawing\n")
        try:
            page = self.client.get("/business-data-sync")
            html = page.get_data(as_text=True)
            self.assertIn('name="include_material_drawings"', html)
            self.assertIn("产品目录只同步启用型号", html)
            self.assertIn("<h3>导出内容</h3>", html)
            self.assertIn("<strong>物料图纸</strong>", html)
            self.assertNotIn("包含物料图纸", html)
            self.assertEqual(html.count('class="sync-option-grid sync-option-grid-unified"'), 1)
            self.assertEqual(html.count('class="sync-option"'), 8)
            for media_field in ("include_drawings", "include_images", "include_material_drawings"):
                media_input = re.search(rf'<input\b[^>]*name="{media_field}"[^>]*>', html)
                self.assertIsNotNone(media_input)
                self.assertNotIn("checked", media_input.group(0))

            response = self.client.post(
                "/business-data-sync/export",
                data={"dataset": "materials", "include_material_drawings": "1"},
            )
            try:
                self.assertEqual(response.status_code, 200)
                package_bytes = response.data
                with tarfile.open(fileobj=io.BytesIO(package_bytes), mode="r:gz") as archive:
                    self.assertIn("data/material_drawings/WEB-SYNC-MATERIAL.pdf", archive.getnames())
                    manifest_file = archive.extractfile("manifest.json")
                    self.assertIsNotNone(manifest_file)
                    manifest = json.loads(manifest_file.read().decode("utf-8"))
            finally:
                response.close()
            self.assertEqual(manifest["version"], 4)
            self.assertTrue(manifest["media"]["material_drawings"])
            self.assertEqual(manifest["media"]["files"]["material_drawings"], 1)

            preview = self.client.post(
                "/business-data-sync/preview",
                data={"package": (io.BytesIO(package_bytes), "web-material-sync.tar.gz")},
                content_type="multipart/form-data",
            )
            preview_html = preview.get_data(as_text=True)
            self.assertEqual(preview.status_code, 200)
            self.assertIn("物料图纸 1 个", preview_html)
            self.assertIn('name="include_material_drawings"', preview_html)
            package_path = re.search(r'name="package_path" value="([^"]+)"', preview_html)
            preview_token = re.search(r'name="preview_token" value="([^"]+)"', preview_html)
            self.assertIsNotNone(package_path)
            self.assertIsNotNone(preview_token)

            drawing.write_bytes(b"local replacement")
            applied = self.client.post(
                "/business-data-sync/apply",
                data={
                    "package_path": package_path.group(1),
                    "preview_token": preview_token.group(1),
                    "include_material_drawings": "1",
                },
                follow_redirects=True,
            )
            self.assertEqual(applied.status_code, 200)
            self.assertIn("业务数据导入完成", applied.get_data(as_text=True))
            self.assertEqual(drawing.read_bytes(), b"%PDF-1.4\nweb material drawing\n")
        finally:
            drawing.unlink(missing_ok=True)

    def test_product_options_endpoint_returns_deduped_picker_candidates(self):
        with self.client.session_transaction() as session:
            session.clear()
        anonymous = self.client.get(
            "/products/options",
            headers={"X-Requested-With": "fetch", "Accept": "application/json"},
        )
        self.assertEqual(anonymous.status_code, 401)

        self.addCleanup(self.cleanup_products, "K-OPTIONS-%")
        self.login()
        for bld_no, series, item, status, active in (
            ("K-OPTIONS-001", "OPTIONBRAND\nOPTIONSECOND", "Option Arm", "1 个球头 2 个衬套", "1"),
            ("K-OPTIONS-002", "OPTIONBRAND", "Option Arm", "1个球头2个衬套", "0"),
            ("K-OPTIONS-003", "", "", "", "1"),
        ):
            saved = self.client.post(
                "/products/save",
                data={
                    "bld_no": bld_no,
                    "series": series,
                    "item": item,
                    "product_status": status,
                    "active": active,
                },
                headers={"X-Requested-With": "fetch", "Accept": "application/json"},
            )
            self.assertEqual(saved.status_code, 200, saved.get_data(as_text=True))

        response = self.client.get("/products/options")
        payload = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["Cache-Control"], "no-store")
        self.assertEqual(sorted(payload.keys()), ["brands", "items", "statuses"])
        self.assertEqual(payload["brands"].count("OPTIONBRAND"), 1)
        self.assertIn("OPTIONSECOND", payload["brands"])
        self.assertEqual(payload["items"].count("Option Arm"), 1)
        self.assertEqual(payload["statuses"].count("1球头 2衬套"), 1)
        for candidates in payload.values():
            self.assertTrue(all(isinstance(value, str) and value for value in candidates))
            lowered = [value.lower() for value in candidates]
            self.assertEqual(len(lowered), len(set(lowered)))

    def test_product_save_registers_managed_option_values(self):
        self.addCleanup(self.cleanup_products, "K-OPTREG-%")
        self.addCleanup(self.cleanup_option_values, "OPTREG%")
        self.login()
        for _ in range(2):
            saved = self.client.post(
                "/products/save",
                data={
                    "bld_no": "K-OPTREG-001",
                    "series": "optreg brand\noptreg second",
                    "item": "Optreg Arm",
                    "product_status": "1 个球头",
                    "active": "1",
                },
                headers={"X-Requested-With": "fetch", "Accept": "application/json"},
            )
            self.assertEqual(saved.status_code, 200, saved.get_data(as_text=True))

        with self.web.connect(self.web.DB_PATH) as conn:
            rows = conn.execute(
                "SELECT kind, value FROM product_option_values WHERE value LIKE 'OPTREG%' OR value = '1球头'"
            ).fetchall()
        registered = {(row["kind"], row["value"]) for row in rows}
        self.assertIn(("brand", "OPTREG BRAND"), registered)
        self.assertIn(("brand", "OPTREG SECOND"), registered)
        self.assertIn(("item", "Optreg Arm"), registered)
        self.assertIn(("product_status", "1球头"), registered)
        brand_count = sum(1 for row in rows if row["kind"] == "brand" and row["value"] == "OPTREG BRAND")
        self.assertEqual(brand_count, 1)

    def test_product_options_admin_page_crud_and_audit(self):
        self.addCleanup(self.cleanup_option_values, "ADMINOPT%")
        self.login()

        page = self.client.get("/product-options")
        html = page.get_data(as_text=True)
        self.assertEqual(page.status_code, 200)
        self.assertIn('data-page="admin.product_options"', html)
        self.assertIn("产品基础信息", html)

        saved = self.client.post(
            "/product-options/save",
            data={"kind": "brand", "value": "adminopt brand"},
            follow_redirects=False,
        )
        self.assertEqual(saved.status_code, 302)
        self.client.post("/product-options/save", data={"kind": "item", "value": "Adminopt Item"})
        self.client.post("/product-options/save", data={"kind": "product_status", "value": "1 个球头 2 个衬套"})
        # 重复新增被拒绝：值保持唯一
        self.client.post("/product-options/save", data={"kind": "brand", "value": "ADMINOPT BRAND"})

        with self.web.connect(self.web.DB_PATH) as conn:
            brand_row = conn.execute(
                "SELECT * FROM product_option_values WHERE kind = 'brand' AND value = 'ADMINOPT BRAND'"
            ).fetchone()
            self.assertIsNotNone(brand_row)
            duplicates = conn.execute(
                "SELECT COUNT(*) FROM product_option_values WHERE kind = 'brand' AND value LIKE 'ADMINOPT BRAND'"
            ).fetchone()[0]
            self.assertEqual(duplicates, 1)
            status_row = conn.execute(
                "SELECT * FROM product_option_values WHERE kind = 'product_status' AND value = '1球头2衬套'"
            ).fetchone()
            self.assertIsNotNone(status_row)

        option_id = brand_row["id"]
        renamed = self.client.post(
            "/product-options/save",
            data={"kind": "brand", "id": str(option_id), "value": "ADMINOPT RENAMED"},
        )
        self.assertEqual(renamed.status_code, 302)
        # 改名规范化出多个值（复合输入）被拒绝
        self.client.post(
            "/product-options/save",
            data={"kind": "brand", "id": str(option_id), "value": "AAA/BBB"},
        )
        with self.web.connect(self.web.DB_PATH) as conn:
            current = conn.execute("SELECT value FROM product_option_values WHERE id = ?", (option_id,)).fetchone()
            self.assertEqual(current["value"], "ADMINOPT RENAMED")

        payload = self.client.get("/products/options").get_json()
        self.assertIn("ADMINOPT RENAMED", payload["brands"])
        self.assertIn("Adminopt Item", payload["items"])
        self.assertIn("1球头 2衬套", payload["statuses"])

        deleted = self.client.post("/product-options/delete", data={"id": str(option_id)})
        self.assertEqual(deleted.status_code, 302)
        with self.web.connect(self.web.DB_PATH) as conn:
            gone = conn.execute("SELECT 1 FROM product_option_values WHERE id = ?", (option_id,)).fetchone()
            self.assertIsNone(gone)
            audit_actions = {
                row["action"]
                for row in conn.execute(
                    "SELECT action FROM audit_logs WHERE target_type = 'product_option_value' AND target_key LIKE '%ADMINOPT%'"
                ).fetchall()
            }
        self.assertIn("新增产品候选值", audit_actions)
        self.assertIn("改名产品候选值", audit_actions)
        self.assertIn("删除产品候选值", audit_actions)

    def test_product_options_admin_page_requires_permission(self):
        from app.modules.admin.persistence import save_user

        self.login()
        with self.web.connect(self.web.DB_PATH) as conn:
            save_user(
                conn,
                {
                    "username": "viewer-options",
                    "display_name": "Viewer Options",
                    "password": "viewer-pw",
                    "role": "viewer",
                    "active": "1",
                },
                actor="tester",
            )
            conn.commit()

        self.client.post("/logout")
        anonymous = self.client.get("/product-options", follow_redirects=False)
        self.assertEqual(anonymous.status_code, 302)
        self.assertIn("/login", anonymous.headers["Location"])

        self.client.post(
            "/login",
            data={"username": "viewer-options", "password": "viewer-pw", "next": "/"},
        )
        denied = self.client.get("/product-options", follow_redirects=False)
        self.assertEqual(denied.status_code, 302)
        self.assertNotIn("/product-options", denied.headers["Location"])
        denied_post = self.client.post(
            "/product-options/save",
            data={"kind": "brand", "value": "DENIED"},
            follow_redirects=False,
        )
        self.assertEqual(denied_post.status_code, 302)
        self.client.post("/logout")

    def test_product_brand_normalization_uses_preview_confirmation_and_backup(self):
        self.login()
        with self.web.connect(self.web.DB_PATH) as connection:
            connection.execute(
                """
                INSERT INTO products (
                  bld_no, series, item, active, source, created_at, updated_at
                ) VALUES (
                  'BRAND-WEB-CLEANUP', 'DODGE RAM', 'Brand cleanup test', 1,
                  'test-fixture', '2026-07-14 10:00:00', '2026-07-14 10:00:00'
                )
                """
            )
            connection.commit()

        from app.modules.products.factory import get_product_service

        preview = get_product_service().preview_brand_normalization()
        self.assertGreaterEqual(preview.changed_count, 1)
        self.assertIn(
            ("BRAND-WEB-CLEANUP", "DODGE RAM", "DODGE"),
            [(change.bld_no, change.before, change.after) for change in preview.changes],
        )
        service = get_product_service()
        with patch.object(
            service,
            "preview_brand_normalization",
            wraps=service.preview_brand_normalization,
        ) as preview_call:
            products_page = self.client.get("/products")
            self.assertEqual(products_page.status_code, 200)
            self.assertIn("检查品牌规范", products_page.get_data(as_text=True))
            self.assertIn(
                'data-products-brand-preview-active="0"',
                products_page.get_data(as_text=True),
            )
            preview_call.assert_not_called()
            products_page = self.client.get("/products", query_string={"brand_preview": "1"})
            preview_call.assert_called_once_with()
        products_html = products_page.get_data(as_text=True)
        self.assertEqual(products_page.status_code, 200)
        self.assertIn(f"规范品牌 {preview.changed_count}", products_html)
        self.assertIn(f'name="snapshot_digest" value="{preview.digest}"', products_html)
        self.assertIn('data-products-brand-preview-active="1"', products_html)

        missing_confirmation = self.client.post(
            "/products/brands/normalize",
            data={"snapshot_digest": preview.digest},
            headers={"Accept": "application/json"},
        )
        self.assertEqual(missing_confirmation.status_code, 400)

        stale_preview = self.client.post(
            "/products/brands/normalize",
            data={
                "confirmation": "normalize-product-brands-v1",
                "snapshot_digest": "0" * 64,
            },
            headers={"Accept": "application/json"},
        )
        self.assertEqual(stale_preview.status_code, 409)
        with self.web.connect(self.web.DB_PATH) as connection:
            unchanged = connection.execute("SELECT series FROM products WHERE bld_no = 'BRAND-WEB-CLEANUP'").fetchone()
        self.assertEqual(unchanged["series"], "DODGE RAM")

        applied_response = self.client.post(
            "/products/brands/normalize",
            data={
                "confirmation": "normalize-product-brands-v1",
                "snapshot_digest": preview.digest,
            },
            headers={"Accept": "application/json"},
        )
        self.assertEqual(applied_response.status_code, 200)
        applied = applied_response.get_json()
        self.assertEqual(applied["changed_count"], preview.changed_count)
        backup_path = self.root / "data" / applied["backup"]
        self.assertTrue(backup_path.is_file())
        with sqlite3.connect(backup_path) as backup:
            self.assertEqual(backup.execute("PRAGMA integrity_check").fetchone()[0], "ok")
            self.assertEqual(
                backup.execute("SELECT series FROM products WHERE bld_no = 'BRAND-WEB-CLEANUP'").fetchone()[0],
                "DODGE RAM",
            )
        with self.web.connect(self.web.DB_PATH) as connection:
            cleaned = connection.execute("SELECT series FROM products WHERE bld_no = 'BRAND-WEB-CLEANUP'").fetchone()
            audit = connection.execute(
                """
                SELECT actor FROM audit_logs
                WHERE action = '清洗产品品牌' AND target_key = 'BRAND-WEB-CLEANUP'
                ORDER BY id DESC LIMIT 1
                """
            ).fetchone()
            connection.execute("DELETE FROM products WHERE bld_no = 'BRAND-WEB-CLEANUP'")
            connection.commit()
        self.assertEqual(cleaned["series"], "DODGE")
        self.assertEqual(audit["actor"], "007")

    def test_product_brand_normalization_returns_stable_operational_error(self):
        self.login()
        with patch("app.modules.products.records_web.get_product_service") as service_factory:
            service_factory.return_value.normalize_brands.side_effect = OSError("private disk path should not leak")
            response = self.client.post(
                "/products/brands/normalize",
                data={
                    "confirmation": "normalize-product-brands-v1",
                    "snapshot_digest": "a" * 64,
                },
                headers={"Accept": "application/json"},
            )
        self.assertEqual(response.status_code, 500)
        payload = response.get_json()
        self.assertEqual(
            payload,
            {"ok": False, "error": "品牌清洗失败，数据未修改，请稍后重试。"},
        )
        self.assertNotIn("private disk path", response.get_data(as_text=True))

    def test_product_brand_preview_stays_hidden_without_import_permission(self):
        from app.modules.admin.persistence import save_user
        from app.modules.products.factory import get_product_service

        with self.web.connect(self.web.DB_PATH) as connection:
            save_user(
                connection,
                {
                    "username": "brand-preview-viewer",
                    "display_name": "Brand Preview Viewer",
                    "password": "viewer-pw",
                    "role": "viewer",
                    "active": "1",
                },
                actor="tester",
            )
            connection.commit()

        self.client.post("/logout")
        self.client.post(
            "/login",
            data={
                "username": "brand-preview-viewer",
                "password": "viewer-pw",
                "next": "/products",
            },
        )
        service = get_product_service()
        with patch.object(
            service,
            "preview_brand_normalization",
            wraps=service.preview_brand_normalization,
        ) as preview_call:
            response = self.client.get("/products", query_string={"brand_preview": "1"})

        self.assertEqual(response.status_code, 200)
        self.assertNotIn("检查品牌规范", response.get_data(as_text=True))
        self.assertNotIn("品牌已规范", response.get_data(as_text=True))
        self.assertIn(
            'data-products-brand-preview-active="0"',
            response.get_data(as_text=True),
        )
        preview_call.assert_not_called()

    def test_product_brand_normalization_respects_global_import_lock(self):
        self.login()
        from app.locks import ImportLockError

        with (
            patch(
                "app.modules.products.records_web.import_lock",
                side_effect=ImportLockError("当前已有用户正在执行导入操作，请稍后再试"),
            ),
            patch("app.modules.products.records_web.get_product_service") as service_factory,
        ):
            response = self.client.post(
                "/products/brands/normalize",
                data={
                    "confirmation": "normalize-product-brands-v1",
                    "snapshot_digest": "a" * 64,
                },
                headers={"Accept": "application/json"},
            )
        self.assertEqual(response.status_code, 409)
        self.assertIn("当前已有用户正在执行导入操作", response.get_json()["error"])
        service_factory.return_value.normalize_brands.assert_not_called()

    def _build_product_sync_package(self, rows: list[dict], *, media: bool = False) -> Path:
        package_path = self.root / "incoming-product-sync.tar.gz"
        work_dir = self.root / "incoming-product-sync"
        if work_dir.exists():
            shutil.rmtree(work_dir)
        if package_path.exists():
            package_path.unlink()
        data_dir = work_dir / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        product_db = data_dir / "products.sqlite3"
        target = sqlite3.connect(product_db)
        try:
            with self.web.connect(self.web.DB_PATH) as source:
                target.row_factory = sqlite3.Row
                schema = source.execute(
                    "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'products'"
                ).fetchone()["sql"]
                target.execute(schema)
                columns = [row["name"] for row in source.execute("PRAGMA table_info(products)")]
                insert_columns = ", ".join(columns)
                placeholders = ", ".join("?" for _ in columns)
                for index, row in enumerate(rows, start=1):
                    values = []
                    for column in columns:
                        if column == "id":
                            values.append(index)
                        else:
                            values.append(row.get(column, "" if column != "price_cny" else None))
                    target.execute(f"INSERT INTO products ({insert_columns}) VALUES ({placeholders})", values)
                target.commit()
        finally:
            target.close()
        (work_dir / "manifest.json").write_text(
            '{"package_type":"bld_product_data","version":1}',
            encoding="utf-8",
        )
        if media:
            image_dir = data_dir / "product_images"
            image_dir.mkdir(parents=True, exist_ok=True)
            (image_dir / "SYNC001.png").write_bytes(b"fake-image")
        with tarfile.open(package_path, "w:gz") as archive:
            archive.add(product_db, arcname="data/products.sqlite3")
            archive.add(work_dir / "manifest.json", arcname="manifest.json")
            if media:
                archive.add(data_dir / "product_images" / "SYNC001.png", arcname="data/product_images/SYNC001.png")
        return package_path

    @unittest.skip("产品数据同步已由业务数据同步取代")
    def test_product_data_sync_exports_products_without_api_keys(self):
        from app.modules.products.persistence import upsert_product

        self.login()
        self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "SYNC-EXPORT",
                    "series": "SYNC",
                    "item": "Export Test",
                    "oe_no_1": "SYNC-EXPORT-OE",
                    "active": "1",
                },
                actor="tester",
            )

        response = self.client.post("/product-data-sync/export")
        self.assertEqual(response.status_code, 200)
        package_path = self.root / "exported-product-data.tar.gz"
        package_path.write_bytes(response.data)
        response.close()
        with tarfile.open(package_path, "r:gz") as archive:
            names = set(archive.getnames())
            self.assertIn("data/products.sqlite3", names)
            self.assertIn("manifest.json", names)
            archive.extract("data/products.sqlite3", self.root / "export-check", filter="data")

        exported_db = self.root / "export-check" / "data" / "products.sqlite3"
        conn = sqlite3.connect(exported_db)
        try:
            tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type = 'table'")}
            count = conn.execute("SELECT COUNT(*) FROM products WHERE bld_no = 'SYNC-EXPORT'").fetchone()[0]
        finally:
            conn.close()
        self.assertEqual(tables, {"products", "sqlite_sequence"})
        self.assertEqual(count, 1)

    @unittest.skip("产品数据同步已由业务数据同步取代")
    def test_product_data_sync_imports_incrementally_and_preserves_api_key(self):
        from app.modules.products.persistence import upsert_product
        from app.platform.api_keys import internal_api_key_status

        self.login()
        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "SYNC001",
                    "series": "LOCAL",
                    "item": "Local Item",
                    "oe_no_1": "OLD-OE",
                    "price_cny": "10",
                    "active": "1",
                },
                actor="tester",
            )

        package_path = self._build_product_sync_package(
            [
                {
                    "bld_no": "SYNC001",
                    "series": "NAS",
                    "item": "NAS Item",
                    "oe_no_1": "NEW-OE",
                    "price_cny": 12.5,
                    "image_path": "data_product_images/SYNC001.png",
                    "active": 1,
                    "source": "nas",
                    "created_at": "2026-05-01 00:00:00",
                    "updated_at": "2099-05-27 10:00:00",
                },
                {
                    "bld_no": "SYNC002",
                    "series": "NAS",
                    "item": "New Product",
                    "oe_no_1": "NEW-ONLY",
                    "price_cny": 20,
                    "active": 1,
                    "source": "nas",
                    "created_at": "2026-05-27 10:00:00",
                    "updated_at": "2099-05-27 10:00:00",
                },
            ],
            media=True,
        )
        response = self.client.post(
            "/product-data-sync/import/preview",
            data={
                "include_images": "1",
                "package": (package_path.open("rb"), package_path.name),
            },
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("新增产品", html)
        self.assertIn("SYNC001", html)
        self.assertIn("SYNC002", html)
        path_match = re.search(r'name="package_path" value="([^"]+)"', html)
        self.assertIsNotNone(path_match)

        response = self.client.post(
            "/product-data-sync/import/apply",
            data={"include_images": "1", "package_path": path_match.group(1)},
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("产品数据导入完成：新增 1 条，更新 1 条", html)
        with self.web.connect(self.web.DB_PATH) as conn:
            row = conn.execute("SELECT * FROM products WHERE bld_no = 'SYNC001'").fetchone()
            new_row = conn.execute("SELECT * FROM products WHERE bld_no = 'SYNC002'").fetchone()
            status = internal_api_key_status(conn)
        self.assertEqual(row["series"], "NAS")
        self.assertEqual(row["oe_no_1"], "NEW-OE")
        self.assertEqual(new_row["item"], "New Product")
        self.assertTrue(status["enabled"])
        self.assertTrue(token.endswith(status["preview"][-6:]))
        self.assertTrue((self.root / "data" / "product_images" / "SYNC001.png").exists())

    @unittest.skip("产品数据同步已由业务数据同步取代")
    def test_product_data_sync_rolls_back_media_when_database_apply_fails(self):
        self.login()
        target = self.root / "data" / "product_images" / "SYNC001.png"
        target.parent.mkdir(parents=True, exist_ok=True)
        target.write_bytes(b"original-image")
        package_path = self._build_product_sync_package(
            [
                {
                    "bld_no": "SYNC-ROLLBACK",
                    "series": "SYNC",
                    "item": "Rollback Test",
                    "image_path": "data_product_images/SYNC001.png",
                    "active": 1,
                    "source": "test",
                    "created_at": "2026-07-11 00:00:00",
                    "updated_at": "2099-07-11 00:00:00",
                }
            ],
            media=True,
        )
        preview = self.client.post(
            "/product-data-sync/import/preview",
            data={"include_images": "1", "package": (package_path.open("rb"), package_path.name)},
            content_type="multipart/form-data",
        )
        path_match = re.search(r'name="package_path" value="([^"]+)"', preview.get_data(as_text=True))
        self.assertIsNotNone(path_match)

        with patch(
            "app.modules.products.sync_repository.SQLiteProductSyncRepository.apply",
            side_effect=RuntimeError("forced apply failure"),
        ):
            response = self.client.post(
                "/product-data-sync/import/apply",
                data={"include_images": "1", "package_path": path_match.group(1)},
                follow_redirects=True,
            )
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("forced apply failure", html)
        self.assertIn("产品数据包导入失败，请稍后重试", html)
        self.assertIn("已恢复本次媒体文件变更", html)
        self.assertEqual(target.read_bytes(), b"original-image")
        backups = sorted((self.root / "data" / "local-backups").glob("*/products.sqlite3"))
        self.assertTrue(backups)
        with sqlite3.connect(backups[-1]) as backup:
            self.assertEqual(backup.execute("PRAGMA integrity_check").fetchone()[0], "ok")

    @unittest.skip("产品数据同步已由业务数据同步取代")
    def test_product_data_sync_skips_older_package_rows(self):
        from app.modules.products.persistence import upsert_product

        self.login()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "SYNC-STALE",
                    "series": "LOCAL",
                    "item": "Local Newer",
                    "oe_no_1": "LOCAL-OE",
                    "price_cny": "30",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "SYNC-BAD-TIME",
                    "series": "LOCAL-VALID-TIME",
                    "item": "Local Protected",
                    "oe_no_1": "LOCAL-VALID-OE",
                    "price_cny": "31",
                    "active": "1",
                },
                actor="tester",
            )
            conn.execute(
                "UPDATE products SET updated_at = ? WHERE bld_no = ?",
                ("2099-01-01 00:00:00", "SYNC-STALE"),
            )
            conn.commit()

        package_path = self._build_product_sync_package(
            [
                {
                    "bld_no": "SYNC-STALE",
                    "series": "OLD-PACKAGE",
                    "item": "Should Not Overwrite",
                    "oe_no_1": "OLD-OE",
                    "price_cny": 1,
                    "active": 1,
                    "source": "old-package",
                    "created_at": "2020-01-01 00:00:00",
                    "updated_at": "2020-01-01 00:00:00",
                },
                {
                    "bld_no": "SYNC-BAD-TIME",
                    "series": "INVALID-PACKAGE-TIME",
                    "item": "Must Not Overwrite",
                    "oe_no_1": "INVALID-TIME-OE",
                    "price_cny": 2,
                    "active": 1,
                    "source": "bad-package-time",
                    "created_at": "2020-01-01 00:00:00",
                    "updated_at": "not-a-timestamp",
                },
            ],
        )
        preview = self.client.post(
            "/product-data-sync/import/preview",
            data={"package": (package_path.open("rb"), package_path.name)},
            content_type="multipart/form-data",
        )
        preview_html = preview.get_data(as_text=True)
        self.assertEqual(preview.status_code, 200)
        self.assertIn("包内旧数据", preview_html)
        path_match = re.search(r'name="package_path" value="([^"]+)"', preview_html)
        self.assertIsNotNone(path_match)

        applied = self.client.post(
            "/product-data-sync/import/apply",
            data={"package_path": path_match.group(1)},
            follow_redirects=True,
        )
        applied_html = applied.get_data(as_text=True)
        self.assertEqual(applied.status_code, 200)
        self.assertIn("跳过包内旧数据 2 条", applied_html)
        with self.web.connect(self.web.DB_PATH) as conn:
            row = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("SYNC-STALE",)).fetchone()
            invalid_time_row = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("SYNC-BAD-TIME",)).fetchone()
        self.assertEqual(row["series"], "LOCAL")
        self.assertEqual(row["oe_no_1"], "LOCAL-OE")
        self.assertEqual(invalid_time_row["series"], "LOCAL-VALID-TIME")
        self.assertEqual(invalid_time_row["oe_no_1"], "LOCAL-VALID-OE")

    @unittest.skip("产品数据同步已由业务数据同步取代")
    def test_product_data_sync_can_deactivate_local_only_rows_after_preview(self):
        from app.modules.products.persistence import upsert_product

        self.login()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "SYNC-LOCAL-ONLY",
                    "series": "LOCAL",
                    "item": "Only On This Machine",
                    "oe_no_1": "LOCAL-ONLY-OE",
                    "price_cny": "30",
                    "active": "1",
                },
                actor="tester",
            )

        package_path = self._build_product_sync_package(
            [
                {
                    "bld_no": "SYNC-PACKAGE-ONLY",
                    "series": "NAS",
                    "item": "Package Product",
                    "oe_no_1": "PACKAGE-OE",
                    "price_cny": 20,
                    "active": 1,
                    "source": "nas",
                    "created_at": "2026-06-01 00:00:00",
                    "updated_at": "2099-06-01 00:00:00",
                }
            ],
        )
        preview = self.client.post(
            "/product-data-sync/import/preview",
            data={"package": (package_path.open("rb"), package_path.name)},
            content_type="multipart/form-data",
        )
        html = preview.get_data(as_text=True)
        self.assertEqual(preview.status_code, 200)
        self.assertIn("当前系统独有", html)
        self.assertIn("SYNC-LOCAL-ONLY", html)
        self.assertIn("deactivate_local_only", html)
        path_match = re.search(r'name="package_path" value="([^"]+)"', html)
        self.assertIsNotNone(path_match)

        response = self.client.post(
            "/product-data-sync/import/apply",
            data={"package_path": path_match.group(1)},
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("停用本机独有 0 条", html)
        with self.web.connect(self.web.DB_PATH) as conn:
            row = conn.execute("SELECT active FROM products WHERE bld_no = ?", ("SYNC-LOCAL-ONLY",)).fetchone()
        self.assertEqual(row["active"], 1)

        response = self.client.post(
            "/product-data-sync/import/apply",
            data={"package_path": path_match.group(1), "deactivate_local_only": "1"},
            follow_redirects=True,
        )
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("停用本机独有", html)
        with self.web.connect(self.web.DB_PATH) as conn:
            row = conn.execute("SELECT active FROM products WHERE bld_no = ?", ("SYNC-LOCAL-ONLY",)).fetchone()
        self.assertEqual(row["active"], 0)

    def test_catalog_import_template_and_conflict_preview_require_explicit_update_choice(self):
        from openpyxl import Workbook, load_workbook
        from app.modules.products.persistence import upsert_product

        self.login()
        with self.web.connect(self.web.DB_PATH) as connection:
            upsert_product(
                connection,
                {
                    "bld_no": "WEB-CATALOG-CONFLICT",
                    "series": "TOYOTA",
                    "item": "Old item",
                    "oe_no_1": "OLD-OE",
                    "models": "CAMRY",
                    "price_cny": "10",
                    "product_status": "1 个球头",
                },
                actor="test",
            )
            upsert_product(
                connection,
                {
                    "bld_no": "WEB-CATALOG-ITEM-CHOICE",
                    "series": "TOYOTA",
                    "item": "New item",
                    "oe_no_1": "NEW-ITEM-OE",
                    "models": "CAMRY",
                    "price_cny": "10",
                    "product_status": "1 个球头",
                },
                actor="test",
            )
        response = self.client.get("/catalog/template")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertIn("filename*=UTF-8''", response.headers["Content-Disposition"])
        template = load_workbook(io.BytesIO(response.data))
        try:
            sheet = template["产品目录"]
            self.assertEqual(sheet.cell(1, 2).value, "SERIES")
            self.assertEqual(sheet.cell(1, 7).value, "SERIES 6")
            self.assertEqual(sheet.cell(1, 8).value, "ITEM")
            self.assertEqual(len(sheet.data_validations.dataValidation), 7)
        finally:
            template.close()
        response.close()
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["BLD NO.", "SERIES", "ITEM", "OE NO.1", "Models", "产品状态", "导入单价"])
        sheet.append(["WEB-CATALOG-CONFLICT", "TOYOTA", "New item", "NEW-OE", "CAMRY", "2 个球头", 88])
        stream = io.BytesIO()
        workbook.save(stream)
        workbook.close()
        stream.seek(0)

        preview_response = self.client.post(
            "/catalog",
            data={"next": "products", "catalog": (stream, "catalog.xlsx")},
            content_type="multipart/form-data",
        )
        html = preview_response.get_data(as_text=True)
        self.assertEqual(preview_response.status_code, 200)
        self.assertIn("确认产品目录导入", html)
        self.assertIn("使用 Excel 更新", html)
        self.assertIn('name="update_bld" value="WEB-CATALOG-CONFLICT"', html)
        preview_id = re.search(r'name="preview_id" value="([^"]+)"', html).group(1)
        digest = re.search(r'name="snapshot_digest" value="([^"]+)"', html).group(1)

        kept = self.client.post(
            "/catalog/confirm",
            data={"preview_id": preview_id, "snapshot_digest": digest},
            follow_redirects=False,
        )
        self.assertEqual(kept.status_code, 302)
        with self.web.connect(self.web.DB_PATH) as connection:
            product = connection.execute(
                "SELECT item FROM products WHERE bld_no = ?",
                ("WEB-CATALOG-CONFLICT",),
            ).fetchone()
            connection.execute("DELETE FROM products WHERE bld_no = ?", ("WEB-CATALOG-CONFLICT",))
            connection.execute("DELETE FROM products WHERE bld_no = ?", ("WEB-CATALOG-ITEM-CHOICE",))
            connection.commit()
        self.assertEqual(product["item"], "Old item")
        (self.root / "data" / "catalog.xlsx").unlink(missing_ok=True)

    def test_catalog_import_recognizes_chinese_brand_number_header(self):
        from app.matcher import ProductCatalog
        from openpyxl import Workbook

        for header in ["品牌号码", "Other Reference"]:
            with self.subTest(header=header):
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["BLD NO.", "品牌", "产品名称", "OE Reference", header, "车型"])
                sheet.append(["K6004CN", "HYUNDAI", "CONTROL ARM", "55270-2Z020", "BRAND-CN-55270", "Sportage"])
                catalog_path = self.root / f"catalog-brand-number-{header.replace(' ', '-').lower()}.xlsx"
                workbook.save(catalog_path)

                catalog = ProductCatalog.from_excel(catalog_path)
                match = catalog.match("", "BRAND-CN-55270")

                self.assertIsNotNone(match)
                self.assertEqual(match.bld_no, "K6004CN")
                self.assertEqual(match.reason, "品牌号码精准命中")

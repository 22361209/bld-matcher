from __future__ import annotations

from tests.web_app_test_base import (
    WebAppTestBase,
    Path,
    json,
)


class TestWebApi(WebAppTestBase):
    def test_internal_api_numbers_generate_openclaw_workbook(self):
        from app.modules.products.persistence import upsert_product
        from openpyxl import load_workbook

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-001",
                    "series": "HYUNDAI",
                    "item": "API CONTROL ARM",
                    "oe_no_1": "API-001",
                    "models": "ApiTester",
                    "price_cny": "88.8",
                    "active": "1",
                },
                actor="tester",
            )

        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={
                "numbers": ["API-001", "NO-MATCH-001"],
                "source_name": "机器人询价结果",
                "price_mode": "net",
                "rows_limit": 10,
                "export": True,
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["mode"], "new-workbook")
        self.assertEqual(payload["matched_count"], 1)
        self.assertEqual(payload["unmatched_count"], 1)
        self.assertEqual(payload["rows"][0]["bld_no"], "K-API-001")
        self.assertEqual(payload["rows"][0]["export_price"], 81)
        self.assertIn("NO-MATCH-001", payload["unmatched_list"])
        self.assertTrue(payload["output_path"].endswith(payload["output_name"]))
        self.assertRegex(payload["output_name"], r"^re\d{6}_机器人询价结果\.xlsx$")
        output_path = Path(payload["output_path"])
        self.assertEqual(output_path.parent.resolve(), (self.root / "outputs" / "openclaw").resolve())
        self.assertTrue(output_path.exists())

        workbook = load_workbook(output_path)
        sheet = workbook.active
        self.assertEqual(sheet.cell(1, 1).value, "OE号")
        self.assertEqual(sheet.cell(1, 2).value, "BLD NO.")
        self.assertEqual(sheet.cell(1, 3).value, "不含税单价")
        self.assertEqual(sheet.cell(2, 2).value, "K-API-001")
        self.assertEqual(sheet.cell(2, 3).value, 81)
        workbook.close()

        rejected_export = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["API-001"], "export": True},
            headers={"Authorization": f"Bearer {token}"},
        )
        rejected_payload = rejected_export.get_json()
        self.assertEqual(rejected_export.status_code, 400)
        self.assertFalse(rejected_payload["ok"])
        self.assertIn("必须传 source_name", rejected_payload["error"])

    def test_internal_api_numbers_use_oe_suffix_variant_matching(self):
        from app.modules.products.persistence import upsert_product

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8041LB",
                    "series": "VW",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "561407151A\n561407151C",
                    "active": "1",
                },
                actor="tester",
            )

        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["561407151D"], "rows_limit": 10},
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["matched_count"], 1)
        self.assertEqual(payload["unmatched_count"], 0)
        self.assertEqual(payload["rows"][0]["bld_no"], "K8041LB")
        self.assertEqual(payload["rows"][0]["match_reason"], "OE 尾字母容错命中")

    def test_internal_api_numbers_use_bld_fragment_lookup(self):
        from app.modules.products.persistence import upsert_product
        from openpyxl import load_workbook

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8072LA",
                    "series": "NISSAN",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "54501-TEST-LA",
                    "oe_no_2": "Moog: TEST-LA",
                    "models": "VERSA TEST",
                    "price_cny": "43",
                    "image_path": "data_product_images/K8072LA.png",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K8072RA",
                    "series": "NISSAN",
                    "item": "Front Right Lower Control Arm",
                    "price_cny": "43",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-OE-ONLY",
                    "series": "TEST",
                    "item": "OE SHOULD NOT WIN BLD SHORTHAND",
                    "oe_no_1": "8072",
                    "active": "1",
                },
                actor="tester",
            )

        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["8072"], "rows_limit": 10},
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["matched_count"], 2)
        self.assertEqual(payload["unmatched_count"], 0)
        self.assertEqual([row["bld_no"] for row in payload["rows"]], ["K8072LA", "K8072RA"])
        self.assertEqual(payload["rows"][0]["original_number"], "8072")
        self.assertEqual(payload["rows"][0]["match_reason"], "BLD NO. 片段命中")
        self.assertEqual(payload["rows"][0]["price_cny"], 43.0)
        self.assertEqual(payload["rows"][0]["product"]["item"], "Front Left Lower Control Arm")
        self.assertEqual(payload["rows"][0]["product"]["oe_no_1"], "54501-TEST-LA")
        self.assertEqual(payload["rows"][0]["product"]["oe_no_2"], "Moog: TEST-LA")
        self.assertEqual(payload["rows"][0]["product"]["models"], "VERSA TEST")
        self.assertEqual(payload["rows"][0]["product"]["image_paths"], ["data_product_images/K8072LA.png"])

        k_response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["K8072"], "rows_limit": 10},
            headers={"Authorization": f"Bearer {token}"},
        )
        k_payload = k_response.get_json()
        self.assertEqual(k_response.status_code, 200)
        self.assertEqual(k_payload["matched_count"], 2)
        self.assertEqual([row["bld_no"] for row in k_payload["rows"]], ["K8072LA", "K8072RA"])

        export_response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["8072"], "source_name": "片段查询", "export": True, "rows_limit": 10},
            headers={"Authorization": f"Bearer {token}"},
        )
        export_payload = export_response.get_json()
        self.assertEqual(export_response.status_code, 200)
        self.assertEqual(export_payload["matched_count"], 2)
        self.assertEqual([row["bld_no"] for row in export_payload["rows"]], ["K8072LA", "K8072RA"])
        generated = load_workbook(Path(export_payload["output_path"]), read_only=True, data_only=True)
        try:
            sheet = generated.active
            self.assertEqual(sheet.cell(2, 1).value, "8072")
            self.assertEqual(sheet.cell(2, 2).value, "K8072LA")
            self.assertEqual(sheet.cell(3, 1).value, "8072")
            self.assertEqual(sheet.cell(3, 2).value, "K8072RA")
        finally:
            generated.close()

    def test_internal_api_numbers_use_psa_352x_dot_rule(self):
        from app.modules.products.persistence import upsert_product

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-PSA-352126",
                    "series": "PEUGEOT\nCITROEN",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "3521.26",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-GM-352126",
                    "series": "GM\nOPEL",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "352126",
                    "active": "1",
                },
                actor="tester",
            )

        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["3521.26"], "rows_limit": 10},
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["matched_count"], 1)
        self.assertEqual(payload["rows"][0]["bld_no"], "K-API-PSA-352126")
        self.assertEqual(payload["rows"][0]["match_reason"], "PSA 号码点号容错命中")

    def test_internal_api_file_augment_and_analyze(self):
        from app.modules.products.persistence import upsert_product
        from openpyxl import Workbook, load_workbook

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-FILE",
                    "series": "HYUNDAI",
                    "item": "API FILE ARM",
                    "oe_no_1": "API-FILE-OE",
                    "models": "ApiFileTester",
                    "price_cny": "79.2",
                    "active": "1",
                },
                actor="tester",
            )

        source_path = self.root / "uploads" / "api-file-source.xlsx"
        source_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["编号", "数量"])
        sheet.append(["API-FILE-OE", 2])
        workbook.save(source_path)
        workbook.close()

        analyze = self.client.post(
            "/api/internal/inquiry/analyze",
            json={"file_path": str(source_path), "match_column": "A", "price_mode": "tax"},
            headers={"Authorization": f"Bearer {token}"},
        )
        analyze_payload = analyze.get_json()
        self.assertEqual(analyze.status_code, 200)
        self.assertTrue(analyze_payload["ok"])
        self.assertEqual(analyze_payload["mode"], "augment-source-workbook")
        self.assertEqual(analyze_payload["summary"]["output_generated"], False)
        self.assertIsNone(analyze_payload["output_path"])
        self.assertEqual(analyze_payload["matched_count"], 1)

        response = self.client.post(
            "/api/internal/inquiry/file",
            json={
                "file_path": str(source_path),
                "match_column": "A",
                "price_mode": "usd",
                "exchange_rate": "7.2",
                "export": True,
            },
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["mode"], "augment-source-workbook")
        self.assertEqual(payload["rows"][0]["export_price"], 10)
        output_path = Path(payload["output_path"])
        self.assertEqual(output_path.parent.resolve(), (self.root / "outputs" / "openclaw").resolve())
        self.assertRegex(payload["output_name"], r"^re\d{6}_api-file-source\.xlsx$")
        self.assertTrue(output_path.exists())

        generated = load_workbook(output_path)
        sheet = generated.active
        self.assertEqual(sheet.cell(1, 3).value, "BLD NO.")
        self.assertEqual(sheet.cell(1, 4).value, "美金价")
        self.assertEqual(sheet.cell(1, 5).value, "Product Status")
        self.assertEqual(sheet.cell(1, 6).value, "匹配说明")
        self.assertEqual(sheet.cell(2, 3).value, "K-API-FILE")
        self.assertEqual(sheet.cell(2, 4).value, 10)
        generated.close()

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-FILE-MULTI",
                    "series": "HYUNDAI",
                    "item": "API MULTI ARM",
                    "oe_no_1": "API-FILE-REF",
                    "active": "1",
                },
                actor="tester",
            )
        multi_source = self.root / "uploads" / "api-file-multi-source.xlsx"
        multi_workbook = Workbook()
        multi_sheet = multi_workbook.active
        multi_sheet.append(["客户号码", "参考号"])
        multi_sheet.append(["NO-HIT-API", "API-FILE-REF"])
        multi_workbook.save(multi_source)
        multi_workbook.close()

        multi_response = self.client.post(
            "/api/internal/inquiry/analyze",
            json={"file_path": str(multi_source), "match_columns": ["A", "B"], "rows_limit": 10},
            headers={"Authorization": f"Bearer {token}"},
        )
        multi_payload = multi_response.get_json()
        self.assertEqual(multi_response.status_code, 200)
        self.assertEqual(multi_payload["matched_count"], 1)
        self.assertEqual(multi_payload["rows"][0]["bld_no"], "K-API-FILE-MULTI")
        self.assertIn("B列：API-FILE-REF", multi_payload["rows"][0]["match_note"])

    def test_internal_api_defaults_to_analysis_and_restricts_file_path(self):
        from app.modules.products.persistence import upsert_product
        from openpyxl import Workbook

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-DEFAULT",
                    "oe_no_1": "API-DEFAULT-OE",
                    "active": "1",
                },
                actor="tester",
            )

        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["API-DEFAULT-OE"]},
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["matched_count"], 1)
        self.assertFalse(payload["summary"]["output_generated"])
        self.assertIsNone(payload["output_path"])
        openclaw_upload_dir = self.root / "uploads" / "openclaw"
        before_uploads = set(openclaw_upload_dir.glob("*")) if openclaw_upload_dir.exists() else set()

        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["API-DEFAULT-OE"], "export": False},
            headers={"Authorization": f"Bearer {token}"},
        )
        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(payload["matched_count"], 1)
        self.assertFalse(payload["summary"]["output_generated"])
        self.assertIsNone(payload["output_path"])
        after_uploads = set(openclaw_upload_dir.glob("*")) if openclaw_upload_dir.exists() else set()
        self.assertEqual(after_uploads, before_uploads)

        outside_path = self.root / "outside-api-source.xlsx"
        workbook = Workbook()
        workbook.active.append(["OE号"])
        workbook.active.append(["API-DEFAULT-OE"])
        workbook.save(outside_path)
        workbook.close()

        rejected = self.client.post(
            "/api/internal/inquiry/file",
            json={"file_path": str(outside_path), "export": True},
            headers={"Authorization": f"Bearer {token}"},
        )
        rejected_payload = rejected.get_json()
        self.assertEqual(rejected.status_code, 400)
        self.assertFalse(rejected_payload["ok"])
        self.assertIn("file_path 不在允许读取范围内", rejected_payload["error"])

    def test_internal_api_requires_api_key(self):
        response = self.client.post(
            "/api/internal/inquiry/numbers",
            json={"numbers": ["API-001"]},
        )
        self.assertEqual(response.status_code, 401)
        self.assertFalse(response.get_json()["ok"])

        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-API-AUTH",
                    "oe_no_1": "API-AUTH-OE",
                    "active": "1",
                },
                actor="tester",
            )
        token = self.create_internal_api_token()
        response = self.client.post(
            "/api/internal/inquiry/analyze",
            json={"numbers": ["API-AUTH-OE"]},
            headers={"Authorization": f"Bearer {token}"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.get_json()["ok"])

    def test_v1_api_index_and_openapi_use_stable_contract(self):
        unauthorized = self.client.get("/api/v1", headers={"X-Request-ID": "v1-unauthorized-1"})
        unauthorized_payload = unauthorized.get_json()
        self.assertEqual(unauthorized.status_code, 401)
        self.assertEqual(unauthorized_payload["api_version"], "1")
        self.assertEqual(unauthorized_payload["request_id"], "v1-unauthorized-1")
        self.assertEqual(unauthorized_payload["error"]["code"], "auth.unauthorized")

        token = self.create_internal_api_token()
        headers = {"Authorization": f"Bearer {token}", "X-Request-ID": "v1-index-1"}
        response = self.client.get("/api/v1", headers=headers)
        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["X-Request-ID"], "v1-index-1")
        self.assertEqual(response.headers["Cache-Control"], "no-store")
        self.assertEqual(payload["api_version"], "1")
        self.assertEqual(payload["request_id"], "v1-index-1")
        self.assertEqual(payload["data"]["name"], "bld-matcher")
        self.assertIn("openapi", payload["data"]["capabilities"])

        document_response = self.client.get("/api/v1/openapi.json", headers=headers)
        document = document_response.get_json()
        self.assertEqual(document_response.status_code, 200)
        self.assertEqual(document["openapi"], "3.1.0")
        self.assertIn("/api/v1", document["paths"])
        self.assertIn("/api/v1/openapi.json", document["paths"])
        self.assertEqual(
            document["paths"]["/api/v1"]["get"]["x-required-scopes"],
            ["api:read"],
        )
        self.assertIn("PlatformInfoEnvelope", document["components"]["schemas"])
        self.assertNotIn("JobPublicData", document["components"]["schemas"])
        self.assertNotIn("/api/v1/jobs/{job_id}/result", document["paths"])

    def test_legacy_quote_api_routes_are_removed(self):
        token = self.create_internal_api_token()
        headers = {"Authorization": f"Bearer {token}"}
        self.assertEqual(self.client.get("/api/quotes", headers=headers).status_code, 404)
        self.assertEqual(
            self.client.post("/api/quotes", json={"customer_name": "ACME"}, headers=headers).status_code,
            404,
        )
        self.assertEqual(self.client.get("/api/quotes/latest", headers=headers).status_code, 404)
        self.assertEqual(
            self.client.put("/api/quotes/1", json={"remark": "x"}, headers=headers).status_code,
            404,
        )

    def test_product_inquiry_v1_and_artifact_consumer_contract(self):
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "V1-INQUIRY-001",
                    "series": "CONTRACT",
                    "item": "Consumer Contract Arm",
                    "oe_no_1": "V1-OE-001",
                    "models": "Contract Model",
                    "price_cny": "110",
                    "active": "1",
                },
                actor="tester",
            )

        token = self.create_internal_api_token(
            scopes=["products:read", "inquiries:run", "artifacts:read"],
            name="WorkBuddy Inquiry V1",
        )
        authorization = {"Authorization": f"Bearer {token}"}

        product_response = self.client.get(
            "/api/v1/products/search",
            query_string={"oe": "V1-OE-001", "limit": 10},
            headers=authorization,
        )
        product_payload = product_response.get_json()
        self.assertEqual(product_response.status_code, 200)
        self.assertEqual(product_payload["data"]["total"], 1)
        self.assertEqual(product_payload["data"]["products"][0]["bld_no"], "V1-INQUIRY-001")

        missing_idempotency = self.client.post(
            "/api/v1/inquiries/analyze",
            json={"numbers": ["V1-OE-001"]},
            headers=authorization,
        )
        self.assertEqual(missing_idempotency.status_code, 400)
        self.assertEqual(missing_idempotency.get_json()["error"]["code"], "idempotency.required")

        analyze_response = self.client.post(
            "/api/v1/inquiries/analyze",
            json={"numbers": ["V1-OE-001", "V1-MISSING"], "price_mode": "net"},
            headers={**authorization, "Idempotency-Key": "inquiry-analyze-v1-001"},
        )
        analyze_payload = analyze_response.get_json()
        self.assertEqual(analyze_response.status_code, 200)
        self.assertEqual(analyze_payload["data"]["summary"]["matched_count"], 1)
        self.assertEqual(analyze_payload["data"]["rows"][0]["bld_no"], "V1-INQUIRY-001")
        self.assertEqual(analyze_payload["data"]["rows"][0]["export_price"], 100)
        self.assertIsNone(analyze_payload["data"]["artifact"])

        legacy_response = self.client.post(
            "/api/internal/inquiry/analyze",
            json={"numbers": ["V1-OE-001"]},
            headers=authorization,
        )
        self.assertEqual(legacy_response.status_code, 200)
        self.assertEqual(legacy_response.get_json()["rows"][0]["bld_no"], "V1-INQUIRY-001")

        export_body = {
            "numbers": ["V1-OE-001"],
            "source_name": "consumer-contract",
            "price_mode": "tax",
        }
        export_headers = {**authorization, "Idempotency-Key": "inquiry-export-v1-001"}
        export_response = self.client.post(
            "/api/v1/inquiries/export",
            json=export_body,
            headers=export_headers,
        )
        export_payload = export_response.get_json()
        self.assertEqual(export_response.status_code, 201)
        artifact = export_payload["data"]["artifact"]
        self.assertTrue(artifact["id"].startswith("art_"))
        self.assertRegex(artifact["filename"], r"^re\d{6}_consumer-contract\.xlsx$")
        self.assertNotIn("output_path", json.dumps(export_payload, ensure_ascii=False))
        self.assertNotIn(str(self.root), json.dumps(export_payload, ensure_ascii=False))

        replay = self.client.post(
            "/api/v1/inquiries/export",
            json=export_body,
            headers=export_headers,
        )
        self.assertEqual(replay.status_code, 201)
        self.assertEqual(replay.headers["Idempotency-Replayed"], "true")
        self.assertEqual(replay.get_json()["data"]["artifact"]["id"], artifact["id"])

        download = self.client.get(artifact["download_url"], headers=authorization)
        self.assertEqual(download.status_code, 200)
        self.assertTrue(download.data.startswith(b"PK"))
        self.assertIn("attachment", download.headers["Content-Disposition"])
        self.assertEqual(download.headers["Cache-Control"], "private, no-store")
        download.close()

        other_token = self.create_internal_api_token(
            scopes=["artifacts:read"],
            name="Other Artifact Consumer",
        )
        denied = self.client.get(
            artifact["download_url"],
            headers={"Authorization": f"Bearer {other_token}"},
        )
        self.assertEqual(denied.status_code, 404)
        self.assertEqual(denied.get_json()["error"]["code"], "artifact.not_found")

        with self.web.connect(self.web.DB_PATH) as conn:
            stored = conn.execute(
                "SELECT owner_id, storage_path, sha256 FROM api_artifacts WHERE id = ?",
                (artifact["id"],),
            ).fetchone()
            audit = conn.execute(
                "SELECT actor FROM audit_logs WHERE action = ? ORDER BY id DESC LIMIT 1",
                ("内部 API 生成号码结果",),
            ).fetchone()
        self.assertIsNotNone(stored)
        self.assertTrue(Path(stored["storage_path"]).resolve().is_relative_to((self.root / "outputs").resolve()))
        self.assertEqual(len(stored["sha256"]), 64)
        self.assertEqual(audit["actor"], "WorkBuddy Inquiry V1")

    def test_product_price_v1_requires_write_scope_and_prevents_stale_updates(self):
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "V1-PRICE-001",
                    "series": "CONTRACT",
                    "item": "Price API Arm",
                    "oe_no_1": "V1-PRICE-OE-001",
                    "models": "Price API Model",
                    "price_cny": "110",
                    "active": "1",
                },
                actor="tester",
            )

        read_token = self.create_internal_api_token(
            scopes=["products:read"],
            name="Product Price Read Only",
        )
        write_token = self.create_internal_api_token(
            scopes=["products:read", "products:write"],
            name="Product Price Writer",
        )
        read_headers = {"Authorization": f"Bearer {read_token}"}
        write_headers = {
            "Authorization": f"Bearer {write_token}",
            "Idempotency-Key": "product-price-v1-001",
        }
        search = self.client.get(
            "/api/v1/products/search",
            query_string={"bld": "V1-PRICE-001", "limit": 10},
            headers=read_headers,
        )
        self.assertEqual(search.status_code, 200)
        product = search.get_json()["data"]["products"][0]
        body = {"price_cny": 128.5, "expected_updated_at": product["updated_at"]}

        denied = self.client.post(
            f"/api/v1/products/{product['id']}/price",
            json=body,
            headers={**read_headers, "Idempotency-Key": "product-price-v1-denied"},
        )
        self.assertEqual(denied.status_code, 403)
        self.assertEqual(denied.get_json()["error"]["code"], "auth.insufficient_scope")

        updated = self.client.post(
            f"/api/v1/products/{product['id']}/price",
            json=body,
            headers=write_headers,
        )
        self.assertEqual(updated.status_code, 200)
        self.assertEqual(updated.get_json()["data"]["product"]["price_cny"], 128.5)

        replayed = self.client.post(
            f"/api/v1/products/{product['id']}/price",
            json=body,
            headers=write_headers,
        )
        self.assertEqual(replayed.status_code, 200)
        self.assertEqual(replayed.headers["Idempotency-Replayed"], "true")

        stale = self.client.post(
            f"/api/v1/products/{product['id']}/price",
            json={"price_cny": 130, "expected_updated_at": "2000-01-01 00:00:00"},
            headers={
                "Authorization": f"Bearer {write_token}",
                "Idempotency-Key": "product-price-v1-stale",
            },
        )
        self.assertEqual(stale.status_code, 412)
        self.assertEqual(stale.get_json()["error"]["code"], "product.version_conflict")

        with self.web.connect(self.web.DB_PATH) as conn:
            audit = conn.execute(
                "SELECT actor FROM audit_logs WHERE action = ? ORDER BY id DESC LIMIT 1",
                ("API 更新产品单价",),
            ).fetchone()
        self.assertIsNotNone(audit)
        self.assertEqual(audit["actor"], "Product Price Writer")

    def test_quote_v1_contract_idempotency_and_optimistic_concurrency(self):
        self.register_customer_and_product("V1 Contract Customer", "V1-BLD-001")
        token = self.create_internal_api_token(
            scopes=["api:read", "quotes:read", "quotes:write"],
            name="WorkBuddy Quote V1",
        )
        authorization = {"Authorization": f"Bearer {token}"}
        create_headers = {**authorization, "Idempotency-Key": "quote-create-v1-001"}
        invalid = self.client.post(
            "/api/v1/quotes",
            json={
                "customer_name": "V1 Contract Customer",
                "bld_no": "V1-BLD-001",
                "tax_price": "12.34",
                "currency": "USD",
                "attachment_path": "/tmp/private-quote.pdf",
            },
            headers=create_headers,
        )
        self.assertEqual(invalid.status_code, 422)
        self.assertEqual(invalid.get_json()["error"]["code"], "request.invalid")
        self.assertNotIn("private-quote.pdf", invalid.get_data(as_text=True))

        legacy_price_only = self.client.post(
            "/api/v1/quotes",
            json={
                "customer_name": "V1 Contract Customer",
                "bld_no": "V1-BLD-001",
                "price": "12.34",
                "currency": "USD",
            },
            headers={**authorization, "Idempotency-Key": "quote-create-v1-legacy-price"},
        )
        self.assertEqual(legacy_price_only.status_code, 422)
        self.assertEqual(legacy_price_only.get_json()["error"]["code"], "request.invalid")

        missing_prices = self.client.post(
            "/api/v1/quotes",
            json={"customer_name": "V1 Contract Customer", "bld_no": "V1-BLD-001", "currency": "USD"},
            headers={**authorization, "Idempotency-Key": "quote-create-v1-missing-price"},
        )
        self.assertEqual(missing_prices.status_code, 422)
        self.assertEqual(missing_prices.get_json()["error"]["code"], "request.invalid")

        create_payload = {
            "customer_name": "V1 Contract Customer",
            "bld_no": "V1-BLD-001",
            "customer_product_code": "V1-CUSTOMER-001",
            "tax_price": "12.34",
            "net_price": "11.22",
            "currency": "USD",
            "quote_date": "2026-07-11",
            "source_type": "wechat",
            "source_text": "V1 quote source",
            "on_behalf_of": "sales-operator",
        }
        created = self.client.post(
            "/api/v1/quotes",
            json=create_payload,
            headers=create_headers,
        )
        self.assertEqual(created.status_code, 201)
        created_body = created.get_json()
        quote = created_body["data"]["quote"]
        self.assertEqual(quote["version"], 1)
        self.assertEqual(quote["quoted_by"], "WorkBuddy Quote V1")
        self.assertEqual(quote["source_type"], "api")
        self.assertRegex(quote["quote_no"], r"^Q\d{9}$")
        self.assertEqual(created.headers["ETag"], '"1"')
        self.assertNotIn("attachment_path", quote)
        quote_id = quote["id"]

        immutable = self.client.patch(
            f"/api/v1/quotes/{quote_id}",
            json={"quoted_by": "spoofed", "source_type": "manual"},
            headers={
                **authorization,
                "Idempotency-Key": "quote-update-v1-immutable",
                "If-Match": '"1"',
            },
        )
        self.assertEqual(immutable.status_code, 422)
        self.assertEqual(immutable.get_json()["error"]["code"], "request.invalid")

        replayed = self.client.post(
            "/api/v1/quotes",
            json=create_payload,
            headers=create_headers,
        )
        self.assertEqual(replayed.status_code, 201)
        self.assertEqual(replayed.headers["Idempotency-Replayed"], "true")
        self.assertEqual(replayed.headers["ETag"], '"1"')
        self.assertEqual(replayed.get_json(), created_body)

        fetched = self.client.get(f"/api/v1/quotes/{quote_id}", headers=authorization)
        self.assertEqual(fetched.status_code, 200)
        self.assertEqual(fetched.headers["ETag"], '"1"')
        self.assertEqual(fetched.get_json()["data"]["quote"]["id"], quote_id)

        listed = self.client.get(
            "/api/v1/quotes?customer_name=V1%20Contract%20Customer&limit=10",
            headers=authorization,
        )
        listed_data = listed.get_json()["data"]
        self.assertEqual(listed.status_code, 200)
        self.assertEqual(listed_data["total"], 1)
        self.assertEqual([item["id"] for item in listed_data["quotes"]], [quote_id])

        listed_by_no = self.client.get(
            f"/api/v1/quotes?quote_no={quote['quote_no']}",
            headers=authorization,
        )
        listed_by_no_data = listed_by_no.get_json()["data"]
        self.assertEqual(listed_by_no.status_code, 200)
        self.assertEqual(listed_by_no_data["total"], 1)
        self.assertEqual(listed_by_no_data["quotes"][0]["id"], quote_id)

        missing_precondition = self.client.patch(
            f"/api/v1/quotes/{quote_id}",
            json={"remark": "first revision"},
            headers={**authorization, "Idempotency-Key": "quote-update-v1-missing"},
        )
        self.assertEqual(missing_precondition.status_code, 428)
        self.assertEqual(missing_precondition.get_json()["error"]["code"], "precondition.required")

        updated = self.client.patch(
            f"/api/v1/quotes/{quote_id}",
            json={"remark": "first revision", "on_behalf_of": "sales-operator"},
            headers={
                **authorization,
                "Idempotency-Key": "quote-update-v1-001",
                "If-Match": '"1"',
            },
        )
        updated_quote = updated.get_json()["data"]["quote"]
        self.assertEqual(updated.status_code, 200)
        self.assertEqual(updated_quote["version"], 2)
        self.assertEqual(updated_quote["remark"], "first revision")
        self.assertEqual(updated.headers["ETag"], '"2"')

        stale = self.client.patch(
            f"/api/v1/quotes/{quote_id}",
            json={"remark": "stale revision"},
            headers={
                **authorization,
                "Idempotency-Key": "quote-update-v1-stale",
                "If-Match": '"1"',
            },
        )
        stale_error = stale.get_json()["error"]
        self.assertEqual(stale.status_code, 412)
        self.assertEqual(stale_error["code"], "quote.version_conflict")
        self.assertEqual(stale_error["details"]["current_version"], 2)

        with self.web.connect(self.web.DB_PATH) as conn:
            quote_count = conn.execute(
                "SELECT COUNT(*) FROM quote_records WHERE customer_name = ?",
                ("V1 Contract Customer",),
            ).fetchone()[0]
            revisions = conn.execute(
                "SELECT changed_by, before_json, after_json FROM quote_record_revisions WHERE quote_id = ?",
                (quote_id,),
            ).fetchall()
            audit = conn.execute(
                "SELECT actor, detail FROM audit_logs WHERE action = 'API mutation' AND target_key = ? ORDER BY id",
                ("quote_v1_api.update_quote_v1",),
            ).fetchall()
        self.assertEqual(quote_count, 1)
        self.assertEqual(len(revisions), 1)
        self.assertEqual(revisions[0]["changed_by"], "WorkBuddy Quote V1")
        self.assertIn('"version": 1', revisions[0]["before_json"])
        self.assertIn('"version": 2', revisions[0]["after_json"])
        self.assertTrue(any(row["actor"] == "WorkBuddy Quote V1" for row in audit))

        document = self.client.get("/api/v1/openapi.json", headers=authorization).get_json()
        self.assertIn("/api/v1/quotes", document["paths"])
        self.assertIn("/api/v1/quotes/{quote_id}", document["paths"])
        create_operation = document["paths"]["/api/v1/quotes"]["post"]
        patch_operation = document["paths"]["/api/v1/quotes/{quote_id}"]["patch"]
        create_schema = document["components"]["schemas"]["QuoteCreateRequest"]
        patch_schema = document["components"]["schemas"]["QuotePatchRequest"]
        self.assertEqual(create_operation["x-required-scopes"], ["quotes:write"])
        self.assertTrue(create_schema["properties"]["quoted_by"]["deprecated"])
        self.assertTrue(create_schema["properties"]["source_type"]["deprecated"])
        self.assertNotIn("price", create_schema["properties"])
        self.assertNotIn("price", patch_schema["properties"])
        self.assertNotIn(
            "price",
            document["components"]["schemas"]["QuoteResponse"]["properties"],
        )
        self.assertNotIn("quoted_by", patch_schema["properties"])
        self.assertNotIn("source_type", patch_schema["properties"])
        self.assertIn("requestBody", create_operation)
        self.assertTrue(any(parameter["name"] == "Idempotency-Key" for parameter in create_operation["parameters"]))
        self.assertTrue(any(parameter["name"] == "If-Match" for parameter in patch_operation["parameters"]))
        self.assertIn("ETag", patch_operation["responses"]["200"]["headers"])
        self.assertIn("/api/v1/docs/{doc_name}", document["paths"])

    def test_api_docs_markdown_endpoint(self):
        token = self.create_internal_api_token(scopes=["api:read"], name="Docs Reader")
        authorization = {"Authorization": f"Bearer {token}"}

        anonymous = self.client.get("/api/v1/docs/quote-v1.md")
        self.assertEqual(anonymous.status_code, 401)

        document = self.client.get("/api/v1/docs/quote-v1.md", headers=authorization)
        self.assertEqual(document.status_code, 200)
        self.assertIn("text/markdown", document.headers["Content-Type"])
        self.assertIn("Quote API v1", document.get_data(as_text=True))

        missing = self.client.get("/api/v1/docs/no-such-doc.md", headers=authorization)
        self.assertEqual(missing.status_code, 404)
        details = missing.get_json()["error"].get("details", {})
        self.assertIn("quote-v1.md", details.get("available_docs", []))

        traversal = self.client.get("/api/v1/docs/..", headers=authorization)
        self.assertEqual(traversal.status_code, 404)

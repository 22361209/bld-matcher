from __future__ import annotations

import json
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import xlrd
import xlwt
from flask import Flask, g
from openpyxl import Workbook, load_workbook

from app.matcher import ProductCatalog
from app.modules.inquiry import download_web, web_helpers
from app.modules.inquiry.adjustments import (
    InquiryAdjustment,
    apply_adjustment,
    parse_adjustments,
)
from app.modules.inquiry.domain import augment_summary_with_bld_fragments
from app.modules.inquiry.excel.export import generate_xls_with_bld, generate_xlsx_with_bld
from app.modules.products import options_web


def _catalog() -> ProductCatalog:
    return ProductCatalog(
        [
            {
                "BLD NO.": "K8053LA",
                "OE NO.1": "OE-LA",
                "OE NO.2": "",
                "price_cny": 80,
                "product_status": "带球头",
            },
            {
                "BLD NO.": "K8053LB",
                "OE NO.1": "OE-LB",
                "OE NO.2": "",
                "price_cny": 58,
                "product_status": "不带球头",
            },
            {
                "BLD NO.": "K9999A",
                "OE NO.1": "OE-9999",
                "OE NO.2": "",
                "price_cny": 40,
                "product_status": "",
            },
        ]
    )


class InquiryAdjustmentParsingTest(unittest.TestCase):
    def test_price_accepts_zero_trailing_zeroes_and_scientific_integers(self) -> None:
        for raw, expected in (("0", "0.00"), ("1.230", "1.23"), ("1E+2", "100.00")):
            with self.subTest(raw=raw):
                parsed = parse_adjustments(
                    {"sheet:1:row:2": {"expected_bld_no": "K8053LB", "tax_price": raw}}
                )
                self.assertEqual(parsed["sheet:1:row:2"].tax_price, Decimal(expected))

    def test_price_rejects_real_third_decimal_and_invalid_shapes_stably(self) -> None:
        with self.assertRaisesRegex(ValueError, "最多保留两位小数"):
            parse_adjustments(
                {"sheet:1:row:2": {"expected_bld_no": "K8053LB", "tax_price": "1.231"}}
            )
        for payload in (
            [],
            {"sheet:1:row:2": "bad-row"},
            {"sheet:1:row:2": {"expected_bld_no": "K8053LB", "tax_price": "not-a-number"}},
        ):
            with self.subTest(payload=payload), self.assertRaises(ValueError):
                parse_adjustments(payload)

    def test_adjustment_requires_expected_match_and_rejects_catalog_drift(self) -> None:
        catalog = _catalog()
        match = catalog.match("", "OE-LB")
        self.assertIsNotNone(match)

        with self.assertRaisesRegex(ValueError, "缺少原匹配产品"):
            parse_adjustments({"sheet:1:row:2": {"tax_price": "12.34"}})
        with self.assertRaisesRegex(ValueError, "查询结果已变化"):
            apply_adjustment(
                catalog,
                match,
                InquiryAdjustment(expected_bld_no="K9999A", tax_price=Decimal("12.34")),
            )


class WorkbookInquiryAdjustmentTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.catalog = _catalog()

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def _xlsx_with_physical_gaps(self) -> Path:
        path = self.root / "stable.xlsx"
        workbook = Workbook()
        first = workbook.active
        first.title = "First"
        first.append(["OE号"])
        first.append(["OE-LB"])
        first.append([None])
        first.append(["OE-9999"])
        second = workbook.create_sheet("Second")
        second.append(["OE号"])
        second.append([None])
        second.append(["OE-LA"])
        workbook.save(path)
        workbook.close()
        return path

    def _xls_with_physical_gaps(self) -> Path:
        path = self.root / "stable.xls"
        workbook = xlwt.Workbook()
        first = workbook.add_sheet("First")
        first.write(0, 0, "OE号")
        first.write(1, 0, "OE-LB")
        first.write(3, 0, "OE-9999")
        second = workbook.add_sheet("Second")
        second.write(0, 0, "OE号")
        second.write(2, 0, "OE-LA")
        workbook.save(str(path))
        return path

    def test_xlsx_and_xls_keys_use_sheet_and_physical_source_row(self) -> None:
        xlsx_summary = generate_xlsx_with_bld(
            self._xlsx_with_physical_gaps(),
            self.root / "unused.xlsx",
            self.catalog,
            write_output=False,
        )
        xls_summary = generate_xls_with_bld(
            self._xls_with_physical_gaps(),
            self.root / "unused.xls",
            self.catalog,
            write_output=False,
        )
        expected = ["sheet:1:row:2", "sheet:1:row:4", "sheet:2:row:3"]
        self.assertEqual([row["adjustment_key"] for row in xlsx_summary["rows"]], expected)
        self.assertEqual([row["adjustment_key"] for row in xls_summary["rows"]], expected)
        self.assertTrue(all(row["adjustment_allowed"] for row in xlsx_summary["rows"]))
        self.assertTrue(all(row["adjustment_allowed"] for row in xls_summary["rows"]))

    def test_zero_override_is_consistent_for_xlsx_xls_tax_net_and_usd(self) -> None:
        xlsx_source = self.root / "zero.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["OE-LB"])
        workbook.save(xlsx_source)
        workbook.close()

        xls_source = self.root / "zero.xls"
        legacy = xlwt.Workbook()
        legacy_sheet = legacy.add_sheet("Sheet1")
        legacy_sheet.write(0, 0, "OE号")
        legacy_sheet.write(1, 0, "OE-LB")
        legacy.save(str(xls_source))

        adjustment = {
            "sheet:1:row:2": InquiryAdjustment(
                expected_bld_no="K8053LB",
                tax_price=Decimal("0.00"),
            )
        }
        for mode in ("tax", "net", "usd"):
            with self.subTest(format="xlsx", mode=mode):
                output = self.root / f"zero-{mode}.xlsx"
                summary = generate_xlsx_with_bld(
                    xlsx_source,
                    output,
                    self.catalog,
                    write_output=True,
                    price_mode=mode,
                    exchange_rate=7.2,
                    adjustments=adjustment,
                )
                generated = load_workbook(output, data_only=True)
                self.assertEqual(generated.active.cell(2, 3).value, 0)
                generated.close()
                self.assertEqual(summary["rows"][0]["price_cny"], 0.0)

            with self.subTest(format="xls", mode=mode):
                output = self.root / f"zero-{mode}.xls"
                summary = generate_xls_with_bld(
                    xls_source,
                    output,
                    self.catalog,
                    write_output=True,
                    price_mode=mode,
                    exchange_rate=7.2,
                    adjustments=adjustment,
                )
                generated = xlrd.open_workbook(str(output))
                self.assertEqual(generated.sheet_by_index(0).cell_value(1, 2), 0)
                self.assertEqual(summary["rows"][0]["price_cny"], 0.0)

    def test_fragment_candidates_are_unique_read_only_and_cannot_be_spoofed(self) -> None:
        source = self.root / "fragments.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["8053"])
        sheet.append(["OE-9999"])
        workbook.save(source)
        workbook.close()

        base = generate_xlsx_with_bld(source, self.root / "unused.xlsx", self.catalog, write_output=False)
        displayed = augment_summary_with_bld_fragments(base, self.catalog)
        fragment_rows = [row for row in displayed["rows"] if row["oe"] == "8053"]
        all_keys = [row["adjustment_key"] for row in displayed["rows"]]

        self.assertEqual(len(fragment_rows), 2)
        self.assertEqual(len(all_keys), len(set(all_keys)))
        self.assertTrue(all(not row["adjustment_allowed"] for row in fragment_rows))
        self.assertEqual(displayed["rows"][-1]["adjustment_key"], "sheet:1:row:3")

        forged = {
            fragment_rows[1]["adjustment_key"]: InquiryAdjustment(
                expected_bld_no=fragment_rows[1]["bld_no"],
                tax_price=Decimal("123.45"),
            )
        }
        with self.assertRaisesRegex(ValueError, "查询结果已变化"):
            generate_xlsx_with_bld(
                source,
                self.root / "forged.xlsx",
                self.catalog,
                write_output=True,
                price_mode="tax",
                adjustments=forged,
            )
        self.assertFalse((self.root / "forged.xlsx").exists())


class ProductLookupAdjustmentTest(unittest.TestCase):
    def test_active_only_preserves_default_and_returns_selector_details(self) -> None:
        app = Flask(__name__)
        app.config.update(TESTING=True, SECRET_KEY="test")

        @app.before_request
        def load_test_user() -> None:
            g.user = {"username": "tester", "role": "viewer", "permissions": set()}

        options_web.register(app)
        records = [
            SimpleNamespace(
                id=1,
                bld_no="K-ACTIVE",
                item="Active Arm",
                series="TEST",
                price_cny=80.0,
                product_status="带球头",
                active=True,
            ),
            SimpleNamespace(
                id=2,
                bld_no="K-INACTIVE",
                item="Inactive Arm",
                series="TEST",
                price_cny=60.0,
                product_status="不带球头",
                active=False,
            ),
        ]
        observed_statuses: list[str] = []

        class FakeProductService:
            def search(self, filters, *, limit, offset):
                observed_statuses.append(filters.status)
                selected = records if filters.status == "all" else [record for record in records if record.active]
                return SimpleNamespace(records=selected)

        client = app.test_client()
        with patch.object(options_web, "get_product_service", return_value=FakeProductService()):
            legacy = client.get("/products/lookup?q=K-")
            details = client.get("/products/lookup?q=K-&active_only=1&details=1")
            active = client.get("/products/lookup?q=K-&active_only=1&details=1&media=1")

        self.assertEqual(observed_statuses, ["all", "active", "active"])
        self.assertEqual([row["bld_no"] for row in legacy.get_json()], ["K-ACTIVE", "K-INACTIVE"])
        self.assertEqual(
            sorted(legacy.get_json()[0]),
            ["bld_no", "id", "item", "series"],
        )
        self.assertNotIn("image_gallery", details.get_json()[0])
        self.assertEqual(
            active.get_json(),
            [
                {
                    "id": 1,
                    "bld_no": "K-ACTIVE",
                    "item": "Active Arm",
                    "series": "TEST",
                    "price_cny": 80.0,
                    "product_status": "带球头",
                    "active": True,
                    "image_gallery": [],
                }
            ],
        )


class InquiryAdjustmentRouteTest(unittest.TestCase):
    @staticmethod
    def _route_app(permissions: set[str]) -> Flask:
        app = Flask(__name__)
        app.config.update(TESTING=True, SECRET_KEY="test")

        @app.before_request
        def load_test_user() -> None:
            g.user = {"username": "tester", "role": "viewer", "permissions": permissions}

        app.add_url_rule("/", endpoint="index", view_func=lambda: "index")
        app.add_url_rule("/download/<path:name>", endpoint="download", view_func=lambda name: name)
        app.add_url_rule("/quotes", endpoint="quote_web.quotes", view_func=lambda: "quotes")
        download_web.register(app)
        return app

    def test_download_rejects_forged_adjustment_without_price_permission(self) -> None:
        app = self._route_app({"generate_match"})
        fake_inquiry = SimpleNamespace(catalog_available=lambda: True)
        payload = json.dumps(
            {
                "sheet:1:row:2": {
                    "expected_bld_no": "K8053LB",
                    "target_bld_no": "K8053LA",
                    "tax_price": "80.00",
                }
            }
        )
        with (
            patch.object(download_web, "get_inquiry_service", return_value=fake_inquiry),
            patch.object(download_web, "validated_user_upload_path", return_value=Path("/tmp/source.xlsx")),
        ):
            response = app.test_client().post(
                "/match/download",
                data={"price_mode": "tax", "inquiry_adjustments": payload},
            )

        self.assertEqual(response.status_code, 302)
        with app.test_client() as client:
            with client.session_transaction() as session:
                session.clear()

    def test_write_quotes_generates_once_returns_download_and_cleans_orphans(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.xlsx"
            source.touch()
            output_root = root / "outputs"
            output_root.mkdir()
            generated = output_root / "unique-generated.xlsx"
            generated_temporary = output_root / ".unique-generated.tmp.xlsx"

            class FakeInquiryService:
                calls = 0

                @staticmethod
                def catalog_available() -> bool:
                    return True

                def analyze_workbook(self, source_path, output_path, **kwargs):
                    self.calls += 1
                    Path(output_path).write_bytes(b"generated")
                    return {
                        "rows": [
                            {
                                "bld_no": "K8053LA",
                                "customer_product_code": "CUST-1",
                                "price_cny": 80.0,
                            }
                        ]
                    }

            class FakeQuoteService:
                def __init__(self, *, write: bool):
                    self.write = write
                    self.rows = []

                def create_many(self, rows, *, actor):
                    self.rows = list(rows)
                    return ([SimpleNamespace(id=1)] if self.write else []), 0, "Q260731001"

            inquiry_service = FakeInquiryService()
            quote_service = FakeQuoteService(write=True)
            app = self._route_app({"manage_customer_prices"})
            client = app.test_client()
            patches = (
                patch.object(download_web, "OUTPUT_DIR", output_root),
                patch.object(download_web, "validated_user_upload_path", return_value=source),
                patch.object(download_web, "get_inquiry_service", return_value=inquiry_service),
                patch.object(
                    download_web,
                    "get_customer_service",
                    return_value=SimpleNamespace(find_by_name=lambda name: SimpleNamespace(name=name)),
                ),
                patch.object(download_web, "get_quote_service", return_value=quote_service),
                patch.object(
                    download_web,
                    "quote_output_paths",
                    return_value=(generated, generated_temporary),
                ),
                patch.object(download_web, "download_name", side_effect=lambda path: path.name),
            )
            with patches[0], patches[1], patches[2], patches[3], patches[4], patches[5], patches[6]:
                response = client.post(
                    "/match/write-quotes",
                    data={
                        "upload_path": str(source),
                        "original_filename": "source.xlsx",
                        "output_name": "stale-shared-name.xlsx",
                        "customer_name": "Test Customer",
                        "price_mode": "tax",
                        "inquiry_adjustments": "{}",
                    },
                    headers={"Accept": "application/json", "X-Requested-With": "fetch"},
                )

            self.assertEqual(response.status_code, 200)
            self.assertEqual(inquiry_service.calls, 1)
            self.assertTrue(generated.exists())
            self.assertEqual(quote_service.rows[0]["attachment_path"], "unique-generated.xlsx")
            self.assertEqual(response.get_json()["download_url"], "/download/unique-generated.xlsx")
            self.assertEqual(response.get_json()["download_filename"], "unique-generated.xlsx")

            orphan = output_root / "orphan.xlsx"
            orphan_temporary = output_root / ".orphan.tmp.xlsx"
            quote_service = FakeQuoteService(write=False)
            with (
                patch.object(download_web, "OUTPUT_DIR", output_root),
                patch.object(download_web, "validated_user_upload_path", return_value=source),
                patch.object(download_web, "get_inquiry_service", return_value=inquiry_service),
                patch.object(
                    download_web,
                    "get_customer_service",
                    return_value=SimpleNamespace(find_by_name=lambda name: SimpleNamespace(name=name)),
                ),
                patch.object(download_web, "get_quote_service", return_value=quote_service),
                patch.object(
                    download_web,
                    "quote_output_paths",
                    return_value=(orphan, orphan_temporary),
                ),
            ):
                failed = client.post(
                    "/match/write-quotes",
                    data={
                        "upload_path": str(source),
                        "original_filename": "source.xlsx",
                        "customer_name": "Test Customer",
                        "price_mode": "tax",
                        "inquiry_adjustments": "{}",
                    },
                    headers={"Accept": "application/json", "X-Requested-With": "fetch"},
                )

            self.assertEqual(failed.status_code, 400)
            self.assertFalse(orphan.exists())
            self.assertFalse(orphan_temporary.exists())

    def test_quote_output_paths_are_request_unique_and_atomically_staged(self) -> None:
        base = Path("/tmp/re260731-source.XLSX")
        with patch.object(web_helpers, "result_output_path", return_value=base):
            first_output, first_temporary = web_helpers.quote_output_paths("source.XLSX", ".XLSX")
            second_output, second_temporary = web_helpers.quote_output_paths("source.XLSX", ".XLSX")

        self.assertNotEqual(first_output, second_output)
        self.assertNotEqual(first_temporary, second_temporary)
        self.assertEqual(first_output.suffix, ".xlsx")
        self.assertEqual(first_temporary.suffix, ".xlsx")
        self.assertTrue(first_temporary.name.startswith("."))
        self.assertIn(first_output.stem, first_temporary.name)

    def test_workbook_output_path_normalizes_uppercase_suffix_for_case_sensitive_hosts(self) -> None:
        source = Path("/tmp/re260731-source.XLSX")
        normalized = web_helpers.normalized_workbook_output_path(source, ".XLSX")
        self.assertEqual(normalized, Path("/tmp/re260731-source.xlsx"))


if __name__ == "__main__":
    unittest.main()

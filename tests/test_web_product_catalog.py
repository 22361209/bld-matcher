from __future__ import annotations

from tests.web_app_test_base import (
    WebAppTestBase,
    PROJECT_ROOT,
    io,
    re,
    unquote,
)


class TestWebProductCatalog(WebAppTestBase):
    def test_product_list_omits_redundant_workspace_header(self):
        self.login()

        response = self.client.get("/products")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('class="data-section products-list-section"', html)
        self.assertIn('class="workspace-command products-command"', html)
        self.assertIn("data-products-results-host", html)
        self.assertIn('data-products-fragment-url="/products/fragment"', html)
        self.assertNotIn('<header class="workspace-header">', html)
        self.assertNotIn("目录规模", html)

    def test_product_fragment_is_authenticated_shell_free_and_keeps_filter_state(self):
        with self.client.session_transaction() as session:
            session.clear()
        anonymous = self.client.get(
            "/products/fragment",
            headers={"X-Requested-With": "fetch", "Accept": "text/html"},
        )
        self.assertEqual(anonymous.status_code, 401)

        self.login()
        response = self.client.get(
            "/products/fragment",
            query_string={"bld": "K8", "status": "all", "brand": "HYUNDAI"},
            headers={"X-Requested-With": "fetch", "Accept": "text/html"},
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["Cache-Control"], "no-store")
        self.assertIn("data-products-results", html)
        self.assertIn("data-canonical-url=", html)
        self.assertIn("brand=HYUNDAI", html)
        self.assertNotIn("<!doctype html>", html.lower())
        self.assertNotIn("data-products-results-host", html)
        self.assertNotIn("新增产品", html)

    def test_product_results_keep_desktop_table_and_mobile_cards_in_fragment_refreshes(self):
        from app.modules.products.persistence import upsert_product

        bld_no = "K-MOBILE-CARD-001"
        self.addCleanup(self.cleanup_products, "K-MOBILE-CARD-%")
        self.login()
        with self.web.connect(self.web.DB_PATH) as connection:
            upsert_product(
                connection,
                {
                    "bld_no": bld_no,
                    "series": "MOBILE BRAND",
                    "item": "Mobile Card Control Arm",
                    "oe_no_1": "MOBILE-OE-001",
                    "models": "Mobile Model",
                    "product_status": "带球头",
                    "price_cny": "91.23",
                    "active": "1",
                },
                actor="tester",
            )
            connection.commit()

        page = self.client.get("/products", query_string={"bld": bld_no})
        fragment = self.client.get(
            "/products/fragment",
            query_string={"bld": bld_no},
            headers={"X-Requested-With": "fetch", "Accept": "text/html"},
        )

        self.assertEqual(page.status_code, 200)
        self.assertEqual(fragment.status_code, 200)
        for html in (page.get_data(as_text=True), fragment.get_data(as_text=True)):
            self.assertIn('id="products-table"', html)
            self.assertIn("data-products-mobile-cards", html)
            self.assertIn("data-product-mobile-card", html)
            self.assertIn(bld_no, html)
            self.assertIn("MOBILE-OE-001", html)
            self.assertIn("data-product-mobile-price", html)

    def test_product_mobile_cards_do_not_leak_prices_or_management_actions_without_permission(self):
        from app.modules.admin.persistence import save_role, save_user
        from app.modules.products.persistence import upsert_product

        bld_no = "K-MOBILE-PRIVATE-001"
        username = "mobile-catalog-readonly"
        role_name = "移动目录只读"
        self.addCleanup(self.cleanup_products, "K-MOBILE-PRIVATE-%")

        def cleanup_access_records():
            self.client.post("/logout")
            with self.web.connect(self.web.DB_PATH) as connection:
                user = connection.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
                if user:
                    connection.execute("DELETE FROM user_permission_overrides WHERE user_id = ?", (user["id"],))
                    connection.execute("DELETE FROM users WHERE id = ?", (user["id"],))
                role = connection.execute("SELECT role_key FROM roles WHERE name = ?", (role_name,)).fetchone()
                if role:
                    connection.execute("DELETE FROM role_permissions WHERE role_key = ?", (role["role_key"],))
                    connection.execute("DELETE FROM roles WHERE role_key = ?", (role["role_key"],))
                connection.commit()

        self.addCleanup(cleanup_access_records)
        self.login()
        with self.web.connect(self.web.DB_PATH) as connection:
            role_key = save_role(
                connection,
                {"name": role_name, "description": "移动产品目录权限回归测试"},
                ["view_products"],
                actor="tester",
            )
            save_user(
                connection,
                {"username": username, "password": "mobile-readonly-pw", "role": role_key, "active": "1"},
                actor="tester",
            )
            upsert_product(
                connection,
                {
                    "bld_no": bld_no,
                    "series": "PRIVATE BRAND",
                    "item": "Private Mobile Arm",
                    "oe_no_1": "PRIVATE-MOBILE-OE",
                    "price_cny": "91.23",
                    "active": "1",
                },
                actor="tester",
            )
            connection.commit()

        self.client.post("/logout")
        login = self.client.post(
            "/login",
            data={"username": username, "password": "mobile-readonly-pw", "next": "/products"},
            follow_redirects=False,
        )
        self.assertEqual(login.status_code, 302)

        page = self.client.get("/products", query_string={"bld": bld_no})
        fragment = self.client.get(
            "/products/fragment",
            query_string={"bld": bld_no},
            headers={"X-Requested-With": "fetch", "Accept": "text/html"},
        )
        self.assertEqual(page.status_code, 200)
        self.assertEqual(fragment.status_code, 200)
        for html in (page.get_data(as_text=True), fragment.get_data(as_text=True)):
            self.assertIn("data-product-mobile-card", html)
            self.assertNotIn("data-product-mobile-price", html)
            self.assertNotIn("91.23", html)
            self.assertNotIn('data-col="price"', html)
            self.assertNotIn("data-open-product-modal", html)
            self.assertNotIn("data-open-edit-product-modal", html)
            self.assertNotIn("data-copy-product-action", html)

    def test_product_only_role_lands_in_catalog_and_cannot_read_drawings_or_other_pages(self):
        from app.modules.admin.persistence import save_role, save_user
        from app.modules.products.persistence import upsert_product

        username = "product-only-viewer"
        role_name = "产品目录查看员回归"
        bld_no = "K-PRODUCT-ONLY-001"

        def cleanup_access_records():
            self.client.post("/logout")
            with self.web.connect(self.web.DB_PATH) as connection:
                user = connection.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
                if user:
                    connection.execute("DELETE FROM user_permission_overrides WHERE user_id = ?", (user["id"],))
                    connection.execute("DELETE FROM users WHERE id = ?", (user["id"],))
                role = connection.execute("SELECT role_key FROM roles WHERE name = ?", (role_name,)).fetchone()
                if role:
                    connection.execute("DELETE FROM role_permissions WHERE role_key = ?", (role["role_key"],))
                    connection.execute("DELETE FROM roles WHERE role_key = ?", (role["role_key"],))
                connection.execute("DELETE FROM products WHERE bld_no = ?", (bld_no,))
                connection.commit()

        self.addCleanup(cleanup_access_records)
        self.login()
        with self.web.connect(self.web.DB_PATH) as connection:
            role_key = save_role(
                connection,
                {"name": role_name, "description": "仅产品目录查看"},
                ["view_products"],
                actor="tester",
            )
            save_user(
                connection,
                {"username": username, "password": "product-only-pw", "role": role_key, "active": "1"},
                actor="tester",
            )
            upsert_product(
                connection,
                {"bld_no": bld_no, "item": "Product Only Arm", "price_cny": "77.88", "active": "1"},
                actor="tester",
            )
            product = connection.execute("SELECT id FROM products WHERE bld_no = ?", (bld_no,)).fetchone()
            drawing_name = f"{bld_no}.pdf"
            drawing_path = self.root / "data" / "drawings" / "pdf" / drawing_name
            drawing_path.parent.mkdir(parents=True, exist_ok=True)
            drawing_path.write_bytes(b"%PDF-1.4\nproduct-only drawing\n%%EOF")
            connection.execute(
                "UPDATE products SET drawing_path = ?, drawing_original_name = ? WHERE id = ?",
                (f"drawings/pdf/{drawing_name}", drawing_name, product["id"]),
            )
            connection.commit()

        self.client.post("/logout")
        login = self.client.post(
            "/login",
            data={"username": username, "password": "product-only-pw"},
            follow_redirects=False,
        )
        self.assertEqual(login.location, "/products")
        page = self.client.get("/products", query_string={"bld": bld_no})
        self.assertEqual(page.status_code, 200)
        html = page.get_data(as_text=True)
        self.assertIn("data-product-detail-template", html)
        self.assertIn("data-product-detail-trigger", html)
        self.assertIn("data-product-detail-dialog", html)
        self.assertIn("无图纸查看权限", html)
        self.assertNotIn("77.88", html)
        self.assertNotIn(f"/products/{product['id']}/drawing", html)
        self.assertNotIn(drawing_name, html)
        self.assertNotIn("data-open-edit-product-modal", html)
        self.assertNotIn('href="/tubes">管件资料</a>', html)
        self.assertNotIn('href="/">询价处理</a>', html)

        account_settings = self.client.get("/account/password").get_data(as_text=True)
        self.assertIn('<option value="products">产品目录</option>', account_settings)
        self.assertNotIn('<option value="index">询价处理</option>', account_settings)
        rejected_default = self.client.post(
            "/account/default-page",
            data={"default_page": "index"},
            follow_redirects=True,
        )
        self.assertIn("该页面不存在或当前账号没有访问权限", rejected_default.get_data(as_text=True))
        with self.web.connect(self.web.DB_PATH) as connection:
            stored_default = connection.execute(
                "SELECT default_page FROM users WHERE username = ?", (username,)
            ).fetchone()[0]
        self.assertEqual(stored_default, "")

        denied_page = self.client.get("/tubes", follow_redirects=False)
        self.assertEqual(denied_page.location, "/products")
        denied_json = self.client.get("/inquiry/quick-search", headers={"Accept": "application/json"})
        self.assertEqual(denied_json.status_code, 403)
        self.assertEqual(denied_json.get_json()["ok"], False)
        drawing = self.client.get(f"/products/{product['id']}/drawing", headers={"Accept": "application/json"})
        self.assertEqual(drawing.status_code, 403)
        self.assertEqual(drawing.get_json()["ok"], False)
        self.assertEqual(self.client.get("/products/lookup").status_code, 200)

        products_js = (PROJECT_ROOT / "static" / "pages" / "products.js").read_text(encoding="utf-8")
        self.assertIn("productDetailDialog.showModal()", products_js)
        self.assertIn("productDetailTrigger?.focus()", products_js)

    def test_product_lookup_returns_matching_products(self):
        from app.modules.products.persistence import upsert_product

        with self.client.session_transaction() as session:
            session.clear()
        anonymous = self.client.get(
            "/products/lookup?q=K-LOOKUP",
            headers={"X-Requested-With": "fetch", "Accept": "application/json"},
        )
        self.assertEqual(anonymous.status_code, 401)

        self.addCleanup(self.cleanup_products, "K-LOOKUP-%")
        self.login()
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-LOOKUP-001",
                    "series": "LOOKUPSERIES",
                    "item": "Lookup Control Arm",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-LOOKUP-002",
                    "series": "JOINTSERIES",
                    "item": "Lookup Ball Joint",
                    "active": "0",
                },
                actor="tester",
            )
            conn.commit()

        response = self.client.get("/products/lookup?q=K-LOOKUP-001")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.headers["Cache-Control"], "no-store")
        payload = response.get_json()
        self.assertIsInstance(payload, list)
        self.assertEqual(len(payload), 1)
        entry = payload[0]
        self.assertEqual(sorted(entry.keys()), ["bld_no", "id", "item", "series"])
        self.assertEqual(entry["bld_no"], "K-LOOKUP-001")
        self.assertEqual(entry["item"], "Lookup Control Arm")
        self.assertEqual(entry["series"], "LOOKUPSERIES")

        filtered = self.client.get("/products/lookup?q=JOINTSERIES").get_json()
        self.assertEqual([entry["bld_no"] for entry in filtered], ["K-LOOKUP-002"])

        empty = self.client.get("/products/lookup?q=NO-SUCH-PRODUCT-XYZ").get_json()
        self.assertEqual(empty, [])

    def test_product_save_fetch_returns_json_for_local_result_refresh(self):
        self.addCleanup(self.cleanup_products, "K-INLINE-SAVE-%")
        self.login()

        response = self.client.post(
            "/products/save",
            data={
                "bld_no": "K-INLINE-SAVE-001",
                "series": "INLINE",
                "item": "Inline Test Arm",
                "oe_no_1": "INLINE-OE-001",
                "oe_no_2": "",
                "models": "Inline Tester",
                "price_cny": "12.50",
                "active": "1",
            },
            headers={"X-Requested-With": "fetch", "Accept": "application/json"},
        )
        payload = response.get_json()

        self.assertEqual(response.status_code, 200)
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["message"], "产品已保存。")
        self.assertIn("bld=K-INLINE-SAVE-001", payload["redirect_url"])

        fragment = self.client.get(
            "/products/fragment",
            query_string={"bld": "K-INLINE-SAVE-001"},
        )
        self.assertEqual(fragment.status_code, 200)
        self.assertIn("K-INLINE-SAVE-001", fragment.get_data(as_text=True))

        invalid = self.client.post(
            "/products/save",
            data={"bld_no": "", "active": "1"},
            headers={"X-Requested-With": "fetch", "Accept": "application/json"},
        )
        self.assertEqual(invalid.status_code, 400)
        self.assertFalse(invalid.get_json()["ok"])

    def test_products_search_uses_results_anchor(self):
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-FILTER-HYUNDAI",
                    "series": "HYUNDAI",
                    "item": "CONTROL ARM",
                    "oe_no_1": "FILTER-001",
                    "models": "Sportage",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-FILTER-HONDA",
                    "series": "HONDA",
                    "item": "CONTROL ARM",
                    "oe_no_1": "FILTER-002",
                    "models": "Civic",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-FILTER-DOT-OE",
                    "series": "PEUGEOT",
                    "item": "CONTROL ARM",
                    "oe_no_1": "3521.R1",
                    "models": "C-CROSSER",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/products")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('id="products-results"', html)
        self.assertIn('action="/products#products-results"', html)
        self.assertIn("按 BLD / 品牌 / 车型搜索", html)
        self.assertIn('<button class="linear-button primary" type="submit">搜索</button>', html)
        self.assertIn('class="toolbar-popover-form catalog-import-actions"', html)
        self.assertIn('href="/catalog/template">下载模板', html)
        self.assertIn("data-catalog-upload-input", html)
        self.assertIn('id="product-modal"', html)
        self.assertIn('id="product-edit-modal"', html)
        self.assertIn("data-draggable-modal-panel", html)

        for query in ["HYUNDAI", "Sportage"]:
            with self.subTest(query=query):
                response = self.client.get("/products", query_string={"bld": query})
                html = response.get_data(as_text=True)
                self.assertIn("K-FILTER-HYUNDAI", html)
                self.assertNotIn("K-FILTER-HONDA", html)

        for query in ["3521.r1", "3521R1", "3521-R1"]:
            with self.subTest(query=query):
                response = self.client.get("/products", query_string={"oe": query})
                html = response.get_data(as_text=True)
                self.assertIn("K-FILTER-DOT-OE", html)

    def test_products_oe_search_psa_352x_dot_does_not_show_gm_exact(self):
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-PRODUCT-PSA-352125",
                    "series": "PEUGEOT\nCITROEN",
                    "item": "CONTROL ARM",
                    "oe_no_1": "3521.25",
                    "models": "C5",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-PRODUCT-GM-352125",
                    "series": "GM\nOPEL",
                    "item": "CONTROL ARM",
                    "oe_no_1": "352125",
                    "models": "OPEL",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        dotted = self.client.get("/products", query_string={"oe": "3521.25"})
        dotted_html = dotted.get_data(as_text=True)
        self.assertEqual(dotted.status_code, 200)
        self.assertIn("K-PRODUCT-PSA-352125", dotted_html)
        self.assertNotIn("K-PRODUCT-GM-352125", dotted_html)

        undotted = self.client.get("/products", query_string={"oe": "352125"})
        undotted_html = undotted.get_data(as_text=True)
        self.assertEqual(undotted.status_code, 200)
        self.assertIn("K-PRODUCT-PSA-352125", undotted_html)
        self.assertIn("K-PRODUCT-GM-352125", undotted_html)

    def test_products_use_bld_natural_order(self):
        from app.bld_sort import bld_sort_key
        from app.modules.products.persistence import upsert_product

        self.assertEqual(
            sorted(["K8274LA", "K8274RA", "K8274LB", "K8274RB"], key=bld_sort_key),
            ["K8274LA", "K8274RA", "K8274LB", "K8274RB"],
        )
        self.assertEqual(
            sorted(["K8058LA-1", "K8058RA-1", "K8058LB", "K8058RB"], key=bld_sort_key),
            ["K8058LA-1", "K8058RA-1", "K8058LB", "K8058RB"],
        )

        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no in ["K8274LA", "K8274RA", "K8274LB", "K8274RB", "K8058LA-1", "K8058RA-1", "K8058LB", "K8058RB"]:
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "SORT",
                        "item": "SORT TEST",
                        "oe_no_1": f"OE-{bld_no}",
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.get("/products", query_string={"bld": "K8274"})
        html = response.get_data(as_text=True)
        self.assertLess(html.index("K8274LA"), html.index("K8274RA"))
        self.assertLess(html.index("K8274RA"), html.index("K8274LB"))
        self.assertLess(html.index("K8274LB"), html.index("K8274RB"))

        response = self.client.get("/products", query_string={"bld": "K8058"})
        html = response.get_data(as_text=True)
        self.assertLess(html.index("K8058LA-1"), html.index("K8058RA-1"))
        self.assertLess(html.index("K8058RA-1"), html.index("K8058LB"))
        self.assertLess(html.index("K8058LB"), html.index("K8058RB"))

    def test_products_are_paginated(self):
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            for index in range(121):
                upsert_product(
                    conn,
                    {
                        "bld_no": f"K-PAGE-{index:03d}",
                        "series": "PAGED",
                        "item": "PAGED PART",
                        "oe_no_1": f"PAGE-{index:03d}",
                        "models": "Batch Tester",
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.get("/products", query_string={"bld": "K-PAGE"})
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertNotIn("第 1-50 条 / 共 121 条", html)
        self.assertIn("1–50</strong><span> / 121", html)
        self.assertEqual(html.count('aria-label="产品分页"'), 1)
        self.assertIn('data-reset-list-for="products"', html)
        self.assertIn('data-grid-page-jump', html)
        self.assertIn('max="3"', html)
        pagination_start = html.index('<nav class="data-grid-pagination"')
        pagination_end = html.index('</nav>', pagination_start)
        pagination_html = html[pagination_start:pagination_end]
        self.assertLess(pagination_html.index('>下一页<'), pagination_html.index('data-grid-page-jump'))
        self.assertIn("K-PAGE-000", html)
        self.assertIn("K-PAGE-049", html)
        self.assertNotIn("K-PAGE-050", html)
        self.assertNotIn("/products/rows", html)

        third_page = self.client.get("/products", query_string={"bld": "K-PAGE", "page": "3"})
        third_html = third_page.get_data(as_text=True)
        self.assertEqual(third_page.status_code, 200)
        self.assertNotIn("第 101-121 条 / 共 121 条", third_html)
        self.assertIn("101–121</strong><span> / 121", third_html)
        self.assertIn("K-PAGE-100", third_html)
        self.assertIn("K-PAGE-120", third_html)
        self.assertNotIn("K-PAGE-099", third_html)

    def test_shared_page_jump_controls_remain_compact(self):
        styles = (PROJECT_ROOT / "static" / "components" / "data_grid.css").read_text(encoding="utf-8")
        self.assertRegex(
            styles,
            r"\.data-grid-page-jump > label\s*\{[^}]*flex:\s*0 0 auto;[^}]*min-width:\s*0;",
        )

    def test_product_column_filters_preserve_multiselect_in_pagination_and_export_form(self):
        from app.modules.products.persistence import upsert_product

        self.addCleanup(self.cleanup_products, "K-WEB-FILTER-%")
        with self.web.connect(self.web.DB_PATH) as conn:
            for index in range(51):
                upsert_product(
                    conn,
                    {
                        "bld_no": f"K-WEB-FILTER-{index:03d}",
                        "series": "WEBBRAND-A\nWEBBRAND-B",
                        "item": "Web Filter Arm",
                        "oe_no_1": f"WEB-FILTER-{index:03d}",
                        "product_status": "1个衬套",
                        "active": "1",
                    },
                    actor="tester",
                )
            upsert_product(
                conn,
                {
                    "bld_no": "K-WEB-FILTER-EXCLUDED",
                    "series": "WEBBRAND-C",
                    "item": "Web Filter Arm",
                    "oe_no_1": "WEB-FILTER-EXCLUDED",
                    "product_status": "1个衬套",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get(
            "/products",
            query_string={
                "bld": "K-WEB-FILTER-",
                "brand": ["WEBBRAND-A", "WEBBRAND-B"],
                "item": "Web Filter Arm",
                "product_status": "1衬套",
            },
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("1–50</strong><span> / 51", html)
        self.assertIn("K-WEB-FILTER-000", html)
        self.assertIn("K-WEB-FILTER-049", html)
        self.assertNotIn("K-WEB-FILTER-050", html)
        self.assertNotIn("K-WEB-FILTER-EXCLUDED", html)
        self.assertEqual(html.count("data-column-filter-trigger"), 3)
        self.assertEqual(html.count("data-column-filter-panel"), 3)
        self.assertEqual(html.count("data-column-filter-select-all"), 3)
        self.assertEqual(html.count("data-column-filter-select-none"), 3)
        self.assertEqual(html.count("data-column-filter-selection"), 3)
        self.assertEqual(html.count("data-column-filter-reset"), 3)
        self.assertEqual(html.count("data-column-filter-apply"), 3)
        self.assertIn("重置筛选", html)
        self.assertIn('aria-label="启用状态筛选"', html)
        self.assertIn('name="brand" value="WEBBRAND-A"', html)
        self.assertIn('name="brand" value="WEBBRAND-B"', html)
        self.assertIn('name="item" value="Web Filter Arm"', html)
        self.assertIn('name="product_status" value="1衬套"', html)
        self.assertIn("data-submit-wait", html)

        next_href_match = re.search(r'href="([^"]*page=2[^"]*)"', html)
        self.assertIsNotNone(next_href_match)
        next_href = next_href_match.group(1).replace("&amp;", "&")
        self.assertIn("brand=WEBBRAND-A&brand=WEBBRAND-B", next_href)
        self.assertIn("item=Web+Filter+Arm", next_href)
        self.assertIn("product_status=1%E8%A1%AC%E5%A5%97", next_href)
        self.assertTrue(next_href.endswith("#products-results"))

    def test_catalog_export_uses_all_web_filters_across_pages_and_handles_empty_results(self):
        from openpyxl import load_workbook

        from app.modules.products.persistence import upsert_product

        self.addCleanup(self.cleanup_products, "K-WEB-EXPORT-%")
        with self.web.connect(self.web.DB_PATH) as conn:
            for index in range(51):
                upsert_product(
                    conn,
                    {
                        "bld_no": f"K-WEB-EXPORT-{index:03d}",
                        "series": "WEBEXPORT",
                        "item": "Web Export Arm",
                        "oe_no_1": f"WEB-EXPORT-{index:03d}",
                        "product_status": "2个衬套\n1个球头",
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.post(
            "/products/export",
            data={
                "bld": "K-WEB-EXPORT-",
                "brand": ["WEBEXPORT"],
                "item": ["Web Export Arm"],
                "product_status": ["2衬套1球头"],
                "status": "active",
                "export_format": "bld",
            },
        )
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(io.BytesIO(response.data), read_only=True, data_only=True)
        sheet = workbook["产品目录"]
        exported = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]
        workbook.close()
        response.close()

        self.assertEqual(len(exported), 51)
        self.assertEqual(exported[0], "K-WEB-EXPORT-000")
        self.assertEqual(exported[-1], "K-WEB-EXPORT-050")

        before = set((self.root / "outputs").glob("u*-007/catalog-export-bld-007-*.xlsx"))
        empty = self.client.post(
            "/products/export",
            data={
                "bld": "K-WEB-EXPORT-NOT-FOUND",
                "brand": ["WEBEXPORT"],
                "status": "active",
                "export_format": "bld",
            },
            follow_redirects=False,
        )
        after = set((self.root / "outputs").glob("u*-007/catalog-export-bld-007-*.xlsx"))
        self.assertEqual(empty.status_code, 302)
        self.assertIn("K-WEB-EXPORT-NOT-FOUND", unquote(empty.headers["Location"]))
        self.assertEqual(after, before)

    def test_product_column_filter_blank_value_is_distinct_and_preserved_in_export(self):
        from openpyxl import load_workbook

        from app.modules.products.persistence import upsert_product

        self.addCleanup(self.cleanup_products, "K-WEB-BLANK-%")
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-WEB-BLANK-EMPTY",
                    "series": "",
                    "item": "",
                    "product_status": "",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-WEB-BLANK-LITERAL",
                    "series": "__blank__",
                    "item": "__blank__",
                    "product_status": "__blank__",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        blank_page = self.client.get(
            "/products",
            query_string=[
                ("bld", "K-WEB-BLANK-"),
                ("brand", ""),
                ("item", ""),
                ("product_status", ""),
            ],
        )
        blank_html = blank_page.get_data(as_text=True)
        self.assertEqual(blank_page.status_code, 200)
        self.assertIn("K-WEB-BLANK-EMPTY", blank_html)
        self.assertNotIn("K-WEB-BLANK-LITERAL", blank_html)
        self.assertGreaterEqual(blank_html.count('name="brand" value=""'), 2)
        self.assertGreaterEqual(blank_html.count('name="item" value=""'), 2)
        self.assertGreaterEqual(blank_html.count('name="product_status" value=""'), 2)

        literal_page = self.client.get(
            "/products",
            query_string={
                "bld": "K-WEB-BLANK-",
                "brand": "__blank__",
                "item": "__blank__",
                "product_status": "__blank__",
            },
        )
        literal_html = literal_page.get_data(as_text=True)
        self.assertEqual(literal_page.status_code, 200)
        self.assertNotIn("K-WEB-BLANK-EMPTY", literal_html)
        self.assertIn("K-WEB-BLANK-LITERAL", literal_html)

        exported = self.client.post(
            "/products/export",
            data={
                "bld": "K-WEB-BLANK-",
                "brand": "",
                "item": "",
                "product_status": "",
                "status": "active",
                "export_format": "bld",
            },
        )
        self.assertEqual(exported.status_code, 200)
        workbook = load_workbook(io.BytesIO(exported.data), read_only=True, data_only=True)
        sheet = workbook["产品目录"]
        exported_bld_numbers = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]
        workbook.close()
        exported.close()
        self.assertEqual(exported_bld_numbers, ["K-WEB-BLANK-EMPTY"])

    def test_product_column_filter_limits_return_400_without_listing_or_exporting(self):
        self.login()
        excessive_values = [("brand", f"UNKNOWN-{index}") for index in range(201)]
        rejected_page = self.client.get("/products", query_string=excessive_values)
        self.assertEqual(rejected_page.status_code, 400)
        self.assertIn("品牌筛选项最多选择 200 个", rejected_page.get_data(as_text=True))
        self.assertNotIn("products-table", rejected_page.get_data(as_text=True))

        before = set((self.root / "outputs").glob("u*-007/catalog-export-bld-007-*.xlsx"))
        rejected_export = self.client.post(
            "/products/export",
            data={
                "item": "X" * 257,
                "status": "active",
                "export_format": "bld",
            },
        )
        after = set((self.root / "outputs").glob("u*-007/catalog-export-bld-007-*.xlsx"))
        self.assertEqual(rejected_export.status_code, 400)
        self.assertIn("产品名称筛选项单项不能超过 256 个字符", rejected_export.get_data(as_text=True))
        self.assertEqual(after, before)

    def test_product_save_can_clear_price_and_reject_invalid_price(self):
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-PRICE-CLEAR",
                    "series": "OLD",
                    "item": "PRICE TEST",
                    "oe_no_1": "PRICE-CLEAR-OE",
                    "models": "Tester",
                    "price_cny": "88.5",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        clear = self.client.post(
            "/products/save",
            data={
                "bld_no": "K-PRICE-CLEAR",
                "series": "CLEARED",
                "item": "PRICE TEST",
                "oe_no_1": "PRICE-CLEAR-OE",
                "oe_no_2": "",
                "models": "Tester",
                "price_cny": "",
                "active": "1",
            },
            follow_redirects=False,
        )
        self.assertEqual(clear.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-PRICE-CLEAR",)).fetchone()
        self.assertEqual(product["series"], "CLEARED")
        self.assertIsNone(product["price_cny"])

        invalid = self.client.post(
            "/products/save",
            data={
                "bld_no": "K-PRICE-CLEAR",
                "series": "BAD",
                "item": "PRICE TEST",
                "oe_no_1": "PRICE-CLEAR-OE",
                "oe_no_2": "",
                "models": "Tester",
                "price_cny": "abc",
                "active": "1",
            },
            follow_redirects=False,
        )
        self.assertEqual(invalid.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-PRICE-CLEAR",)).fetchone()
        self.assertEqual(product["series"], "CLEARED")
        self.assertIsNone(product["price_cny"])

    def test_product_status_can_be_edited_and_shown_in_catalog(self):
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-STATUS-001",
                    "series": "TEST",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "STATUS-OE-001",
                    "models": "Tester",
                    "price_cny": "45",
                    "product_status": "1 个球头 2 个衬套",
                    "active": "1",
                },
                actor="tester",
            )
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-STATUS-001",)).fetchone()

        self.login()
        page = self.client.get("/products", query_string={"bld": "K-STATUS-001"})
        html = page.get_data(as_text=True)
        self.assertEqual(page.status_code, 200)
        self.assertIn('data-col="product-status"', html)
        self.assertIn('data-column-label="产品状态"', html)
        self.assertIn('aria-label="筛选产品状态', html)
        self.assertIn("1 ball joint 2 bushings", html)

        edit = self.client.get(f"/products/{product['id']}/edit")
        edit_html = edit.get_data(as_text=True)
        self.assertEqual(edit.status_code, 200)
        self.assertIn('name="product_status"', edit_html)
        self.assertIn("1 个球头 2 个衬套", edit_html)

        save = self.client.post(
            "/products/save",
            data={
                "bld_no": "K-STATUS-001",
                "series": "TEST",
                "item": "Front Left Lower Control Arm",
                "oe_no_1": "STATUS-OE-001",
                "oe_no_2": "",
                "models": "Tester",
                "price_cny": "45",
                "product_status": "0 个球头 1 个衬套",
                "active": "1",
            },
            follow_redirects=False,
        )
        self.assertEqual(save.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            updated = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-STATUS-001",)).fetchone()
        self.assertEqual(updated["product_status"], "0 个球头 1 个衬套")

    def test_product_save_rejects_invalid_image_before_updating_fields(self):
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-IMAGE-FAIL",
                    "series": "OLD",
                    "item": "IMAGE TEST",
                    "oe_no_1": "IMAGE-FAIL-OE",
                    "models": "Tester",
                    "price_cny": "55",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.post(
            "/products/save",
            data={
                "bld_no": "K-IMAGE-FAIL",
                "series": "NEW",
                "item": "IMAGE TEST UPDATED",
                "oe_no_1": "IMAGE-FAIL-OE",
                "oe_no_2": "",
                "models": "Tester",
                "price_cny": "66",
                "active": "1",
                "product_image_1": (io.BytesIO(b"not really a png"), "K-IMAGE-FAIL.png"),
            },
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-IMAGE-FAIL",)).fetchone()
        self.assertEqual(product["series"], "OLD")
        self.assertEqual(product["item"], "IMAGE TEST")
        self.assertEqual(product["price_cny"], 55)
        self.assertEqual(product["image_path"], "")

    def test_product_image_table_uses_generated_thumbnail(self):
        from PIL import Image

        from app.modules.products.persistence import upsert_product

        image_dir = self.root / "data" / "product_images"
        image_dir.mkdir(parents=True, exist_ok=True)
        image_path = image_dir / "K-THUMB-001.png"
        Image.new("RGB", (640, 360), "white").save(image_path)

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-THUMB-001",
                    "series": "TEST",
                    "item": "THUMB PART",
                    "oe_no_1": "THUMB-001",
                    "models": "Tester",
                    "image_path": "data_product_images/K-THUMB-001.png",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/products", query_string={"bld": "K-THUMB-001"})
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("/product-image-thumbs/K-THUMB-001.png", html)
        self.assertIn("/product-images/K-THUMB-001.png", html)
        self.assertNotIn('src="/product-images/K-THUMB-001.png"', html)

        thumb = self.client.get("/product-image-thumbs/K-THUMB-001.png")
        self.assertEqual(thumb.status_code, 200)
        with Image.open(io.BytesIO(thumb.get_data())) as generated:
            self.assertEqual(generated.format, "WEBP")
            self.assertLessEqual(generated.width, 320)
            self.assertLessEqual(generated.height, 240)
        thumb.close()
        self.assertTrue((image_dir / "thumbs" / "K-THUMB-001.webp").exists())

    def test_missing_product_thumbnail_is_a_small_safe_image_response(self):
        self.login()
        missing = self.client.get(
            "/product-image-thumbs/DOES-NOT-EXIST.png",
            follow_redirects=False,
        )

        self.assertEqual(missing.status_code, 200)
        self.assertEqual(missing.mimetype, "image/svg+xml")
        self.assertLess(len(missing.get_data()), 1024)
        self.assertTrue(missing.get_data().startswith(b"<svg "))
        self.assertNotIn("Location", missing.headers)
        self.assertEqual(missing.headers["Cache-Control"], "no-store")
        self.assertEqual(missing.headers["X-Content-Type-Options"], "nosniff")

        self.client.post("/logout")
        anonymous = self.client.get(
            "/product-image-thumbs/DOES-NOT-EXIST.png",
            follow_redirects=False,
        )
        self.assertEqual(anonymous.status_code, 302)
        self.assertIn("/login", anonymous.headers["Location"])

    def test_product_edit_can_delete_product(self):
        from app.modules.inquiry.persistence import save_alias
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-DELETE-001",
                    "series": "TEST",
                    "item": "DELETE TARGET",
                    "oe_no_1": "DELETE-001",
                    "models": "Tester",
                    "active": "1",
                },
                actor="tester",
            )
            save_alias(conn, "DELETE-ALIAS-001", "K-DELETE-001", actor="tester")
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-DELETE-001",)).fetchone()

        self.login()
        edit = self.client.get(f"/products/{product['id']}/edit")
        edit_html = edit.get_data(as_text=True)
        self.assertEqual(edit.status_code, 200)
        self.assertIn("删除产品", edit_html)
        self.assertIn(f'formaction="/products/{product["id"]}/delete"', edit_html)
        self.assertIn('data-confirm="确认删除 K-DELETE-001', edit_html)

        delete = self.client.post(f"/products/{product['id']}/delete", follow_redirects=False)
        self.assertEqual(delete.status_code, 302)
        self.assertTrue(delete.headers["Location"].endswith("/products"))

        with self.web.connect(self.web.DB_PATH) as conn:
            deleted = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K-DELETE-001",)).fetchone()
            alias = conn.execute("SELECT * FROM aliases WHERE source_code = ?", ("DELETEALIAS001",)).fetchone()
            log = conn.execute(
                "SELECT * FROM audit_logs WHERE action = ? AND target_key = ? ORDER BY id DESC LIMIT 1",
                ("删除产品", "K-DELETE-001"),
            ).fetchone()

        self.assertIsNone(deleted)
        self.assertEqual(alias["active"], 0)
        self.assertIsNotNone(log)
        self.assertIn("DELETE TARGET", log["detail"])

from __future__ import annotations

from tests.web_app_test_base import (
    WebAppTestBase,
    Path,
    io,
    json,
    pollute_xlsx_tail,
    re,
    strip_xlsx_dimension,
    zipfile,
)


class TestWebInquiryExport(WebAppTestBase):
    def test_uploaded_inquiry_can_export_tax_price(self):
        from PIL import Image
        from app.modules.products.persistence import upsert_product
        from openpyxl import Workbook, load_workbook

        image_dir = self.root / "data" / "product_images"
        image_dir.mkdir(parents=True, exist_ok=True)
        Image.new("RGB", (320, 180), "white").save(image_dir / "KPRICE01.png")
        Image.new("RGB", (320, 180), "blue").save(image_dir / "KPRICE02.png")
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "KPRICE01",
                    "series": "HYUNDAI",
                    "item": "CONTROL ARM",
                    "oe_no_1": "PRICE-001",
                    "models": "Elantra",
                    "price_cny": "88.8",
                    "image_path": "data_product_images/KPRICE01.png",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "KPRICE02",
                    "series": "HYUNDAI",
                    "item": "CONTROL ARM WITH BALL JOINT",
                    "oe_no_1": "PRICE-002",
                    "models": "Elantra",
                    "price_cny": "99.9",
                    "product_status": "1个球头",
                    "image_path": "data_product_images/KPRICE02.png",
                    "active": "1",
                },
                actor="tester",
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["PRICE-001"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        self.login()
        lookup = self.client.get("/products/lookup?q=KPRICE02&details=1&active_only=1&media=1")
        lookup_payload = lookup.get_json()
        self.assertEqual(lookup.status_code, 200)
        self.assertEqual(lookup_payload[0]["bld_no"], "KPRICE02")
        self.assertEqual(lookup_payload[0]["image_gallery"][0]["url"], "/product-images/KPRICE02.png")
        self.assertEqual(
            lookup_payload[0]["image_gallery"][0]["thumb"],
            "/product-image-thumbs/KPRICE02.png",
        )
        response = self.client.post(
            "/match",
            data={"inquiry": (buffer, "price-export.xlsx")},
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("选择匹配列", html)
        self.assertIn("没有识别到明确的 OE 号码表头", html)
        self.assertNotIn("KPRICE01", html)
        self.assertNotIn('id="download-excel-modal"', html)

        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)
        upload_path = upload_match.group(1)
        output_name = output_match.group(1)
        output_path = self.root / "outputs" / "u1-007" / output_name

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_path,
                "original_filename": "price-export.xlsx",
                "output_name": output_name,
                "match_column": "0",
            },
        )
        result_html = result.get_data(as_text=True)

        self.assertEqual(result.status_code, 200)
        self.assertIn("KPRICE01", result_html)
        self.assertIn('value="88.80" data-inquiry-tax-price', result_html)
        self.assertNotIn('<span aria-hidden="true">¥</span>', result_html)
        self.assertIn('value="KPRICE01"', result_html)
        self.assertIn('data-inquiry-bld-input', result_html)
        self.assertIn('id="inquiry-bld-options"', result_html)
        self.assertIn('data-product-lookup-url="/products/lookup"', result_html)
        self.assertIn('data-col="image"', result_html)
        self.assertIn('/product-image-thumbs/KPRICE01.png', result_html)
        self.assertIn('/product-images/KPRICE01.png', result_html)
        self.assertNotIn('data-open-product-adjustment', result_html)
        self.assertNotIn('id="product-adjustment-modal"', result_html)
        self.assertIn('id="download-excel-modal"', result_html)
        self.assertIn('value="net">带不含税单价', result_html)
        self.assertIn("返回上一步", result_html)

        download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_path,
                "original_filename": "price-export.xlsx",
                "output_name": output_name,
                "match_column": "0",
                "price_mode": "tax",
            },
        )
        self.assertEqual(download.status_code, 200)
        download.close()
        self.assertTrue(output_path.exists())

        generated = load_workbook(output_path)
        sheet = generated.active
        self.assertEqual(len(sheet._images), 0)
        self.assertNotIn("产品图片", [cell.value for cell in sheet[1]])
        self.assertEqual(sheet.cell(1, 2).value, "BLD NO.")
        self.assertEqual(sheet.cell(1, 3).value, "含税单价")
        self.assertEqual(sheet.cell(1, 4).value, "产品状态")
        self.assertEqual(sheet.cell(1, 5).value, "匹配说明")
        self.assertEqual(sheet.cell(2, 2).value, "KPRICE01")
        self.assertEqual(sheet.cell(2, 3).value, 88.8)
        generated.close()

        net_download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_path,
                "original_filename": "price-export.xlsx",
                "output_name": output_name,
                "match_column": "0",
                "price_mode": "net",
            },
        )
        self.assertEqual(net_download.status_code, 200)
        net_download.close()

        generated = load_workbook(output_path)
        sheet = generated.active
        self.assertEqual(sheet.cell(1, 3).value, "不含税单价")
        self.assertEqual(sheet.cell(2, 3).value, 81)
        self.assertEqual(sheet.cell(2, 3).number_format, "0")
        generated.close()

        adjusted_download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_path,
                "original_filename": "price-export.xlsx",
                "output_name": output_name,
                "match_column": "0",
                "price_mode": "tax",
                "inquiry_adjustments": json.dumps(
                    {
                        "sheet:1:row:2": {
                            "expected_bld_no": "KPRICE01",
                            "target_bld_no": "KPRICE02",
                            "tax_price": "123.45",
                        }
                    }
                ),
            },
        )
        self.assertEqual(adjusted_download.status_code, 200)
        adjusted_download.close()

        generated = load_workbook(output_path)
        sheet = generated.active
        self.assertEqual(len(sheet._images), 0)
        self.assertEqual(sheet.cell(2, 2).value, "KPRICE02")
        self.assertEqual(sheet.cell(2, 3).value, 123.45)
        self.assertIn("KPRICE01 → KPRICE02", sheet.cell(2, 5).value)
        generated.close()
        with self.web.connect(self.web.DB_PATH) as conn:
            original_price = conn.execute(
                "SELECT price_cny FROM products WHERE bld_no = ?", ("KPRICE01",)
            ).fetchone()["price_cny"]
        self.assertEqual(original_price, 88.8)

    def test_uploaded_polluted_xlsx_uses_cleaned_copy_without_skipping_late_rows(self):
        from app.modules.products.persistence import upsert_product
        from openpyxl import Workbook, load_workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no, oe_no in [("KCLEAN01", "CLEAN-002"), ("KCLEAN02", "CLEAN-250")]:
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "HYUNDAI",
                        "item": "CLEAN TEST ARM",
                        "oe_no_1": oe_no,
                        "active": "1",
                    },
                    actor="tester",
                )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["CLEAN-002"])
        sheet.cell(250, 1).value = "CLEAN-250"
        polluted_path = self.root / "uploads" / "polluted-inquiry.xlsx"
        polluted_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(polluted_path)
        workbook.close()
        pollute_xlsx_tail(polluted_path, declared_rows=2000, after_row=251)

        self.login()
        with polluted_path.open("rb") as handle:
            response = self.client.post(
                "/match",
                data={"inquiry": (io.BytesIO(handle.read()), "polluted-inquiry.xlsx")},
                content_type="multipart/form-data",
            )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("已自动清理 Excel 尾部空白格式", html)
        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)
        self.assertIn("inquiry-cleaned", upload_match.group(1))
        cleaned_workbook = load_workbook(Path(upload_match.group(1)), read_only=True, data_only=True)
        self.assertEqual(cleaned_workbook.active.cell(2, 1).value, "CLEAN-002")
        cleaned_workbook.close()

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "polluted-inquiry.xlsx",
                "output_name": output_match.group(1),
                "match_column": "0",
            },
        )
        result_html = result.get_data(as_text=True)

        self.assertEqual(result.status_code, 200)
        self.assertIn("共 2 行，命中 2 行，未找到 0 行", result_html)
        self.assertIn("KCLEAN01", result_html)
        self.assertIn("KCLEAN02", result_html)
        self.assertIn('<td data-col="row">250</td>', result_html)

    def test_uploaded_inquiry_can_match_multiple_selected_columns(self):
        from app.modules.products.persistence import upsert_product
        from openpyxl import Workbook, load_workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "KMULTI02",
                    "series": "HYUNDAI",
                    "item": "MULTI COLUMN ARM",
                    "oe_no_1": "REF-MULTI-002",
                    "price_cny": "77",
                    "active": "1",
                },
                actor="tester",
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["客户OE", "参考号"])
        sheet.append(["NO-HIT-001", "REF-MULTI-002"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        self.login()
        response = self.client.post(
            "/match",
            data={"inquiry": (buffer, "multi-column.xlsx")},
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('name="match_columns" value="0"', html)
        self.assertIn('name="match_columns" value="1"', html)
        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "multi-column.xlsx",
                "output_name": output_match.group(1),
                "match_columns": ["0", "1"],
            },
        )
        result_html = result.get_data(as_text=True)

        self.assertEqual(result.status_code, 200)
        self.assertIn("KMULTI02", result_html)
        self.assertIn("命中列：B列：REF-MULTI-002", result_html)

        output_path = self.root / "outputs" / "u1-007" / output_match.group(1)
        download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "multi-column.xlsx",
                "output_name": output_match.group(1),
                "match_columns": ["0", "1"],
                "price_mode": "tax",
            },
        )
        self.assertEqual(download.status_code, 200)
        download.close()

        generated = load_workbook(output_path)
        generated_sheet = generated.active
        self.assertEqual(generated_sheet.cell(1, 3).value, "BLD NO.")
        self.assertEqual(generated_sheet.cell(2, 3).value, "KMULTI02")
        self.assertIn("命中列：B列：REF-MULTI-002", generated_sheet.cell(2, 6).value)
        generated.close()

    def test_match_result_can_write_quotes_with_customer_code_column(self):
        from app.modules.products.persistence import upsert_product
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no, oe_code, price in (
                ("KWQ01", "WQ-OE-001", "100"),
                ("KWQ02", "WQ-OE-002", "55"),
                ("KWQ03", "WQ-OE-003", ""),
            ):
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "TEST",
                        "item": "WRITE QUOTE ARM",
                        "oe_no_1": oe_code,
                        "price_cny": price,
                        "active": "1",
                    },
                    actor="tester",
                )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["客户编码", "参考号"])
        sheet.append(["CUST-A1", "WQ-OE-001"])
        sheet.append(["CUST-A1", "WQ-OE-001"])
        sheet.append(["CUST-B2", "WQ-OE-001"])
        sheet.append(["CUST-C3", "WQ-OE-002"])
        sheet.append(["CUST-D4", "WQ-OE-003"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        self.login()
        self.register_customer_and_product("测试客户WQ", "KWQ01")
        response = self.client.post(
            "/match",
            data={"inquiry": (buffer, "write-quotes.xlsx")},
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn('name="customer_code_column"', html)
        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "write-quotes.xlsx",
                "output_name": output_match.group(1),
                "match_columns": ["1"],
                "customer_code_column": "0",
            },
        )
        result_html = result.get_data(as_text=True)
        self.assertEqual(result.status_code, 200)
        self.assertIn('data-col="customer-code"', result_html)
        self.assertNotIn('data-col="name"', result_html)
        self.assertIn("CUST-A1", result_html)
        self.assertIn("写入报价", result_html)
        self.assertIn("重新查询", result_html)
        result_rows = re.findall(r"<tr[^>]*>.*?</tr>", result_html, flags=re.DOTALL)
        no_catalog_price_row = next(
            (row for row in result_rows if 'data-default-bld="KWQ03"' in row),
            "",
        )
        self.assertTrue(no_catalog_price_row)
        self.assertIn('value="KWQ03"', no_catalog_price_row)
        self.assertIn("data-inquiry-bld-input", no_catalog_price_row)
        self.assertIn("data-inquiry-tax-price", no_catalog_price_row)

        write = self.client.post(
            "/match/write-quotes",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "write-quotes.xlsx",
                "match_columns": ["1"],
                "customer_code_column": "0",
                "price_mode": "tax",
                "customer_name": "测试客户WQ",
                "remark": "询价写入",
            },
        )
        self.assertEqual(write.status_code, 302)
        self.assertIn("/quotes", write.headers["Location"])

        with self.web.connect(self.web.DB_PATH) as conn:
            rows = conn.execute(
                """
                SELECT bld_no, customer_product_code, tax_price, net_price, currency, remark, source_type
                FROM quote_records
                WHERE customer_name = '测试客户WQ'
                ORDER BY bld_no, customer_product_code
                """
            ).fetchall()

        self.assertEqual(len(rows), 3)
        self.assertEqual(rows[0]["bld_no"], "KWQ01")
        self.assertEqual(rows[0]["customer_product_code"], "CUST-A1")
        self.assertEqual(rows[0]["tax_price"], 100)
        self.assertIsNone(rows[0]["net_price"])
        self.assertEqual(rows[0]["currency"], "CNY")
        self.assertEqual(rows[0]["remark"], "询价写入")
        self.assertEqual(rows[0]["source_type"], "excel")
        self.assertEqual(rows[1]["customer_product_code"], "CUST-B2")
        self.assertEqual(rows[2]["bld_no"], "KWQ02")
        self.assertEqual(rows[2]["tax_price"], 55)
        self.assertFalse(any(row["bld_no"] == "KWQ03" for row in rows))

        write_json = self.client.post(
            "/match/write-quotes",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "write-quotes.xlsx",
                "match_columns": ["1"],
                "customer_code_column": "0",
                "price_mode": "tax",
                "customer_name": "测试客户WQ",
                "remark": "询价写入",
            },
            headers={"X-Requested-With": "fetch", "Accept": "application/json"},
        )
        self.assertEqual(write_json.status_code, 200)
        payload = write_json.get_json()
        self.assertTrue(payload["ok"])
        self.assertEqual(payload["written"], 3)
        quote_no = payload["quote_no"]
        self.assertRegex(quote_no, r"^Q\d{9}$")
        self.assertIn(quote_no, payload["quotes_url"])
        self.assertIn("/download/", payload["download_url"])
        self.assertTrue(payload["download_filename"].endswith(".xlsx"))

        with self.web.connect(self.web.DB_PATH) as conn:
            batch = conn.execute(
                "SELECT id, bld_no, customer_product_code, quote_no, attachment_path FROM quote_records WHERE quote_no = ? ORDER BY id",
                (quote_no,),
            ).fetchall()
        self.assertEqual(len(batch), 3)
        self.assertTrue(all(row["quote_no"] == quote_no for row in batch))
        self.assertTrue(all(row["attachment_path"] for row in batch))

        adjusted_write = self.client.post(
            "/match/write-quotes",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "write-quotes.xlsx",
                "match_columns": ["1"],
                "customer_code_column": "0",
                "price_mode": "tax",
                "customer_name": "测试客户WQ",
                "remark": "调整后报价",
                "inquiry_adjustments": json.dumps(
                    {
                        "sheet:1:row:2": {
                            "expected_bld_no": "KWQ01",
                            "target_bld_no": "KWQ02",
                            "tax_price": "123.45",
                        }
                    }
                ),
            },
            headers={"X-Requested-With": "fetch", "Accept": "application/json"},
        )
        self.assertEqual(adjusted_write.status_code, 200)
        adjusted_payload = adjusted_write.get_json()
        adjusted_quote_no = adjusted_payload["quote_no"]
        with self.web.connect(self.web.DB_PATH) as conn:
            adjusted = conn.execute(
                "SELECT bld_no, tax_price, attachment_path FROM quote_records WHERE quote_no = ? AND customer_product_code = ?",
                (adjusted_quote_no, "CUST-A1"),
            ).fetchone()
        self.assertEqual(adjusted["bld_no"], "KWQ02")
        self.assertEqual(adjusted["tax_price"], 123.45)
        self.assertNotEqual(adjusted["attachment_path"], batch[0]["attachment_path"])
        self.assertTrue((self.root / "outputs" / batch[0]["attachment_path"]).is_file())
        self.assertTrue((self.root / "outputs" / adjusted["attachment_path"]).is_file())

        no_catalog_price_write = self.client.post(
            "/match/write-quotes",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "write-quotes.xlsx",
                "match_columns": ["1"],
                "customer_code_column": "0",
                "price_mode": "tax",
                "customer_name": "测试客户WQ",
                "remark": "无目录价临时报价",
                "inquiry_adjustments": json.dumps(
                    {
                        "sheet:1:row:6": {
                            "expected_bld_no": "KWQ03",
                            "tax_price": "66.60",
                        }
                    }
                ),
            },
            headers={"X-Requested-With": "fetch", "Accept": "application/json"},
        )
        self.assertEqual(no_catalog_price_write.status_code, 200)
        no_catalog_price_quote_no = no_catalog_price_write.get_json()["quote_no"]
        with self.web.connect(self.web.DB_PATH) as conn:
            no_catalog_price_quote = conn.execute(
                "SELECT bld_no, tax_price FROM quote_records WHERE quote_no = ? AND customer_product_code = ?",
                (no_catalog_price_quote_no, "CUST-D4"),
            ).fetchone()
        self.assertEqual(no_catalog_price_quote["bld_no"], "KWQ03")
        self.assertEqual(no_catalog_price_quote["tax_price"], 66.6)

        detail = self.client.get(f"/quotes/number/{quote_no}")
        detail_html = detail.get_data(as_text=True)
        self.assertEqual(detail.status_code, 200)
        self.assertIn("CUST-A1", detail_html)
        self.assertIn("CUST-B2", detail_html)
        self.assertIn("CUST-C3", detail_html)
        self.assertIn("报价文件", detail_html)
        self.assertIn('class="quote-number-table-scroll"', detail_html)
        self.assertIn('data-quote-number-table-scroll', detail_html)
        self.assertIn('class="quote-contract-settings"', detail_html)
        self.assertIn('class="quote-contract-language-options"', detail_html)
        self.assertIn('class="quote-contract-actions"', detail_html)
        self.assertIn("销售合同版本", detail_html)
        self.assertIn('type="checkbox" name="language" value="zh-CN"', detail_html)
        self.assertNotIn('type="radio" name="language"', detail_html)
        attachment_url = re.search(r'href="(/quotes/number/[^\"]+/attachment/\d+)"', detail_html)
        self.assertIsNotNone(attachment_url)
        attachment = self.client.get(attachment_url.group(1))
        self.assertEqual(attachment.status_code, 200)
        self.assertIn("attachment", attachment.headers.get("Content-Disposition", ""))
        attachment.close()

        list_page = self.client.get(f"/quotes?quote_no={quote_no}")
        list_html = list_page.get_data(as_text=True)
        self.assertEqual(list_page.status_code, 200)
        self.assertLess(list_html.index("CUST-A1"), list_html.index("CUST-B2"))
        self.assertLess(list_html.index("CUST-B2"), list_html.index("CUST-C3"))

    def test_match_write_quotes_requires_price_mode_and_customer_name(self):
        from app.modules.products.persistence import upsert_product
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "KWQ09",
                    "series": "TEST",
                    "item": "WRITE QUOTE GUARD",
                    "oe_no_1": "WQ-OE-009",
                    "price_cny": "88",
                    "active": "1",
                },
                actor="tester",
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["参考号"])
        sheet.append(["WQ-OE-009"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        self.login()
        self.register_customer_and_product("测试客户WQ2", "KWQ09")
        response = self.client.post(
            "/match",
            data={"inquiry": (buffer, "write-quotes-guard.xlsx")},
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)
        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)

        missing_customer = self.client.post(
            "/match/write-quotes",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "write-quotes-guard.xlsx",
                "match_columns": ["0"],
                "price_mode": "tax",
            },
            follow_redirects=True,
        )
        self.assertIn("写入报价前请填写客户名称", missing_customer.get_data(as_text=True))

        missing_price_mode = self.client.post(
            "/match/write-quotes",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "write-quotes-guard.xlsx",
                "match_columns": ["0"],
                "price_mode": "none",
                "customer_name": "测试客户WQ2",
            },
            follow_redirects=True,
        )
        self.assertIn("未选择单价方式", missing_price_mode.get_data(as_text=True))

        with self.web.connect(self.web.DB_PATH) as conn:
            count = conn.execute("SELECT COUNT(*) AS total FROM quote_records WHERE bld_no = 'KWQ09'").fetchone()
        self.assertEqual(count["total"], 0)

    def test_item_header_with_code_values_prompts_for_match_column(self):
        from app.modules.products.persistence import upsert_product
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "KPIKA01",
                    "series": "HYUNDAI",
                    "item": "LOWER ARM",
                    "oe_no_1": "TST545012B000",
                    "models": "SANTA FE 2006",
                    "price_cny": "66",
                    "active": "1",
                },
                actor="tester",
            )

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["PIKA BOOKING ORDER-200426"])
        sheet.append([])
        sheet.append(["SN", "ITEM", "DESCRIPTION", "QTY", "PRICE", "PICTURE", "BRAND"])
        sheet.append(["1", "TST545012B000 ", "LOWER ARM-HYUNDAI SANTA FE 2006", 60, None, None, "L-TGL"])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        self.login()
        response = self.client.post(
            "/match",
            data={"inquiry": (buffer, "pika-order.xlsx")},
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("选择匹配列", html)
        self.assertIn("没有识别到明确的 OE 号码表头", html)
        self.assertIn("match-preview-table", html)
        self.assertIn("match-preview-cell", html)
        self.assertIn("<th>源行</th>", html)

        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "pika-order.xlsx",
                "output_name": output_match.group(1),
                "match_column": "1",
            },
        )
        result_html = result.get_data(as_text=True)

        self.assertEqual(result.status_code, 200)
        self.assertIn("共 1 行，命中 1 行，未找到 0 行", result_html)
        self.assertIn("KPIKA01", result_html)
        self.assertIn('value="66.00" data-inquiry-tax-price', result_html)
        self.assertIn('<td data-col="row">4</td>', result_html)
        self.assertNotIn('<td data-col="row">3</td>', result_html)

    def test_xlsx_without_dimension_can_preview_match_columns(self):
        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["فحمات", None])
        sheet.append(["رقم", "كمية"])
        sheet.append([446633220, 300])
        sheet.append(["58101F2A00", 300])
        inquiry_path = self.root / "uploads" / "arabic-no-dimension.xlsx"
        inquiry_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(inquiry_path)
        workbook.close()
        strip_xlsx_dimension(inquiry_path)

        self.login()
        with inquiry_path.open("rb") as handle:
            response = self.client.post(
                "/match",
                data={"inquiry": (handle, "هونداي و تويوتا.xlsx")},
                content_type="multipart/form-data",
            )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("选择匹配列", html)
        self.assertIn("فحمات", html)
        self.assertIn("رقم", html)
        self.assertIn('name="match_columns" value="0"', html)
        self.assertIn('name="match_columns" value="1"', html)
        self.assertNotIn("生成失败", html)

    def test_segmented_merged_headers_do_not_count_as_inquiry_rows(self):
        from app.modules.products.persistence import upsert_product
        from openpyxl import Workbook, load_workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            for index in range(1, 5):
                upsert_product(
                    conn,
                    {
                        "bld_no": f"KSEG{index:02d}",
                        "series": "TEST",
                        "item": "SEGMENTED ARM",
                        "oe_no_1": f"SEG-OE-{index:03d}",
                        "active": "1",
                    },
                    actor="tester",
                )

        workbook = Workbook()
        sheet = workbook.active
        row = 1
        for section in range(2):
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            sheet.cell(row, 1).value = f"分段 {section + 1}"
            row += 1
            sheet.append(["序号", "OE号", "数量"])
            row += 1
            for item in range(2):
                number = section * 2 + item + 1
                sheet.append([number, f"SEG-OE-{number:03d}", 10])
                row += 1
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        buffer.seek(0)

        self.login()
        response = self.client.post(
            "/match",
            data={"inquiry": (buffer, "segmented-merged.xlsx")},
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "segmented-merged.xlsx",
                "output_name": output_match.group(1),
                "match_column": "1",
            },
        )
        result_html = result.get_data(as_text=True)

        self.assertEqual(result.status_code, 200)
        self.assertIn("共 4 行，命中 4 行，未找到 0 行", result_html)
        self.assertNotIn('<td data-col="row">2</td>', result_html)
        self.assertNotIn('<td data-col="row">6</td>', result_html)
        for index in range(1, 5):
            self.assertIn(f"KSEG{index:02d}", result_html)

        download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_match.group(1),
                "original_filename": "segmented-merged.xlsx",
                "output_name": output_match.group(1),
                "match_column": "1",
            },
        )
        self.assertEqual(download.status_code, 200)
        download.close()
        generated = load_workbook(self.root / "outputs" / "u1-007" / output_match.group(1), data_only=True)
        generated_sheet = generated.active
        self.assertEqual(generated_sheet.cell(2, 4).value, "BLD NO.")
        self.assertEqual(generated_sheet.cell(3, 4).value, "KSEG01")
        self.assertEqual(generated_sheet.cell(4, 4).value, "KSEG02")
        self.assertIsNone(generated_sheet.cell(6, 4).value)
        self.assertEqual(generated_sheet.cell(7, 4).value, "KSEG03")
        self.assertEqual(generated_sheet.cell(8, 4).value, "KSEG04")
        generated.close()

    def test_manual_column_result_defers_excel_until_download(self):
        from app.modules.products.persistence import upsert_product
        from openpyxl import Workbook, load_workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K6004LC",
                    "series": "HYUNDAI",
                    "item": "CONTROL ARM",
                    "oe_no_1": "55270-2Z001",
                    "models": "Sportage",
                    "price_cny": "55",
                    "active": "1",
                },
                actor="tester",
            )
            product = conn.execute("SELECT * FROM products WHERE bld_no = ?", ("K6004LC",)).fetchone()

        self.login()
        drawing_upload = self.client.post(
            f"/products/{product['id']}/drawing",
            data={"drawing": (io.BytesIO(b"%PDF-1.4\nK6004LC drawing\n%%EOF"), "K6004LC.pdf")},
            content_type="multipart/form-data",
            follow_redirects=False,
        )
        self.assertEqual(drawing_upload.status_code, 302)

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["客户号码", "数量"])
        sheet.append(["55270-2Z001", 1])
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = self.client.post(
            "/match",
            data={"inquiry": (buffer, "manual-column.xlsx")},
            content_type="multipart/form-data",
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("选择匹配列", html)
        self.assertIn("返回上一步", html)
        self.assertNotIn("返回首页", html)
        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)
        upload_path = upload_match.group(1)
        output_name = output_match.group(1)
        output_path = self.root / "outputs" / "u1-007" / output_name

        result = self.client.post(
            "/match/column",
            data={
                "upload_path": upload_path,
                "original_filename": "manual-column.xlsx",
                "output_name": output_name,
                "match_column": "0",
            },
        )
        result_html = result.get_data(as_text=True)

        self.assertEqual(result.status_code, 200)
        self.assertIn("Excel 文件将在点击下载时生成", result_html)
        self.assertIn("下载 Excel", result_html)
        self.assertIn("下载图纸包", result_html)
        self.assertIn("返回上一步", result_html)
        self.assertNotIn("返回首页", result_html)
        self.assertIn("K6004LC", result_html)
        self.assertIn('value="55.00" data-inquiry-tax-price', result_html)
        self.assertIn('id="download-excel-modal"', result_html)
        self.assertIn('name="price_mode"', result_html)
        self.assertFalse(output_path.exists())

        drawing_zip = self.client.post(
            "/match/drawings/download",
            data={
                "upload_path": upload_path,
                "original_filename": "manual-column.xlsx",
                "match_column": "0",
            },
        )
        self.assertEqual(drawing_zip.status_code, 200)
        with zipfile.ZipFile(io.BytesIO(drawing_zip.get_data())) as archive:
            self.assertIn("K6004LC_55270-2Z001.pdf", archive.namelist())
        drawing_zip.close()

        back = self.client.post(
            "/match/column/back",
            data={
                "upload_path": upload_path,
                "original_filename": "manual-column.xlsx",
                "output_name": output_name,
                "match_column": "0",
            },
        )
        back_html = back.get_data(as_text=True)
        self.assertEqual(back.status_code, 200)
        self.assertIn("选择匹配列", back_html)
        self.assertIn("返回上一步", back_html)
        self.assertNotIn("返回首页", back_html)
        self.assertRegex(back_html, r'name="match_columns" value="0"[^>]*checked')

        download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_path,
                "original_filename": "manual-column.xlsx",
                "output_name": output_name,
                "match_column": "0",
                "price_mode": "tax",
            },
        )
        self.assertEqual(download.status_code, 200)
        download.close()
        self.assertTrue(output_path.exists())

        generated = load_workbook(output_path)
        generated_sheet = generated.active
        self.assertEqual(generated_sheet.cell(1, 3).value, "BLD NO.")
        self.assertEqual(generated_sheet.cell(1, 4).value, "含税单价")
        self.assertEqual(generated_sheet.cell(1, 5).value, "产品状态")
        self.assertEqual(generated_sheet.cell(1, 6).value, "匹配说明")
        self.assertEqual(generated_sheet.cell(2, 3).value, "K6004LC")
        self.assertEqual(generated_sheet.cell(2, 4).value, 55)
        generated.close()

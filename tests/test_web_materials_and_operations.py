from __future__ import annotations

import json
import zipfile
from html import unescape

from tests.web_app_test_base import (
    WebAppTestBase,
    PROJECT_ROOT,
    SimpleNamespace,
    io,
    patch,
    re,
    subprocess,
    sys,
)


class TestWebMaterialsAndOperations(WebAppTestBase):
    def test_primary_data_grid_footers_render_statistics_and_zero_ranges(self):
        self.login()
        product_service = SimpleNamespace(
            search=lambda filters, limit, offset: SimpleNamespace(records=[], total=0),
            stats=lambda: SimpleNamespace(as_dict=lambda: {"products": 17, "active": 13, "inactive": 4, "aliases": 2}),
            filter_options=lambda filters: SimpleNamespace(
                web_payload=lambda: {"brand": [], "item": [], "product_status": []}
            ),
            preview_brand_normalization=lambda: None,
        )
        quote_service = SimpleNamespace(
            list_records=lambda filters, limit, offset: SimpleNamespace(records=[], total=0),
            stats=lambda: SimpleNamespace(as_dict=lambda: {"total": 19, "customers": 5, "models": 11}),
            filter_options=lambda filters: {},
        )

        class FooterTubeService:
            def list_items(self, *, filters, limit, offset):
                return {
                    "records": [],
                    "total": 0,
                    "counts": {"sentinel": 23},
                    "blank_length_options": [],
                    "inner_tolerance_options": [],
                    "purchase_base_options": [],
                    "tolerance_options": [],
                    "consumption_options": [],
                }

        material_service = SimpleNamespace(
            list_items=lambda query, status, limit, offset, column_filters=None: SimpleNamespace(
                records=[],
                total=0,
                stats={"items": 29, "active": 21, "inactive": 8, "models": 9},
            ),
            filter_options=lambda query, status, column_filters=None: {},
            source_stats=lambda: {},
            source_path=lambda: None,
        )
        cases = (
            (
                "/products",
                {"bld": "__FOOTER_ZERO__"},
                "产品",
                "总产品 17 · 启用 13 · 停用 4",
            ),
            (
                "/quotes",
                {"customer_name": "__FOOTER_ZERO__"},
                "报价记录",
                "总记录 19 · 客户 5 · BLD号 11",
            ),
            (
                "/tubes",
                {"q": "__FOOTER_ZERO__"},
                "管件明细",
                "总管件 23",
            ),
            (
                "/materials/items",
                {"q": "__FOOTER_ZERO__"},
                "材料明细",
                "总明细 29 · 启用 21 · 停用 8",
            ),
        )
        with (
            patch(
                "app.modules.products.catalog_web.get_product_service",
                return_value=product_service,
            ),
            patch(
                "app.modules.quotes.web.get_quote_service",
                return_value=quote_service,
            ),
            patch(
                "app.modules.tubes.web.get_tube_service",
                return_value=FooterTubeService(),
            ),
            patch(
                "app.modules.materials.web.get_material_service",
                return_value=material_service,
            ),
        ):
            for path, query_string, label, statistics in cases:
                with self.subTest(path=path):
                    response = self.client.get(path, query_string=query_string)
                    self.assertEqual(response.status_code, 200)
                    html = response.get_data(as_text=True)
                    footer_match = re.search(r'<footer class="data-grid-footer">.*?</footer>', html, re.S)
                    self.assertIsNotNone(footer_match)
                    footer = footer_match.group()
                    self.assertIn(
                        f'<span class="data-grid-visually-hidden">{label}统计：</span>{statistics}',
                        footer,
                    )
                    self.assertRegex(
                        footer,
                        rf'<span class="data-grid-visually-hidden">{label}当前范围：</span>\s*'
                        r"<strong>0</strong><span>条</span>",
                    )

    def test_tube_page_refetches_after_clamping_an_out_of_range_page(self):
        self.login()
        offsets: list[int] = []
        last_page_record = SimpleNamespace(
            id=178,
            code="TUBE-LAST-PAGE",
            tube_type="测试管件",
            spec_text="Ø35 × 30",
            blank_length_text="100",
            inner_diameter_tolerance=None,
            purchase_base=1,
            weight_kg=1.25,
            tolerance_mm=0.1,
            consumption_mm=2.0,
            borrowed_codes="",
            borrowed_from="",
        )

        class FakeTubeService:
            def list_items(self, *, filters, limit, offset):
                offsets.append(offset)
                return {
                    "records": [last_page_record] if offset == 100 else [],
                    "total": 178,
                    "counts": {},
                    "blank_length_options": [],
                    "inner_tolerance_options": [],
                    "purchase_base_options": [],
                    "tolerance_options": [],
                    "consumption_options": [],
                }

        with patch("app.modules.tubes.web.get_tube_service", return_value=FakeTubeService()):
            response = self.client.get("/tubes?page=999")

        self.assertEqual(response.status_code, 200)
        body = response.get_data(as_text=True)
        self.assertEqual(offsets, [99800, 100])
        self.assertIn("TUBE-LAST-PAGE", body)
        self.assertIn("101–178", body)

    def test_material_drawings_page_lists_codes_and_previews_pdf(self):
        self.login()
        drawing_dir = self.root / "data" / "material_drawings"
        drawing_dir.mkdir(parents=True, exist_ok=True)
        (drawing_dir / "QD1000.pdf").write_bytes(b"%PDF-1.4\n% test drawing\n")
        (drawing_dir / "QD999.pdf").write_bytes(b"%PDF-1.4\n% test drawing\n")

        response = self.client.get("/material-drawings?q=1000&category=%E7%90%83%E9%94%80")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("当前 1 个 / 共 2 个", html)
        self.assertIn("<strong>QD1000</strong>", html)
        self.assertIn("球销", html)
        self.assertIn("data-material-drawing-select", html)
        self.assertIn("data-material-drawing-frame", html)
        self.assertIn("/material-drawings/preview/QD1000.pdf", html)
        self.assertIn("/material-drawings/preview/QD1000.pdf#page=1&zoom=100", html)
        self.assertNotIn("<strong>QD999</strong>", html)

        selected_page = self.client.get("/material-drawings?selected=QD999.pdf")
        selected_html = selected_page.get_data(as_text=True)
        self.assertEqual(selected_page.status_code, 200)
        self.assertIn("data-material-drawing-current-code>QD999</h2>", selected_html)
        self.assertIn('data-material-drawing-current-download href="/material-drawings/QD999.pdf"', selected_html)

        preview = self.client.get("/material-drawings/preview/QD1000.pdf")
        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.mimetype, "application/pdf")
        self.assertIn("inline", preview.headers.get("Content-Disposition", ""))
        preview.close()

    def test_system_updates_page_reads_handoff_notes(self):
        self.login()
        response = self.client.get("/system-updates")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("系统更新", html)
        self.assertIn('data-page="admin.system_updates"', html)
        self.assertIn('data-page-type="system-admin"', html)
        self.assertIn("当前最近重要变更", html)
        self.assertIn("项目交接说明.md", html)
        self.assertIn("建立长期项目治理基线并加固安全边界", html)
        self.assertIn("2026-07-10", html)
        self.assertIn("补齐系统更新记录并设为强制提交要求", html)
        self.assertIn("d1ab621", html)
        self.assertIn("改进 OpenClaw 询价命令行启动和调用体验", html)
        self.assertIn("ac3aa1a", html)
        self.assertIn("新增系统更新页面", html)
        self.assertNotIn(">unreleased<", html)
        self.assertIn("当前版本", html)

    def test_quote_and_material_column_filters_apply_on_server_and_keep_table_headers(self):
        self.login()
        self.register_customer_and_product("列筛报价甲", "QF-QUOTE-A")
        self.register_customer_and_product("列筛报价乙", "QF-QUOTE-B")
        for customer_name, bld_no, currency in (
            ("列筛报价甲", "QF-QUOTE-A", "CNY"),
            ("列筛报价乙", "QF-QUOTE-B", "USD"),
        ):
            response = self.client.post(
                "/quotes/save",
                data={
                    "customer_name": customer_name,
                    "bld_no": bld_no,
                    "tax_price": "12.34",
                    "currency": currency,
                    "quote_date": "2026-07-29",
                },
                follow_redirects=False,
            )
            self.assertEqual(response.status_code, 302)

        response = self.client.get("/quotes", query_string=[("qf_customer_name", "列筛报价甲")])
        quote_html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn('id="quotes-table"', quote_html)
        self.assertIn('data-filter-key="qf_customer_name"', quote_html)
        self.assertIn('aria-label="筛选报价单号"', quote_html)
        self.assertIn('data-col="quote-no"', quote_html)
        self.assertIn("<strong>列筛报价甲</strong>", quote_html)
        self.assertNotIn("<strong>列筛报价乙</strong>", quote_html)
        filter_state_match = re.search(r"data-quote-filter-state='([^']*)'", quote_html)
        self.assertIsNotNone(filter_state_match)
        quote_filter_state = json.loads(unescape(filter_state_match.group(1)))
        self.assertEqual(quote_filter_state["selected"]["customer_name"], ["列筛报价甲"])
        from app.modules.quotes.factory import get_quote_service

        with self.web.app.app_context():
            expected_options = get_quote_service().filter_options(
                {"column_filters": {"customer_name": ("列筛报价甲",)}}
            )
        self.assertEqual(len(quote_filter_state["options"]), 11)
        self.assertEqual(quote_filter_state["options"], expected_options)
        self.assertNotIn("data-column-filter-option><input", quote_html)

        for model, code, category in (
            ("QF-MATERIAL-A", "QF-MA", "类别甲"),
            ("QF-MATERIAL-B", "QF-MB", "类别乙"),
        ):
            response = self.client.post(
                "/materials/items/save",
                data={
                    "model": model,
                    "code": code,
                    "category": category,
                    "part": "列筛测试件",
                    "pieces": "2",
                    "spec_text": "2 100 200",
                    "active": "1",
                },
                follow_redirects=False,
            )
            self.assertEqual(response.status_code, 302)

        response = self.client.get("/materials/items", query_string=[("mf_model", "QF-MATERIAL-A")])
        material_html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn('id="materials-table"', material_html)
        self.assertIn('data-filter-key="mf_model"', material_html)
        self.assertIn('aria-label="筛选母件编码"', material_html)
        self.assertIn('data-col="model"', material_html)
        self.assertIn("<strong>QF-MATERIAL-A</strong>", material_html)
        self.assertNotIn("<strong>QF-MATERIAL-B</strong>", material_html)
        self.assertIn('value="QF-MATERIAL-A" checked', material_html)

    def test_new_material_item_uses_modal(self):
        self.login()
        response = self.client.get("/materials")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn('class="material-landing"', html)
        self.assertIn('href="/materials/items"', html)
        self.assertNotIn('id="materials-results"', html)
        self.assertIn('class="embedded-submit" type="submit">生成并下载', html)
        self.assertIn('data-file-drop-zone data-file-drop-accept=".xlsx"', html)
        self.assertIn('class="file-picker-clear" type="button" disabled>清除', html)
        self.assertIn('class="material-template-link" href="/materials/template">下载生产计划模板</a>', html)
        self.assertNotIn("生产计划 Excel", html)
        self.assertNotIn("材料文件：", html)
        self.assertNotIn("上传生产计划，按启用材料明细生成采购和下料所需料单。", html)

        response = self.client.get("/materials/items")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("data-open-material-modal", html)
        self.assertIn('id="material-modal"', html)
        self.assertIn('action="/materials/items/save"', html)
        self.assertIn('id="materials-results"', html)
        self.assertIn('class="workspace-command material-items-command"', html)
        self.assertIn('action="/materials/items#materials-results"', html)
        self.assertIn("data-enter-navigation", html)
        self.assertIn('name="spec_text"', html)
        self.assertIn('<button class="linear-button" type="submit">搜索</button>', html)
        self.assertIn('class="embedded-submit" type="submit">确认导入', html)
        self.assertIn("母件编码", html)
        self.assertIn("零件编码", html)
        self.assertRegex(html, r'<input name="code"[^>]*required')
        self.assertRegex(html, r'<input name="part"[^>]*required')
        self.assertIn('aria-label="筛选母件编码"', html)
        self.assertIn('aria-label="筛选零件编码"', html)
        self.assertIn('aria-label="筛选单件重量kg"', html)
        self.assertNotIn("<th>型号</th>", html)
        self.assertNotIn("<th>编码</th>", html)
        self.assertNotIn('name="thickness"', html)
        self.assertNotIn('name="width"', html)
        self.assertNotIn('name="length"', html)
        self.assertNotIn('href="/materials/items/new"', html)
        self.assertLess(html.index('name="part"'), html.index('name="pieces"'))
        self.assertLess(html.index('name="spec_text"'), html.index('name="category"'))
        self.assertLess(html.index('name="category"'), html.index('name="car"'))

    def test_admin_materials_page_shows_all_recent_material_files_with_operator(self):
        output_root = self.root / "outputs"
        current_user_dir = output_root / "u1-007"
        other_user_dir = output_root / "u88-other-material"
        current_user_dir.mkdir(parents=True, exist_ok=True)
        other_user_dir.mkdir(parents=True, exist_ok=True)
        current_file = current_user_dir / "current-user-260701料单.xlsx"
        other_file = other_user_dir / "other-user-260701料单.xlsx"
        unrelated_file = other_user_dir / "other-user-quote.xlsx"
        current_file.write_bytes(b"current")
        other_file.write_bytes(b"other")
        unrelated_file.write_bytes(b"quote")

        self.login()
        response = self.client.get("/materials")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("最近料单", html)
        self.assertIn("操作用户", html)
        self.assertIn("current-user-260701料单.xlsx", html)
        self.assertIn("other-user-260701料单.xlsx", html)
        self.assertIn("other-material", html)
        self.assertNotIn("other-user-quote.xlsx", html)

        response = self.client.get("/materials", query_string={"material_history_q": "other-material"})
        html = response.get_data(as_text=True)
        self.assertIn('class="materials-history-drawer" open', html)
        self.assertIn("other-user-260701料单.xlsx", html)
        self.assertNotIn("current-user-260701料单.xlsx", html)

    def test_material_item_save_calculates_spec_text(self):
        self.login()
        examples = [
            ("T-SPEC-WEB-SPACE", "2.5 357 1260", "2.5×357×1260"),
            ("T-SPEC-WEB-STAR", "2.5*357*1260", "2.5×357×1260"),
            ("T-SPEC-WEB-DASH", "2.5-357-1260", "2.5×357×1260"),
            ("T-SPEC-WEB-SLASH", "2.5/357/1260", "2.5×357×1260"),
        ]
        for model, spec_text, expected in examples:
            with self.subTest(spec_text=spec_text):
                response = self.client.post(
                    "/materials/items/save",
                    data={
                        "model": model,
                        "code": "KA-TEST",
                        "category": "测试类别",
                        "car": "测试车型",
                        "part": "测试零件",
                        "pieces": "2",
                        "spec_text": spec_text,
                        "active": "1",
                    },
                    follow_redirects=False,
                )
                self.assertEqual(response.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            rows = conn.execute(
                "SELECT model, spec_text, thickness, width, length FROM material_items WHERE model LIKE 'T-SPEC-WEB-%'"
            ).fetchall()
        saved = {row["model"]: row for row in rows}
        for model, _, expected in examples:
            self.assertIn(model, saved)
            self.assertEqual(saved[model]["spec_text"], expected)
            self.assertEqual(saved[model]["thickness"], 2.5)
            self.assertEqual(saved[model]["width"], 357)
            self.assertEqual(saved[model]["length"], 1260)

        response = self.client.get("/materials/items?q=T-SPEC-WEB-SPACE")
        html = response.get_data(as_text=True)
        self.assertIn("单件重量kg", html)
        self.assertIn("4.41", html)
        for query in ["357", "2.5 357", "357/1260", "2.5-1260", "2.5*357*1260"]:
            with self.subTest(query=query):
                response = self.client.get("/materials/items", query_string={"q": query})
                html = response.get_data(as_text=True)
                self.assertIn("T-SPEC-WEB-SPACE", html)
        response = self.client.get("/materials/items", query_string={"q": "2.5 999"})
        html = response.get_data(as_text=True)
        self.assertNotIn("T-SPEC-WEB-SPACE", html)

    def test_material_item_requires_code_and_part(self):
        self.login()
        for field, data in [
            (
                "code",
                {
                    "model": "T-SPEC-REQUIRED-CODE",
                    "part": "测试零件",
                    "pieces": "2",
                    "spec_text": "2.5 357 1260",
                    "active": "1",
                },
            ),
            (
                "part",
                {
                    "model": "T-SPEC-REQUIRED-PART",
                    "code": "KA-TEST",
                    "pieces": "2",
                    "spec_text": "2.5 357 1260",
                    "active": "1",
                },
            ),
        ]:
            with self.subTest(field=field):
                response = self.client.post("/materials/items/save", data=data, follow_redirects=False)
                self.assertEqual(response.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            count = conn.execute(
                "SELECT COUNT(*) FROM material_items WHERE model IN (?, ?)",
                ("T-SPEC-REQUIRED-CODE", "T-SPEC-REQUIRED-PART"),
            ).fetchone()[0]
        self.assertEqual(count, 0)

    def test_material_import_calculates_spec_text_from_dimensions(self):
        from app.modules.materials.persistence import import_materials_from_excel
        from openpyxl import Workbook

        path = self.root / "stale-material-spec.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "材料数据"
        sheet.append(
            ["型号", "型号", "类别", "车型", "零件名称", "规格尺寸", "下料只数", "单重", "规格1", "规格2", "规格3"]
        )
        sheet.append(["T-SPEC-IMPORT", "KA-IMPORT", "测试类别", "测试车型", "测试零件", "旧规格", 3, "", 4, 92.5, 1260])
        workbook.save(path)
        workbook.close()

        with self.web.connect(self.web.DB_PATH) as conn:
            imported = import_materials_from_excel(conn, path, replace=False, actor="tester")
            row = conn.execute("SELECT spec_text FROM material_items WHERE model = ?", ("T-SPEC-IMPORT",)).fetchone()

        self.assertEqual(imported, 1)
        self.assertIsNotNone(row)
        self.assertEqual(row["spec_text"], "4.0×92.5×1260")

    def test_material_source_sync_rewrites_spec_text_column(self):
        from app.material_sheet import sync_material_specs_from_dimensions
        from openpyxl import Workbook, load_workbook

        path = self.root / "sync-material-spec.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "材料数据"
        sheet.append(
            ["型号", "型号", "类别", "车型", "零件名称", "规格尺寸", "下料只数", "单重", "规格1", "规格2", "规格3"]
        )
        sheet.append(
            ["T-SPEC-SOURCE", "KA-SOURCE", "测试类别", "测试车型", "测试零件", "旧规格", 3, "", 2.5, 312, 1260]
        )
        workbook.save(path)
        workbook.close()

        changed = sync_material_specs_from_dimensions(path)
        synced = load_workbook(path, read_only=True, data_only=True)
        try:
            self.assertEqual(changed, 1)
            self.assertEqual(synced["材料数据"].cell(2, 6).value, "2.5×312×1260")
        finally:
            synced.close()

    def test_upload_limits_keep_product_sync_headroom(self):
        self.assertEqual(self.web.MAX_UPLOAD_MB, 20)
        self.assertEqual(self.web.PRODUCT_SYNC_MAX_UPLOAD_MB, 512)
        self.assertEqual(self.web.app.config["MAX_CONTENT_LENGTH"], 512 * 1024 * 1024)

    def test_business_sync_preview_allows_large_package(self):
        self.login()
        big_file = io.BytesIO(b"x" * (21 * 1024 * 1024))
        try:
            response = self.client.post(
                "/business-data-sync/preview",
                data={"package": (big_file, "big.tar.gz")},
                content_type="multipart/form-data",
                follow_redirects=False,
            )
            # 大于 20MB 的普通上传限额时不应再被 413 拦截，而是进入预览流程后因内容无效重定向。
            self.assertEqual(response.status_code, 302)
            self.assertTrue(response.headers["Location"].endswith("/business-data-sync"))
            request_stream = response.request.environ.get("wsgi.input")
            response.close()
            if request_stream is not None:
                request_stream.close()
        finally:
            big_file.close()

    def test_oversized_upload_redirects(self):
        self.login()
        original_limit = self.web.app.config["MAX_CONTENT_LENGTH"]
        self.web.app.config["MAX_CONTENT_LENGTH"] = 10
        try:
            big_file = io.BytesIO(b"x" * 11)
            response = self.client.post(
                "/catalog",
                data={"catalog": (big_file, "big.xlsx")},
                content_type="multipart/form-data",
                follow_redirects=False,
            )
            self.assertEqual(response.status_code, 302)
            self.assertTrue(response.headers["Location"].endswith("/products"))
            response.close()
        finally:
            self.web.app.config["MAX_CONTENT_LENGTH"] = original_limit

    def test_migrations_are_recorded(self):
        with self.web.connect(self.web.DB_PATH) as conn:
            rows = conn.execute("SELECT id FROM schema_migrations ORDER BY id").fetchall()
        self.assertEqual(
            [row["id"] for row in rows],
            [
                "001_audit_log_actor",
                "002_product_price_and_image",
                "003_product_drawings",
                "004_product_image_slots",
                "005_internal_api_keys",
                "006_shipment_recognition_jobs",
                "007_product_status",
                "008_internal_api_key_plaintext",
                "009_quote_records",
                "010_quote_record_bld_prices",
                "011_customer_price_bld_index",
                "012_scrub_internal_api_key_plaintext",
                "013_api_principal_scopes_and_idempotency",
                "014_quote_record_version",
                "015_idempotency_response_headers",
                "016_api_artifacts",
                "017_runtime_jobs_ai_and_health",
                "018_tube_items",
                "019_tube_dimensions",
                "020_tube_manufacturing_fields",
                "021_flatten_tube_borrowing",
                "022_cross_device_sync_keys",
                "023_rekey_cross_device_sync_keys",
                "024_drop_quote_record_price",
                "025_create_product_option_values",
                "026_quote_record_quote_no",
                "027_customers",
                "028_customer_profiles_documents_and_quote_contracts",
                "029_customer_workspace_integrity",
                "030_repair_customer_workspace_integrity",
                "031_editable_roles_and_user_permission_overrides",
                "032_grant_view_product_prices",
                "033_revoke_view_product_prices",
                "034_split_granular_permissions",
                "035_customer_drawings",
                "036_customer_products",
                "037_customer_identity_and_material_drawing_permissions",
                "038_product_option_value_sort_order",
                "039_product_view_permissions",
                "040_user_default_page",
                "041_user_mobile_default_page",
            ],
        )

    def test_migration_scrubs_historical_api_key_plaintext(self):
        from app.migrations import run_migrations
        from app.platform.api_keys import verify_internal_api_token
        from app.platform.api_principal import LEGACY_COMPATIBILITY_SCOPES

        token = self.create_internal_api_token()
        with self.web.connect(self.web.DB_PATH) as conn:
            conn.execute("ALTER TABLE internal_api_keys ADD COLUMN token_plain TEXT DEFAULT ''")
            conn.execute(
                "UPDATE internal_api_keys SET token_plain = ? WHERE id = (SELECT MIN(id) FROM internal_api_keys)",
                ("historical-plaintext",),
            )
            conn.execute("DELETE FROM schema_migrations WHERE id = '012_scrub_internal_api_key_plaintext'")
            conn.execute("UPDATE internal_api_keys SET scopes = '[]'")
            conn.execute("DELETE FROM schema_migrations WHERE id = '013_api_principal_scopes_and_idempotency'")
            conn.commit()
            run_migrations(conn)
            columns = {row["name"] for row in conn.execute("PRAGMA table_info(internal_api_keys)").fetchall()}
            principal = verify_internal_api_token(conn, token)
        self.assertNotIn("token_plain", columns)
        self.assertEqual(principal.integration_name, "OpenClaw Test")
        self.assertEqual(principal.scopes, LEGACY_COMPATIBILITY_SCOPES)

    def test_concurrent_database_initialization_is_process_safe(self):
        database_path = self.root / "concurrent-init.sqlite3"
        gate_path = self.root / "concurrent-init.go"
        script = """
import sys
import time
from pathlib import Path
from app.database import connect

database_path = Path(sys.argv[1])
gate_path = Path(sys.argv[2])
while not gate_path.exists():
    time.sleep(0.005)
with connect(database_path) as conn:
    conn.execute("SELECT COUNT(*) FROM schema_migrations").fetchone()
"""
        processes = [
            subprocess.Popen(
                [sys.executable, "-c", script, str(database_path), str(gate_path)],
                cwd=PROJECT_ROOT,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
            )
            for _ in range(6)
        ]
        gate_path.touch()
        results = [process.communicate(timeout=20) for process in processes]
        failures = [stderr for process, (_stdout, stderr) in zip(processes, results) if process.returncode != 0]
        self.assertEqual(failures, [])

    def test_default_admin_requires_explicit_password_and_never_rewrites_existing_hash(self):
        from app.config import DEFAULT_ADMIN_PASSWORD_PLACEHOLDER
        from app.modules.admin.persistence import ensure_default_admin
        from app.platform.clock import now_text

        with self.web.connect(self.web.DB_PATH) as conn:
            with self.assertRaisesRegex(RuntimeError, "DEFAULT_ADMIN_PASSWORD"):
                ensure_default_admin(
                    conn,
                    username="placeholder-admin",
                    password=DEFAULT_ADMIN_PASSWORD_PLACEHOLDER,
                )
            legacy_hash = "scrypt:32768:8:1$legacy$unsupported-hash"
            stamp = now_text()
            conn.execute(
                """
                INSERT INTO users (username, display_name, password_hash, role, active, created_at, updated_at)
                VALUES (?, '', ?, 'admin', 1, ?, ?)
                """,
                ("legacy-admin", legacy_hash, stamp, stamp),
            )
            conn.commit()
            ensure_default_admin(conn, username="legacy-admin", password="replacement-password")
            stored = conn.execute(
                "SELECT password_hash FROM users WHERE username = ?",
                ("legacy-admin",),
            ).fetchone()["password_hash"]
        self.assertEqual(stored, legacy_hash)

    def test_generated_files_are_scoped_to_user(self):
        self.login()
        user_files = set((self.root / "outputs").glob("u*-007/catalog-export-bld-007-*.xlsx"))
        response = self.client.post("/products/export", data={"status": "active", "export_format": "bld"})
        self.assertEqual(response.status_code, 200)
        response.close()

        files = set((self.root / "outputs").glob("u*-007/catalog-export-bld-007-*.xlsx")) - user_files
        self.assertEqual(len(files), 1)
        self.assertFalse(list((self.root / "outputs").glob("catalog-export-bld-007-*.xlsx")))

    def test_catalog_export_is_admin_only(self):
        from app.modules.admin.persistence import save_user

        with self.web.connect(self.web.DB_PATH) as conn:
            save_user(
                conn,
                {
                    "username": "editor-export",
                    "display_name": "Editor Export",
                    "password": "editor-pw",
                    "role": "editor",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        admin_page = self.client.get("/products").get_data(as_text=True)
        self.assertIn("导出目录", admin_page)

        self.client.post("/logout")
        login = self.client.post(
            "/login",
            data={"username": "editor-export", "password": "editor-pw", "next": "/"},
            follow_redirects=False,
        )
        self.assertEqual(login.status_code, 302)

        editor_page = self.client.get("/products").get_data(as_text=True)
        self.assertNotIn("导出目录", editor_page)
        self.assertNotIn('action="/products/export"', editor_page)

        response = self.client.post("/products/export", data={"status": "active", "export_format": "bld"})
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/products"))
        self.assertFalse(list((self.root / "outputs").glob("**/catalog-export-bld-editor-export-*.xlsx")))
        self.client.post("/logout")

    def test_product_export_embeds_generated_thumbnail_instead_of_main_image(self):
        from openpyxl import load_workbook
        from PIL import Image

        from app.modules.products.persistence import upsert_product

        image_dir = self.root / "data" / "product_images"
        image_dir.mkdir(parents=True, exist_ok=True)
        image_path = image_dir / "K-EXPORT-IMG.png"
        Image.new("RGB", (1600, 1200), "white").save(image_path)

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-EXPORT-IMG",
                    "series": "TEST",
                    "item": "EXPORT IMAGE",
                    "oe_no_1": "EXPORT-IMAGE-001",
                    "models": "Tester",
                    "image_path": "data_product_images/K-EXPORT-IMG.png",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.post("/products/export", data={"status": "active", "export_format": "bld"})
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(io.BytesIO(response.data))
        sheet = workbook["产品目录"]
        row_index = next(row[0].row for row in sheet.iter_rows(min_row=2) if row[0].value == "K-EXPORT-IMG")

        self.assertIsNone(sheet.cell(row_index, 7).value)
        self.assertGreaterEqual(len(sheet._images), 1)
        self.assertGreaterEqual(sheet.row_dimensions[row_index].height, 62)
        workbook.close()
        with zipfile.ZipFile(io.BytesIO(response.data)) as archive:
            embedded_names = [name for name in archive.namelist() if name.startswith("xl/media/")]
            self.assertTrue(embedded_names)
            with Image.open(io.BytesIO(archive.read(embedded_names[0]))) as embedded:
                self.assertLessEqual(embedded.width, 320)
                self.assertLessEqual(embedded.height, 240)
        self.assertTrue((image_dir / "thumbs" / "K-EXPORT-IMG.webp").exists())
        response.close()

    def test_catalog_export_uses_bld_natural_order(self):
        from openpyxl import load_workbook

        from app.modules.products.persistence import upsert_product

        expected = ["K8274LA", "K8274RA", "K8274LB", "K8274RB"]
        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no in expected:
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "SORT EXPORT",
                        "item": "SORT EXPORT TEST",
                        "oe_no_1": f"OE-{bld_no}",
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.post("/products/export", data={"status": "active", "export_format": "bld"})
        self.assertEqual(response.status_code, 200)

        workbook = load_workbook(io.BytesIO(response.data), read_only=True)
        sheet = workbook["产品目录"]
        exported = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] in expected]

        self.assertEqual(exported, expected)
        workbook.close()
        response.close()

    def test_admin_homepage_shows_all_recent_outputs(self):
        output_root = self.root / "outputs"
        other_user_dir = output_root / "u99-other"
        other_user_dir.mkdir(parents=True, exist_ok=True)
        root_file = output_root / "old-root-result.xlsx"
        other_user_file = other_user_dir / "other-user-result.xlsx"
        catalog_file = output_root / "catalog-export-bld-history-sample.xlsx"
        material_file = output_root / "26年4月冲压生产计划260423料单.xlsx"
        root_file.write_bytes(b"legacy")
        other_user_file.write_bytes(b"other")
        catalog_file.write_bytes(b"catalog")
        material_file.write_bytes(b"materials")

        self.login()
        response = self.client.get("/")
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("操作用户", html)
        self.assertIn("data-history-loader", html)
        self.assertIn("展开后加载", html)
        self.assertIn("data-file-drop-zone", html)
        self.assertIn("可拖入询价文件", html)
        self.assertIn("输入 OE或 BLD 号", html)
        self.assertIn("file-picker-clear", html)
        self.assertNotIn("old-root-result.xlsx", html)
        self.assertNotIn("other-user-result.xlsx", html)

        response = self.client.get("/history-files")
        payload = response.get_json()
        self.assertEqual(response.status_code, 200)
        self.assertIsNotNone(payload)
        names = [item["name"] for item in payload["rows"]]
        self.assertIn("old-root-result.xlsx", names)
        self.assertIn("other-user-result.xlsx", names)
        operators = {item["name"]: item["operator"] for item in payload["rows"]}
        self.assertEqual(operators["other-user-result.xlsx"], "other")
        self.assertNotIn("catalog-export-bld-history-sample.xlsx", names)
        self.assertNotIn("26年4月冲压生产计划260423料单.xlsx", names)

        response = self.client.get("/?history_q=other-user")
        html = response.get_data(as_text=True)
        self.assertIn("other-user-result.xlsx", html)
        self.assertNotIn("old-root-result.xlsx", html)

        response = self.client.get("/?history_q=other")
        html = response.get_data(as_text=True)
        self.assertIn("other-user-result.xlsx", html)
        self.assertNotIn("old-root-result.xlsx", html)

        response = self.client.get("/history-files?history_q=other")
        payload = response.get_json()
        names = [item["name"] for item in payload["rows"]]
        self.assertIn("other-user-result.xlsx", names)
        self.assertNotIn("old-root-result.xlsx", names)

    def test_price_import_web_routes_are_removed(self):
        self.login()
        self.assertEqual(self.client.get("/prices/import").status_code, 404)
        self.assertEqual(self.client.post("/prices/import/preview").status_code, 404)

    def test_import_lock_blocks_parallel_imports(self):
        from app.locks import ImportLockError, import_lock

        with import_lock("tester", "测试导入"):
            with self.assertRaises(ImportLockError):
                with import_lock("other", "第二个导入"):
                    pass

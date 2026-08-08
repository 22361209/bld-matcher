from __future__ import annotations

from tests.web_app_test_base import (
    WebAppTestBase,
    re,
)


class TestWebInquiryMatching(WebAppTestBase):
    def test_pasted_multiple_codes_generates_match_excel(self):
        from app.modules.products.persistence import upsert_product
        from openpyxl import load_workbook

        products = [
            ("K54500L", "54500-2D000", "79.2"),
            ("K54501L", "54501-2D000", "39.6"),
            ("K54501A", "54501-A0000", "118.8"),
        ]
        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no, oe_no, price_cny in products:
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "HYUNDAI",
                        "item": "CONTROL ARM",
                        "oe_no_1": oe_no,
                        "models": "Elantra",
                        "price_cny": price_cny,
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.post(
            "/match",
            data={"quick_oe": "54500-2d000 54501-2d000 54501-a0000"},
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("<title>查询结果</title>", html)
        self.assertIn("下载 Excel", html)
        self.assertIn("K54500L", html)
        self.assertIn("K54501L", html)
        self.assertIn("K54501A", html)
        self.assertIn("粘贴号码询价.xlsx", html)
        self.assertIn("含税单价", html)
        self.assertIn('value="79.20" data-inquiry-tax-price', html)
        self.assertIn('data-adjustment-key="sheet:1:row:2"', html)
        self.assertIn('<td data-col="row">1</td>', html)
        self.assertIn('<td data-col="row">2</td>', html)
        self.assertIn('<td data-col="row">3</td>', html)
        self.assertNotIn('<td data-col="row">4</td>', html)
        self.assertIn('id="download-excel-modal"', html)
        self.assertIn('action="/match/download"', html)
        self.assertNotIn("返回上一步", html)

        upload_match = re.search(r'name="upload_path" value="([^"]+)"', html)
        output_match = re.search(r'name="output_name" value="([^"]+)"', html)
        self.assertIsNotNone(upload_match)
        self.assertIsNotNone(output_match)
        upload_path = upload_match.group(1)
        output_name = output_match.group(1)
        output_path = self.root / "outputs" / "u1-007" / output_name

        download = self.client.post(
            "/match/download",
            data={
                "upload_path": upload_path,
                "original_filename": "粘贴号码询价.xlsx",
                "output_name": output_name,
                "match_column": "",
                "price_mode": "usd",
                "exchange_rate": "7.2",
            },
        )
        self.assertEqual(download.status_code, 200)
        download.close()
        self.assertTrue(output_path.exists())

        generated = load_workbook(output_path)
        sheet = generated.active
        self.assertEqual(sheet.cell(1, 1).value, "OE号")
        self.assertEqual(sheet.cell(1, 2).value, "BLD NO.")
        self.assertEqual(sheet.cell(1, 3).value, "美金价")
        self.assertEqual(sheet.cell(1, 4).value, "Product Status")
        self.assertEqual(sheet.cell(1, 5).value, "匹配说明")
        self.assertEqual(sheet.cell(2, 1).value, "54500-2d000")
        self.assertEqual(sheet.cell(2, 2).value, "K54500L")
        self.assertEqual(sheet.cell(2, 3).value, 10)
        self.assertEqual(sheet.cell(3, 2).value, "K54501L")
        self.assertEqual(sheet.cell(3, 3).value, 5)
        self.assertEqual(sheet.cell(4, 2).value, "K54501A")
        self.assertEqual(sheet.cell(4, 3).value, 15)
        generated.close()

    def test_pasted_combined_oe_prefix_stays_one_query(self):
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8282RA",
                    "series": "FORD",
                    "item": "Front Right Lower Control Arm",
                    "oe_no_1": "F1F1-3A423-AAA\nF1F1-3A423-AAB",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K8235RA",
                    "series": "FORD",
                    "item": "Front Right Lower Control Arm",
                    "oe_no_1": "JX61\n3A423\nAPB",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.post("/match", data={"quick_oe": "F1F1 3A423"})
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("查询结果", html)
        self.assertIn("K8282RA", html)
        self.assertIn("OE 组合前缀命中", html)
        self.assertNotIn("K8235RA", html)
        self.assertIn('<td data-col="row">1</td>', html)
        self.assertNotIn('<td data-col="row">2</td>', html)

    def test_uploaded_inquiry_combined_oe_prefix_matches_before_fragments(self):
        from app.modules.products.persistence import upsert_product
        from app.excel_io import generate_excel_with_bld
        from app.helpers import load_catalog
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8282RA",
                    "series": "FORD",
                    "item": "Front Right Lower Control Arm",
                    "oe_no_1": "F1F1-3A423-AAA\nF1F1-3A423-AAB",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K8235RA",
                    "series": "FORD",
                    "item": "Front Right Lower Control Arm",
                    "oe_no_1": "JX61\n3A423\nAPB",
                    "active": "1",
                },
                actor="tester",
            )

        inquiry_path = self.root / "uploads" / "combined-prefix.xlsx"
        inquiry_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["F1F1 3A423"])
        workbook.save(inquiry_path)
        workbook.close()

        summary = generate_excel_with_bld(
            inquiry_path,
            self.root / "outputs" / "combined-prefix-result.xlsx",
            load_catalog(),
            write_output=False,
        )

        self.assertEqual(summary["total"], 1)
        self.assertEqual(summary["matched"], 1)
        self.assertEqual(summary["rows"][0]["bld_no"], "K8282RA")
        self.assertEqual(summary["rows"][0]["reason"], "OE 组合前缀命中")

    def test_uploaded_inquiry_integer_decimal_text_matches_prefix(self):
        from app.modules.products.persistence import upsert_product
        from app.excel_io import generate_excel_with_bld
        from app.helpers import load_catalog
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8041LB",
                    "series": "VW",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "561407151A\n561407151C",
                    "active": "1",
                },
                actor="tester",
            )

        inquiry_path = self.root / "uploads" / "integer-decimal-text.xlsx"
        inquiry_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["561407151.0"])
        workbook.save(inquiry_path)
        workbook.close()

        summary = generate_excel_with_bld(
            inquiry_path,
            self.root / "outputs" / "integer-decimal-text-result.xlsx",
            load_catalog(),
            write_output=False,
        )

        self.assertEqual(summary["total"], 1)
        self.assertEqual(summary["matched"], 1)
        self.assertEqual(summary["rows"][0]["bld_no"], "K8041LB")
        self.assertEqual(summary["rows"][0]["reason"], "OE 组合前缀命中")

    def test_uploaded_inquiry_oe_suffix_variant_matches_unique_base(self):
        from app.modules.products.persistence import upsert_product
        from app.excel_io import generate_excel_with_bld
        from app.helpers import load_catalog
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8041LB",
                    "series": "VW",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "561407151A\n561407151C",
                    "active": "1",
                },
                actor="tester",
            )

        inquiry_path = self.root / "uploads" / "oe-suffix-variant.xlsx"
        inquiry_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["561407151D"])
        workbook.save(inquiry_path)
        workbook.close()

        summary = generate_excel_with_bld(
            inquiry_path,
            self.root / "outputs" / "oe-suffix-variant-result.xlsx",
            load_catalog(),
            write_output=False,
        )

        self.assertEqual(summary["total"], 1)
        self.assertEqual(summary["matched"], 1)
        self.assertEqual(summary["rows"][0]["bld_no"], "K8041LB")
        self.assertEqual(summary["rows"][0]["reason"], "OE 尾字母容错命中")

    def test_uploaded_inquiry_split_oe_suffix_variants_match_same_product(self):
        from app.modules.products.persistence import upsert_product
        from app.excel_io import generate_excel_with_bld
        from app.helpers import load_catalog
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K8321LB",
                    "series": "VW",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "2QD407151",
                    "active": "1",
                },
                actor="tester",
            )

        inquiry_path = self.root / "uploads" / "split-oe-suffix-variants.xlsx"
        inquiry_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["2QD407151A;2QD407151C"])
        workbook.save(inquiry_path)
        workbook.close()

        summary = generate_excel_with_bld(
            inquiry_path,
            self.root / "outputs" / "split-oe-suffix-variants-result.xlsx",
            load_catalog(),
            write_output=False,
        )

        self.assertEqual(summary["total"], 1)
        self.assertEqual(summary["matched"], 1)
        self.assertEqual(summary["rows"][0]["bld_no"], "K8321LB")
        self.assertEqual(summary["rows"][0]["reason"], "OE 尾字母容错命中")
        self.assertEqual(summary["rows"][0]["matched_oe_codes"], ["2QD407151A", "2QD407151C"])
        self.assertIn("命中号码：2QD407151A, 2QD407151C", summary["rows"][0]["match_note"])

    def test_psa_352x_dot_matches_psa_before_gm_exact(self):
        from app.matcher import ProductCatalog

        catalog = ProductCatalog(
            [
                {
                    "BLD NO.": "K-PSA-352123",
                    "SERIES": "PEUGEOT\nCITROEN",
                    "ITEM": "Front Left Lower Control Arm",
                    "OE NO.1": "3521.23",
                },
                {
                    "BLD NO.": "K-GM-352123",
                    "SERIES": "GM\nOPEL",
                    "ITEM": "Front Left Lower Control Arm",
                    "OE NO.1": "352123",
                },
            ]
        )

        match = catalog.match("", "3521.23")

        self.assertIsNotNone(match)
        self.assertEqual(match.bld_no, "K-PSA-352123")
        self.assertEqual(match.reason, "PSA 号码点号容错命中")

    def test_psa_352x_without_dot_is_ambiguous_when_gm_exact_also_exists(self):
        from app.matcher import ProductCatalog

        catalog = ProductCatalog(
            [
                {
                    "BLD NO.": "K-PSA-352123",
                    "SERIES": "PEUGEOT\nCITROEN",
                    "ITEM": "Front Left Lower Control Arm",
                    "OE NO.1": "3521.23",
                },
                {
                    "BLD NO.": "K-GM-352123",
                    "SERIES": "GM\nOPEL",
                    "ITEM": "Front Left Lower Control Arm",
                    "OE NO.1": "352123",
                },
            ]
        )

        match = catalog.match("", "352123")

        self.assertIsNotNone(match)
        self.assertIn("K-PSA-352123", match.bld_no)
        self.assertIn("K-GM-352123", match.bld_no)
        self.assertEqual(match.reason, "3520/3521 号码同时命中 PSA 与其他品牌，请人工确认")

    def test_psa_352x_without_dot_still_matches_gm_when_no_psa_exists(self):
        from app.matcher import ProductCatalog

        catalog = ProductCatalog(
            [
                {
                    "BLD NO.": "K-GM-352023",
                    "SERIES": "GM\nOPEL",
                    "ITEM": "Front Left Lower Control Arm",
                    "OE NO.1": "352023",
                }
            ]
        )

        match = catalog.match("", "352023")

        self.assertIsNotNone(match)
        self.assertEqual(match.bld_no, "K-GM-352023")
        self.assertEqual(match.reason, "OE 精准命中")

    def test_uploaded_inquiry_psa_352x_dot_does_not_match_gm_exact(self):
        from app.modules.products.persistence import upsert_product
        from app.excel_io import generate_excel_with_bld
        from app.helpers import load_catalog
        from openpyxl import Workbook

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-PSA-352124-FILE",
                    "series": "PEUGEOT\nCITROEN",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "3521.24",
                    "active": "1",
                },
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-GM-352124-FILE",
                    "series": "GM\nOPEL",
                    "item": "Front Left Lower Control Arm",
                    "oe_no_1": "352124",
                    "active": "1",
                },
                actor="tester",
            )

        inquiry_path = self.root / "uploads" / "psa-352x-dot.xlsx"
        inquiry_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["OE号"])
        sheet.append(["3521.24"])
        workbook.save(inquiry_path)
        workbook.close()

        summary = generate_excel_with_bld(
            inquiry_path,
            self.root / "outputs" / "psa-352x-dot-result.xlsx",
            load_catalog(),
            write_output=False,
        )

        self.assertEqual(summary["total"], 1)
        self.assertEqual(summary["matched"], 1)
        self.assertEqual(summary["rows"][0]["bld_no"], "K-PSA-352124-FILE")
        self.assertEqual(summary["rows"][0]["reason"], "PSA 号码点号容错命中")

    def test_pasted_multiple_bld_codes_generates_match_excel(self):
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no in ["K-BLD-BATCH-1", "K-BLD-BATCH-2"]:
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "HYUNDAI",
                        "item": "CONTROL ARM",
                        "oe_no_1": f"{bld_no}-OE",
                        "models": "Elantra",
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.post("/match", data={"quick_oe": "K-BLD-BATCH-1 K-BLD-BATCH-2"})
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        self.assertIn("查询结果", html)
        self.assertIn("K-BLD-BATCH-1", html)
        self.assertIn("K-BLD-BATCH-2", html)
        self.assertIn("BLD NO. 精准命中", html)

    def test_pasted_bld_fragment_candidates_are_read_only(self):
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            for bld_no, price, status in (
                ("K6004LB", "58", "2个衬套"),
                ("K6004RB", "80", "2个衬套1个球头"),
                ("K-EXACT-ADJUST-ROW", "66.6", "唯一命中"),
            ):
                upsert_product(
                    conn,
                    {
                        "bld_no": bld_no,
                        "series": "READONLY",
                        "item": "FRAGMENT CONTROL ARM",
                        "oe_no_1": f"{bld_no}-OE",
                        "price_cny": price,
                        "product_status": status,
                        "active": "1",
                    },
                    actor="tester",
                )

        self.login()
        response = self.client.post(
            "/match",
            data={"quick_oe": "6004,K-EXACT-ADJUST-ROW"},
        )
        html = response.get_data(as_text=True)

        self.assertEqual(response.status_code, 200)
        result_rows = re.findall(r"<tr[^>]*>.*?</tr>", html, flags=re.DOTALL)
        fragment_row = next(
            (row for row in result_rows if "<strong data-current-bld>K6004LB</strong>" in row),
            "",
        )
        self.assertTrue(fragment_row)
        self.assertNotIn("data-adjustment-key", fragment_row)
        self.assertNotIn("data-inquiry-bld-input", fragment_row)
        self.assertNotIn("data-inquiry-tax-price", fragment_row)

        exact_row = next(
            (row for row in result_rows if 'data-default-bld="K-EXACT-ADJUST-ROW"' in row),
            "",
        )
        self.assertTrue(exact_row)
        self.assertIn('data-adjustment-key="sheet:1:row:3"', exact_row)
        self.assertIn('value="K-EXACT-ADJUST-ROW"', exact_row)
        self.assertIn("data-inquiry-bld-input", exact_row)
        self.assertNotIn("data-open-product-adjustment", exact_row)
        self.assertIn("data-inquiry-tax-price", exact_row)

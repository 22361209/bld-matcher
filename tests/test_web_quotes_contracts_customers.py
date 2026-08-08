from __future__ import annotations

import json
from html import unescape

from tests.web_app_test_base import (
    WebAppTestBase,
    io,
    re,
)


class TestWebQuotesContractsCustomers(WebAppTestBase):
    def test_product_price_visibility_follows_quote_write_permission(self):
        from app.modules.admin.persistence import save_role, save_user
        from app.modules.products.persistence import upsert_product
        from openpyxl import load_workbook

        self.addCleanup(self.cleanup_products, "K-PRICE-VIS-%")
        username = "price-hide-user"
        role_name = "PRICE HIDE 角色"

        def cleanup_access_records():
            self.client.post("/logout")
            with self.web.connect(self.web.DB_PATH) as connection:
                user = connection.execute(
                    "SELECT id FROM users WHERE username = ?", (username,)
                ).fetchone()
                if user:
                    connection.execute(
                        "DELETE FROM user_permission_overrides WHERE user_id = ?", (user["id"],)
                    )
                    connection.execute("DELETE FROM users WHERE id = ?", (user["id"],))
                role = connection.execute(
                    "SELECT role_key FROM roles WHERE name = ?", (role_name,)
                ).fetchone()
                if role:
                    connection.execute(
                        "DELETE FROM role_permissions WHERE role_key = ?", (role["role_key"],)
                    )
                    connection.execute("DELETE FROM roles WHERE role_key = ?", (role["role_key"],))
                connection.commit()

        self.addCleanup(cleanup_access_records)
        self.login()
        with self.web.connect(self.web.DB_PATH) as conn:
            role_key = save_role(
                conn,
                {"name": role_name, "description": "看不到产品单价但可以维护产品的角色"},
                ["generate_match", "export_catalog", "generate_contract", "edit_products"],
                actor="tester",
            )
            save_user(
                conn,
                {"username": username, "password": "price-hide-pw", "role": role_key, "active": "1"},
                actor="tester",
            )
            upsert_product(
                conn,
                {
                    "bld_no": "K-PRICE-VIS-001",
                    "series": "TEST",
                    "item": "Price Vis Arm",
                    "oe_no_1": "PRICE-VIS-OE",
                    "price_cny": "80.00",
                    "active": "1",
                },
                actor="tester",
            )
            conn.commit()
        self.client.post("/logout")
        login = self.client.post(
            "/login",
            data={"username": username, "password": "price-hide-pw", "next": "/"},
            follow_redirects=False,
        )
        self.assertEqual(login.status_code, 302)

        products_page = self.client.get("/products")
        products_html = products_page.get_data(as_text=True)
        self.assertEqual(products_page.status_code, 200)
        self.assertNotIn('data-col="price"', products_html)
        self.assertNotIn("price_cny", products_html)

        lookup = self.client.get("/products/lookup?q=K-PRICE-VIS&details=1")
        self.assertEqual(lookup.status_code, 200)
        self.assertTrue(lookup.get_json())
        self.assertTrue(all("price_cny" not in row for row in lookup.get_json()))

        contract_lookup = self.client.get("/purchase-contracts/product-lookup?bld=K-PRICE-VIS-001")
        self.assertEqual(contract_lookup.status_code, 200)
        self.assertNotIn("price_cny", contract_lookup.get_json())

        result_page = self.client.post("/match", data={"quick_oe": "PRICE-VIS-OE\nPRICE-VIS-UNMATCHED"})
        result_html = result_page.get_data(as_text=True)
        self.assertEqual(result_page.status_code, 200)
        self.assertNotIn('data-col="price"', result_html)

        quick = self.client.get(
            "/inquiry/quick-search",
            query_string={"quick_oe": "PRICE-VIS-OE"},
            headers={"Accept": "text/html", "X-Requested-With": "fetch"},
        )
        self.assertEqual(quick.status_code, 200)
        self.assertNotIn("¥", quick.get_data(as_text=True))

        export_response = self.client.post(
            "/products/export",
            data={"export_format": "bld", "status": "all"},
        )
        self.assertEqual(export_response.status_code, 200)
        export_bytes = export_response.get_data()
        export_response.close()
        workbook = load_workbook(io.BytesIO(export_bytes))
        try:
            headers = [cell.value for cell in workbook.active[1]]
        finally:
            workbook.close()
        self.assertNotIn("Unit Price", headers)

        saved = self.client.post(
            "/products/save",
            data={
                "bld_no": "K-PRICE-VIS-001",
                "series": "TEST",
                "item": "Price Vis Arm Renamed",
                "active": "1",
            },
            follow_redirects=False,
        )
        self.assertEqual(saved.status_code, 302)
        with self.web.connect(self.web.DB_PATH) as conn:
            row = conn.execute(
                "SELECT price_cny, item FROM products WHERE bld_no = ?", ("K-PRICE-VIS-001",)
            ).fetchone()
        self.assertEqual(float(row["price_cny"]), 80.0)
        self.assertEqual(row["item"], "Price Vis Arm Renamed")
        self.client.post("/logout")

    def test_purchase_contract_can_generate_pdf(self):
        from PIL import Image
        from app.modules.products.persistence import upsert_product

        image_dir = self.root / "data" / "product_images"
        image_dir.mkdir(parents=True, exist_ok=True)
        Image.new("RGB", (96, 64), "white").save(image_dir / "K-OUT-001.png")
        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-OUT-001",
                    "oe_no_1": "OE-CATALOG-001",
                    "item": "目录外购支架",
                    "models": "目录车型",
                    "price_cny": "45.5",
                    "image_path": "data_product_images/K-OUT-001.png",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/contracts")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("合同管理", html)
        self.assertNotIn("采购合同和销售合同分别生成、分别归档。", html)
        self.assertIn("采购合同", html)
        self.assertIn("销售合同", html)
        self.assertNotIn('class="contract-entry-switch"', html)
        self.assertIn('class="contract-history-drawer" id="contract-history">', html)
        self.assertIn("历史合同记录", html)
        self.assertNotIn("销售合同模板后续接入", html)
        self.assertIn("玉环博莱德机械有限公司", html)
        self.assertIn("浙江省玉环市金汇路11号", html)
        self.assertIn("月结 30 天", html)
        self.assertIn('name="product_code[]"', html)
        self.assertIn('name="oe_no[]"', html)
        self.assertIn('name="models[]"', html)
        self.assertIn("data-add-purchase-row", html)
        self.assertIn("data-supplier-sign-name", html)
        self.assertIn("data-purchase-confirm-modal", html)
        self.assertIn("确认生成 PDF", html)
        self.assertNotIn("统一社会信用代码", html)
        self.assertIn('name="buyer_signature_address"', html)
        self.assertIn('name="supplier_signature_address"', html)
        self.assertIn('name="buyer_bank"', html)
        self.assertIn('name="supplier_bank_account"', html)
        self.assertIn('name="buyer_signature_date"', html)
        self.assertIn("supplier-detail-line", html)

        lookup = self.client.get("/purchase-contracts/product-lookup", query_string={"bld": "K-OUT-001"})
        self.assertEqual(lookup.status_code, 200)
        payload = lookup.get_json()
        self.assertTrue(payload["found"])
        self.assertEqual(payload["oe_no"], "OE-CATALOG-001")
        self.assertEqual(payload["product_name"], "目录外购支架")
        self.assertEqual(payload["models"], "目录车型")
        self.assertEqual(payload["price_cny"], 45.5)
        self.assertIn("product-image-thumbs", payload["thumb_url"])

        response = self.client.post(
            "/purchase-contracts/generate",
            data={
                "contract_no": "CG-TEST-001",
                "contract_date": "2026-05-05",
                "buyer_name": "玉环博莱德机械有限公司",
                "buyer_contact": "李四",
                "buyer_phone": "13900000000",
                "supplier_name": "外购供应商",
                "supplier_contact": "张三",
                "supplier_phone": "13800000000",
                "buyer_signature_address": "甲方签章地址",
                "supplier_signature_address": "乙方签章地址",
                "buyer_signature_phone": "0576-11111111",
                "supplier_signature_phone": "0576-22222222",
                "buyer_bank": "甲方开户行",
                "supplier_bank": "乙方开户行",
                "buyer_bank_account": "11112222",
                "supplier_bank_account": "33334444",
                "buyer_signature_date": "2026-05-06",
                "supplier_signature_date": "2026-05-07",
                "delivery_address": "浙江省玉环市",
                "price_note": "以上价格为含税价（增值税税率13%），含包装费及运费，送达甲方指定地点。",
                "payment_terms": "月结",
                "quality_terms": "按图纸执行",
                "product_code[]": ["K-OUT-001", ""],
                "oe_no[]": ["手填OE", ""],
                "product_name[]": ["手填名称", ""],
                "models[]": ["手填车型", ""],
                "quantity[]": ["10", ""],
                "unit_price[]": ["25.5", ""],
                "delivery_date[]": ["2026-05-20", ""],
                "item_note[]": ["加急", ""],
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/pdf")
        self.assertIn("CG-TEST-001", response.headers["Content-Disposition"])
        self.assertTrue(response.get_data().startswith(b"%PDF-"))
        response.close()
        files = list((self.root / "outputs").glob("u*-007/采购合同/外购供应商/CG-TEST-001外购供应商.pdf"))
        self.assertEqual(len(files), 1)
        response = self.client.get("/contracts", query_string={"contract_q": "CG-TEST-001"})
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("已生成合同", html)
        self.assertIn('class="contract-history-drawer" id="contract-history" open>', html)
        self.assertIn("外购供应商", html)
        self.assertIn("CG-TEST-001外购供应商.pdf", html)
        self.assertIn("下载", html)

    def test_sales_contract_can_generate_pdf_with_customer_code(self):
        from app.modules.products.persistence import upsert_product

        with self.web.connect(self.web.DB_PATH) as conn:
            upsert_product(
                conn,
                {
                    "bld_no": "K-SALE-001",
                    "oe_no_1": "OE-SALE-001",
                    "item": "销售控制臂",
                    "models": "销售车型",
                    "price_cny": "88.8",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        response = self.client.get("/contracts/sales")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("产 品 销 售 合 同", html)
        self.assertIn('class="contract-history-drawer" id="contract-history">', html)
        self.assertIn("供方（甲方）", html)
        self.assertIn("需方（乙方）", html)
        self.assertIn("客户编码", html)
        self.assertIn('name="customer_code[]"', html)
        self.assertIn('action="/sales-contracts/generate"', html)
        self.assertIn("甲方按行业通用标准及乙方要求进行包装", html)
        self.assertIn("增值税专用发票（税率 13%）的开具时间由双方另行约定", html)

        response = self.client.post(
            "/sales-contracts/generate",
            data={
                "contract_no": "XS-TEST-001",
                "contract_date": "2026-05-06",
                "buyer_name": "玉环博莱德机械有限公司",
                "buyer_contact": "李四",
                "buyer_phone": "13900000000",
                "supplier_name": "销售客户",
                "supplier_contact": "王五",
                "supplier_phone": "13700000000",
                "delivery_address": "客户仓库",
                "price_note": "以上价格为含税价。",
                "payment_terms": "月结 30 天",
                "quality_terms": "按封样执行",
                "product_code[]": ["K-SALE-001"],
                "customer_code[]": ["CUST-001"],
                "oe_no[]": ["手填销售OE"],
                "product_name[]": ["手填销售名称"],
                "models[]": ["手填销售车型"],
                "quantity[]": ["3"],
                "unit_price[]": ["88.8"],
                "delivery_date[]": ["2026-05-25"],
                "item_note[]": ["销售备注"],
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.mimetype, "application/pdf")
        self.assertIn("XS-TEST-001", response.headers["Content-Disposition"])
        self.assertTrue(response.get_data().startswith(b"%PDF-"))
        response.close()
        files = list((self.root / "outputs").glob("u*-007/销售合同/销售客户/XS-TEST-001销售客户.pdf"))
        self.assertEqual(len(files), 1)
        response = self.client.get("/contracts", query_string={"contract_type": "sales", "contract_q": "销售客户"})
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("已生成合同", html)
        self.assertIn("销售合同", html)
        self.assertIn("销售客户", html)
        self.assertIn("XS-TEST-001销售客户.pdf", html)

    def test_purchase_contract_signature_fields_are_optional(self):
        from app.purchase_contract import purchase_contract_from_form

        class FormData(dict):
            def getlist(self, key):
                value = self.get(key, [])
                return value if isinstance(value, list) else [value]

        contract = purchase_contract_from_form(
            FormData(
                {
                    "contract_no": "CG-OPTIONAL-SIGN",
                    "contract_date": "2026-05-06",
                    "buyer_name": "甲方公司",
                    "supplier_name": "乙方公司",
                    "product_code[]": ["K-OPTIONAL-001"],
                    "quantity[]": ["1"],
                    "unit_price[]": ["2.50"],
                }
            )
        )

        self.assertEqual(contract["buyer_signature_address"], "")
        self.assertEqual(contract["supplier_signature_phone"], "")
        self.assertEqual(contract["buyer_bank"], "")
        self.assertEqual(contract["supplier_bank_account"], "")
        self.assertEqual(contract["buyer_signature_date"], "")

    def test_purchase_contracts_are_admin_only(self):
        from app.modules.admin.persistence import save_user

        with self.web.connect(self.web.DB_PATH) as conn:
            save_user(
                conn,
                {
                    "username": "editor-contracts",
                    "display_name": "Editor Contracts",
                    "password": "editor-pw",
                    "role": "editor",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        admin_page = self.client.get("/").get_data(as_text=True)
        self.assertIn("合同管理", admin_page)
        self.client.post("/logout")

        login = self.client.post(
            "/login",
            data={"username": "editor-contracts", "password": "editor-pw", "next": "/"},
            follow_redirects=False,
        )
        self.assertEqual(login.status_code, 302)

        editor_page = self.client.get("/").get_data(as_text=True)
        self.assertNotIn("合同管理", editor_page)
        for path in ["/contracts", "/contracts/sales", "/purchase-contracts"]:
            with self.subTest(path=path):
                response = self.client.get(path, follow_redirects=False)
                self.assertEqual(response.status_code, 302)
                self.assertTrue(response.headers["Location"].endswith("/"))

        response = self.client.get(
            "/purchase-contracts/product-lookup", query_string={"bld": "K-OUT-001"}, follow_redirects=False
        )
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/"))

        for path in ["/purchase-contracts/generate", "/sales-contracts/generate"]:
            with self.subTest(path=path):
                response = self.client.post(path, follow_redirects=False)
                self.assertEqual(response.status_code, 302)
                self.assertTrue(response.headers["Location"].endswith("/"))
        self.client.post("/logout")

    def test_quote_and_tube_result_fragments_keep_list_state_without_a_document_shell(self):
        self.login()
        cases = (
            (
                "/quotes/fragment",
                {"customer_name": "博世", "page": "2"},
                "data-quote-results",
                "data-quote-results-host",
            ),
            ("/tubes/fragment", {"q": "KE8036", "page": "2"}, "data-tube-results", "data-tube-results-host"),
        )
        for path, query_string, result_marker, host_marker in cases:
            with self.subTest(path=path):
                response = self.client.get(
                    path,
                    query_string=query_string,
                    headers={"Accept": "text/html", "X-Requested-With": "fetch"},
                )
                html = response.get_data(as_text=True)
                self.assertEqual(response.status_code, 200)
                self.assertIn(result_marker, html)
                self.assertNotIn(host_marker, html)
                self.assertNotIn("<!doctype html>", html.lower())
                self.assertEqual(response.headers["Cache-Control"], "no-store")

        tube_page = self.client.get("/tubes")
        self.assertIn("data-tube-results-host", tube_page.get_data(as_text=True))
        self.assertIn('src="/static/pages/tubes.js?v=', tube_page.get_data(as_text=True))

    def test_quote_records_page_can_create_search_and_edit(self):
        self.login()
        self.register_customer_and_product("博世", "K48620")
        response = self.client.get("/quotes")
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("报价记录", html)
        self.assertNotIn("集中保存人工和 Hermes 录入的客户报价", html)
        search_form = re.search(r'<form class="search-form quote-search".*?</form>', html, re.S).group()
        self.assertGreater(html.index('class="search-form quote-search"'), html.index('id="quote-results"'))
        self.assertLess(html.index('class="search-form quote-search"'), html.index('class="data-table-scroll"'))
        self.assertIn('<summary class="linear-button primary">新增报价</summary>', html)
        self.assertIn('<summary class="linear-button primary">导入报价记录</summary>', html)
        self.assertIn("data-quote-results-host", html)
        self.assertIn('data-quote-results-fragment-url="/quotes/fragment"', html)
        self.assertIn('action="/quotes" data-quote-search-form', html)
        new_quote_form = re.search(r'<form action="/quotes/save".*?</form>', html, re.S).group()
        self.assertIn('name="customer_name"', html)
        self.assertIn('name="bld_no"', html)
        self.assertIn('name="customer_product_code"', html)
        self.assertIn('name="tax_price"', html)
        self.assertIn('name="net_price"', html)
        for system_field in ("quoted_by", "source_type", "source_text", "attachment_path"):
            self.assertNotIn(f'name="{system_field}"', new_quote_form)
        self.assertNotIn("<th>原文</th>", html)
        self.assertNotIn("附件路径", html)
        self.assertNotIn('name="date_from"', search_form)
        self.assertNotIn('name="date_to"', search_form)
        self.assertNotIn('name="currency"', search_form)
        self.assertNotIn('name="quoted_by"', search_form)
        self.assertIn('name="currency"', html)
        self.assertIn("导入报价记录", html)
        self.assertIn('action="/quotes/import/preview"', html)
        self.assertIn('name="customer_name"', html)
        self.assertIn('name="currency"', html)
        self.assertIn('name="quote_file"', html)
        self.assertNotIn("MOQ", html)
        html_without_edit_dialogs = re.sub(r'<dialog class="quote-edit-dialog".*?</dialog>', "", html, flags=re.S)
        self.assertNotIn("删除", html_without_edit_dialogs)

        response = self.client.post(
            "/quotes/save",
            data={
                "customer_name": "博世",
                "bld_no": "K48620",
                "customer_product_code": "48620-0K040",
                "tax_price": "5.35",
                "net_price": "4.73",
                "currency": "USD",
                "quote_date": "2026-06-10",
                "quoted_by": "spoofed-web-user",
                "source_type": "image",
                "source_text": "不应由网页保存的原文",
                "attachment_path": "/tmp/should-not-be-saved.pdf",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

        response = self.client.post(
            "/quotes/save",
            data={
                "customer_name": "博世",
                "bld_no": "K48620",
                "customer_product_code": "48620-0K040",
                "tax_price": "5.55",
                "net_price": "4.91",
                "currency": "USD",
                "quote_date": "2026-06-11",
                "quoted_by": "sales",
                "source_type": "wechat",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)

        response = self.client.get("/quotes", query_string={"customer_name": "博世", "bld_no": "K48620"})
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertIn("最近一次报价", html)
        self.assertIn("含税 USD 5.5500", html)
        self.assertIn("不含税 USD 4.9100", html)
        self.assertNotIn("不应由网页保存的原文", html)
        self.assertIn("修正", html)
        self.assertNotIn("data-open-customer-price-delete", html)
        self.assertIn('data-open-quote-edit data-quote-edit-record=', html)
        self.assertEqual(html.count('class="quote-edit-dialog"'), 1)
        self.assertEqual(html.count("data-quote-edit-form"), 1)
        self.assertIn('class="dialog-panel quote-number-dialog"', html)
        self.assertIn('class="dialog-header quote-number-dialog-header"', html)
        self.assertIn('class="dialog-close" type="button" data-close-quote-number', html)
        self.assertNotIn('class="quote-edit-dialog quote-number-dialog"', html)
        self.assertIn('id="quote-number-dialog-title">报价单明细 <span data-quote-number-label>', html)
        self.assertIn('data-quote-number-detail', html)
        self.assertIn('src="/static/pages/quotes.js?', html)
        edit_dialog = re.search(r'<dialog class="quote-edit-dialog".*?</dialog>', html, re.S).group()
        edit_form = re.search(r'<form\b[^>]*data-quote-edit-form.*?</form>', edit_dialog, re.S).group()
        self.assertEqual(edit_form.count('name="csrf_token"'), 1)
        for system_field in ("quoted_by", "source_type", "source_text", "attachment_path"):
            self.assertNotIn(f'name="{system_field}"', edit_form)
        edit_payloads = [
            json.loads(unescape(value))
            for value in re.findall(r"data-quote-edit-record='([^']*)'", html)
        ]
        self.assertEqual(len(edit_payloads), 2)
        for payload in edit_payloads:
            self.assertEqual(
                set(payload),
                {
                    "id",
                    "version",
                    "customer_name",
                    "bld_no",
                    "customer_product_code",
                    "tax_price",
                    "net_price",
                    "currency",
                    "quote_date",
                    "remark",
                    "edit_url",
                    "delete_url",
                },
            )

        with self.web.connect(self.web.DB_PATH) as conn:
            latest_record = conn.execute(
                "SELECT id FROM quote_records WHERE customer_name = ? AND bld_no = ? ORDER BY quote_date DESC, id DESC",
                ("博世", "K48620"),
            ).fetchone()
            first_record = conn.execute(
                "SELECT quoted_by, source_type, source_text, attachment_path FROM quote_records "
                "WHERE customer_name = ? AND bld_no = ? ORDER BY quote_date, id LIMIT 1",
                ("博世", "K48620"),
            ).fetchone()
        quote_id = latest_record["id"]
        self.assertEqual(
            dict(first_record),
            {"quoted_by": "007", "source_type": "manual", "source_text": "", "attachment_path": ""},
        )

        response = self.client.post(
            f"/quotes/{quote_id}/edit",
            data={
                "customer_name": "博世",
                "bld_no": "K48620",
                "customer_product_code": "BOSCH-K48620",
                "tax_price": "5.65",
                "net_price": "5.00",
                "currency": "USD",
                "quote_date": "2026-06-11",
                "quoted_by": "spoofed-editor",
                "source_type": "pdf",
                "source_text": "spoofed-edit-source",
                "attachment_path": "/tmp/spoofed-edit.pdf",
                "remark": "人工复核",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        with self.web.connect(self.web.DB_PATH) as conn:
            revised = conn.execute("SELECT * FROM quote_records WHERE id = ?", (quote_id,)).fetchone()
            revisions = conn.execute(
                "SELECT COUNT(*) FROM quote_record_revisions WHERE quote_id = ?", (quote_id,)
            ).fetchone()[0]
        self.assertEqual(revised["bld_no"], "K48620")
        self.assertEqual(revised["customer_product_code"], "BOSCH-K48620")
        self.assertEqual(revised["tax_price"], 5.65)
        self.assertEqual(revised["net_price"], 5.00)
        self.assertEqual(revised["quoted_by"], "007")
        self.assertEqual(revised["source_type"], "manual")
        self.assertEqual(revised["source_text"], "")
        self.assertEqual(revised["attachment_path"], "")
        self.assertEqual(revisions, 1)

        response = self.client.post(
            f"/quotes/{quote_id}/edit",
            data={
                "customer_name": "博世",
                "bld_no": "K48620",
                "tax_price": "5.65",
                "net_price": "",
                "currency": "USD",
                "quote_date": "2026-06-11",
                "quoted_by": "sales",
                "source_type": "wechat",
                "source_text": "",
                "attachment_path": "",
                "remark": "",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 302)
        with self.web.connect(self.web.DB_PATH) as conn:
            cleared = conn.execute("SELECT * FROM quote_records WHERE id = ?", (quote_id,)).fetchone()
        self.assertIsNone(cleared["net_price"])
        self.assertEqual(cleared["remark"], "")

        response = self.client.get("/quotes", query_string={"customer_name": "博世", "bld_no": "K48620"})
        html = response.get_data(as_text=True)
        edit_dialog = re.search(r'<dialog class="quote-edit-dialog".*?</dialog>', html, re.S).group()
        self.assertIn("data-quote-edit-delete", edit_dialog)
        self.assertNotIn('formaction="/quotes/', edit_dialog)
        self.assertNotIn('data-confirm="确认删除这条报价记录', edit_dialog)
        refreshed_payloads = [
            json.loads(unescape(value))
            for value in re.findall(r"data-quote-edit-record='([^']*)'", html)
        ]
        refreshed_payload = next(payload for payload in refreshed_payloads if payload["id"] == quote_id)
        self.assertEqual(refreshed_payload["delete_url"], f"/quotes/{quote_id}/delete")
        self.assertEqual(refreshed_payload["net_price"], None)
        self.assertEqual(refreshed_payload["remark"], "")

        response = self.client.post(f"/quotes/{quote_id}/delete", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        with self.web.connect(self.web.DB_PATH) as conn:
            deleted = conn.execute("SELECT * FROM quote_records WHERE id = ?", (quote_id,)).fetchone()
            leftover_revisions = conn.execute(
                "SELECT COUNT(*) FROM quote_record_revisions WHERE quote_id = ?", (quote_id,)
            ).fetchone()[0]
        self.assertIsNone(deleted)
        self.assertEqual(leftover_revisions, 0)

        old_path = self.client.get("/customer-prices", follow_redirects=False)
        self.assertEqual(old_path.status_code, 302)
        self.assertTrue(old_path.headers["Location"].endswith("/quotes"))

    def test_quote_edit_payload_is_safe_and_trimmed_by_permission(self):
        from app.modules.admin.persistence import save_role, save_user

        customer_name = "载荷客户 </script><script>alert('x')</script> & \"双引号\" '单引号'"
        bld_no = "QDOM-SPECIAL-001"
        usernames = ("quote-edit-no-delete", "quote-view-only")
        role_keys = []

        def cleanup_records():
            self.client.post("/logout")
            with self.web.connect(self.web.DB_PATH) as connection:
                connection.execute("DELETE FROM quote_records WHERE bld_no = ?", (bld_no,))
                connection.execute("DELETE FROM customers WHERE name = ?", (customer_name,))
                connection.execute("DELETE FROM products WHERE bld_no = ?", (bld_no,))
                for username in usernames:
                    user = connection.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
                    if user:
                        connection.execute("DELETE FROM user_permission_overrides WHERE user_id = ?", (user["id"],))
                        connection.execute("DELETE FROM users WHERE id = ?", (user["id"],))
                for role_key in role_keys:
                    connection.execute("DELETE FROM role_permissions WHERE role_key = ?", (role_key,))
                    connection.execute("DELETE FROM roles WHERE role_key = ?", (role_key,))
                connection.commit()

        self.addCleanup(cleanup_records)
        self.login()
        self.register_customer_and_product(customer_name, bld_no)
        special_remark = "备注 </script><script>alert('remark')</script> & \"quoted\" 'single'"
        saved = self.client.post(
            "/quotes/save",
            data={
                "customer_name": customer_name,
                "bld_no": bld_no,
                "customer_product_code": "CODE<&>\"'",
                "tax_price": "12.34",
                "currency": "USD",
                "quote_date": "2026-08-08",
                "remark": special_remark,
            },
            follow_redirects=False,
        )
        self.assertEqual(saved.status_code, 302)

        admin_html = self.client.get("/quotes", query_string={"bld_no": bld_no}).get_data(as_text=True)
        self.assertNotIn("</script><script>alert('x')</script>", admin_html)
        self.assertNotIn("</script><script>alert('remark')</script>", admin_html)
        payload_match = re.search(r"data-quote-edit-record='([^']*)'", admin_html)
        self.assertIsNotNone(payload_match)
        admin_payload = json.loads(unescape(payload_match.group(1)))
        self.assertEqual(admin_payload["customer_name"], customer_name)
        self.assertEqual(admin_payload["remark"], special_remark)
        self.assertEqual(admin_payload["customer_product_code"], "CODE<&>\"'")
        self.assertIn("delete_url", admin_payload)
        self.assertFalse(
            {"quoted_by", "source_type", "source_text", "attachment_path", "created_at", "updated_at"}
            & set(admin_payload)
        )

        with self.web.connect(self.web.DB_PATH) as connection:
            edit_role = save_role(
                connection,
                {"name": "报价修正无删除", "description": "载荷权限裁剪测试"},
                ["view_customer_prices", "edit_customer_prices"],
                actor="tester",
            )
            view_role = save_role(
                connection,
                {"name": "报价只读", "description": "载荷权限裁剪测试"},
                ["view_customer_prices"],
                actor="tester",
            )
            role_keys.extend((edit_role, view_role))
            save_user(
                connection,
                {
                    "username": usernames[0],
                    "password": "quote-edit-pw",
                    "role": edit_role,
                    "active": "1",
                },
                actor="tester",
            )
            save_user(
                connection,
                {
                    "username": usernames[1],
                    "password": "quote-view-pw",
                    "role": view_role,
                    "active": "1",
                },
                actor="tester",
            )

        self.client.post("/logout")
        self.client.post(
            "/login",
            data={"username": usernames[0], "password": "quote-edit-pw", "next": "/quotes"},
        )
        edit_html = self.client.get("/quotes", query_string={"bld_no": bld_no}).get_data(as_text=True)
        edit_payload = json.loads(unescape(re.search(r"data-quote-edit-record='([^']*)'", edit_html).group(1)))
        self.assertNotIn("delete_url", edit_payload)
        self.assertIn('id="quote-edit-dialog"', edit_html)
        self.assertNotIn("data-quote-edit-delete", edit_html)

        self.client.post("/logout")
        self.client.post(
            "/login",
            data={"username": usernames[1], "password": "quote-view-pw", "next": "/quotes"},
        )
        view_html = self.client.get("/quotes", query_string={"bld_no": bld_no}).get_data(as_text=True)
        self.assertNotIn("data-quote-edit-record", view_html)
        self.assertNotIn('id="quote-edit-dialog"', view_html)

    def test_quote_list_initial_markup_stays_slim_at_full_page_size(self):
        from app.modules.quotes.factory import get_quote_service

        customer_name = "报价 DOM 百行测试客户"
        bld_no = "QDOM-100-ROWS"

        def cleanup_records():
            with self.web.connect(self.web.DB_PATH) as connection:
                connection.execute("DELETE FROM quote_records WHERE bld_no = ?", (bld_no,))
                connection.execute("DELETE FROM customers WHERE name = ?", (customer_name,))
                connection.execute("DELETE FROM products WHERE bld_no = ?", (bld_no,))
                connection.commit()

        self.addCleanup(cleanup_records)
        self.login()
        self.register_customer_and_product(customer_name, bld_no)
        with self.web.app.app_context():
            service = get_quote_service()
            for index in range(100):
                service.create(
                    {
                        "customer_name": customer_name,
                        "bld_no": bld_no,
                        "customer_product_code": "QDOM-CODE",
                        "tax_price": "88.88",
                        "currency": "CNY",
                        "quote_date": "2026-08-08",
                        "remark": "百行初始 DOM 指标",
                    },
                    actor="007",
                )

        response = self.client.get("/quotes", query_string={"customer_name": customer_name, "bld_no": bld_no})
        html = response.get_data(as_text=True)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(html.count("data-quote-edit-record="), 100)
        self.assertEqual(html.count('class="quote-edit-dialog"'), 1)
        self.assertEqual(html.count("data-quote-edit-form"), 1)
        edit_dialog = re.search(r'<dialog class="quote-edit-dialog".*?</dialog>', html, re.S).group()
        self.assertEqual(edit_dialog.count('name="csrf_token"'), 1)
        self.assertEqual(html.count("data-column-filter-panel"), 11)
        self.assertEqual(html.count("data-quote-filter-options></div>"), 11)
        self.assertNotIn("data-column-filter-option><input", html)
        self.assertEqual(html.count("data-quote-filter-state="), 1)
        for unused_stylesheet in (
            "components/import-preview.css",
            "components/search_clear.css",
            "components/inline_search_command.css",
        ):
            self.assertNotIn(unused_stylesheet, html)
        self.assertLess(
            len(response.data),
            220_000,
            f"100 行报价首屏 HTML 回归为 {len(response.data)} 字节，应保持低于 220000 字节",
        )

    def test_quote_records_can_import_excel_into_quote_table(self):
        from openpyxl import Workbook

        self.login()
        self.register_customer_and_product("导入客户", "K-IMPORT-QUOTE")
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["BLD号", "客户产品编码", "含税单价", "不含税单价", "报价日期", "报价人", "来源", "原文", "备注"])
        sheet.append(
            [
                "K-IMPORT-QUOTE",
                "CUST-IMPORT",
                12.34,
                10.92,
                "2026-07-01",
                "importer",
                "excel",
                "导入客户 K-IMPORT-QUOTE USD 12.34",
                "批量导入",
            ]
        )
        buffer = io.BytesIO()
        workbook.save(buffer)
        workbook.close()
        buffer.seek(0)

        preview = self.client.post(
            "/quotes/import/preview",
            data={"customer_name": "导入客户", "currency": "USD", "quote_file": (buffer, "quotes.xlsx")},
            content_type="multipart/form-data",
        )
        html = preview.get_data(as_text=True)
        self.assertEqual(preview.status_code, 200)
        self.assertIn("导入报价记录", html)
        self.assertIn("K-IMPORT-QUOTE", html)
        payload_match = re.search(r'name="payload" value="([^"]+)"', html)
        self.assertIsNotNone(payload_match)

        payload = __import__("html").unescape(payload_match.group(1))
        apply = self.client.post(
            "/quotes/import/apply",
            data={"payload": payload},
            follow_redirects=False,
        )
        self.assertEqual(apply.status_code, 302)

        with self.web.connect(self.web.DB_PATH) as conn:
            row = conn.execute("SELECT * FROM quote_records WHERE bld_no = ?", ("K-IMPORT-QUOTE",)).fetchone()
        self.assertIsNotNone(row)
        self.assertEqual(row["customer_name"], "导入客户")
        self.assertEqual(row["customer_product_code"], "CUST-IMPORT")
        self.assertEqual(row["currency"], "USD")
        self.assertEqual(row["tax_price"], 12.34)
        self.assertEqual(row["net_price"], 10.92)
        self.assertEqual(row["quoted_by"], "007")
        self.assertEqual(row["source_type"], "excel")
        self.assertEqual(row["source_text"], "")

    def test_customers_page_lookup_and_v1_endpoint(self):
        self.login()
        page = self.client.get("/customers")
        self.assertEqual(page.status_code, 200)
        self.assertIn("客户信息", page.get_data(as_text=True))

        save = self.client.post("/customers/save", data={"name": "宁波多迦"}, follow_redirects=False)
        self.assertEqual(save.status_code, 302)
        duplicate = self.client.post("/customers/save", data={"name": "宁波多迦"}, follow_redirects=True)
        self.assertIn("已存在", duplicate.get_data(as_text=True))

        lookup = self.client.get("/customers/lookup?q=多迦")
        self.assertEqual(lookup.status_code, 200)
        self.assertEqual([row["name"] for row in lookup.get_json()], ["宁波多迦"])

        token = self.create_internal_api_token(scopes=["quotes:read"], name="Customers Read")
        v1 = self.client.get("/api/v1/customers?q=多迦", headers={"Authorization": f"Bearer {token}"})
        self.assertEqual(v1.status_code, 200)
        data = v1.get_json()["data"]
        self.assertEqual(data["customers"], [{"name": "宁波多迦"}])
        self.assertEqual(data["total"], 1)

    def test_customer_page_filters_by_owner_label_and_bounds_long_queries(self):
        self.login()
        with self.web.connect(self.web.DB_PATH) as connection:
            owner = connection.execute(
                "SELECT display_name FROM users WHERE username = '007'"
            ).fetchone()
            connection.execute(
                """
                INSERT INTO customers (name, owner_username, sync_id)
                VALUES ('负责人页面筛选客户', '007', 'customer-owner-filter-test')
                """
            )
            connection.commit()

        self.assertIsNotNone(owner)
        display_name = str(owner["display_name"] or "")
        self.assertTrue(display_name)

        long_query = self.client.get("/customers", query_string={"q": "x" * 201})
        by_display_name = self.client.get("/customers", query_string={"q": display_name})
        by_username = self.client.get("/customers", query_string={"q": "007"})

        self.assertEqual(long_query.status_code, 200)
        self.assertIn('maxlength="200"', long_query.get_data(as_text=True))
        self.assertIn("负责人页面筛选客户", by_display_name.get_data(as_text=True))
        self.assertIn("负责人页面筛选客户", by_username.get_data(as_text=True))

    def test_quote_api_oversized_request_returns_json(self):
        response = self.client.post(
            "/api/v1/quotes",
            data=b"x" * (21 * 1024 * 1024),
            content_type="application/json",
            follow_redirects=False,
        )
        payload = response.get_json()
        self.assertEqual(response.status_code, 413)
        self.assertEqual(payload["error"]["code"], "request.too_large")
        self.assertIn("上传文件不能超过", payload["error"]["message"])

    def test_quotes_are_admin_only(self):
        from app.modules.admin.persistence import save_user

        with self.web.connect(self.web.DB_PATH) as conn:
            save_user(
                conn,
                {
                    "username": "editor-prices",
                    "display_name": "Editor Prices",
                    "password": "editor-pw",
                    "role": "editor",
                    "active": "1",
                },
                actor="tester",
            )

        self.login()
        admin_page = self.client.get("/").get_data(as_text=True)
        self.assertIn("报价记录", admin_page)
        self.client.post("/logout")

        login = self.client.post(
            "/login",
            data={"username": "editor-prices", "password": "editor-pw", "next": "/"},
            follow_redirects=False,
        )
        self.assertEqual(login.status_code, 302)

        editor_page = self.client.get("/").get_data(as_text=True)
        self.assertNotIn("报价记录", editor_page)
        response = self.client.get("/quotes", follow_redirects=False)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(response.headers["Location"].endswith("/"))
        self.client.post("/logout")

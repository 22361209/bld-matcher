from __future__ import annotations

import gc
import io
import os
import sys
import tempfile
import unittest
from html.parser import HTMLParser
from importlib.util import module_from_spec, spec_from_file_location
from pathlib import Path
from typing import Any, ClassVar

from flask.testing import FlaskClient
from openpyxl import load_workbook
from werkzeug.datastructures import MultiDict


PROJECT_ROOT = Path(__file__).resolve().parents[1]


class ExportFormParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.action = ""
        self.fields: list[tuple[str, str]] = []
        self._inside_export_form = False

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        attributes = dict(attrs)
        if tag == "form":
            action = attributes.get("action") or ""
            self._inside_export_form = action.endswith("/products/export")
            if self._inside_export_form:
                self.action = action
            return
        if tag != "input" or not self._inside_export_form:
            return
        name = attributes.get("name")
        if name:
            self.fields.append((name, attributes.get("value") or ""))

    def handle_endtag(self, tag: str) -> None:
        if tag == "form" and self._inside_export_form:
            self._inside_export_form = False


def load_isolated_web_module() -> Any:
    module_name = "bld_matcher_catalog_export_form_test_web"
    spec = spec_from_file_location(module_name, PROJECT_ROOT / "app.py")
    assert spec and spec.loader
    module = module_from_spec(spec)
    sys.modules[module_name] = module
    spec.loader.exec_module(module)
    return module


class ProductCatalogExportFormTest(unittest.TestCase):
    temporary: ClassVar[tempfile.TemporaryDirectory[str]]
    root: ClassVar[Path]
    web: ClassVar[Any]
    client: ClassVar[FlaskClient]

    @classmethod
    def setUpClass(cls) -> None:
        cls.temporary = tempfile.TemporaryDirectory()
        cls.root = Path(cls.temporary.name)
        os.environ.update(
            {
                "SECRET_KEY": "catalog-export-form-test-secret",
                "MAX_UPLOAD_MB": "20",
                "PRODUCT_SYNC_MAX_UPLOAD_MB": "512",
                "BLD_DATA_DIR": str(cls.root / "data"),
                "BLD_UPLOAD_DIR": str(cls.root / "uploads"),
                "BLD_OUTPUT_DIR": str(cls.root / "outputs"),
                "DEFAULT_ADMIN_PASSWORD": "test-admin-pw",
                "INTERNAL_API_TOKEN": "",
            }
        )
        for module_name in [name for name in sys.modules if name == "app" or name.startswith("app.")]:
            sys.modules.pop(module_name, None)
        cls.web = load_isolated_web_module()
        if not cls.web.DB_PATH.resolve().is_relative_to(cls.root.resolve()):
            raise RuntimeError(f"Test database escaped the isolated root: {cls.web.DB_PATH}")
        cls.web.app.config["TESTING"] = True
        cls.client = cls.web.app.test_client()

    @classmethod
    def tearDownClass(cls) -> None:
        gc.collect()
        cls.temporary.cleanup()

    def test_rendered_export_form_submits_current_multiselect_filters(self) -> None:
        from app.modules.products.persistence import upsert_product

        fixtures = (
            ("K-FORM-001", "FORM-A", "Rendered Form Arm", "1个衬套"),
            ("K-FORM-002", "FORM-B", "Rendered Form Arm", "1 个衬套"),
            ("K-FORM-003", "FORM-C", "Rendered Form Arm", "1个衬套"),
            ("K-FORM-004", "FORM-A", "Different Arm", "1个衬套"),
        )
        with self.web.connect(self.web.DB_PATH) as connection:
            for bld_no, series, item, product_status in fixtures:
                upsert_product(
                    connection,
                    {
                        "bld_no": bld_no,
                        "series": series,
                        "item": item,
                        "product_status": product_status,
                        "active": "1",
                    },
                    actor="catalog-export-form-test",
                )

        login = self.client.post(
            "/login",
            data={"username": "007", "password": "test-admin-pw", "next": "/"},
            follow_redirects=False,
        )
        self.assertEqual(login.status_code, 302)

        page = self.client.get(
            "/products",
            query_string={
                "bld": "K-FORM-",
                "brand": ["FORM-A", "FORM-B"],
                "item": ["Rendered Form Arm"],
                "product_status": ["1衬套"],
                "status": "active",
            },
        )
        self.assertEqual(page.status_code, 200)

        parser = ExportFormParser()
        parser.feed(page.get_data(as_text=True))
        form_data = MultiDict(parser.fields)

        self.assertEqual(parser.action, "/products/export")
        self.assertEqual(form_data.get("bld"), "K-FORM-")
        self.assertEqual(form_data.get("oe"), "")
        self.assertEqual(form_data.get("status"), "active")
        self.assertEqual(form_data.getlist("brand"), ["FORM-A", "FORM-B"])
        self.assertEqual(form_data.getlist("item"), ["Rendered Form Arm"])
        self.assertEqual(form_data.getlist("product_status"), ["1衬套"])

        form_data.setlist("export_format", ["bld"])
        exported_response = self.client.post(parser.action, data=form_data)
        self.assertEqual(exported_response.status_code, 200)

        workbook = load_workbook(io.BytesIO(exported_response.data), read_only=True, data_only=True)
        try:
            sheet = workbook["产品目录"]
            exported_bld_numbers = [row[0] for row in sheet.iter_rows(min_row=2, values_only=True)]
        finally:
            workbook.close()
            exported_response.close()

        self.assertEqual(exported_bld_numbers, ["K-FORM-001", "K-FORM-002"])


if __name__ == "__main__":
    unittest.main()

from __future__ import annotations

import sqlite3
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.catalog_export import export_products_xlsx
from app.product_status import format_product_status


class ProductStatusFormatTest(unittest.TestCase):
    def test_bushing_and_ball_joint_are_formatted_on_separate_lines(self) -> None:
        self.assertEqual(format_product_status("2个衬套1个球头", "zh"), "2个衬套\n1个球头")
        self.assertEqual(format_product_status("2个衬套1个球头", "en"), "2 bushings\n1 ball joint")
        self.assertEqual(
            format_product_status("2个衬套1个球头", "en", multiline=False),
            "2 bushings 1 ball joint",
        )
        self.assertEqual(format_product_status("1个衬套", "zh"), "1个衬套")
        self.assertEqual(format_product_status("特殊状态", "zh"), "特殊状态")

    def test_product_catalog_export_wraps_formatted_status(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output_path = Path(temporary) / "product-catalog.xlsx"
            connection = sqlite3.connect(":memory:")
            connection.row_factory = sqlite3.Row
            connection.execute(
                """
                CREATE TABLE products (
                    bld_no TEXT, series TEXT, item TEXT, oe_no_1 TEXT, oe_no_2 TEXT,
                    models TEXT, image_path TEXT, price_cny REAL, product_status TEXT,
                    active INTEGER, updated_at TEXT
                )
                """
            )
            connection.execute(
                "INSERT INTO products VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                ("K-STATUS-001", "TEST", "Control Arm", "", "", "", "", 10, "2个衬套1个球头", 1, "2026-07-15"),
            )
            try:
                rows = connection.execute("SELECT * FROM products").fetchall()
                export_products_xlsx(connection, output_path, export_format="bld", product_rows=rows)
            finally:
                connection.close()

            workbook = load_workbook(output_path, read_only=False, data_only=True)
            try:
                sheet = workbook.active
                self.assertEqual(sheet.cell(2, 9).value, "2个衬套\n1个球头")
                self.assertGreaterEqual(sheet.row_dimensions[2].height, 30)
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()

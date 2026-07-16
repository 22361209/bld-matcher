from __future__ import annotations

import sqlite3
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, urlsplit
from unittest.mock import patch

from flask import Flask
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image

from app.database import connect
from app.matcher import normalize_code
from app.modules.products.persistence import upsert_product
import app.modules.products.persistence as product_persistence
import app.modules.products.repository as product_repository
from app.modules.products import catalog_web
from app.modules.products.catalog_import import CatalogImportStorage
from app.modules.products.domain import ProductFilters, ProductFilterValidationError
from app.migrations import run_migrations
from app.modules.inquiry.domain import InquiryValidationError, extract_numbers
from app.modules.inquiry.infrastructure import WorkbookInquiryEngine
from app.modules.inquiry.repository import SQLiteInquiryUnitOfWork
from app.modules.inquiry.service import InquiryService
from app.modules.products.repository import SQLiteProductUnitOfWork
from app.modules.products.service import ProductService
from app.platform.artifacts import ArtifactNotFoundError, SQLiteArtifactStore


class ProductInquiryModuleTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.database_path = self.root / "data" / "products.sqlite3"
        self.upload_dir = self.root / "uploads"
        self.output_dir = self.root / "outputs"
        self.product_service = ProductService(
            lambda: SQLiteProductUnitOfWork(self.database_path),
            lambda: None,
            lambda: {},
            CatalogImportStorage(
                self.root / "data" / "catalog.xlsx",
                self.root / "data" / "product_images",
                self.root / "data" / "product_images" / "thumbs",
            ),
        )
        self.artifact_store = SQLiteArtifactStore(self.database_path, (self.output_dir,))
        self.inquiry_service = InquiryService(
            self.product_service,
            WorkbookInquiryEngine(
                base_dir=self.root,
                upload_dir=self.upload_dir,
                output_dir=self.output_dir,
            ),
            lambda: SQLiteInquiryUnitOfWork(self.database_path),
            self.artifact_store,
        )

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def add_product(
        self,
        bld_no: str,
        oe_no: str,
        *,
        item: str = "Test Arm",
        series: str = "",
        product_status: str = "",
        active: bool = True,
    ) -> None:
        with connect(self.database_path) as connection:
            upsert_product(
                connection,
                {
                    "bld_no": bld_no,
                    "oe_no_1": oe_no,
                    "item": item,
                    "series": series,
                    "product_status": product_status,
                    "price_cny": "55",
                    "active": "1" if active else "0",
                },
                actor="module-test",
            )

    def test_oversized_numeric_inquiry_code_is_rejected_without_integer_conversion_error(self) -> None:
        oversized_digits = "1" * 4301

        self.assertEqual(normalize_code(oversized_digits), "")
        self.assertEqual(normalize_code("1e500"), "")
        self.assertEqual(normalize_code("001234.0"), "1234")
        self.assertEqual(extract_numbers({"numbers": [oversized_digits]}), ([], [oversized_digits]))

    def test_catalog_import_requires_template_fields_and_imports_price_status_and_image(self) -> None:
        catalog_path = self.root / "catalog-import.xlsx"
        image_path = self.root / "catalog-product.png"
        Image.new("RGB", (320, 240), "white").save(image_path)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "产品目录"
        sheet.append(
            [
                "BLD NO.",
                "SERIES",
                "ITEM",
                "OE NO.1",
                "Models",
                "产品状态",
                "导入单价",
                "OE NO.2",
                "图片",
            ]
        )
        sheet.append(
            [
                "K-IMPORT-001",
                "TOYOTA",
                "BALL JOINT",
                "43330-09070",
                "CAMRY",
                "1 个球头",
                68.5,
                "4333009070",
                "",
            ]
        )
        sheet.add_image(ExcelImage(BytesIO(image_path.read_bytes())), "I2")
        workbook.save(catalog_path)
        workbook.close()

        preview = self.product_service.preview_catalog_import(catalog_path)
        result = self.product_service.apply_catalog_import(
            catalog_path,
            expected_digest=preview.digest,
            update_bld_nos=set(),
            actor="module-test",
        )

        self.assertEqual(result.created_count, 1)
        image_dir = self.root / "data" / "product_images"
        with connect(self.database_path) as connection:
            product = connection.execute(
                "SELECT * FROM products WHERE bld_no = ?",
                ("K-IMPORT-001",),
            ).fetchone()
        self.assertIsNotNone(product)
        self.assertEqual(product["price_cny"], 68.5)
        self.assertEqual(product["product_status"], "1 个球头")
        self.assertEqual(product["image_path"], "data_product_images/K-IMPORT-001.png")
        self.assertTrue((image_dir / "K-IMPORT-001.png").is_file())

    def test_catalog_import_rejects_missing_required_value_before_writing(self) -> None:
        catalog_path = self.root / "catalog-invalid.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["BLD NO.", "SERIES", "ITEM", "OE NO.1", "Models", "产品状态", "导入单价"])
        sheet.append(["K-IMPORT-INVALID", "TOYOTA", "", "43330", "CAMRY", "1 个球头", 68.5])
        workbook.save(catalog_path)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "第 2 行缺少：ITEM"):
            self.product_service.preview_catalog_import(catalog_path)
        with connect(self.database_path) as connection:
            count = connection.execute(
                "SELECT COUNT(*) FROM products WHERE bld_no = ?",
                ("K-IMPORT-INVALID",),
            ).fetchone()[0]
        self.assertEqual(count, 0)

    def test_catalog_import_rejects_nonempty_row_without_bld_no(self) -> None:
        catalog_path = self.root / "catalog-missing-bld.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["BLD NO.", "SERIES", "ITEM", "OE NO.1", "Models", "产品状态", "导入单价"])
        sheet.append(["", "TOYOTA", "BALL JOINT", "43330", "CAMRY", "1 个球头", 68.5])
        workbook.save(catalog_path)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "第 2 行缺少：BLD NO."):
            self.product_service.preview_catalog_import(catalog_path)

    def test_catalog_import_requires_per_product_selection_for_conflicts(self) -> None:
        self.add_product("K-CONFLICT-001", "OLD-OE", item="Old item", series="TOYOTA", product_status="1 个球头")
        catalog_path = self.root / "catalog-conflict.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["BLD NO.", "SERIES", "ITEM", "OE NO.1", "Models", "产品状态", "导入单价"])
        sheet.append(["K-CONFLICT-001", "TOYOTA", "New item", "NEW-OE", "CAMRY", "2 个球头", 88])
        workbook.save(catalog_path)
        workbook.close()

        preview = self.product_service.preview_catalog_import(catalog_path)
        self.assertEqual(len(preview.conflicts), 1)
        kept = self.product_service.apply_catalog_import(
            catalog_path,
            expected_digest=preview.digest,
            update_bld_nos=set(),
            actor="module-test",
        )
        self.assertEqual(kept.updated_count, 0)
        self.assertEqual(kept.kept_count, 1)
        with connect(self.database_path) as connection:
            product = connection.execute("SELECT * FROM products WHERE bld_no = ?", ("K-CONFLICT-001",)).fetchone()
        self.assertEqual(product["item"], "Old item")

        applied = self.product_service.apply_catalog_import(
            catalog_path,
            expected_digest=preview.digest,
            update_bld_nos={"K-CONFLICT-001"},
            actor="module-test",
        )
        self.assertEqual(applied.updated_count, 1)
        with connect(self.database_path) as connection:
            product = connection.execute("SELECT * FROM products WHERE bld_no = ?", ("K-CONFLICT-001",)).fetchone()
        self.assertEqual(product["item"], "New item")
        self.assertEqual(product["price_cny"], 88)

    def test_catalog_import_restores_media_and_catalog_when_confirmation_fails(self) -> None:
        image_dir = self.root / "data" / "product_images"
        image_dir.mkdir(parents=True)
        old_image = image_dir / "K-ROLLBACK-001.png"
        Image.new("RGB", (20, 20), "red").save(old_image)
        old_bytes = old_image.read_bytes()
        with connect(self.database_path) as connection:
            upsert_product(
                connection,
                {
                    "bld_no": "K-ROLLBACK-001",
                    "series": "TOYOTA",
                    "item": "Old item",
                    "oe_no_1": "OLD-OE",
                    "models": "CAMRY",
                    "price_cny": "10",
                    "product_status": "1 个球头",
                    "image_path": "data_product_images/K-ROLLBACK-001.png",
                },
                actor="module-test",
            )
        new_image = self.root / "new-image.png"
        Image.new("RGB", (20, 20), "blue").save(new_image)
        catalog_path = self.root / "catalog-rollback.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["BLD NO.", "SERIES", "ITEM", "OE NO.1", "Models", "产品状态", "导入单价", "图片"])
        sheet.append(["K-ROLLBACK-001", "TOYOTA", "New item", "NEW-OE", "CAMRY", "2 个球头", 88, ""])
        sheet.add_image(ExcelImage(BytesIO(new_image.read_bytes())), "H2")
        workbook.save(catalog_path)
        workbook.close()

        preview = self.product_service.preview_catalog_import(catalog_path)
        catalog_target = self.root / "data" / "catalog.xlsx"
        self.assertFalse(catalog_target.exists())
        with patch.object(product_repository.SQLiteProductRepository, "export_catalog_source", side_effect=RuntimeError("fail export")):
            with self.assertRaisesRegex(RuntimeError, "fail export"):
                self.product_service.apply_catalog_import(
                    catalog_path,
                    expected_digest=preview.digest,
                    update_bld_nos={"K-ROLLBACK-001"},
                    actor="module-test",
                )

        with connect(self.database_path) as connection:
            product = connection.execute("SELECT * FROM products WHERE bld_no = ?", ("K-ROLLBACK-001",)).fetchone()
        self.assertEqual(product["item"], "Old item")
        self.assertEqual(old_image.read_bytes(), old_bytes)
        self.assertFalse(catalog_target.exists())

    def test_catalog_column_filters_are_exact_multiselect_facets(self) -> None:
        self.add_product(
            "K-FACET-001",
            "FACET-OE-001",
            item="Control Arm",
            series="HONDA\nKIA",
            product_status="2 个衬套\n1 个球头",
        )
        self.add_product(
            "K-FACET-002",
            "FACET-OE-002",
            item="Tie Rod",
            series="KIA",
            product_status="2个衬套1个球头",
        )
        self.add_product(
            "K-FACET-003",
            "FACET-OE-003",
            item="Tie Rod",
            series="HONDA",
        )
        self.add_product("K-FACET-004", "FACET-OE-004", item="", series="")
        self.add_product(
            "K-FACET-005",
            "FACET-OE-005",
            item="Custom Part",
            series="CUSTOM",
            product_status="1个螺丝",
        )
        self.add_product(
            "K-FACET-006",
            "FACET-OE-006",
            item="Leading Zero Part",
            series="CUSTOM",
            product_status="01\u00a0个衬套",
        )
        self.add_product(
            "K-FACET-007",
            "FACET-OE-007",
            item="__blank__",
            series="__blank__",
            product_status="__blank__",
        )

        exact = self.product_service.search(
            {
                "brand": ["HONDA"],
                "item": ["Control Arm"],
                "product_status": ["2个衬套1个球头"],
            },
            limit=10,
        )
        self.assertEqual([record.bld_no for record in exact.records], ["K-FACET-001"])
        self.assertEqual(self.product_service.search({"brand": ["HOND"]}, limit=10).total, 0)

        multi = self.product_service.search(
            {"brand": ["HONDA", "KIA"], "item": ["Control Arm", "Tie Rod"]},
            limit=10,
        )
        self.assertEqual(multi.total, 3)
        blank = self.product_service.search(
            {"brand": [""], "product_status": [""]},
            limit=10,
        )
        self.assertEqual([record.bld_no for record in blank.records], ["K-FACET-004"])
        mixed_with_blank = self.product_service.search(
            {"brand": ["HONDA", ""]},
            limit=10,
        )
        self.assertEqual(
            [record.bld_no for record in mixed_with_blank.records],
            ["K-FACET-001", "K-FACET-003", "K-FACET-004"],
        )
        typed_blank = self.product_service.search(ProductFilters(brand_blank=True), limit=10)
        self.assertEqual([record.bld_no for record in typed_blank.records], ["K-FACET-004"])
        literal_blank_text = self.product_service.search(
            {
                "brand": ["__blank__"],
                "item": ["__blank__"],
                "product_status": ["__blank__"],
            },
            limit=10,
        )
        self.assertEqual([record.bld_no for record in literal_blank_text.records], ["K-FACET-007"])
        custom_status = self.product_service.search(
            {"product_status": ["1个螺丝"]},
            limit=10,
        )
        self.assertEqual([record.bld_no for record in custom_status.records], ["K-FACET-005"])
        leading_zero_status = self.product_service.search(
            {"product_status": ["1衬套"]},
            limit=10,
        )
        self.assertIn("K-FACET-006", [record.bld_no for record in leading_zero_status.records])

        options = self.product_service.filter_options(
            {"brand": ["HONDA"], "item": ["Control Arm"]}
        ).web_payload()
        item_options = {option["value"]: option["count"] for option in options["item"]}
        self.assertEqual(item_options["Control Arm"], 1)
        self.assertEqual(item_options["Tie Rod"], 1)
        brand_options = {option["value"]: option["count"] for option in options["brand"]}
        self.assertEqual(brand_options["HONDA"], 1)
        self.assertEqual(brand_options["KIA"], 1)

        status_options = self.product_service.filter_options({}).web_payload()["product_status"]
        equivalent = [option for option in status_options if option["label"] == "2 bushings 1 ball joint"]
        self.assertEqual(equivalent, [{"value": "2衬套1球头", "label": "2 bushings 1 ball joint", "count": 2}])
        self.assertEqual(
            next(option["count"] for option in status_options if option["value"] == ""),
            2,
        )
        self.assertEqual(
            next(option["count"] for option in status_options if option["value"] == "__blank__"),
            1,
        )

        missing = self.product_service.filter_options(
            {"brand": ["MISSING"], "item": ["Control Arm"]}
        ).web_payload()
        self.assertIn(
            {"value": "MISSING", "label": "MISSING", "count": 0},
            missing["brand"],
        )
        self.assertIn(
            {"value": "Control Arm", "label": "Control Arm", "count": 0},
            missing["item"],
        )

        with self.assertRaisesRegex(ProductFilterValidationError, "品牌筛选项最多选择 200 个"):
            self.product_service.search(
                {"brand": [f"UNKNOWN-{index}" for index in range(201)]},
                limit=10,
            )
        with self.assertRaisesRegex(ProductFilterValidationError, "品牌筛选项最多选择 200 个"):
            self.product_service.search(
                ProductFilters(brands=tuple(f"UNKNOWN-{index}" for index in range(201))),
                limit=10,
            )
        with self.assertRaisesRegex(ProductFilterValidationError, "产品名称筛选项单项不能超过 256 个字符"):
            self.product_service.search({"item": ["X" * 257]}, limit=10)

        invalid_export_path = self.output_dir / "catalog-invalid-filter.xlsx"
        with self.assertRaises(ProductFilterValidationError):
            self.product_service.export_catalog(
                invalid_export_path,
                filters={"product_status": [f"STATUS-{index}" for index in range(201)]},
                export_format="bld",
                actor="module-test",
            )
        self.assertFalse(invalid_export_path.exists())

    def test_catalog_filter_url_preserves_repeated_values_and_page(self) -> None:
        app = Flask(__name__)
        app.add_url_rule("/products", endpoint="products", view_func=lambda: "")
        with app.test_request_context(
            "/products?bld=ignored&oe=OE-001&status=inactive"
            "&brand=HONDA&brand=KIA&brand=&item=Control+Arm"
            "&product_status=2+%E4%B8%AA%E8%A1%AC%E5%A5%97%0A1+%E4%B8%AA%E7%90%83%E5%A4%B4"
        ):
            filters = catalog_web._product_query_args()
            page_url = catalog_web._product_page_url(filters, 3)

        self.assertEqual(filters.bld_query, "")
        self.assertEqual(filters.oe_query, "OE-001")
        self.assertEqual(filters.status, "inactive")
        self.assertEqual(filters.brands, ("HONDA", "KIA"))
        self.assertTrue(filters.brand_blank)
        self.assertEqual(filters.items, ("Control Arm",))
        self.assertEqual(filters.product_statuses, ("2衬套1球头",))
        parsed = parse_qs(urlsplit(page_url).query, keep_blank_values=True)
        self.assertEqual(parsed["oe"], ["OE-001"])
        self.assertEqual(parsed["status"], ["inactive"])
        self.assertEqual(parsed["brand"], ["HONDA", "KIA", ""])
        self.assertEqual(parsed["item"], ["Control Arm"])
        self.assertEqual(parsed["product_status"], ["2衬套1球头"])
        self.assertEqual(parsed["page"], ["3"])
        self.assertEqual(urlsplit(page_url).fragment, "products-results")

    def test_catalog_export_reuses_filters_for_both_formats_and_all_activity_states(self) -> None:
        for index in range(55):
            self.add_product(
                f"K-EXPORT-A-{index:03d}",
                f"EXPORT-A-{index:03d}",
                series="HONDA",
            )
        for index in range(2):
            self.add_product(
                f"K-EXPORT-I-{index:03d}",
                f"EXPORT-I-{index:03d}",
                series="HONDA",
                active=False,
            )
        self.add_product("K-EXPORT-OTHER", "EXPORT-OTHER", series="KIA")

        expected_counts = {"active": 55, "inactive": 2, "all": 57}
        for export_format in ("bld", "brand"):
            for status, expected_count in expected_counts.items():
                with self.subTest(export_format=export_format, status=status):
                    output_path = self.output_dir / f"catalog-{export_format}-{status}.xlsx"
                    exported = self.product_service.export_catalog(
                        output_path,
                        filters={"status": status, "brand": ["HONDA"]},
                        export_format=export_format,
                        actor="module-test",
                    )
                    self.assertEqual(exported, expected_count)
                    workbook = load_workbook(output_path, read_only=True, data_only=True)
                    sheet = workbook["产品目录"]
                    bld_column = 3 if export_format == "brand" else 1
                    exported_bld_numbers = [
                        row[bld_column - 1]
                        for row in sheet.iter_rows(min_row=2, values_only=True)
                    ]
                    workbook.close()
                    self.assertEqual(len(exported_bld_numbers), expected_count)
                    self.assertTrue(all(str(value).startswith("K-EXPORT-") for value in exported_bld_numbers))
                    self.assertNotIn("K-EXPORT-OTHER", exported_bld_numbers)

        empty_path = self.output_dir / "catalog-empty.xlsx"
        exported = self.product_service.export_catalog(
            empty_path,
            filters={"status": "active", "brand": ["MISSING"]},
            export_format="bld",
            actor="module-test",
        )
        self.assertEqual(exported, 0)
        self.assertFalse(empty_path.exists())

        audit_failure_path = self.output_dir / "catalog-audit-failure.xlsx"
        with patch.object(
            product_repository,
            "log_event",
            side_effect=RuntimeError("audit failed"),
        ):
            with self.assertRaisesRegex(RuntimeError, "audit failed"):
                self.product_service.export_catalog(
                    audit_failure_path,
                    filters={"status": "active", "brand": ["HONDA"]},
                    export_format="bld",
                    actor="module-test",
                )
        self.assertFalse(audit_failure_path.exists())

    def test_product_repository_search_and_catalog_observe_wal_updates(self) -> None:
        self.add_product("K-MODULE-001", "MODULE-OE-001")
        page = self.product_service.search({"oe": "MODULE-OE-001"}, limit=10)
        self.assertEqual(page.total, 1)
        self.assertEqual(page.records[0].bld_no, "K-MODULE-001")
        first_catalog = self.product_service.catalog()
        self.assertIsNotNone(first_catalog)
        self.assertIsNotNone(first_catalog.match("", "MODULE-OE-001"))

        with connect(self.database_path) as connection:
            upsert_product(
                connection,
                {
                    "bld_no": "K-MODULE-001",
                    "oe_no_1": "MODULE-OE-002",
                    "item": "Updated in the same second",
                    "active": "1",
                },
                actor="module-test",
            )
        refreshed = self.product_service.catalog()
        self.assertIsNotNone(refreshed)
        self.assertIsNone(refreshed.match("", "MODULE-OE-001"))
        self.assertEqual(refreshed.match("", "MODULE-OE-002").bld_no, "K-MODULE-001")

    def test_inquiry_service_is_shared_and_artifacts_are_owner_scoped(self) -> None:
        self.add_product("K-MODULE-API", "MODULE-API-OE")
        analysis = self.inquiry_service.run_numbers(
            {"numbers": ["MODULE-API-OE"], "price_mode": "net"},
            export=False,
            actor="consumer-a",
        )
        self.assertEqual(analysis.summary["matched"], 1)
        self.assertEqual(analysis.api_payload()["rows"][0]["bld_no"], "K-MODULE-API")
        self.assertEqual(
            self.inquiry_service.quick_search("MODULE-API-OE")[0]["product"]["bld_no"],
            "K-MODULE-API",
        )

        exported = self.inquiry_service.run_numbers(
            {"numbers": ["MODULE-API-OE"], "source_name": "module-contract"},
            export=True,
            actor="consumer-a",
            artifact_owner="key:101",
        )
        self.assertIsNotNone(exported.artifact)
        artifact = self.artifact_store.get(exported.artifact.id, owner_id="key:101")
        self.assertTrue(artifact.storage_path.is_file())
        with self.assertRaises(ArtifactNotFoundError):
            self.artifact_store.get(exported.artifact.id, owner_id="key:202")

        with connect(self.database_path) as connection:
            connection.execute(
                "UPDATE api_artifacts SET expires_at = '2000-01-01 00:00:00' WHERE id = ?",
                (artifact.id,),
            )
            connection.commit()
        with self.assertRaises(ArtifactNotFoundError):
            self.artifact_store.get(artifact.id, owner_id="key:101")
        self.assertEqual(self.artifact_store.purge_expired(), 1)

    def test_inquiry_validation_and_alias_transaction(self) -> None:
        self.add_product("K-MODULE-ALIAS", "MODULE-PRIMARY")
        with self.assertRaises(InquiryValidationError) as context:
            self.inquiry_service.run_numbers({}, export=False, actor="module-test")
        self.assertEqual(context.exception.code, "inquiry.numbers_required")

        appended = self.inquiry_service.save_alias(
            "MODULE-CUSTOMER",
            "K-MODULE-ALIAS",
            "consumer mapping",
            "oe",
            actor="module-test",
        )
        self.assertTrue(appended)
        result = self.inquiry_service.run_numbers(
            {"numbers": ["MODULE-CUSTOMER"]},
            export=False,
            actor="module-test",
        )
        self.assertEqual(result.summary["rows"][0]["bld_no"], "K-MODULE-ALIAS")
        with connect(self.database_path) as connection:
            alias = connection.execute(
                "SELECT bld_no, active FROM aliases WHERE source_code = ?",
                ("MODULECUSTOMER",),
            ).fetchone()
            audit_actions = {
                row["action"] for row in connection.execute("SELECT action FROM audit_logs").fetchall()
            }
        self.assertEqual(alias["bld_no"], "K-MODULE-ALIAS")
        self.assertEqual(alias["active"], 1)
        self.assertIn("新增人工映射", audit_actions)
        self.assertIn("追加OE 号", audit_actions)

    def test_historical_database_adds_artifact_table(self) -> None:
        with connect(self.database_path) as connection:
            connection.execute("DROP TABLE api_artifacts")
            connection.execute("DELETE FROM schema_migrations WHERE id = '016_api_artifacts'")
            connection.commit()
        raw = sqlite3.connect(self.database_path)
        raw.row_factory = sqlite3.Row
        try:
            run_migrations(raw)
            columns = {
                row["name"] for row in raw.execute("PRAGMA table_info(api_artifacts)").fetchall()
            }
            migration = raw.execute(
                "SELECT id FROM schema_migrations WHERE id = '016_api_artifacts'"
            ).fetchone()
        finally:
            raw.close()
        self.assertIn("storage_path", columns)
        self.assertIn("expires_at", columns)
        self.assertIsNotNone(migration)

    def test_catalog_import_rolls_back_the_whole_batch(self) -> None:
        catalog_path = self.root / "catalog.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["BLD NO.", "OE NO.1", "ITEM"])
        sheet.append(["K-IMPORT-001", "IMPORT-OE-001", "First"])
        sheet.append(["K-IMPORT-002", "IMPORT-OE-002", "Second"])
        workbook.save(catalog_path)
        workbook.close()

        original_upsert = product_persistence.upsert_product
        calls = 0

        def fail_on_second(connection, data, *args, **kwargs):
            nonlocal calls
            calls += 1
            if calls == 2:
                raise RuntimeError("simulated batch failure")
            return original_upsert(connection, data, *args, **kwargs)

        with patch.object(product_persistence, "upsert_product", side_effect=fail_on_second):
            with self.assertRaises(RuntimeError):
                self.product_service.import_catalog(catalog_path, actor="module-test")

        with connect(self.database_path) as connection:
            rows = connection.execute(
                "SELECT bld_no FROM products WHERE bld_no LIKE 'K-IMPORT-%'"
            ).fetchall()
            logs = connection.execute(
                "SELECT id FROM audit_logs WHERE target_key = ?",
                (catalog_path.name,),
            ).fetchall()
        self.assertEqual(rows, [])
        self.assertEqual(logs, [])


if __name__ == "__main__":
    unittest.main()

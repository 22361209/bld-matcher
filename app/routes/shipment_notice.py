from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from copy import copy
from datetime import datetime
from pathlib import Path

from flask import flash, redirect, render_template, request, url_for
from openpyxl import load_workbook

from app.config import DATA_DIR, DB_PATH
from app.database import connect, log_event
from app.helpers import download_name, safe_upload_name, unique_prefixed_path, user_output_dir, user_recent_outputs, user_upload_dir, user_upload_path
from app.security import actor_name, permission_required


TEMPLATE_ROOT = DATA_DIR / "shipping_notice_templates"
TEMPLATE_FILE_DIR = TEMPLATE_ROOT / "files"
MANIFEST_PATH = TEMPLATE_ROOT / "manifest.json"
ALLOWED_TEMPLATE_SUFFIXES = {".xlsx", ".xlsm"}
ALLOWED_SHIPMENT_SUFFIXES = {".xlsx", ".csv"}
CODE_ALIASES = {"商品编码", "产品编码", "货号", "型号", "物料编码", "客户编码", "BLD NO.", "BLD号", "ITEM CODE", "PART NO", "PART NO.", "SKU"}
QTY_ALIASES = {"数量", "发货数量", "出货数量", "QTY", "QUANTITY"}


def _now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _template_id() -> str:
    return datetime.now().strftime("%Y%m%d%H%M%S%f")


def _norm(value: object) -> str:
    return re.sub(r"[^A-Z0-9\u4E00-\u9FFF]+", "", str(value or "").strip().upper())


CODE_KEYS = {_norm(alias) for alias in CODE_ALIASES}
QTY_KEYS = {_norm(alias) for alias in QTY_ALIASES}


def _load_manifest() -> list[dict]:
    if not MANIFEST_PATH.is_file():
        return []
    try:
        data = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []
    return data if isinstance(data, list) else []


def _save_manifest(templates: list[dict]) -> None:
    TEMPLATE_ROOT.mkdir(parents=True, exist_ok=True)
    MANIFEST_PATH.write_text(json.dumps(templates, ensure_ascii=False, indent=2), encoding="utf-8")


def _template_file_path(template: dict) -> Path | None:
    file_name = str(template.get("file_name") or "").strip()
    if not file_name:
        return None
    path = (TEMPLATE_FILE_DIR / file_name).resolve()
    if TEMPLATE_FILE_DIR.resolve() not in path.parents:
        return None
    return path


def _find_template(template_id: str) -> dict | None:
    return next((item for item in _load_manifest() if item.get("id") == template_id), None)


def _template_choices() -> tuple[list[str], list[dict]]:
    templates = sorted(
        [item for item in _load_manifest() if item.get("status") == "ready"],
        key=lambda item: (str(item.get("customer") or ""), str(item.get("name") or "")),
    )
    customers = sorted({str(item.get("customer") or "").strip() for item in templates if str(item.get("customer") or "").strip()})
    return customers, templates


def _preview_workbook(path: Path, *, max_rows: int = 8, max_cols: int = 8) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = []
        for values in sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
            rows.append(list(values[:max_cols]))
        cols = min(max((len(row) for row in rows), default=0), max_cols)
        rows = [row + [None] * (cols - len(row)) for row in rows]
        return {"sheet": sheet.title, "columns": [chr(65 + index) for index in range(cols)], "rows": rows}
    finally:
        workbook.close()


def _detect_template_mapping(path: Path) -> dict:
    workbook = load_workbook(path)
    try:
        sheet = workbook.active
        for row in sheet.iter_rows(max_row=min(sheet.max_row, 40)):
            code_col = qty_col = None
            for cell in row:
                text = str(cell.value or "").strip()
                if text in {"{{商品编码}}", "{商品编码}", "{{产品编码}}"}:
                    code_col = cell.column
                elif text in {"{{数量}}", "{数量}", "{{发货数量}}"}:
                    qty_col = cell.column
            if code_col and qty_col:
                return {"mode": "placeholder", "sheet": sheet.title, "row": row[0].row, "code_col": code_col, "qty_col": qty_col}

        for row in sheet.iter_rows(max_row=min(sheet.max_row, 40), values_only=False):
            code_col = qty_col = None
            for cell in row:
                key = _norm(cell.value)
                if key in CODE_KEYS:
                    code_col = cell.column
                if key in QTY_KEYS:
                    qty_col = cell.column
            if code_col and qty_col:
                return {"mode": "header", "sheet": sheet.title, "header_row": row[0].row, "row": row[0].row + 1, "code_col": code_col, "qty_col": qty_col}
    finally:
        workbook.close()
    raise ValueError("模板里没有找到商品编码和数量列。")


def _save_template_file(file, *, customer: str, name: str) -> dict:
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in ALLOWED_TEMPLATE_SUFFIXES:
        raise ValueError("模板文件仅支持 .xlsx 或 .xlsm。")
    template_id = _template_id()
    safe_base = safe_upload_name(f"{customer}-{name}{suffix}")
    file_name = f"{template_id}-{safe_base}"
    TEMPLATE_FILE_DIR.mkdir(parents=True, exist_ok=True)
    destination = TEMPLATE_FILE_DIR / file_name
    file.save(destination)
    mapping = _detect_template_mapping(destination)
    preview = _preview_workbook(destination, max_rows=5, max_cols=6)
    return {
        "id": template_id,
        "customer": customer,
        "name": name,
        "status": "ready",
        "file_name": file_name,
        "original_name": file.filename,
        "mapping": mapping,
        "preview_sheet": preview["sheet"],
        "created_at": _now_text(),
        "updated_at": _now_text(),
        "created_by": actor_name(),
        "note": "",
    }


def _append_template(template: dict) -> None:
    templates = _load_manifest()
    templates.append(template)
    _save_manifest(templates)


def _parse_filename_defaults(filename: str) -> tuple[str, str]:
    stem = Path(filename).stem.strip()
    parts = [part.strip() for part in re.split(r"[-_－—]+", stem, maxsplit=1) if part.strip()]
    if len(parts) == 2:
        return parts[0], parts[1]
    return "未分组客户", stem or "未命名模板"


def _shipment_rows_from_xlsx(path: Path) -> tuple[list[dict], dict]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        cached_rows = list(sheet.iter_rows(max_row=40, values_only=True))
        header_index = code_index = qty_index = None
        for row_index, row in enumerate(cached_rows, start=1):
            for col_index, value in enumerate(row):
                key = _norm(value)
                if key in CODE_KEYS:
                    code_index = col_index
                    header_index = row_index
                if key in QTY_KEYS:
                    qty_index = col_index
                    header_index = row_index if header_index is None else header_index
            if code_index is not None and qty_index is not None and header_index is not None:
                break
        if code_index is None or qty_index is None or header_index is None:
            raise ValueError("发货数据里没有找到商品编码和数量列。")
        rows = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=header_index + 1, values_only=True), start=header_index + 1):
            code = str(row[code_index] if code_index < len(row) else "" or "").strip()
            qty = row[qty_index] if qty_index < len(row) else ""
            if not code and not str(qty or "").strip():
                continue
            rows.append({"row": row_number, "code": code, "quantity": qty})
        return rows, {"sheet": sheet.title, "header_row": header_index}
    finally:
        workbook.close()


def _shipment_rows_from_csv(path: Path) -> tuple[list[dict], dict]:
    raw = path.read_bytes()
    text = raw.decode("utf-8-sig", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    all_rows = list(reader)
    code_index = qty_index = header_index = None
    for row_index, row in enumerate(all_rows[:40]):
        for col_index, value in enumerate(row):
            key = _norm(value)
            if key in CODE_KEYS:
                code_index = col_index
                header_index = row_index
            if key in QTY_KEYS:
                qty_index = col_index
                header_index = row_index if header_index is None else header_index
        if code_index is not None and qty_index is not None and header_index is not None:
            break
    if code_index is None or qty_index is None or header_index is None:
        raise ValueError("发货数据里没有找到商品编码和数量列。")
    rows = []
    for row_number, row in enumerate(all_rows[header_index + 1 :], start=header_index + 2):
        code = str(row[code_index] if code_index < len(row) else "").strip()
        qty = row[qty_index] if qty_index < len(row) else ""
        if not code and not str(qty or "").strip():
            continue
        rows.append({"row": row_number, "code": code, "quantity": qty})
    return rows, {"sheet": "CSV", "header_row": header_index + 1}


def _shipment_rows(path: Path) -> tuple[list[dict], dict]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _shipment_rows_from_xlsx(path)
    if suffix == ".csv":
        return _shipment_rows_from_csv(path)
    raise ValueError("发货数据仅支持 .xlsx 或 .csv。")


def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
    if source_row == target_row:
        return
    for col_index in range(1, sheet.max_column + 1):
        source = sheet.cell(source_row, col_index)
        target = sheet.cell(target_row, col_index)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.font:
            target.font = copy(source.font)


def _generate_notice(template: dict, shipment_rows: list[dict]) -> Path:
    template_path = _template_file_path(template)
    if not template_path or not template_path.is_file():
        raise ValueError("当前模板还没有上传 Excel 文件。")
    mapping = template.get("mapping") or _detect_template_mapping(template_path)
    workbook = load_workbook(template_path)
    try:
        sheet = workbook[mapping.get("sheet")] if mapping.get("sheet") in workbook.sheetnames else workbook.active
        start_row = int(mapping["row"])
        code_col = int(mapping["code_col"])
        qty_col = int(mapping["qty_col"])
        for offset, item in enumerate(shipment_rows):
            target_row = start_row + offset
            _copy_row_style(sheet, start_row, target_row)
            sheet.cell(target_row, code_col).value = item["code"]
            sheet.cell(target_row, qty_col).value = item["quantity"]
        output_dir = user_output_dir() / "发货通知"
        safe_customer = safe_upload_name(template.get("customer") or "客户")
        output_path = unique_prefixed_path(output_dir, f"shipping-notice-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{safe_customer}.xlsx")
        workbook.save(output_path)
    finally:
        workbook.close()
    return output_path


def _latest_outputs() -> list[dict]:
    rows = []
    for path in user_recent_outputs("发货通知/*.xlsx", limit=20):
        rows.append({"name": path.name, "updated_at": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"), "download_name": download_name(path)})
    return rows


def _page_context(*, selected_template_id: str = "", template_preview=None, generate_preview=None) -> dict:
    customers, templates = _template_choices()
    selected_template = _find_template(selected_template_id) if selected_template_id else None
    if selected_template and not template_preview:
        path = _template_file_path(selected_template)
        if path and path.is_file():
            template_preview = _preview_workbook(path)
    return {
        "customers": customers,
        "templates": templates,
        "selected_template_id": selected_template_id,
        "selected_template": selected_template,
        "template_preview": template_preview,
        "generate_preview": generate_preview,
        "latest_outputs": _latest_outputs(),
    }


def register(app) -> None:
    @app.get("/shipping-notices")
    @permission_required("generate_shipping_notice")
    def shipping_notice():
        return render_template("shipping_notice.html", **_page_context(selected_template_id=request.args.get("template_id", "")))

    @app.post("/shipping-notices/templates/upload")
    @permission_required("generate_shipping_notice")
    def upload_shipping_notice_template():
        customer = request.form.get("customer", "").strip()
        name = request.form.get("template_name", "").strip()
        file = request.files.get("template")
        if not customer or not name or not file or not file.filename:
            flash("请填写客户、模板名称，并选择单个模板文件。", "error")
            return redirect(url_for("shipping_notice"))
        try:
            template = _save_template_file(file, customer=customer, name=name)
            _append_template(template)
            with connect(DB_PATH) as conn:
                log_event(conn, "上传发货通知模板", "shipping_notice", template["file_name"], f"{customer} / {name}", actor=actor_name())
                conn.commit()
        except Exception as exc:
            flash(f"模板上传失败：{exc}", "error")
            return redirect(url_for("shipping_notice"))
        flash("模板已上传。", "success")
        return redirect(url_for("shipping_notice", template_id=template["id"]))

    @app.post("/shipping-notices/templates/batch")
    @permission_required("generate_shipping_notice")
    def batch_upload_shipping_notice_templates():
        file = request.files.get("template_zip")
        if not file or not file.filename:
            flash("请选择模板压缩包。", "error")
            return redirect(url_for("shipping_notice"))
        if Path(file.filename).suffix.lower() != ".zip":
            flash("批量模板请上传 .zip 文件。", "error")
            return redirect(url_for("shipping_notice"))
        upload_path = user_upload_path(file.filename, prefix="shipping-template-batch")
        file.save(upload_path)
        imported = 0
        errors = []
        try:
            with zipfile.ZipFile(upload_path) as archive:
                for member in archive.infolist():
                    if member.is_dir() or Path(member.filename).suffix.lower() not in ALLOWED_TEMPLATE_SUFFIXES:
                        continue
                    customer, name = _parse_filename_defaults(Path(member.filename).name)
                    data = archive.read(member)
                    storage = type("UploadedTemplate", (), {"filename": Path(member.filename).name, "save": lambda self, dest, payload=data: Path(dest).write_bytes(payload)})()
                    try:
                        _append_template(_save_template_file(storage, customer=customer, name=name))
                        imported += 1
                    except Exception as exc:
                        errors.append(f"{Path(member.filename).name}: {exc}")
        except zipfile.BadZipFile:
            flash("模板压缩包无法读取。", "error")
            return redirect(url_for("shipping_notice"))
        if errors:
            flash(f"已导入 {imported} 个模板，{len(errors)} 个失败。首个失败：{errors[0]}", "error")
        else:
            flash(f"已导入 {imported} 个模板。", "success")
        return redirect(url_for("shipping_notice"))

    @app.post("/shipping-notices/preview")
    @permission_required("generate_shipping_notice")
    def preview_shipping_notice():
        template_id = request.form.get("template_id", "").strip()
        template = _find_template(template_id)
        file = request.files.get("shipment_data")
        if not template:
            flash("请选择客户模板。", "error")
            return redirect(url_for("shipping_notice"))
        if template.get("status") != "ready":
            flash("这个模板还只是需求记录，需要先上传模板 Excel。", "error")
            return redirect(url_for("shipping_notice", template_id=template_id))
        if not file or not file.filename:
            flash("请选择发货数据文件。", "error")
            return redirect(url_for("shipping_notice", template_id=template_id))
        suffix = Path(file.filename).suffix.lower()
        if suffix not in ALLOWED_SHIPMENT_SUFFIXES:
            flash("发货数据仅支持 .xlsx 或 .csv。", "error")
            return redirect(url_for("shipping_notice", template_id=template_id))
        upload_path = user_upload_path(file.filename, prefix="shipping-data")
        file.save(upload_path)
        try:
            rows, source = _shipment_rows(upload_path)
        except Exception as exc:
            flash(f"发货数据读取失败：{exc}", "error")
            return redirect(url_for("shipping_notice", template_id=template_id))
        generate_preview = {
            "template": template,
            "upload_path": str(upload_path),
            "source": source,
            "row_count": len(rows),
            "rows": rows[:20],
        }
        return render_template("shipping_notice_preview.html", **_page_context(selected_template_id=template_id, generate_preview=generate_preview))

    @app.post("/shipping-notices/generate")
    @permission_required("generate_shipping_notice")
    def generate_shipping_notice():
        template_id = request.form.get("template_id", "").strip()
        template = _find_template(template_id)
        upload_path = Path(request.form.get("upload_path", "")).expanduser().resolve()
        upload_root = user_upload_dir(create=False).resolve()
        if not template or template.get("status") != "ready":
            flash("请选择可用模板。", "error")
            return redirect(url_for("shipping_notice"))
        if not upload_path.is_file() or upload_root not in upload_path.parents:
            flash("发货数据路径无效，请重新上传预览。", "error")
            return redirect(url_for("shipping_notice", template_id=template_id))
        try:
            rows, _source = _shipment_rows(upload_path)
            output_path = _generate_notice(template, rows)
            with connect(DB_PATH) as conn:
                log_event(conn, "生成发货通知", "shipping_notice", output_path.name, f"{template.get('customer')} / {template.get('name')} / {len(rows)} 行", actor=actor_name())
                conn.commit()
        except Exception as exc:
            flash(f"发货通知生成失败：{exc}", "error")
            return redirect(url_for("shipping_notice", template_id=template_id))
        flash("发货通知 Excel 已生成。", "success")
        return redirect(url_for("shipping_notice", template_id=template_id, generated=download_name(output_path)))

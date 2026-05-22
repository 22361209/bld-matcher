from __future__ import annotations

import re
from math import ceil
from pathlib import Path

from flask import flash, redirect, render_template, request, url_for
from openpyxl import load_workbook

from app.config import DB_PATH
from app.database import (
    add_customer_price_record,
    connect,
    count_customer_price_customers,
    count_customer_price_records,
    customer_price_stats,
    delete_customer_price_record,
    list_customer_price_customer_summaries,
    list_customer_price_model_comparisons,
    list_customer_price_records,
    log_event,
)
from app.helpers import clean_original_filename, user_upload_path
from app.security import actor_name, permission_required, safe_referrer


PRICE_RECORD_PAGE_SIZE = 100
TYPE_LABELS = {"quote": "报价", "order": "成交"}

HEADER_ALIASES = {
    "customer_name": ("客户", "客户名称", "客户名", "customer", "customername"),
    "record_date": ("日期", "报价日期", "订单日期", "date", "recorddate", "quotedate", "orderdate"),
    "document_no": ("单据号", "报价单号", "订单号", "客户订单号", "documentno", "quoteno", "orderno"),
    "source_name": ("来源", "来源名称", "source", "sourcename"),
    "source_code": ("客户号码", "客户编码", "编码", "号码", "sourcecode", "customercode", "customerpartno"),
    "oe_no": ("oe", "oe号", "oeno", "oenumber"),
    "bld_no": ("bld", "bldno", "bld号", "bld no", "bld no."),
    "item": ("产品名称", "物料名称", "名称", "item", "product", "productname"),
    "models": ("车型", "适用车型", "models", "model", "application"),
    "price_cny": ("含税单价", "人民币", "人民币单价", "rmb", "cny", "pricecny", "unitprice"),
    "price_usd": ("美金价", "美元价", "usd", "priceusd"),
    "exchange_rate": ("汇率", "exchangerate", "rate"),
    "note": ("备注", "说明", "note", "remark"),
}


def _normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    return re.sub(r"[\s._\-:/\\]+", "", text).lower()


HEADER_LOOKUP = {
    _normalize_header(alias): field
    for field, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def _request_page() -> int:
    try:
        return max(1, int(request.args.get("page", "1") or 1))
    except ValueError:
        return 1


def _filters() -> dict[str, str]:
    record_type = request.args.get("record_type", "")
    if record_type not in {"", "quote", "order"}:
        record_type = ""
    return {
        "customer_q": request.args.get("customer_q", "").strip(),
        "customer": request.args.get("customer", "").strip(),
        "bld_no": request.args.get("bld_no", "").strip(),
        "source_code": request.args.get("source_code", "").strip(),
        "record_type": record_type,
    }


def _page_url(filters: dict[str, str], page: int) -> str:
    params = {key: value for key, value in filters.items() if value}
    if page > 1:
        params["page"] = str(page)
    return f"{url_for('customer_prices', **params)}#customer-price-results"


def _pagination(filters: dict[str, str], page: int, total: int) -> dict[str, object]:
    total_pages = max(1, ceil(total / PRICE_RECORD_PAGE_SIZE))
    page = min(max(1, page), total_pages)
    window = {1, total_pages, page - 1, page, page + 1}
    pages = sorted(item for item in window if 1 <= item <= total_pages)
    links = []
    previous_page = 0
    for item in pages:
        if previous_page and item - previous_page > 1:
            links.append({"gap": True})
        links.append({"page": item, "url": _page_url(filters, item), "current": item == page})
        previous_page = item
    return {
        "page": page,
        "total_pages": total_pages,
        "has_prev": page > 1,
        "has_next": page < total_pages,
        "prev_url": _page_url(filters, page - 1) if page > 1 else "",
        "next_url": _page_url(filters, page + 1) if page < total_pages else "",
        "links": links,
    }


def _header_map(values: tuple[object, ...]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for index, value in enumerate(values):
        field = HEADER_LOOKUP.get(_normalize_header(value))
        if field and field not in mapped:
            mapped[field] = index
    return mapped


def _cell_value(row: tuple[object, ...], index: int | None) -> object:
    if index is None or index >= len(row):
        return ""
    return row[index]


def _import_price_records(path: Path, *, record_type: str, default_customer: str, source_file: str) -> tuple[int, int, list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    imported = 0
    skipped = 0
    errors: list[str] = []
    try:
        sheet = workbook[workbook.sheetnames[0]]
        rows = list(sheet.iter_rows(values_only=True))
        header_index = -1
        mapping: dict[str, int] = {}
        for index, row in enumerate(rows[:10]):
            mapping = _header_map(tuple(row))
            if len(mapping) >= 3 and ("price_cny" in mapping or "price_usd" in mapping):
                header_index = index
                break
        if header_index < 0:
            raise ValueError("没有识别到表头。")

        with connect(DB_PATH) as conn:
            for excel_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
                if not any(value not in (None, "") for value in row):
                    continue
                data = {
                    "record_type": record_type,
                    "customer_name": _cell_value(row, mapping.get("customer_name")) or default_customer,
                    "record_date": _cell_value(row, mapping.get("record_date")),
                    "document_no": _cell_value(row, mapping.get("document_no")),
                    "source_name": _cell_value(row, mapping.get("source_name")),
                    "source_code": _cell_value(row, mapping.get("source_code")),
                    "oe_no": _cell_value(row, mapping.get("oe_no")),
                    "bld_no": _cell_value(row, mapping.get("bld_no")),
                    "item": _cell_value(row, mapping.get("item")),
                    "models": _cell_value(row, mapping.get("models")),
                    "price_cny": _cell_value(row, mapping.get("price_cny")),
                    "price_usd": _cell_value(row, mapping.get("price_usd")),
                    "exchange_rate": _cell_value(row, mapping.get("exchange_rate")),
                    "note": _cell_value(row, mapping.get("note")),
                    "source_file": source_file,
                }
                try:
                    add_customer_price_record(conn, data, actor=actor_name(), audit=False, commit=False)
                    imported += 1
                except Exception as exc:
                    skipped += 1
                    if len(errors) < 5:
                        errors.append(f"第 {excel_index} 行：{exc}")
            log_event(
                conn,
                "导入价格维护记录",
                "customer_price",
                source_file,
                f"{TYPE_LABELS[record_type]}记录 {imported} 条，跳过 {skipped} 条",
                actor=actor_name(),
            )
            conn.commit()
    finally:
        workbook.close()
    return imported, skipped, errors


def register(app) -> None:
    @app.get("/customer-prices")
    @permission_required("view_customer_prices")
    def customer_prices():
        filters = _filters()
        detail_filters = {
            "customer": filters["customer"],
            "bld_no": filters["bld_no"],
            "source_code": filters["source_code"],
            "record_type": filters["record_type"],
        }
        show_detail = bool(filters["customer"] or filters["bld_no"] or filters["source_code"])
        with connect(DB_PATH) as conn:
            total = (
                count_customer_price_records(conn, **detail_filters)
                if show_detail
                else count_customer_price_customers(
                    conn,
                    customer_query=filters["customer_q"],
                    record_type=filters["record_type"],
                )
            )
            pagination = _pagination(filters, _request_page(), total)
            if show_detail:
                records = list_customer_price_records(
                    conn,
                    **detail_filters,
                    limit=PRICE_RECORD_PAGE_SIZE,
                    offset=(int(pagination["page"]) - 1) * PRICE_RECORD_PAGE_SIZE,
                )
                customer_summaries = []
                model_comparisons = (
                    list_customer_price_model_comparisons(conn, **detail_filters)
                    if filters["bld_no"] or filters["source_code"]
                    else []
                )
            else:
                records = []
                model_comparisons = []
                customer_summaries = list_customer_price_customer_summaries(
                    conn,
                    customer_query=filters["customer_q"],
                    record_type=filters["record_type"],
                    limit=PRICE_RECORD_PAGE_SIZE,
                    offset=(int(pagination["page"]) - 1) * PRICE_RECORD_PAGE_SIZE,
                )
            stats = customer_price_stats(conn)
        return render_template(
            "customer_prices.html",
            records=records,
            customer_summaries=customer_summaries,
            model_comparisons=model_comparisons,
            show_detail=show_detail,
            filters=filters,
            total_records=total,
            stats=stats,
            pagination=pagination,
            page_size=PRICE_RECORD_PAGE_SIZE,
            type_labels=TYPE_LABELS,
        )

    @app.post("/customer-prices/save")
    @permission_required("manage_customer_prices")
    def save_customer_price():
        try:
            with connect(DB_PATH) as conn:
                add_customer_price_record(conn, request.form, actor=actor_name())
        except Exception as exc:
            flash(f"保存失败：{exc}", "error")
            return redirect(url_for("customer_prices"))
        flash("价格记录已保存。", "success")
        return redirect(url_for("customer_prices", customer=request.form.get("customer_name", "")) + "#customer-price-results")

    @app.post("/customer-prices/import")
    @permission_required("manage_customer_prices")
    def import_customer_prices():
        record_type = request.form.get("record_type", "quote")
        if record_type not in TYPE_LABELS:
            flash("记录类型不正确。", "error")
            return redirect(url_for("customer_prices"))
        file = request.files.get("price_file")
        if not file or not file.filename:
            flash("请选择价格记录 Excel 文件。", "error")
            return redirect(url_for("customer_prices"))
        if Path(file.filename).suffix.lower() != ".xlsx":
            flash("价格记录请使用 .xlsx 文件。", "error")
            return redirect(url_for("customer_prices"))

        original_name = clean_original_filename(file.filename, fallback_suffix=".xlsx")
        upload_path = user_upload_path(file.filename, prefix="customer-price")
        file.save(upload_path)
        try:
            imported, skipped, errors = _import_price_records(
                upload_path,
                record_type=record_type,
                default_customer=request.form.get("default_customer", ""),
                source_file=original_name,
            )
        except Exception as exc:
            flash(f"导入失败：{exc}", "error")
            return redirect(url_for("customer_prices"))
        if errors:
            flash(f"已导入 {imported} 条，跳过 {skipped} 条。{'；'.join(errors)}", "warning")
        else:
            flash(f"已导入 {imported} 条{TYPE_LABELS[record_type]}记录。", "success")
        return redirect(url_for("customer_prices") + "#customer-price-results")

    @app.post("/customer-prices/<int:record_id>/delete")
    @permission_required("manage_customer_prices")
    def delete_customer_price(record_id: int):
        with connect(DB_PATH) as conn:
            row = delete_customer_price_record(conn, record_id, actor=actor_name())
        if not row:
            flash("价格记录不存在。", "error")
        else:
            flash("价格记录已删除。", "success")
        return redirect(safe_referrer(url_for("customer_prices") + "#customer-price-results"))

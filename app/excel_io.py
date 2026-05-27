from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import re
from string import ascii_uppercase
import zipfile

from openpyxl import load_workbook
import xlrd
from xlutils.copy import copy as copy_xls

from .matcher import ProductCatalog, normalize_code, split_codes


INQUIRY_HEADERS = {
    "name": {"物料名称", "产品名称", "名称", "ITEM", "PART"},
    "oe": {"OE号", "OE", "OE NO", "OE NO.", "OE号码", "OE REFERENCE", "OE REF", "号码", "查询号码", "客户号码", "BLD号", "BLD NO", "BLD NO."},
    "description": {"物料描述", "描述", "DESCRIPTION", "DESC"},
}


def _norm_header(value: object) -> str:
    return str(value or "").strip().upper().replace(" ", "")


PRICE_EXPORT_MODES = {"none", "tax", "net", "usd"}
ROW_DIMENSION_CLEANUP_SLACK = 100
POLLUTED_WORKSHEET_ROW_GAP = 1000
RISKY_XLSX_CLEANUP_PREFIXES = (
    "xl/charts/",
    "xl/comments",
    "xl/drawings/",
    "xl/tables/",
    "xl/threadedComments/",
    "xl/vmlDrawings/",
)
_DIMENSION_RE = re.compile(rb'(<dimension\b[^>]*\bref=")([^"]+)(")')
_DIMENSION_ROW_RE = re.compile(r"\$?[A-Z]{1,4}\$?(\d+)", re.IGNORECASE)
_ROW_ATTR_RE = re.compile(rb'\br="(\d+)"')
_ROW_ELEMENT_RE = re.compile(
    rb"<row\b(?P<attrs_self>[^>]*)/>|<row\b(?P<attrs>[^>]*)>(?P<body>.*?)</row>",
    re.DOTALL,
)
_ROW_PAYLOAD_MARKERS = (b"<v", b"<f", b"<is", b"<t")
MANUAL_HEADER_ALIASES = {
    _norm_header(alias)
    for aliases in INQUIRY_HEADERS.values()
    for alias in aliases
} | {
    _norm_header(alias)
    for alias in {
        "SN",
        "NO",
        "NO.",
        "序号",
        "数量",
        "客户号码",
        "客户编码",
        "号码",
        "料号",
        "ITEM NO",
        "ITEM NO.",
        "PART NO",
        "PART NO.",
        "QTY",
        "QUANTITY",
        "رقم",
        "كمية",
    }
}


@dataclass(frozen=True)
class WorkbookCleanupResult:
    path: Path
    cleaned: bool = False
    original_path: Path | None = None
    message: str = ""
    details: dict[str, int] = field(default_factory=dict)


def _looks_like_match_code(value: object) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    upper = text.upper()
    if re.search(r"[A-Z]{3,}\s+[A-Z]{3,}", upper):
        return False
    parts = split_codes(text)
    return any(len(normalize_code(part)) >= 5 and any(char.isdigit() for char in normalize_code(part)) for part in parts)


def _looks_like_manual_header(value: object) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    return _norm_header(text) in MANUAL_HEADER_ALIASES or not _looks_like_match_code(text)


def _find_inquiry_columns(sheet) -> tuple[int, dict[str, int]]:
    aliases = {
        _norm_header(alias): key
        for key, names in INQUIRY_HEADERS.items()
        for alias in names
    }
    for row_index in range(min(sheet.nrows, 20)):
        columns: dict[str, int] = {}
        for col_index in range(sheet.ncols):
            key = aliases.get(_norm_header(sheet.cell_value(row_index, col_index)))
            if key:
                columns[key] = col_index
        if "oe" in columns:
            return row_index, columns
    raise ValueError("询价表没有找到可识别表头，需要包含 OE号。")


def _column_label(index: int) -> str:
    label = ""
    index += 1
    while index:
        index, remainder = divmod(index - 1, 26)
        label = ascii_uppercase[remainder] + label
    return label


def preview_inquiry_columns(inquiry_path: Path, max_rows: int = 8, max_cols: int = 12) -> dict:
    if inquiry_path.suffix.lower() == ".xls":
        book = xlrd.open_workbook(inquiry_path, formatting_info=True, ignore_workbook_corruption=True)
        sheet = book.sheet_by_index(0)
        cols = min(sheet.ncols, max_cols)
        rows = []
        for row_index in range(min(sheet.nrows, max_rows)):
            rows.append([sheet.cell_value(row_index, col_index) for col_index in range(cols)])
        return {
            "sheet": sheet.name,
            "columns": [{"index": i, "label": _column_label(i)} for i in range(cols)],
            "rows": rows,
        }

    workbook = load_workbook(inquiry_path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = []
        for values in sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
            rows.append(list(values[:max_cols]))
        cols = min(max((len(row) for row in rows), default=0), max_cols)
        rows = [row + [None] * (cols - len(row)) for row in rows]
        return {
            "sheet": sheet.title,
            "columns": [{"index": i, "label": _column_label(i)} for i in range(cols)],
            "rows": rows,
        }
    finally:
        workbook.close()


def _norm_openpyxl_cell(value: object) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def _find_xlsx_inquiry_columns(sheet) -> tuple[int, dict[str, int]]:
    aliases = {
        _norm_header(alias): key
        for key, names in INQUIRY_HEADERS.items()
        for alias in names
    }
    for row_index in range(1, min(sheet.max_row, 20) + 1):
        columns: dict[str, int] = {}
        for col_index in range(1, sheet.max_column + 1):
            key = aliases.get(_norm_openpyxl_cell(sheet.cell(row_index, col_index).value))
            if key:
                columns[key] = col_index
        if "oe" in columns:
            return row_index, columns
    raise ValueError("询价表没有找到可识别表头，需要包含 OE号。")


def _manual_match_columns(match_column: object) -> list[int]:
    if match_column is None:
        return []
    if isinstance(match_column, int):
        values = [match_column]
    elif isinstance(match_column, str):
        values = [match_column]
    else:
        try:
            values = list(match_column)
        except TypeError:
            values = [match_column]

    columns: list[int] = []
    seen = set()
    for value in values:
        try:
            index = int(value)
        except (TypeError, ValueError):
            continue
        if index < 0 or index in seen:
            continue
        seen.add(index)
        columns.append(index)
    return columns


def _combined_match_text(match_values: list[tuple[str, object]]) -> str:
    return "\n".join(str(value).strip() for _, value in match_values if str(value or "").strip())


def _selected_match_row_is_header(match_values: list[tuple[str, object]]) -> bool:
    values = [value for _, value in match_values if str(value or "").strip()]
    if not values:
        return False
    return all(_norm_header(value) in MANUAL_HEADER_ALIASES for value in values)


def _annotate_row_summary_with_match_columns(row_summary: dict, match_values: list[tuple[str, object]], match) -> dict:
    if not match or not match_values or not match.matched_codes:
        return row_summary

    matched_keys = {normalize_code(code) for code in match.matched_codes if normalize_code(code)}
    if not matched_keys:
        return row_summary

    hits = []
    for label, value in match_values:
        for part in split_codes(value) or ([value] if normalize_code(value) else []):
            if normalize_code(part) in matched_keys:
                hits.append(f"{label}列：{part}")

    if not hits:
        return row_summary

    prefix = "命中列：" + "，".join(dict.fromkeys(hits))
    row_summary["match_note"] = f"{prefix}；{row_summary['match_note']}" if row_summary.get("match_note") else prefix
    return row_summary


def _xls_match_values(sheet, row_index: int, match_columns: list[int]) -> list[tuple[str, object]]:
    values = []
    for column_index in match_columns:
        if 0 <= column_index < sheet.ncols:
            values.append((_column_label(column_index), sheet.cell_value(row_index, column_index)))
    return values


def _xlsx_match_values(sheet, row_index: int, match_columns: list[int]) -> list[tuple[str, object]]:
    values = []
    for column_index in match_columns:
        if 0 < column_index <= sheet.max_column:
            values.append((_column_label(column_index - 1), sheet.cell(row_index, column_index).value))
    return values


def _row_match_values(values: tuple, match_columns: list[int]) -> list[tuple[str, object]]:
    row_values = []
    for column_index in match_columns:
        row_values.append((_column_label(column_index - 1), _cell_from_values(values, column_index)))
    return row_values


def _find_xls_selected_header_row(sheet, match_columns: int | list[int]) -> int:
    selected_columns = _manual_match_columns(match_columns)
    max_scan = min(sheet.nrows, 20)
    for row_index in range(max_scan):
        if not any(
            column_index < sheet.ncols and _looks_like_manual_header(sheet.cell_value(row_index, column_index))
            for column_index in selected_columns
        ):
            continue
        for next_row in range(row_index + 1, min(sheet.nrows, row_index + 5)):
            if any(
                column_index < sheet.ncols and _looks_like_match_code(sheet.cell_value(next_row, column_index))
                for column_index in selected_columns
            ):
                return row_index
    return 0


def _find_xlsx_selected_header_row(sheet, match_columns: int | list[int]) -> int:
    selected_columns = _manual_match_columns(match_columns)
    max_scan = min(sheet.max_row, 20)
    for row_index in range(1, max_scan + 1):
        if not any(
            column_index <= sheet.max_column and _looks_like_manual_header(sheet.cell(row_index, column_index).value)
            for column_index in selected_columns
        ):
            continue
        for next_row in range(row_index + 1, min(sheet.max_row, row_index + 4) + 1):
            if any(
                column_index <= sheet.max_column and _looks_like_match_code(sheet.cell(next_row, column_index).value)
                for column_index in selected_columns
            ):
                return row_index
    return 1


def _cell_from_values(row: tuple, column_index: int) -> object:
    return row[column_index - 1] if 0 < column_index <= len(row) else ""


def _trim_unused_row_dimensions(sheet, keep_until: int) -> None:
    if len(sheet.row_dimensions) <= keep_until + ROW_DIMENSION_CLEANUP_SLACK:
        return
    for row_index in [index for index in sheet.row_dimensions if index > keep_until]:
        del sheet.row_dimensions[row_index]


def _xlsx_row_index(attrs: bytes | None) -> int | None:
    if not attrs:
        return None
    match = _ROW_ATTR_RE.search(attrs)
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _xlsx_row_has_payload(body: bytes | None) -> bool:
    if not body:
        return False
    return any(marker in body for marker in _ROW_PAYLOAD_MARKERS)


def _dimension_max_row(sheet_xml: bytes) -> int:
    match = _DIMENSION_RE.search(sheet_xml)
    if not match:
        return 0
    ref = match.group(2).decode("utf-8", errors="ignore")
    rows = [int(value) for value in _DIMENSION_ROW_RE.findall(ref)]
    return max(rows, default=0)


def _dimension_ref_with_max_row(ref: str, max_row: int) -> str:
    max_row = max(max_row, 1)
    parts = ref.split(":", 1)
    if len(parts) == 1:
        start = parts[0].strip() or "A1"
        col_match = re.match(r"(\$?[A-Z]{1,4})", start, re.IGNORECASE)
        col = col_match.group(1) if col_match else "A"
        return f"{col}1" if max_row == 1 else f"{col}1:{col}{max_row}"

    start, end = parts[0].strip() or "A1", parts[1].strip()
    col_match = re.match(r"(\$?[A-Z]{1,4})", end, re.IGNORECASE)
    end_col = col_match.group(1) if col_match else "A"
    return f"{start}:{end_col}{max_row}"


def _update_dimension_max_row(sheet_xml: bytes, max_row: int) -> tuple[bytes, bool]:
    def replace(match: re.Match[bytes]) -> bytes:
        original_ref = match.group(2).decode("utf-8", errors="ignore")
        updated_ref = _dimension_ref_with_max_row(original_ref, max_row)
        if updated_ref == original_ref:
            return match.group(0)
        return match.group(1) + updated_ref.encode("utf-8") + match.group(3)

    updated, count = _DIMENSION_RE.subn(replace, sheet_xml, count=1)
    return updated, bool(count and updated != sheet_xml)


def _cleanup_worksheet_xml(sheet_xml: bytes) -> tuple[bytes, dict[str, int]]:
    meaningful_max_row = 0
    max_row_tag = 0
    row_tag_count = 0

    for match in _ROW_ELEMENT_RE.finditer(sheet_xml):
        attrs = match.group("attrs") or match.group("attrs_self")
        row_index = _xlsx_row_index(attrs)
        if row_index is None:
            continue
        row_tag_count += 1
        max_row_tag = max(max_row_tag, row_index)
        if _xlsx_row_has_payload(match.group("body")):
            meaningful_max_row = max(meaningful_max_row, row_index)

    keep_until = max(meaningful_max_row, 1)
    declared_max_row = max(_dimension_max_row(sheet_xml), max_row_tag)
    if declared_max_row - keep_until < POLLUTED_WORKSHEET_ROW_GAP:
        return sheet_xml, {}

    removed_row_tags = 0

    def remove_empty_tail_rows(match: re.Match[bytes]) -> bytes:
        nonlocal removed_row_tags
        attrs = match.group("attrs") or match.group("attrs_self")
        row_index = _xlsx_row_index(attrs)
        if row_index and row_index > keep_until and not _xlsx_row_has_payload(match.group("body")):
            removed_row_tags += 1
            return b""
        return match.group(0)

    cleaned_xml = _ROW_ELEMENT_RE.sub(remove_empty_tail_rows, sheet_xml)
    cleaned_xml, dimension_updated = _update_dimension_max_row(cleaned_xml, keep_until)
    if not removed_row_tags and not dimension_updated:
        return sheet_xml, {}

    return cleaned_xml, {
        "effective_rows": keep_until,
        "declared_rows": declared_max_row,
        "removed_tail_rows": max(0, declared_max_row - keep_until),
        "removed_row_tags": removed_row_tags,
        "row_tag_count": row_tag_count,
    }


def sanitize_inquiry_workbook_if_needed(inquiry_path: Path, output_path: Path | None = None) -> WorkbookCleanupResult:
    inquiry_path = Path(inquiry_path)
    if inquiry_path.suffix.lower() != ".xlsx":
        return WorkbookCleanupResult(path=inquiry_path)

    try:
        with zipfile.ZipFile(inquiry_path, "r") as source:
            entries = source.infolist()
            names = [entry.filename for entry in entries]
            if any(name.startswith(RISKY_XLSX_CLEANUP_PREFIXES) for name in names):
                return WorkbookCleanupResult(path=inquiry_path)

            cleaned_entries: list[tuple[zipfile.ZipInfo, bytes]] = []
            changed_sheets = 0
            max_effective_rows = 0
            removed_tail_rows = 0
            removed_row_tags = 0
            for entry in entries:
                data = source.read(entry.filename)
                if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", entry.filename):
                    cleaned_data, details = _cleanup_worksheet_xml(data)
                    if details:
                        changed_sheets += 1
                        max_effective_rows = max(max_effective_rows, details["effective_rows"])
                        removed_tail_rows += details["removed_tail_rows"]
                        removed_row_tags += details["removed_row_tags"]
                    data = cleaned_data
                cleaned_entries.append((entry, data))
    except zipfile.BadZipFile:
        return WorkbookCleanupResult(path=inquiry_path)

    if not changed_sheets:
        return WorkbookCleanupResult(path=inquiry_path)

    cleaned_path = output_path or inquiry_path.with_name(f"{inquiry_path.stem}-cleaned{inquiry_path.suffix}")
    cleaned_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = cleaned_path.with_name(f".{cleaned_path.name}.tmp")
    with zipfile.ZipFile(temporary_path, "w", compression=zipfile.ZIP_DEFLATED) as target:
        for entry, data in cleaned_entries:
            target.writestr(entry, data)
    temporary_path.replace(cleaned_path)

    details = {
        "cleaned_sheets": changed_sheets,
        "effective_rows": max_effective_rows,
        "removed_tail_rows": removed_tail_rows,
        "removed_row_tags": removed_row_tags,
    }
    message = f"已自动清理 Excel 尾部空白格式：实际有效行 {max_effective_rows}，清理尾部空白格式约 {removed_tail_rows} 行。"
    return WorkbookCleanupResult(
        path=cleaned_path,
        cleaned=True,
        original_path=inquiry_path,
        message=message,
        details=details,
    )


def _find_xlsx_inquiry_columns_from_rows(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    aliases = {
        _norm_header(alias): key
        for key, names in INQUIRY_HEADERS.items()
        for alias in names
    }
    for row_index, row in enumerate(rows[:20], start=1):
        columns: dict[str, int] = {}
        for col_index, value in enumerate(row, start=1):
            key = aliases.get(_norm_openpyxl_cell(value))
            if key:
                columns[key] = col_index
        if "oe" in columns:
            return row_index, columns
    raise ValueError("询价表没有找到可识别表头，需要包含 OE号。")


def _find_xlsx_selected_header_row_from_rows(rows: list[tuple], match_columns: int | list[int]) -> int:
    selected_columns = _manual_match_columns(match_columns)
    max_scan = min(len(rows), 20)
    for row_index in range(1, max_scan + 1):
        if not any(_looks_like_manual_header(_cell_from_values(rows[row_index - 1], column_index)) for column_index in selected_columns):
            continue
        for next_row in range(row_index + 1, min(max_scan, row_index + 4) + 1):
            if any(_looks_like_match_code(_cell_from_values(rows[next_row - 1], column_index)) for column_index in selected_columns):
                return row_index
    return 1


def _summary_row(row_number: int, inquiry_oe: object, inquiry_name: object, match) -> dict:
    parts = split_codes(inquiry_oe)
    match_note = ""
    price_cny = None
    if match and " / " not in (match.bld_no or ""):
        price_cny = _numeric_price(match.row.get("price_cny"))

    if len(parts) > 1:
        if match and match.matched_codes:
            notes = [f"命中号码：{', '.join(match.matched_codes)}"]
            if match.unmatched_codes:
                notes.append(f"未命中号码：{', '.join(match.unmatched_codes)}")
            if " / " in match.bld_no:
                notes.append(f"命中 BLD：{match.bld_no}")
            match_note = "；".join(notes)
        elif not match:
            match_note = f"多个号码均未命中：{', '.join(parts)}"

    row = {
        "row": row_number,
        "oe": inquiry_oe,
        "name": inquiry_name,
        "bld_no": match.bld_no if match else "",
        "price_cny": price_cny,
        "reason": match.reason if match else "未找到",
        "score": match.score if match else 0,
        "match_note": match_note,
        "matched_oe_codes": [],
        "unmatched_oe_codes": [],
    }
    if match and len(parts) > 1 and match.matched_codes:
        row["matched_oe_codes"] = list(match.matched_codes)
        row["unmatched_oe_codes"] = list(match.unmatched_codes)
    return row


def _price_export_header(price_mode: str) -> str:
    if price_mode == "tax":
        return "含税单价"
    if price_mode == "net":
        return "不含税单价"
    if price_mode == "usd":
        return "美金价"
    return ""


def _decimal_price(value: object) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return None


def _numeric_price(value: object) -> float | None:
    price = _decimal_price(value)
    if price is None:
        return None
    return float(price)


def _match_export_price(match, price_mode: str, exchange_rate: float | None) -> float | None:
    if price_mode not in {"tax", "net", "usd"} or not match:
        return None
    if " / " in (match.bld_no or ""):
        return None

    raw_price = match.row.get("price_cny")
    decimal_price = _decimal_price(raw_price)
    if decimal_price is None:
        return None
    if price_mode == "tax":
        price = float(decimal_price)
        return round(price, 2)
    if price_mode == "net":
        net_price = (decimal_price / Decimal("1.1")).quantize(Decimal("1"), rounding=ROUND_HALF_UP)
        return int(net_price)
    if not exchange_rate or exchange_rate <= 0:
        return None
    price = float(decimal_price)
    return round(price / 1.1 / exchange_rate, 2)


def generate_xls_with_bld(
    inquiry_path: Path,
    output_path: Path,
    catalog: ProductCatalog,
    match_column: object = None,
    write_output: bool = True,
    price_mode: str = "none",
    exchange_rate: float | None = None,
) -> dict:
    book = xlrd.open_workbook(inquiry_path, formatting_info=True, ignore_workbook_corruption=True)
    writable = copy_xls(book) if write_output else None
    summary = {"total": 0, "matched": 0, "unmatched": 0, "rows": []}
    manual_match_columns = _manual_match_columns(match_column)

    for sheet_index in range(book.nsheets):
        source_sheet = book.sheet_by_index(sheet_index)
        target_sheet = writable.get_sheet(sheet_index) if writable else None
        if source_sheet.nrows == 0:
            continue

        if not manual_match_columns:
            header_row, columns = _find_inquiry_columns(source_sheet)
            selected_match_columns = [columns["oe"]]
        else:
            header_row = _find_xls_selected_header_row(source_sheet, manual_match_columns)
            columns = {}
            selected_match_columns = manual_match_columns
        output_col = source_sheet.ncols
        price_header = _price_export_header(price_mode)
        note_col = output_col + 2 if price_header else output_col + 1
        if target_sheet:
            target_sheet.write(header_row, output_col, "BLD NO.")
            if price_header:
                target_sheet.write(header_row, output_col + 1, price_header)
            target_sheet.write(header_row, note_col, "匹配说明")

        for row_index in range(header_row + 1, source_sheet.nrows):
            inquiry_name = source_sheet.cell_value(row_index, columns["name"]) if "name" in columns else ""
            match_values = _xls_match_values(source_sheet, row_index, selected_match_columns)
            if _selected_match_row_is_header(match_values):
                continue
            inquiry_oe = _combined_match_text(match_values)
            inquiry_desc = source_sheet.cell_value(row_index, columns["description"]) if "description" in columns else ""
            if not str(inquiry_name).strip() and not str(inquiry_oe).strip():
                continue

            match = catalog.match(inquiry_name, inquiry_oe, inquiry_desc)
            summary["total"] += 1
            if match:
                if target_sheet:
                    target_sheet.write(row_index, output_col, match.bld_no)
                    if price_header:
                        price = _match_export_price(match, price_mode, exchange_rate)
                        target_sheet.write(row_index, output_col + 1, "" if price is None else price)
                summary["matched"] += 1
                row_summary = _annotate_row_summary_with_match_columns(
                    _summary_row(row_index + 1, inquiry_oe, inquiry_name, match),
                    match_values,
                    match,
                )
                if target_sheet:
                    target_sheet.write(row_index, note_col, row_summary["match_note"])
                summary["rows"].append(row_summary)
            else:
                if target_sheet:
                    target_sheet.write(row_index, output_col, "")
                    if price_header:
                        target_sheet.write(row_index, output_col + 1, "")
                summary["unmatched"] += 1
                row_summary = _summary_row(row_index + 1, inquiry_oe, inquiry_name, None)
                if target_sheet:
                    target_sheet.write(row_index, note_col, row_summary["match_note"])
                summary["rows"].append(row_summary)

    if writable:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        writable.save(str(output_path))
    return summary


def generate_xlsx_with_bld(
    inquiry_path: Path,
    output_path: Path,
    catalog: ProductCatalog,
    match_column: object = None,
    write_output: bool = True,
    price_mode: str = "none",
    exchange_rate: float | None = None,
) -> dict:
    if not write_output:
        return analyze_xlsx_with_bld(
            inquiry_path,
            catalog,
            match_column=match_column,
            price_mode=price_mode,
            exchange_rate=exchange_rate,
        )

    workbook = load_workbook(inquiry_path)
    try:
        summary = {"total": 0, "matched": 0, "unmatched": 0, "rows": []}

        for sheet in workbook.worksheets:
            if sheet.max_row == 0:
                continue
            data_end_row = sheet.max_row

            manual_match_columns = _manual_match_columns(match_column)
            if not manual_match_columns:
                header_row, columns = _find_xlsx_inquiry_columns(sheet)
                selected_match_columns = [columns["oe"]]
            else:
                selected_match_columns = [column + 1 for column in manual_match_columns]
                header_row = _find_xlsx_selected_header_row(sheet, selected_match_columns)
                columns = {}
            output_col = sheet.max_column + 1
            price_header = _price_export_header(price_mode)
            note_col = output_col + 2 if price_header else output_col + 1
            if write_output:
                sheet.cell(header_row, output_col).value = "BLD NO."
                if price_header:
                    sheet.cell(header_row, output_col + 1).value = price_header
                sheet.cell(header_row, note_col).value = "匹配说明"

            for row_index in range(header_row + 1, sheet.max_row + 1):
                inquiry_name = sheet.cell(row_index, columns["name"]).value if "name" in columns else ""
                match_values = _xlsx_match_values(sheet, row_index, selected_match_columns)
                if _selected_match_row_is_header(match_values):
                    continue
                inquiry_oe = _combined_match_text(match_values)
                inquiry_desc = sheet.cell(row_index, columns["description"]).value if "description" in columns else ""
                if not str(inquiry_name or "").strip() and not str(inquiry_oe or "").strip():
                    continue

                match = catalog.match(inquiry_name, inquiry_oe, inquiry_desc)
                summary["total"] += 1
                if match:
                    if write_output:
                        sheet.cell(row_index, output_col).value = match.bld_no
                        if price_header:
                            price = _match_export_price(match, price_mode, exchange_rate)
                            price_cell = sheet.cell(row_index, output_col + 1)
                            price_cell.value = price
                            price_cell.number_format = "0" if price_mode == "net" else "0.00"
                    summary["matched"] += 1
                    row_summary = _annotate_row_summary_with_match_columns(
                        _summary_row(row_index, inquiry_oe, inquiry_name, match),
                        match_values,
                        match,
                    )
                    if write_output:
                        sheet.cell(row_index, note_col).value = row_summary["match_note"]
                    summary["rows"].append(row_summary)
                else:
                    if write_output:
                        sheet.cell(row_index, output_col).value = ""
                        if price_header:
                            sheet.cell(row_index, output_col + 1).value = None
                    summary["unmatched"] += 1
                    row_summary = _summary_row(row_index, inquiry_oe, inquiry_name, None)
                    if write_output:
                        sheet.cell(row_index, note_col).value = row_summary["match_note"]
                    summary["rows"].append(row_summary)
            if write_output:
                _trim_unused_row_dimensions(sheet, data_end_row)

        if write_output:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(output_path)
        return summary
    finally:
        workbook.close()


def analyze_xlsx_with_bld(
    inquiry_path: Path,
    catalog: ProductCatalog,
    match_column: object = None,
    price_mode: str = "none",
    exchange_rate: float | None = None,
) -> dict:
    workbook = load_workbook(inquiry_path, read_only=True, data_only=True)
    try:
        summary = {"total": 0, "matched": 0, "unmatched": 0, "rows": []}

        for sheet in workbook.worksheets:
            first_rows = list(sheet.iter_rows(min_row=1, max_row=20, values_only=True))
            if not first_rows:
                continue

            manual_match_columns = _manual_match_columns(match_column)
            if not manual_match_columns:
                header_row, columns = _find_xlsx_inquiry_columns_from_rows(first_rows)
                selected_match_columns = [columns["oe"]]
            else:
                selected_match_columns = [column + 1 for column in manual_match_columns]
                header_row = _find_xlsx_selected_header_row_from_rows(first_rows, selected_match_columns)
                columns = {}

            for row_index, values in enumerate(sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                inquiry_name = _cell_from_values(values, columns["name"]) if "name" in columns else ""
                match_values = _row_match_values(values, selected_match_columns)
                if _selected_match_row_is_header(match_values):
                    continue
                inquiry_oe = _combined_match_text(match_values)
                inquiry_desc = _cell_from_values(values, columns["description"]) if "description" in columns else ""
                if not str(inquiry_name or "").strip() and not str(inquiry_oe or "").strip():
                    continue

                match = catalog.match(inquiry_name, inquiry_oe, inquiry_desc)
                summary["total"] += 1
                if match:
                    summary["matched"] += 1
                    summary["rows"].append(
                        _annotate_row_summary_with_match_columns(
                            _summary_row(row_index, inquiry_oe, inquiry_name, match),
                            match_values,
                            match,
                        )
                    )
                else:
                    summary["unmatched"] += 1
                    summary["rows"].append(_summary_row(row_index, inquiry_oe, inquiry_name, None))

        return summary
    finally:
        workbook.close()


def generate_excel_with_bld(
    inquiry_path: Path,
    output_path: Path,
    catalog: ProductCatalog,
    match_column: object = None,
    write_output: bool = True,
    price_mode: str = "none",
    exchange_rate: float | None = None,
) -> dict:
    suffix = inquiry_path.suffix.lower()
    if suffix == ".xls":
        return generate_xls_with_bld(
            inquiry_path,
            output_path.with_suffix(".xls"),
            catalog,
            match_column=match_column,
            write_output=write_output,
            price_mode=price_mode,
            exchange_rate=exchange_rate,
        )
    if suffix == ".xlsx":
        return generate_xlsx_with_bld(
            inquiry_path,
            output_path.with_suffix(".xlsx"),
            catalog,
            match_column=match_column,
            write_output=write_output,
            price_mode=price_mode,
            exchange_rate=exchange_rate,
        )
    raise ValueError("客户询价文件仅支持 .xls 或 .xlsx。")

from __future__ import annotations

import sqlite3
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import BASE_DIR, PRODUCT_IMAGE_DATA_PREFIX, PRODUCT_IMAGE_DIR


BLD_HEADERS = ["BLD NO.", "品牌", "产品名称", "OE Reference", "Other Reference", "车型", "Product Image", "Unit Price", "状态", "更新时间"]
BRAND_HEADERS = ["NO.", "SERIES", "BLD NO.", "ITEM", "OE Reference", "Other Reference", "Models", "Product Image", "Unit Price", "状态", "更新时间"]
CHINESE_RE = re.compile(r"[\u4e00-\u9fff]")
IMAGE_MAX_WIDTH = 112
IMAGE_MAX_HEIGHT = 72
IMAGE_ROW_HEIGHT = 62
IMAGE_EXTENSIONS = ("jpg", "jpeg", "png", "webp")


def _font_for(value, *, bold: bool = False) -> Font:
    text = "" if value is None else str(value)
    return Font(name="微软雅黑" if CHINESE_RE.search(text) else "Arial", size=10, bold=bold)


def _setup_sheet(sheet, headers: list[str]) -> None:
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = _font_for(cell.value, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _finish_sheet(sheet, widths: list[int], image_rows: set[int] | None = None) -> None:
    image_rows = image_rows or set()
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=2):
        max_lines = 1
        for cell in row:
            cell.font = _font_for(cell.value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, str):
                max_lines = max(max_lines, cell.value.count("\n") + 1)
        row_height = max(24, max_lines * 15)
        if row[0].row in image_rows:
            row_height = max(row_height, IMAGE_ROW_HEIGHT)
        sheet.row_dimensions[row[0].row].height = min(180, row_height)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def _existing_path(path: Path) -> Path | None:
    try:
        resolved = path.expanduser().resolve()
    except OSError:
        return None
    return resolved if resolved.is_file() else None


def _path_from_explicit_image(value: str) -> Path | None:
    explicit = value.strip()
    if not explicit or explicit.startswith(("http://", "https://")):
        return None

    if explicit.startswith(PRODUCT_IMAGE_DATA_PREFIX):
        return _existing_path(PRODUCT_IMAGE_DIR / Path(explicit[len(PRODUCT_IMAGE_DATA_PREFIX) :]).name)

    if explicit.startswith("/product-images/"):
        return _existing_path(PRODUCT_IMAGE_DIR / Path(explicit.removeprefix("/product-images/")).name)

    if explicit.startswith("/static/"):
        return _existing_path(BASE_DIR / "static" / explicit.removeprefix("/static/"))

    relative = explicit.lstrip("/")
    for candidate in (
        BASE_DIR / "static" / relative,
        BASE_DIR / relative,
        PRODUCT_IMAGE_DIR / Path(relative).name,
    ):
        path = _existing_path(candidate)
        if path:
            return path
    return None


def _image_path(row) -> Path | None:
    explicit = row["image_path"] if "image_path" in row.keys() else ""
    if explicit:
        path = _path_from_explicit_image(str(explicit))
        if path:
            return path

    bld_no = str(row["bld_no"] or "").strip()
    if not bld_no:
        return None
    for suffix in IMAGE_EXTENSIONS:
        for candidate in (
            PRODUCT_IMAGE_DIR / f"{bld_no}.{suffix}",
            BASE_DIR / "static" / "product_images" / f"{bld_no}.{suffix}",
        ):
            path = _existing_path(candidate)
            if path:
                return path
    return None


def _build_excel_image(path: Path) -> ExcelImage | None:
    try:
        image = ExcelImage(str(path))
    except Exception:
        return None

    if not image.width or not image.height:
        return None

    scale = min(IMAGE_MAX_WIDTH / image.width, IMAGE_MAX_HEIGHT / image.height)
    image.width = max(1, int(image.width * scale))
    image.height = max(1, int(image.height * scale))
    return image


def _add_sheet_images(sheet, image_refs: list[tuple[int, int, Path]]) -> None:
    for row_index, column_index, path in image_refs:
        image = _build_excel_image(path)
        if not image:
            continue
        sheet.add_image(image, f"{get_column_letter(column_index)}{row_index}")


def export_products_xlsx(
    conn: sqlite3.Connection,
    output_path: Path,
    include_inactive: bool = True,
    export_format: str = "bld",
) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "产品目录"

    sql = "SELECT * FROM products"
    if not include_inactive:
        sql += " WHERE active = 1"
    if export_format == "brand":
        sql += " ORDER BY series, bld_no"
        _setup_sheet(sheet, BRAND_HEADERS)
    else:
        sql += " ORDER BY bld_no"
        _setup_sheet(sheet, BLD_HEADERS)

    image_col = 8 if export_format == "brand" else 7
    image_refs: list[tuple[int, int, Path]] = []
    image_rows: set[int] = set()

    for number, row in enumerate(conn.execute(sql), start=1):
        image_path = _image_path(row)
        if export_format == "brand":
            sheet.append(
                [
                    number,
                    row["series"],
                    row["bld_no"],
                    row["item"],
                    row["oe_no_1"],
                    row["oe_no_2"],
                    row["models"],
                    "",
                    row["price_cny"],
                    "启用" if row["active"] else "停用",
                    row["updated_at"],
                ]
            )
        else:
            sheet.append(
                [
                    row["bld_no"],
                    row["series"],
                    row["item"],
                    row["oe_no_1"],
                    row["oe_no_2"],
                    row["models"],
                    "",
                    row["price_cny"],
                    "启用" if row["active"] else "停用",
                    row["updated_at"],
                ]
            )
        if image_path:
            image_refs.append((sheet.max_row, image_col, image_path))
            image_rows.add(sheet.max_row)

    widths = [10, 18, 16, 28, 34, 28, 34, 22, 12, 10, 20] if export_format == "brand" else [14, 18, 28, 34, 28, 34, 22, 12, 10, 20]
    _finish_sheet(sheet, widths, image_rows)
    _add_sheet_images(sheet, image_refs)
    price_col = 9 if export_format == "brand" else 8
    for cell in sheet.iter_cols(min_col=price_col, max_col=price_col, min_row=2):
        for item in cell:
            item.number_format = '¥#,##0.00'

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path

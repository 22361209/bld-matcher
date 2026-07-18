from __future__ import annotations

import json
import re
from pathlib import Path

import xlrd
from openpyxl import load_workbook

from .matcher import compact_text


HEADER_ALIASES = {
    "bld_no": {"BLD号", "BLD NO.", "BLD NO", "BLD"},
    "customer_product_code": {"客户产品编码", "客户编码", "客户料号", "CUSTOMER CODE", "CUSTOMER PART NO"},
    "tax_price": {"含税单价", "含税价格", "TAX PRICE", "TAX INCLUDED PRICE"},
    "net_price": {"不含税单价", "未税单价", "不含税价格", "NET PRICE"},
    "quote_date": {"日期", "报价日期", "DATE", "QUOTE DATE"},
    "remark": {"备注", "NOTE", "REMARK"},
}

REQUIRED_COLUMNS = {"bld_no"}


def _norm_header(value: object) -> str:
    return re.sub(r"[^A-Z0-9\u4e00-\u9fff]+", "", str(value or "").upper())


def _parse_price(value: object) -> float | None:
    text = compact_text(value).replace("¥", "").replace("￥", "").replace(",", "")
    if not text:
        return None
    try:
        return round(float(text), 4)
    except ValueError:
        return None


def _read_rows(path: Path) -> list[list[object]]:
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.worksheets[0]
            return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
    if path.suffix.lower() == ".xls":
        book = xlrd.open_workbook(path, ignore_workbook_corruption=True)
        sheet = book.sheet_by_index(0)
        return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    raise ValueError("报价记录导入仅支持 .xls 或 .xlsx。")


def _find_columns(rows: list[list[object]]) -> tuple[int, dict[str, int]]:
    aliases = {
        field: {_norm_header(alias) for alias in values}
        for field, values in HEADER_ALIASES.items()
    }
    for row_index, row in enumerate(rows[:20]):
        columns: dict[str, int] = {}
        for col_index, value in enumerate(row):
            key = _norm_header(value)
            for field, keys in aliases.items():
                if key in keys:
                    columns[field] = col_index
        if REQUIRED_COLUMNS <= columns.keys() and ("tax_price" in columns or "net_price" in columns):
            return row_index, columns
    raise ValueError("没有找到 BLD号、含税单价/不含税单价表头。")


def _cell(row: list[object], columns: dict[str, int], field: str) -> object:
    index = columns.get(field)
    if index is None or index >= len(row):
        return ""
    return row[index]


def parse_quote_import_file(path: Path, *, customer_name: str, currency: str) -> dict:
    customer_name = compact_text(customer_name)
    currency = compact_text(currency).upper()
    if not customer_name:
        raise ValueError("客户不能为空。")
    if currency not in {"CNY", "USD", "EUR"}:
        raise ValueError("币种只允许 CNY/USD/EUR。")

    rows = _read_rows(path)
    header_row, columns = _find_columns(rows)
    preview = []
    counts = {"total": 0, "valid": 0, "invalid": 0}

    for row_number, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        bld_no = compact_text(_cell(row, columns, "bld_no"))
        tax_price = _parse_price(_cell(row, columns, "tax_price"))
        net_price = _parse_price(_cell(row, columns, "net_price"))
        if not bld_no and tax_price is None and net_price is None:
            continue

        errors = []
        if not bld_no:
            errors.append("BLD号不能为空")
        if tax_price is None and net_price is None:
            errors.append("含税单价或不含税单价至少填写一个")
        data = {
            "customer_name": customer_name,
            "bld_no": bld_no,
            "customer_product_code": compact_text(_cell(row, columns, "customer_product_code")),
            "tax_price": tax_price,
            "net_price": net_price,
            "currency": currency,
            "quote_date": compact_text(_cell(row, columns, "quote_date")),
            "source_type": "excel",
            "remark": compact_text(_cell(row, columns, "remark")),
        }
        status = "invalid" if errors else "valid"
        counts[status] += 1
        counts["total"] += 1
        preview.append({"row": row_number, "status": status, "error": "；".join(errors), **data})
    return {"counts": counts, "rows": preview}


def encode_rows(rows: list[dict]) -> str:
    return json.dumps(rows, ensure_ascii=False)


def decode_rows(payload: str) -> list[dict]:
    data = json.loads(payload)
    if not isinstance(data, list):
        raise ValueError("导入数据无效。")
    return data

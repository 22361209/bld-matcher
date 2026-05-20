from __future__ import annotations

import re
import sqlite3
import threading
import hashlib
import hmac
import secrets
from datetime import datetime
from pathlib import Path
from collections import defaultdict

from werkzeug.security import generate_password_hash

from .bld_sort import compare_bld_no
from .migrations import run_migrations
from .matcher import PSA_352X_BRANDS, ProductCatalog, compact_text, normalize_code, psa_352x_key, split_codes


PASSWORD_HASH_METHOD = "pbkdf2:sha256"
_INIT_LOCK = threading.Lock()
_INITIALIZED_DB_PATHS: set[Path] = set()
_BOOTSTRAPPED_SOURCES: set[tuple[str, Path, Path, int | None]] = set()


def hash_password(password: str) -> str:
    return generate_password_hash(password, method=PASSWORD_HASH_METHOD)


SCHEMA = """
CREATE TABLE IF NOT EXISTS products (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  bld_no TEXT NOT NULL UNIQUE,
  series TEXT DEFAULT '',
  item TEXT DEFAULT '',
  oe_no_1 TEXT DEFAULT '',
  oe_no_2 TEXT DEFAULT '',
  models TEXT DEFAULT '',
  price_cny REAL,
  image_path TEXT DEFAULT '',
  image_path_2 TEXT DEFAULT '',
  image_path_3 TEXT DEFAULT '',
  image_path_4 TEXT DEFAULT '',
  image_path_5 TEXT DEFAULT '',
  drawing_path TEXT DEFAULT '',
  drawing_original_name TEXT DEFAULT '',
  drawing_updated_at TEXT DEFAULT '',
  active INTEGER NOT NULL DEFAULT 1,
  source TEXT DEFAULT '',
  created_at TEXT NOT NULL,
  updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS aliases (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  source_code TEXT NOT NULL UNIQUE,
  bld_no TEXT NOT NULL,
  note TEXT DEFAULT '',
  active INTEGER NOT NULL DEFAULT 1,
  created_at TEXT NOT NULL,
  updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS audit_logs (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  action TEXT NOT NULL,
  target_type TEXT NOT NULL,
  target_key TEXT NOT NULL,
  actor TEXT DEFAULT '',
  detail TEXT DEFAULT '',
  created_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS users (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  username TEXT NOT NULL UNIQUE,
  display_name TEXT DEFAULT '',
  password_hash TEXT NOT NULL,
  role TEXT NOT NULL DEFAULT 'viewer',
  active INTEGER NOT NULL DEFAULT 1,
  created_at TEXT NOT NULL,
  updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS material_items (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  model TEXT NOT NULL,
  code TEXT DEFAULT '',
  category TEXT DEFAULT '',
  car TEXT DEFAULT '',
  part TEXT DEFAULT '',
  spec_text TEXT DEFAULT '',
  pieces REAL NOT NULL,
  thickness REAL NOT NULL,
  width REAL NOT NULL,
  length REAL NOT NULL,
  active INTEGER NOT NULL DEFAULT 1,
  source TEXT DEFAULT '',
  source_row INTEGER DEFAULT 0,
  created_at TEXT NOT NULL,
  updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS customer_price_records (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  record_type TEXT NOT NULL DEFAULT 'quote',
  customer_name TEXT NOT NULL,
  record_date TEXT NOT NULL,
  document_no TEXT DEFAULT '',
  source_name TEXT DEFAULT '',
  source_code TEXT DEFAULT '',
  oe_no TEXT DEFAULT '',
  bld_no TEXT DEFAULT '',
  item TEXT DEFAULT '',
  models TEXT DEFAULT '',
  price_cny REAL,
  price_usd REAL,
  currency TEXT DEFAULT 'CNY',
  exchange_rate REAL,
  tax_included INTEGER NOT NULL DEFAULT 1,
  note TEXT DEFAULT '',
  source_file TEXT DEFAULT '',
  created_by TEXT DEFAULT '',
  created_at TEXT NOT NULL,
  updated_at TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_customer_price_records_customer ON customer_price_records(customer_name);
CREATE INDEX IF NOT EXISTS idx_customer_price_records_bld ON customer_price_records(bld_no);
CREATE INDEX IF NOT EXISTS idx_customer_price_records_date ON customer_price_records(record_date);

CREATE TABLE IF NOT EXISTS internal_api_keys (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  name TEXT NOT NULL DEFAULT 'OpenClaw',
  token_hash TEXT NOT NULL UNIQUE,
  token_prefix TEXT DEFAULT '',
  token_suffix TEXT DEFAULT '',
  active INTEGER NOT NULL DEFAULT 1,
  created_by TEXT DEFAULT '',
  created_at TEXT NOT NULL,
  updated_at TEXT NOT NULL,
  last_used_at TEXT DEFAULT ''
);
CREATE INDEX IF NOT EXISTS idx_internal_api_keys_active ON internal_api_keys(active);

CREATE TABLE IF NOT EXISTS shipment_recognition_jobs (
  id TEXT PRIMARY KEY,
  owner TEXT NOT NULL,
  payload TEXT NOT NULL,
  created_at TEXT NOT NULL,
  updated_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_shipment_recognition_jobs_owner ON shipment_recognition_jobs(owner);
CREATE INDEX IF NOT EXISTS idx_shipment_recognition_jobs_updated ON shipment_recognition_jobs(updated_at);
"""


def connect(db_path: Path) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    needs_init = not db_path.exists()
    conn = sqlite3.connect(db_path, timeout=5.0)
    conn.row_factory = sqlite3.Row
    conn.create_collation("BLD_NATURAL", compare_bld_no)
    # 多人并发匹配/导入时减少 "database is locked"。
    # WAL 一次性设置,后续连接都受益;synchronous=NORMAL 在 WAL 下安全且更快;
    # busy_timeout 让短暂冲突自动重试,而不是立刻抛 OperationalError。
    try:
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA busy_timeout=5000")
    except sqlite3.DatabaseError:
        # 某些网络挂载文件系统不支持 WAL,退回默认行为不影响功能。
        pass
    resolved = db_path.resolve()
    if needs_init or resolved not in _INITIALIZED_DB_PATHS:
        with _INIT_LOCK:
            if needs_init or resolved not in _INITIALIZED_DB_PATHS:
                conn.executescript(SCHEMA)
                run_migrations(conn)
                _INITIALIZED_DB_PATHS.add(resolved)
    return conn


def _bootstrap_key(kind: str, db_path: Path, source_path: Path) -> tuple[str, Path, Path, int | None]:
    source_mtime = source_path.stat().st_mtime_ns if source_path.exists() else None
    return (kind, db_path.resolve(), source_path.resolve(), source_mtime)


def now_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def clean_multiline(value: object) -> str:
    text = "" if value is None else str(value)
    lines = [compact_text(line) for line in text.replace("\r", "\n").split("\n")]
    return "\n".join(line for line in lines if line)


def clean_oe_list(value: object) -> str:
    return "\n".join(split_codes(value))


def merge_unique_lines(values: list[object], *, oe: bool = False) -> str:
    seen: list[str] = []
    for value in values:
        text = clean_oe_list(value) if oe else clean_multiline(value)
        for line in text.split("\n"):
            if line and line not in seen:
                seen.append(line)
    return "\n".join(seen)


def _parse_price(value: object) -> float | None:
    text = compact_text(value)
    if not text:
        return None
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def _parse_required_float(value: object, label: str) -> float:
    text = compact_text(value)
    if not text:
        raise ValueError(f"{label}不能为空。")
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"{label}必须是数字：{text}") from exc


def _format_material_int(value: float) -> int | float:
    return int(value) if float(value).is_integer() else value


def _format_material_thickness(value: float) -> str:
    text = f"{value:.2f}".rstrip("0")
    if text.endswith("."):
        return text + "0"
    return text


def _format_material_spec(thickness: float, width: float, length: float) -> str:
    return f"{_format_material_thickness(thickness)}×{_format_material_int(width)}×{_format_material_int(length)}"


def _parse_material_spec_text(value: object) -> tuple[float, float, float]:
    text = compact_text(value)
    if not text:
        raise ValueError("规格尺寸不能为空。")
    normalized = re.sub(r"[×xX*＊/／\\\-－—]+", " ", text)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    parts = normalized.split(" ") if normalized else []
    if len(parts) != 3:
        raise ValueError("规格尺寸请按“厚度 宽度 长度”填写，例如 2.5 357 1260。")
    try:
        return tuple(float(part) for part in parts)  # type: ignore[return-value]
    except ValueError as exc:
        raise ValueError(f"规格尺寸必须包含 3 个数字：{text}") from exc


def _parse_material_spec_query(value: object) -> list[float]:
    text = compact_text(value)
    if not text:
        return []
    normalized = re.sub(r"[×xX*＊/／\\\-－—]+", " ", text)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    parts = normalized.split(" ") if normalized else []
    if not 1 <= len(parts) <= 3:
        return []
    try:
        return [float(part) for part in parts]
    except ValueError:
        return []


def _material_spec_search(tokens: list[float]) -> tuple[str, list[object]]:
    if not tokens:
        return "", []
    epsilon = 0.000001

    def equal_expr(field: str) -> str:
        return f"ABS({field} - ?) < ?"

    if len(tokens) == 1:
        value = tokens[0]
        return (
            f"({equal_expr('thickness')} OR {equal_expr('width')} OR {equal_expr('length')})",
            [value, epsilon, value, epsilon, value, epsilon],
        )
    if len(tokens) == 2:
        first, second = tokens
        return (
            "("
            f"({equal_expr('thickness')} AND {equal_expr('width')}) OR "
            f"({equal_expr('thickness')} AND {equal_expr('length')}) OR "
            f"({equal_expr('width')} AND {equal_expr('length')})"
            ")",
            [
                first,
                epsilon,
                second,
                epsilon,
                first,
                epsilon,
                second,
                epsilon,
                first,
                epsilon,
                second,
                epsilon,
            ],
        )
    first, second, third = tokens
    return (
        f"({equal_expr('thickness')} AND {equal_expr('width')} AND {equal_expr('length')})",
        [first, epsilon, second, epsilon, third, epsilon],
    )


def log_event(conn: sqlite3.Connection, action: str, target_type: str, target_key: str, detail: str = "", actor: str = "") -> None:
    conn.execute(
        "INSERT INTO audit_logs (action, target_type, target_key, actor, detail, created_at) VALUES (?, ?, ?, ?, ?, ?)",
        (action, target_type, target_key, actor, detail, now_text()),
    )


def _hash_api_token(token: str) -> str:
    return hashlib.sha256(str(token or "").encode("utf-8")).hexdigest()


def _new_api_token() -> str:
    return f"bld_sk_{secrets.token_urlsafe(32)}"


def _api_key_preview(row: sqlite3.Row | None) -> str:
    if not row:
        return ""
    prefix = str(row["token_prefix"] or "bld_sk")
    suffix = str(row["token_suffix"] or "")
    return f"{prefix}****{suffix}" if suffix else f"{prefix}****"


def internal_api_key_status(conn: sqlite3.Connection) -> dict:
    active = conn.execute(
        """
        SELECT * FROM internal_api_keys
        WHERE active = 1
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    latest = conn.execute(
        """
        SELECT * FROM internal_api_keys
        ORDER BY id DESC
        LIMIT 1
        """
    ).fetchone()
    row = active or latest
    return {
        "enabled": bool(active),
        "preview": _api_key_preview(active),
        "name": row["name"] if row else "OpenClaw",
        "created_by": row["created_by"] if row else "",
        "created_at": row["created_at"] if row else "",
        "updated_at": row["updated_at"] if row else "",
        "last_used_at": row["last_used_at"] if row else "",
    }


def create_internal_api_key(conn: sqlite3.Connection, *, actor: str = "", name: str = "OpenClaw") -> str:
    token = _new_api_token()
    timestamp = now_text()
    label = compact_text(name) or "OpenClaw"
    conn.execute("UPDATE internal_api_keys SET active = 0, updated_at = ? WHERE active = 1", (timestamp,))
    conn.execute(
        """
        INSERT INTO internal_api_keys
          (name, token_hash, token_prefix, token_suffix, active, created_by, created_at, updated_at)
        VALUES (?, ?, ?, ?, 1, ?, ?, ?)
        """,
        (label, _hash_api_token(token), "bld_sk_", token[-6:], actor, timestamp, timestamp),
    )
    log_event(conn, "生成内部 API Key", "internal_api_key", label, "旧 Key 已自动停用，新 Key 只在生成页面显示一次。", actor=actor)
    conn.commit()
    return token


def disable_internal_api_key(conn: sqlite3.Connection, *, actor: str = "") -> bool:
    timestamp = now_text()
    cursor = conn.execute(
        "UPDATE internal_api_keys SET active = 0, updated_at = ? WHERE active = 1",
        (timestamp,),
    )
    changed = cursor.rowcount > 0
    if changed:
        log_event(conn, "停用内部 API Key", "internal_api_key", "OpenClaw", "内部 API 已停用。", actor=actor)
        conn.commit()
    return changed


def verify_internal_api_token(conn: sqlite3.Connection, token: str) -> bool:
    token_hash = _hash_api_token(token)
    rows = conn.execute(
        "SELECT id, token_hash FROM internal_api_keys WHERE active = 1"
    ).fetchall()
    for row in rows:
        if hmac.compare_digest(str(row["token_hash"]), token_hash):
            conn.execute(
                "UPDATE internal_api_keys SET last_used_at = ? WHERE id = ?",
                (now_text(), row["id"]),
            )
            conn.commit()
            return True
    return False


def _field_changes(before: sqlite3.Row | None, after: dict) -> list[str]:
    labels = {
        "series": "品牌",
        "item": "产品名称",
        "oe_no_1": "OE 号",
        "oe_no_2": "品牌号码",
        "models": "车型",
        "price_cny": "含税单价",
        "image_path": "产品图片",
        "active": "状态",
    }
    if before is None:
        return ["新增产品"]
    changes = []
    for field, label in labels.items():
        old = str(before[field] or "")
        new = str(after[field] or "")
        if old != new:
            changes.append(f"{label}: {old or '(空)'} -> {new or '(空)'}")
    return changes


def upsert_product(conn: sqlite3.Connection, data: dict, source: str = "manual", audit: bool = True, actor: str = "") -> None:
    timestamp = now_text()
    bld_no = compact_text(data.get("bld_no") or data.get("BLD NO."))
    if not bld_no:
        raise ValueError("BLD NO. 不能为空。")

    values = {
        "bld_no": bld_no,
        "series": clean_multiline(data.get("series") or data.get("SERIES")),
        "item": clean_multiline(data.get("item") or data.get("ITEM")),
        "oe_no_1": clean_oe_list(data.get("oe_no_1") or data.get("OE NO.1")),
        "oe_no_2": clean_oe_list(data.get("oe_no_2") or data.get("OE NO.2")),
        "models": clean_multiline(data.get("models") or data.get("Models")),
        "price_cny": _parse_price(data.get("price_cny")),
        "image_path": compact_text(data.get("image_path")),
        "active": 1 if str(data.get("active", "1")) != "0" else 0,
        "source": source,
        "created_at": timestamp,
        "updated_at": timestamp,
    }
    before = conn.execute("SELECT * FROM products WHERE bld_no = ?", (bld_no,)).fetchone()
    conn.execute(
        """
        INSERT INTO products
          (bld_no, series, item, oe_no_1, oe_no_2, models, price_cny, image_path, active, source, created_at, updated_at)
        VALUES
          (:bld_no, :series, :item, :oe_no_1, :oe_no_2, :models, :price_cny, :image_path, :active, :source, :created_at, :updated_at)
        ON CONFLICT(bld_no) DO UPDATE SET
          series=excluded.series,
          item=excluded.item,
          oe_no_1=excluded.oe_no_1,
          oe_no_2=excluded.oe_no_2,
          models=excluded.models,
          price_cny=COALESCE(excluded.price_cny, products.price_cny),
          image_path=CASE WHEN excluded.image_path != '' THEN excluded.image_path ELSE products.image_path END,
          active=excluded.active,
          source=excluded.source,
          updated_at=excluded.updated_at
        """,
        values,
    )
    if audit:
        changes = _field_changes(before, values)
        if changes:
            action = "新增产品" if before is None else "编辑产品"
            log_event(conn, action, "product", bld_no, "\n".join(changes[:20]), actor=actor)
    conn.commit()


def import_catalog(conn: sqlite3.Connection, catalog_path: Path, replace: bool = False, actor: str = "") -> int:
    catalog = ProductCatalog.from_excel(catalog_path)
    if replace:
        conn.execute("DELETE FROM products")
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in catalog.rows:
        bld_no = compact_text(row.get("BLD NO."))
        if bld_no:
            grouped[bld_no].append(row)

    for bld_no, rows in grouped.items():
        merged = {
            "BLD NO.": bld_no,
            "SERIES": merge_unique_lines([row.get("SERIES") for row in rows]),
            "ITEM": merge_unique_lines([row.get("ITEM") for row in rows]),
            "OE NO.1": merge_unique_lines([row.get("OE NO.1") for row in rows], oe=True),
            "OE NO.2": merge_unique_lines([row.get("OE NO.2") for row in rows], oe=True),
            "Models": merge_unique_lines([row.get("Models") for row in rows]),
            "price_cny": None,
            "image_path": "",
        }
        upsert_product(conn, merged, source=catalog_path.name, audit=False)
    log_event(conn, "导入目录", "catalog", catalog_path.name, f"汇总导入 {len(grouped)} 个唯一 BLD 号", actor=actor)
    conn.commit()
    return len(grouped)


def bootstrap_from_excel(db_path: Path, catalog_path: Path) -> None:
    key = _bootstrap_key("catalog", db_path, catalog_path)
    if key in _BOOTSTRAPPED_SOURCES:
        return
    with connect(db_path) as conn:
        existing = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        if existing == 0 and catalog_path.exists():
            import_catalog(conn, catalog_path, replace=True)
    _BOOTSTRAPPED_SOURCES.add(key)


def _normalized_code_sql(column: str) -> str:
    expression = f"UPPER({column})"
    for item in ("-", " ", ".", "/", "_"):
        expression = f"REPLACE({expression}, '{item}', '')"
    for char_code, replacement in ((9, ""), (10, "|"), (13, "|")):
        expression = f"REPLACE({expression}, CHAR({char_code}), '{replacement}')"
    return expression


def _psa_product_clause() -> tuple[str, list[str]]:
    checks = []
    params = []
    for brand in PSA_352X_BRANDS:
        checks.append("UPPER(series) LIKE ?")
        checks.append("UPPER(models) LIKE ?")
        params.extend([f"%{brand}%", f"%{brand}%"])
    return "(" + " OR ".join(checks) + ")", params


def _product_filter_clauses(
    query: str = "",
    include_inactive: bool = False,
    only_inactive: bool = False,
    bld_query: str = "",
    oe_query: str = "",
    series_query: str = "",
    model_query: str = "",
) -> tuple[list[str], list[object]]:
    params: list[object] = []
    clauses = []
    if only_inactive:
        clauses.append("active = 0")
    elif not include_inactive:
        clauses.append("active = 1")
    if query.strip() and not bld_query.strip() and not oe_query.strip():
        bld_query = query
    if bld_query.strip():
        product_key = f"%{bld_query.strip().upper()}%"
        clauses.append("(UPPER(bld_no) LIKE ? OR UPPER(series) LIKE ? OR UPPER(models) LIKE ?)")
        params.extend([product_key, product_key, product_key])
    if oe_query.strip():
        psa_probe = psa_352x_key(oe_query)
        norm_key = f"%{psa_probe[0] if psa_probe else normalize_code(oe_query)}%"
        oe1 = _normalized_code_sql("oe_no_1")
        oe2 = _normalized_code_sql("oe_no_2")
        oe_clause = f"({oe1} LIKE ? OR {oe2} LIKE ?)"
        if psa_probe and psa_probe[1]:
            psa_clause, psa_params = _psa_product_clause()
            clauses.append(f"({oe_clause} AND {psa_clause})")
            params.extend([norm_key, norm_key, *psa_params])
        else:
            clauses.append(oe_clause)
            params.extend([norm_key, norm_key])
    if series_query.strip():
        clauses.append("UPPER(series) LIKE ?")
        params.append(f"%{series_query.strip().upper()}%")
    if model_query.strip():
        clauses.append("UPPER(models) LIKE ?")
        params.append(f"%{model_query.strip().upper()}%")
    return clauses, params


def list_products(
    conn: sqlite3.Connection,
    query: str = "",
    include_inactive: bool = False,
    only_inactive: bool = False,
    limit: int = 3000,
    bld_query: str = "",
    oe_query: str = "",
    series_query: str = "",
    model_query: str = "",
    offset: int = 0,
) -> list[sqlite3.Row]:
    sql = "SELECT * FROM products"
    clauses, params = _product_filter_clauses(
        query=query,
        include_inactive=include_inactive,
        only_inactive=only_inactive,
        bld_query=bld_query,
        oe_query=oe_query,
        series_query=series_query,
        model_query=model_query,
    )
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    sql += " ORDER BY bld_no COLLATE BLD_NATURAL LIMIT ? OFFSET ?"
    params.extend([max(0, limit), max(0, offset)])
    return conn.execute(sql, params).fetchall()


def count_products(
    conn: sqlite3.Connection,
    query: str = "",
    include_inactive: bool = False,
    only_inactive: bool = False,
    bld_query: str = "",
    oe_query: str = "",
    series_query: str = "",
    model_query: str = "",
) -> int:
    sql = "SELECT COUNT(*) FROM products"
    clauses, params = _product_filter_clauses(
        query=query,
        include_inactive=include_inactive,
        only_inactive=only_inactive,
        bld_query=bld_query,
        oe_query=oe_query,
        series_query=series_query,
        model_query=model_query,
    )
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    return int(conn.execute(sql, params).fetchone()[0] or 0)


def get_product(conn: sqlite3.Connection, product_id: int) -> sqlite3.Row | None:
    return conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()


def deactivate_product(conn: sqlite3.Connection, product_id: int, actor: str = "") -> None:
    row = get_product(conn, product_id)
    conn.execute("UPDATE products SET active = 0, updated_at = ? WHERE id = ?", (now_text(), product_id))
    if row:
        log_event(conn, "停用产品", "product", row["bld_no"], "状态: 启用 -> 停用", actor=actor)
    conn.commit()


def delete_product(conn: sqlite3.Connection, product_id: int, actor: str = "") -> sqlite3.Row | None:
    row = get_product(conn, product_id)
    if not row:
        return None

    alias_count = conn.execute(
        "SELECT COUNT(*) FROM aliases WHERE bld_no = ? AND active = 1",
        (row["bld_no"],),
    ).fetchone()[0]
    conn.execute("DELETE FROM products WHERE id = ?", (product_id,))
    if alias_count:
        conn.execute(
            "UPDATE aliases SET active = 0, updated_at = ? WHERE bld_no = ? AND active = 1",
            (now_text(), row["bld_no"]),
        )
    detail = f"品牌: {row['series'] or '(空)'}；产品名称: {row['item'] or '(空)'}"
    if alias_count:
        detail += f"；同步停用人工映射 {alias_count} 条"
    log_event(conn, "删除产品", "product", row["bld_no"], detail, actor=actor)
    conn.commit()
    return row


def product_stats(conn: sqlite3.Connection) -> dict[str, int]:
    row = conn.execute(
        "SELECT COUNT(*) AS total, SUM(active = 1) AS active, SUM(active = 0) AS inactive FROM products"
    ).fetchone()
    alias_count = conn.execute("SELECT COUNT(*) FROM aliases WHERE active = 1").fetchone()[0]
    return {
        "products": row["total"] or 0,
        "active": row["active"] or 0,
        "inactive": row["inactive"] or 0,
        "aliases": alias_count or 0,
    }


PRICE_RECORD_TYPES = {"quote", "order"}


def _parse_optional_float(value: object) -> float | None:
    text = compact_text(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"数字格式不正确：{text}") from exc


def _parse_record_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = compact_text(value)
    if not text:
        return datetime.now().strftime("%Y-%m-%d")
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    raise ValueError(f"日期格式不正确：{text}")


def _customer_price_values(data: dict, *, actor: str = "") -> dict:
    record_type = compact_text(data.get("record_type")).lower() or "quote"
    if record_type not in PRICE_RECORD_TYPES:
        raise ValueError("记录类型不正确。")
    customer_name = clean_multiline(data.get("customer_name"))
    if not customer_name:
        raise ValueError("客户名称不能为空。")

    price_cny = _parse_optional_float(data.get("price_cny"))
    price_usd = _parse_optional_float(data.get("price_usd"))
    if price_cny is None and price_usd is None:
        raise ValueError("含税单价或美金价至少填写一个。")

    source_code = clean_oe_list(data.get("source_code"))
    oe_no = clean_oe_list(data.get("oe_no"))
    bld_no = compact_text(data.get("bld_no")).upper()
    item = clean_multiline(data.get("item"))
    models = clean_multiline(data.get("models"))
    if not any([source_code, oe_no, bld_no, item]):
        raise ValueError("客户号码、OE 号、BLD NO. 或产品名称至少填写一个。")

    timestamp = now_text()
    return {
        "record_type": record_type,
        "customer_name": customer_name,
        "record_date": _parse_record_date(data.get("record_date")),
        "document_no": compact_text(data.get("document_no")),
        "source_name": clean_multiline(data.get("source_name")),
        "source_code": source_code,
        "oe_no": oe_no,
        "bld_no": bld_no,
        "item": item,
        "models": models,
        "price_cny": round(price_cny, 2) if price_cny is not None else None,
        "price_usd": round(price_usd, 4) if price_usd is not None else None,
        "currency": compact_text(data.get("currency")).upper() or ("USD" if price_usd is not None and price_cny is None else "CNY"),
        "exchange_rate": _parse_optional_float(data.get("exchange_rate")),
        "tax_included": 0 if str(data.get("tax_included", "1")) == "0" else 1,
        "note": clean_multiline(data.get("note")),
        "source_file": compact_text(data.get("source_file")),
        "created_by": actor,
        "created_at": timestamp,
        "updated_at": timestamp,
    }


def add_customer_price_record(
    conn: sqlite3.Connection,
    data: dict,
    *,
    actor: str = "",
    audit: bool = True,
    commit: bool = True,
) -> int:
    values = _customer_price_values(data, actor=actor)
    cursor = conn.execute(
        """
        INSERT INTO customer_price_records
          (record_type, customer_name, record_date, document_no, source_name, source_code, oe_no, bld_no,
           item, models, price_cny, price_usd, currency, exchange_rate, tax_included,
           note, source_file, created_by, created_at, updated_at)
        VALUES
          (:record_type, :customer_name, :record_date, :document_no, :source_name, :source_code, :oe_no, :bld_no,
           :item, :models, :price_cny, :price_usd, :currency, :exchange_rate, :tax_included,
           :note, :source_file, :created_by, :created_at, :updated_at)
        """,
        values,
    )
    record_id = int(cursor.lastrowid)
    if audit:
        label = "成交记录" if values["record_type"] == "order" else "报价记录"
        key = values["bld_no"] or values["source_code"] or values["oe_no"] or str(record_id)
        log_event(conn, "新增价格维护记录", "customer_price", key, f"{label}；客户: {values['customer_name']}", actor=actor)
    if commit:
        conn.commit()
    return record_id


def delete_customer_price_record(conn: sqlite3.Connection, record_id: int, *, actor: str = "") -> sqlite3.Row | None:
    row = conn.execute("SELECT * FROM customer_price_records WHERE id = ?", (record_id,)).fetchone()
    if not row:
        return None
    conn.execute("DELETE FROM customer_price_records WHERE id = ?", (record_id,))
    key = row["bld_no"] or row["source_code"] or row["oe_no"] or str(record_id)
    log_event(conn, "删除价格维护记录", "customer_price", key, f"客户: {row['customer_name']}", actor=actor)
    conn.commit()
    return row


def _customer_price_filter_clauses(
    customer: str = "",
    bld_no: str = "",
    source_code: str = "",
    record_type: str = "",
) -> tuple[list[str], list[object]]:
    clauses: list[str] = []
    params: list[object] = []
    if customer.strip():
        clauses.append("customer_name = ?")
        params.append(customer.strip())
    if record_type in PRICE_RECORD_TYPES:
        clauses.append("record_type = ?")
        params.append(record_type)
    if bld_no.strip():
        clauses.append("UPPER(bld_no) LIKE ?")
        params.append(f"%{bld_no.strip().upper()}%")
    if source_code.strip():
        normalized = f"%{normalize_code(source_code)}%"
        clauses.append("REPLACE(REPLACE(REPLACE(REPLACE(UPPER(source_code), '-', ''), ' ', ''), CHAR(10), '|'), CHAR(13), '|') LIKE ?")
        params.append(normalized)
    return clauses, params


def list_customer_price_records(
    conn: sqlite3.Connection,
    *,
    customer: str = "",
    bld_no: str = "",
    source_code: str = "",
    record_type: str = "",
    limit: int = 100,
    offset: int = 0,
) -> list[sqlite3.Row]:
    sql = "SELECT * FROM customer_price_records"
    clauses, params = _customer_price_filter_clauses(customer, bld_no, source_code, record_type)
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    sql += " ORDER BY record_date DESC, id DESC LIMIT ? OFFSET ?"
    params.extend([max(0, limit), max(0, offset)])
    return conn.execute(sql, params).fetchall()


def count_customer_price_records(
    conn: sqlite3.Connection,
    *,
    customer: str = "",
    bld_no: str = "",
    source_code: str = "",
    record_type: str = "",
) -> int:
    sql = "SELECT COUNT(*) FROM customer_price_records"
    clauses, params = _customer_price_filter_clauses(customer, bld_no, source_code, record_type)
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    return int(conn.execute(sql, params).fetchone()[0] or 0)


def list_customer_price_customer_summaries(
    conn: sqlite3.Connection,
    *,
    customer_query: str = "",
    record_type: str = "",
    limit: int = 100,
    offset: int = 0,
) -> list[sqlite3.Row]:
    clauses = []
    params: list[object] = []
    if customer_query.strip():
        clauses.append("UPPER(customer_name) LIKE ?")
        params.append(f"%{customer_query.strip().upper()}%")
    if record_type in PRICE_RECORD_TYPES:
        clauses.append("record_type = ?")
        params.append(record_type)
    where_sql = "WHERE " + " AND ".join(clauses) if clauses else ""
    return conn.execute(
        f"""
        SELECT
          customer_name,
          COUNT(*) AS total_records,
          COUNT(DISTINCT NULLIF(bld_no, '')) AS model_count,
          SUM(record_type = 'quote') AS quote_records,
          COUNT(DISTINCT CASE WHEN record_type = 'quote' THEN NULLIF(bld_no, '') END) AS quote_models,
          SUM(record_type = 'order') AS order_records,
          COUNT(DISTINCT CASE WHEN record_type = 'order' THEN NULLIF(bld_no, '') END) AS order_models,
          MAX(CASE WHEN record_type = 'quote' THEN record_date END) AS latest_quote_date,
          MAX(CASE WHEN record_type = 'order' THEN record_date END) AS latest_order_date,
          MAX(record_date) AS latest_date
        FROM customer_price_records
        {where_sql}
        GROUP BY customer_name
        ORDER BY latest_date DESC, customer_name
        LIMIT ? OFFSET ?
        """,
        [*params, max(0, limit), max(0, offset)],
    ).fetchall()


def count_customer_price_customers(conn: sqlite3.Connection, *, customer_query: str = "", record_type: str = "") -> int:
    clauses = []
    params: list[object] = []
    if customer_query.strip():
        clauses.append("UPPER(customer_name) LIKE ?")
        params.append(f"%{customer_query.strip().upper()}%")
    if record_type in PRICE_RECORD_TYPES:
        clauses.append("record_type = ?")
        params.append(record_type)
    sql = "SELECT COUNT(DISTINCT customer_name) FROM customer_price_records"
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    return int(conn.execute(sql, params).fetchone()[0] or 0)


def list_customer_price_model_comparisons(
    conn: sqlite3.Connection,
    *,
    customer: str = "",
    bld_no: str = "",
    source_code: str = "",
    record_type: str = "",
    limit: int = 200,
) -> list[sqlite3.Row]:
    clauses, params = _customer_price_filter_clauses(customer, bld_no, source_code, record_type)
    where_sql = " WHERE " + " AND ".join(clauses) if clauses else ""
    model_key = "COALESCE(NULLIF(bld_no, ''), NULLIF(source_code, ''), NULLIF(oe_no, ''))"
    sql = f"""
        WITH filtered AS (
          SELECT *
          FROM customer_price_records
          {where_sql}
        ),
        ranked AS (
          SELECT
            *,
            {model_key} AS model_key,
            ROW_NUMBER() OVER (
              PARTITION BY customer_name, {model_key}
              ORDER BY record_date DESC, id DESC
            ) AS rn
          FROM filtered
          WHERE {model_key} IS NOT NULL
        )
        SELECT
          customer_name,
          model_key,
          MAX(CASE WHEN rn = 1 THEN bld_no END) AS bld_no,
          MAX(CASE WHEN rn = 1 THEN source_code END) AS source_code,
          MAX(CASE WHEN rn = 1 THEN oe_no END) AS oe_no,
          MAX(CASE WHEN rn = 1 THEN item END) AS item,
          MAX(CASE WHEN rn = 1 THEN models END) AS models,
          MAX(CASE WHEN rn = 1 THEN record_date END) AS latest_date,
          MAX(CASE WHEN rn = 1 THEN price_cny END) AS latest_price_cny,
          MIN(price_cny) AS min_price_cny,
          MAX(price_cny) AS max_price_cny,
          COUNT(*) AS total_records
        FROM ranked
        GROUP BY customer_name, model_key
        ORDER BY model_key, customer_name
        LIMIT ?
    """
    return conn.execute(sql, [*params, max(0, limit)]).fetchall()


def customer_price_stats(conn: sqlite3.Connection) -> dict[str, int]:
    row = conn.execute(
        """
        SELECT
          COUNT(*) AS total,
          SUM(record_type = 'quote') AS quotes,
          SUM(record_type = 'order') AS orders,
          COUNT(DISTINCT customer_name) AS customers
        FROM customer_price_records
        """
    ).fetchone()
    return {
        "total": row["total"] or 0,
        "quotes": row["quotes"] or 0,
        "orders": row["orders"] or 0,
        "customers": row["customers"] or 0,
    }


def save_alias(conn: sqlite3.Connection, source_code: str, bld_no: str, note: str = "", actor: str = "") -> None:
    timestamp = now_text()
    key = normalize_code(source_code)
    before = conn.execute("SELECT * FROM aliases WHERE source_code = ?", (key,)).fetchone()
    conn.execute(
        """
        INSERT INTO aliases (source_code, bld_no, note, active, created_at, updated_at)
        VALUES (?, ?, ?, 1, ?, ?)
        ON CONFLICT(source_code) DO UPDATE SET
          bld_no=excluded.bld_no,
          note=excluded.note,
          active=1,
          updated_at=excluded.updated_at
        """,
        (key, compact_text(bld_no), compact_text(note), timestamp, timestamp),
    )
    action = "新增人工映射" if before is None else "编辑人工映射"
    log_event(conn, action, "alias", key, f"{key} -> {compact_text(bld_no)}", actor=actor)
    conn.commit()


def append_product_code(conn: sqlite3.Connection, bld_no: str, code_value: str, target: str = "oe", actor: str = "") -> bool:
    product = conn.execute("SELECT * FROM products WHERE bld_no = ?", (compact_text(bld_no),)).fetchone()
    if not product:
        return False

    code = compact_text(code_value)
    if not code:
        return False
    field = "oe_no_2" if target == "brand_code" else "oe_no_1"
    label = "品牌号码" if target == "brand_code" else "OE 号"

    existing = [line for line in clean_oe_list(product[field]).split("\n") if line]
    existing_keys = {normalize_code(line) for line in existing}
    if normalize_code(code) in existing_keys:
        return False

    updated = "\n".join(existing + [code])
    conn.execute(
        f"UPDATE products SET {field} = ?, updated_at = ? WHERE id = ?",
        (updated, now_text(), product["id"]),
    )
    log_event(conn, f"追加{label}", "product", product["bld_no"], f"{label}新增: {code}", actor=actor)
    conn.commit()
    return True


def append_product_oe(conn: sqlite3.Connection, bld_no: str, oe_code: str, actor: str = "") -> bool:
    return append_product_code(conn, bld_no, oe_code, target="oe", actor=actor)


def delete_alias(conn: sqlite3.Connection, source_code: str, actor: str = "") -> None:
    key = normalize_code(source_code)
    conn.execute("UPDATE aliases SET active = 0, updated_at = ? WHERE source_code = ?", (now_text(), key))
    log_event(conn, "删除人工映射", "alias", key, "", actor=actor)
    conn.commit()


def list_aliases(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute("SELECT * FROM aliases WHERE active = 1 ORDER BY source_code").fetchall()


def rows_for_catalog(conn: sqlite3.Connection) -> tuple[list[dict], dict[str, str]]:
    products = []
    for row in conn.execute("SELECT * FROM products WHERE active = 1 ORDER BY bld_no COLLATE BLD_NATURAL"):
        products.append(
            {
                "BLD NO.": row["bld_no"],
                "SERIES": row["series"],
                "ITEM": row["item"],
                "OE NO.1": row["oe_no_1"],
                "OE NO.2": row["oe_no_2"],
                "Models": row["models"],
                "price_cny": row["price_cny"],
                "image_path": row["image_path"],
                "image_path_2": row["image_path_2"] if "image_path_2" in row.keys() else "",
                "image_path_3": row["image_path_3"] if "image_path_3" in row.keys() else "",
                "image_path_4": row["image_path_4"] if "image_path_4" in row.keys() else "",
                "image_path_5": row["image_path_5"] if "image_path_5" in row.keys() else "",
            }
        )
    aliases = {
        row["source_code"]: row["bld_no"]
        for row in conn.execute("SELECT source_code, bld_no FROM aliases WHERE active = 1")
    }
    return products, aliases


def _material_values_from_data(
    data: dict,
    *,
    source: str = "web",
    source_row: int = 0,
    require_detail_fields: bool = True,
) -> dict:
    model = compact_text(data.get("model"))
    if not model:
        raise ValueError("母件编码不能为空。")
    code = compact_text(data.get("code"))
    if require_detail_fields and not code:
        raise ValueError("零件编码不能为空。")
    part = compact_text(data.get("part"))
    if require_detail_fields and not part:
        raise ValueError("零件名称不能为空。")
    pieces = _parse_required_float(data.get("pieces"), "下料只数")
    if compact_text(data.get("spec_text")):
        thickness, width, length = _parse_material_spec_text(data.get("spec_text"))
    else:
        thickness = _parse_required_float(data.get("thickness"), "规格1")
        width = _parse_required_float(data.get("width"), "规格2")
        length = _parse_required_float(data.get("length"), "规格3")
    if pieces <= 0:
        raise ValueError("下料只数必须大于 0。")
    spec_text = _format_material_spec(thickness, width, length)
    return {
        "model": model,
        "code": code,
        "category": compact_text(data.get("category")),
        "car": compact_text(data.get("car")),
        "part": part,
        "spec_text": spec_text,
        "pieces": pieces,
        "thickness": thickness,
        "width": width,
        "length": length,
        "active": 1 if str(data.get("active", "1")) != "0" else 0,
        "source": source,
        "source_row": int(source_row or 0),
    }


def _material_changes(before: sqlite3.Row | None, after: dict) -> list[str]:
    labels = {
        "model": "母件编码",
        "code": "零件编码",
        "category": "类别",
        "car": "车型",
        "part": "零件名称",
        "spec_text": "规格尺寸",
        "pieces": "下料只数",
        "thickness": "规格1",
        "width": "规格2",
        "length": "规格3",
        "active": "状态",
    }
    if before is None:
        return ["新增材料明细"]
    changes = []
    for field, label in labels.items():
        old = str(before[field] or "")
        new = str(after[field] or "")
        if old != new:
            changes.append(f"{label}: {old or '(空)'} -> {new or '(空)'}")
    return changes


def upsert_material_item(conn: sqlite3.Connection, data: dict, actor: str = "", source: str = "web") -> int:
    timestamp = now_text()
    item_id = data.get("id")
    before = get_material_item(conn, int(item_id)) if item_id else None
    values = _material_values_from_data(data, source=source, source_row=before["source_row"] if before else 0)
    values.update({"updated_at": timestamp})

    if before:
        values["id"] = int(item_id)
        conn.execute(
            """
            UPDATE material_items
            SET model=:model, code=:code, category=:category, car=:car, part=:part,
                spec_text=:spec_text, pieces=:pieces, thickness=:thickness, width=:width,
                length=:length, active=:active, source=:source, updated_at=:updated_at
            WHERE id=:id
            """,
            values,
        )
        item_key = f"{values['model']}-{values['code'] or values['part'] or values['id']}"
        changes = _material_changes(before, values)
        if changes:
            log_event(conn, "编辑材料明细", "material_item", item_key, "\n".join(changes[:20]), actor=actor)
        saved_id = int(item_id)
    else:
        values.update({"created_at": timestamp})
        cursor = conn.execute(
            """
            INSERT INTO material_items
              (model, code, category, car, part, spec_text, pieces, thickness, width, length,
               active, source, source_row, created_at, updated_at)
            VALUES
              (:model, :code, :category, :car, :part, :spec_text, :pieces, :thickness, :width, :length,
               :active, :source, :source_row, :created_at, :updated_at)
            """,
            values,
        )
        saved_id = int(cursor.lastrowid)
        item_key = f"{values['model']}-{values['code'] or values['part'] or saved_id}"
        log_event(conn, "新增材料明细", "material_item", item_key, "新增材料明细", actor=actor)
    conn.commit()
    return saved_id


def import_materials_from_excel(conn: sqlite3.Connection, material_path: Path, replace: bool = True, actor: str = "") -> int:
    from openpyxl import load_workbook

    wb = load_workbook(material_path, read_only=True, data_only=True)
    try:
        if "材料数据" not in wb.sheetnames:
            raise ValueError("材料数据文件里找不到工作表：材料数据")
        ws = wb["材料数据"]
        timestamp = now_text()
        rows: list[dict] = []
        for row_number, values in enumerate(ws.iter_rows(min_row=2, max_col=11, values_only=True), start=2):
            data = {
                "model": values[0],
                "code": values[1],
                "category": values[2],
                "car": values[3],
                "part": values[4],
                "spec_text": "",
                "pieces": values[6],
                "thickness": values[8],
                "width": values[9],
                "length": values[10],
                "active": 1,
            }
            if not compact_text(data["model"]):
                continue
            if any(value in (None, "") for value in [data["pieces"], data["thickness"], data["width"], data["length"]]):
                continue
            row = _material_values_from_data(
                data,
                source=material_path.name,
                source_row=row_number,
                require_detail_fields=False,
            )
            row.update({"created_at": timestamp, "updated_at": timestamp})
            rows.append(row)
    finally:
        wb.close()

    if not rows:
        raise ValueError("材料数据里没有可导入的明细。")
    if replace:
        conn.execute("DELETE FROM material_items")
    conn.executemany(
        """
        INSERT INTO material_items
          (model, code, category, car, part, spec_text, pieces, thickness, width, length,
           active, source, source_row, created_at, updated_at)
        VALUES
          (:model, :code, :category, :car, :part, :spec_text, :pieces, :thickness, :width, :length,
           :active, :source, :source_row, :created_at, :updated_at)
        """,
        rows,
    )
    log_event(conn, "导入材料数据", "material_data", material_path.name, f"导入 {len(rows)} 行材料明细", actor=actor)
    conn.commit()
    return len(rows)


def bootstrap_materials_from_excel(db_path: Path, material_path: Path) -> None:
    key = _bootstrap_key("materials", db_path, material_path)
    if key in _BOOTSTRAPPED_SOURCES:
        return
    with connect(db_path) as conn:
        existing = conn.execute("SELECT COUNT(*) FROM material_items").fetchone()[0]
        if existing == 0 and material_path.exists():
            import_materials_from_excel(conn, material_path, replace=True, actor="system")
    _BOOTSTRAPPED_SOURCES.add(key)


def list_material_items(
    conn: sqlite3.Connection,
    query: str = "",
    include_inactive: bool = False,
    only_inactive: bool = False,
    limit: int = 3000,
) -> list[sqlite3.Row]:
    sql = """
        SELECT *,
               (width * length * 7.85 * thickness / pieces / 1000000.0) AS unit_weight
        FROM material_items
    """
    params: list[object] = []
    clauses = []
    if only_inactive:
        clauses.append("active = 0")
    elif not include_inactive:
        clauses.append("active = 1")
    if query.strip():
        key = f"%{query.strip()}%"
        search_clauses = [
            "(model LIKE ? OR code LIKE ? OR category LIKE ? OR car LIKE ? OR part LIKE ? OR spec_text LIKE ?)"
        ]
        search_params: list[object] = [key, key, key, key, key, key]
        spec_clause, spec_params = _material_spec_search(_parse_material_spec_query(query))
        if spec_clause:
            search_clauses.append(spec_clause)
            search_params.extend(spec_params)
        clauses.append(
            "(" + " OR ".join(search_clauses) + ")"
        )
        params.extend(search_params)
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    sql += " ORDER BY model, code, id LIMIT ?"
    params.append(limit)
    return conn.execute(sql, params).fetchall()


def get_material_item(conn: sqlite3.Connection, item_id: int) -> sqlite3.Row | None:
    return conn.execute("SELECT * FROM material_items WHERE id = ?", (item_id,)).fetchone()


def deactivate_material_item(conn: sqlite3.Connection, item_id: int, actor: str = "") -> None:
    row = get_material_item(conn, item_id)
    conn.execute("UPDATE material_items SET active = 0, updated_at = ? WHERE id = ?", (now_text(), item_id))
    if row:
        key = f"{row['model']}-{row['code'] or row['part'] or row['id']}"
        log_event(conn, "停用材料明细", "material_item", key, "状态: 启用 -> 停用", actor=actor)
    conn.commit()


def material_item_stats(conn: sqlite3.Connection) -> dict[str, int]:
    row = conn.execute(
        "SELECT COUNT(*) AS total, SUM(active = 1) AS active, SUM(active = 0) AS inactive FROM material_items"
    ).fetchone()
    model_count = conn.execute("SELECT COUNT(DISTINCT model) FROM material_items WHERE active = 1").fetchone()[0]
    return {
        "items": row["total"] or 0,
        "active": row["active"] or 0,
        "inactive": row["inactive"] or 0,
        "models": model_count or 0,
    }


def rows_for_material_sheet(conn: sqlite3.Connection) -> dict[str, list[dict]]:
    rows_by_model: dict[str, list[dict]] = defaultdict(list)
    for row in conn.execute("SELECT * FROM material_items WHERE active = 1 ORDER BY model, code, id"):
        rows_by_model[row["model"]].append(
            {
                "source_row": row["source_row"],
                "model": row["model"],
                "code": row["code"],
                "category": row["category"],
                "car": row["car"],
                "part": row["part"],
                "spec_text": row["spec_text"],
                "pieces": float(row["pieces"]),
                "thickness": float(row["thickness"]),
                "width": float(row["width"]),
                "length": float(row["length"]),
            }
        )
    return rows_by_model


def list_audit_logs(conn: sqlite3.Connection, query: str = "", actor: str = "", limit: int = 300) -> list[sqlite3.Row]:
    sql = "SELECT * FROM audit_logs"
    params: list[object] = []
    clauses = []
    if query.strip():
        key = f"%{query.strip()}%"
        clauses.append("(target_key LIKE ? OR detail LIKE ? OR action LIKE ? OR actor LIKE ?)")
        params.extend([key, key, key, key])
    if actor.strip():
        clauses.append("actor = ?")
        params.append(actor.strip())
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    sql += " ORDER BY id DESC LIMIT ?"
    params.append(limit)
    return conn.execute(sql, params).fetchall()


def list_log_actors(conn: sqlite3.Connection) -> list[str]:
    return [
        row["actor"]
        for row in conn.execute(
            "SELECT DISTINCT actor FROM audit_logs WHERE actor IS NOT NULL AND actor != '' ORDER BY actor"
        )
    ]


def ensure_default_admin(
    conn: sqlite3.Connection,
    username: str | None = None,
    password: str | None = None,
) -> None:
    # 默认值从配置中读取,允许通过环境变量在首启前覆盖。
    # 已经存在的管理员不会被这里改密。
    from .config import DEFAULT_ADMIN_PASSWORD, DEFAULT_ADMIN_USERNAME

    username = username or DEFAULT_ADMIN_USERNAME
    password = password or DEFAULT_ADMIN_PASSWORD
    existing = conn.execute("SELECT id, password_hash FROM users WHERE username = ?", (username,)).fetchone()
    if existing:
        if str(existing["password_hash"] or "").startswith("scrypt:"):
            conn.execute(
                "UPDATE users SET password_hash = ?, updated_at = ? WHERE id = ?",
                (hash_password(password), now_text(), existing["id"]),
            )
            log_event(conn, "迁移管理员密码", "user", username, "切换为兼容的密码哈希算法", actor="system")
            conn.commit()
        return
    timestamp = now_text()
    conn.execute(
        """
        INSERT INTO users (username, display_name, password_hash, role, active, created_at, updated_at)
        VALUES (?, ?, ?, 'admin', 1, ?, ?)
        """,
        (username, "管理员", hash_password(password), timestamp, timestamp),
    )
    log_event(conn, "初始化管理员", "user", username, "创建默认管理员账号", actor="system")
    conn.commit()


def get_user_by_username(conn: sqlite3.Connection, username: str) -> sqlite3.Row | None:
    return conn.execute("SELECT * FROM users WHERE username = ?", (username.strip(),)).fetchone()


def get_user(conn: sqlite3.Connection, user_id: int) -> sqlite3.Row | None:
    return conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()


def list_users(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute("SELECT * FROM users ORDER BY active DESC, username").fetchall()


def save_user(conn: sqlite3.Connection, data: dict, actor: str = "") -> None:
    timestamp = now_text()
    user_id = data.get("id")
    username = compact_text(data.get("username"))
    if not username:
        raise ValueError("登录名不能为空。")
    role = data.get("role") or "viewer"
    if role not in {"admin", "editor", "user", "viewer"}:
        raise ValueError("角色无效。")
    active = 1 if str(data.get("active", "1")) != "0" else 0
    display_name = compact_text(data.get("display_name"))
    password = str(data.get("password") or "")

    if user_id:
        before = get_user(conn, int(user_id))
        if not before:
            raise ValueError("用户不存在。")
        params = {
            "id": int(user_id),
            "username": username,
            "display_name": display_name,
            "role": role,
            "active": active,
            "updated_at": timestamp,
        }
        password_sql = ""
        if password:
            params["password_hash"] = hash_password(password)
            password_sql = ", password_hash=:password_hash"
        conn.execute(
            f"""
            UPDATE users
            SET username=:username, display_name=:display_name, role=:role, active=:active,
                updated_at=:updated_at {password_sql}
            WHERE id=:id
            """,
            params,
        )
        changes = []
        for field, label in {"username": "登录名", "display_name": "显示名", "role": "角色", "active": "状态"}.items():
            if str(before[field] or "") != str(params[field] or ""):
                changes.append(f"{label}: {before[field]} -> {params[field]}")
        if password:
            changes.append("密码已重置")
        if changes:
            log_event(conn, "编辑账号", "user", username, "\n".join(changes), actor=actor)
    else:
        if not password:
            raise ValueError("新增用户必须设置密码。")
        conn.execute(
            """
            INSERT INTO users (username, display_name, password_hash, role, active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (username, display_name, hash_password(password), role, active, timestamp, timestamp),
        )
        log_event(conn, "新增账号", "user", username, f"角色: {role}", actor=actor)
    conn.commit()

from __future__ import annotations

import json
import math
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


CATALOG_HEADER_ALIASES = {
    "BLD NO.": {"BLD NO.", "BLD NO", "BLDNO", "BOLAIDE NO", "PART NO"},
    "OE NO.1": {"OE NO.1", "OE NO1", "OE1", "OE NO.", "OE NO", "OE REFERENCE", "OE REF", "OE 号", "OE号"},
    "OE NO.2": {"OE NO.2", "OE NO2", "OE2", "OTHER REFERENCE", "OTHER REF", "BRAND NO", "BRAND NUMBER", "品牌号码", "品牌号"},
    "SERIES": {"SERIES", "BRAND", "品牌"},
    "ITEM": {"ITEM", "DESCRIPTION", "PRODUCT", "产品名称", "品名"},
    "Models": {"MODELS", "MODEL", "APPLICATION", "车型", "适用车型"},
}

LOOKALIKE_TRANSLATION = str.maketrans(
    {
        "А": "A",
        "В": "B",
        "Е": "E",
        "К": "K",
        "М": "M",
        "Н": "H",
        "О": "O",
        "Р": "P",
        "С": "C",
        "Т": "T",
        "У": "Y",
        "Х": "X",
        "а": "A",
        "в": "B",
        "е": "E",
        "к": "K",
        "м": "M",
        "н": "H",
        "о": "O",
        "р": "P",
        "с": "C",
        "т": "T",
        "у": "Y",
        "х": "X",
    }
)

PSA_352X_BRANDS = ("PEUGEOT", "CITROEN", "CITROËN", "PSA", "标致", "雪铁龙")


@dataclass(frozen=True)
class CatalogMatch:
    bld_no: str
    score: int
    reason: str
    row: dict
    matched_codes: tuple[str, ...] = ()
    unmatched_codes: tuple[str, ...] = ()


def normalize_code(value: object) -> str:
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        value = int(value)
    text = "" if value is None else str(value).strip()
    if isinstance(value, str) and re.fullmatch(r"\d+(?:\.0+|(?:\.\d+)?[Ee][+-]?\d+)", text):
        try:
            numeric_value = Decimal(text)
        except InvalidOperation:
            pass
        else:
            if numeric_value == numeric_value.to_integral_value():
                text = str(int(numeric_value))
    text = text.translate(LOOKALIKE_TRANSLATION)
    return re.sub(r"[^A-Z0-9]", "", text.upper())


def zero_o_key(value: object) -> str:
    return normalize_code(value).replace("O", "0")


def psa_352x_key(value: object) -> tuple[str, bool] | None:
    text = compact_text(value).translate(LOOKALIKE_TRANSLATION).upper()
    match = re.fullmatch(r"(352[01])\s*[.．]\s*([A-Z0-9]{2})", text)
    if match:
        return f"{match.group(1)}{match.group(2)}", True

    key = normalize_code(value)
    if re.fullmatch(r"352[01][A-Z0-9]{2}", key):
        return key, False
    return None


def split_codes(value: object) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if not text:
        return []
    parts = re.split(r"[\n\r\t,;/，；、]+", text)
    return [part.strip() for part in parts if normalize_code(part)]


def brand_code_aliases(code: str) -> list[str]:
    aliases = [code]
    text = str(code).strip()
    if not text:
        return []

    colon_parts = [part.strip() for part in re.split(r"[:：]", text)[1:] if normalize_code(part)]
    aliases.extend(colon_parts)

    for alias in list(aliases):
        key = normalize_code(alias)
        match = re.fullmatch(r"[A-Z]{1,4}(\d{4,})", key)
        if match:
            aliases.append(match.group(1))

    unique = []
    seen = set()
    for alias in aliases:
        key = normalize_code(alias)
        if key and key not in seen:
            seen.add(key)
            unique.append(alias)
    return unique


def compact_text(value: object) -> str:
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", "" if value is None else str(value)).strip()


def _canonical_header(value: object) -> str:
    return re.sub(r"[^A-Z0-9\u4E00-\u9FFF]+", "", "" if value is None else str(value).upper())


def _find_header_row(rows: list[tuple], max_scan: int = 20) -> tuple[int, list[str]]:
    alias_lookup = {
        _canonical_header(alias): canonical
        for canonical, aliases in CATALOG_HEADER_ALIASES.items()
        for alias in aliases
    }

    best_index = -1
    best_headers: list[str] = []
    best_score = 0
    for index, row in enumerate(rows[:max_scan]):
        headers: list[str] = []
        score = 0
        for cell in row:
            canonical = alias_lookup.get(_canonical_header(cell), compact_text(cell))
            headers.append(canonical)
            if canonical in CATALOG_HEADER_ALIASES:
                score += 1
        if score > best_score:
            best_index = index
            best_headers = headers
            best_score = score

    if best_score < 3:
        raise ValueError("没有在产品目录里找到可识别的表头，需要包含 BLD NO. 和 OE NO. 等字段。")
    return best_index, best_headers


class ProductCatalog:
    def __init__(self, rows: list[dict], manual_map: dict[str, str] | None = None):
        self.rows = rows
        self.manual_map = manual_map or {}
        self.by_bld: dict[str, dict] = {}
        self.by_oe: dict[str, list[dict]] = {}
        self.by_oe_zero_o: dict[str, list[dict]] = {}
        self.by_oe_fields: dict[str, set[str]] = {}
        self.by_oe_zero_o_fields: dict[str, set[str]] = {}
        self.by_psa_352x: dict[str, list[dict]] = {}

        for row in rows:
            bld_no = compact_text(row.get("BLD NO."))
            if bld_no:
                self.by_bld[normalize_code(bld_no)] = row
            for field in ("OE NO.1", "OE NO.2"):
                for code in split_codes(row.get(field)):
                    aliases = brand_code_aliases(code) if field == "OE NO.2" else [code]
                    for alias in aliases:
                        code_key = normalize_code(alias)
                        tolerant_key = zero_o_key(alias)
                        self.by_oe.setdefault(code_key, []).append(row)
                        self.by_oe_zero_o.setdefault(tolerant_key, []).append(row)
                        self.by_oe_fields.setdefault(code_key, set()).add(field)
                        self.by_oe_zero_o_fields.setdefault(tolerant_key, set()).add(field)
                        psa_key = psa_352x_key(alias)
                        if psa_key and self._row_is_psa_352x(row):
                            self.by_psa_352x.setdefault(psa_key[0], []).append(row)

    @classmethod
    def from_excel(cls, path: Path, manual_map: dict[str, str] | None = None) -> "ProductCatalog":
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            raw_rows = list(worksheet.iter_rows(values_only=True))
            header_index, headers = _find_header_row(raw_rows)

            rows: list[dict] = []
            for raw in raw_rows[header_index + 1 :]:
                row = {}
                for header, value in zip(headers, raw):
                    if header:
                        row[header] = value
                if compact_text(row.get("BLD NO.")):
                    rows.append(row)
            return cls(rows, manual_map=manual_map)
        finally:
            workbook.close()

    def match(self, inquiry_name: object, inquiry_oe: object, inquiry_desc: object = "") -> CatalogMatch | None:
        oe_key = normalize_code(inquiry_oe)
        name_key = normalize_code(inquiry_name)
        oe_parts = split_codes(inquiry_oe)

        manual_bld = self.manual_map.get(oe_key) or self.manual_map.get(name_key)
        if manual_bld:
            row = self.by_bld.get(normalize_code(manual_bld))
            if row:
                matched_codes = tuple(oe_parts) if len(oe_parts) > 1 else ((compact_text(inquiry_oe),) if compact_text(inquiry_oe) else ())
                return CatalogMatch(compact_text(row.get("BLD NO.")), 100, "人工确认映射", row, matched_codes=matched_codes)

        if oe_key and oe_key in self.by_bld:
            row = self.by_bld[oe_key]
            return CatalogMatch(compact_text(row.get("BLD NO.")), 92, "BLD NO. 精准命中", row, matched_codes=((compact_text(inquiry_oe),) if compact_text(inquiry_oe) else ()))

        split_matches = self._match_split_oe_parts(oe_parts)
        if split_matches:
            return split_matches

        psa_probe = psa_352x_key(inquiry_oe)
        psa_match = self._match_psa_352x(inquiry_oe)
        if psa_match:
            return psa_match
        if psa_probe and psa_probe[1]:
            return None

        if oe_key and oe_key in self.by_oe:
            row = self.by_oe[oe_key][0]
            return CatalogMatch(compact_text(row.get("BLD NO.")), 95, self._exact_reason(oe_key), row, matched_codes=((compact_text(inquiry_oe),) if compact_text(inquiry_oe) else ()))

        prefix_match = self._match_unique_oe_prefix(oe_key, inquiry_oe)
        if prefix_match:
            return prefix_match

        suffix_variant_match = self._match_oe_suffix_variant(oe_key, inquiry_oe)
        if suffix_variant_match:
            return suffix_variant_match

        oe_zero_o_key = zero_o_key(inquiry_oe)
        if oe_zero_o_key and oe_zero_o_key != oe_key and oe_zero_o_key in self.by_oe_zero_o:
            rows = self._unique_rows(self.by_oe_zero_o[oe_zero_o_key])
            if len(rows) == 1:
                row = rows[0]
                return CatalogMatch(compact_text(row.get("BLD NO.")), 88, self._tolerant_reason(oe_zero_o_key), row, matched_codes=((compact_text(inquiry_oe),) if compact_text(inquiry_oe) else ()))

        if name_key and name_key in self.by_bld:
            row = self.by_bld[name_key]
            return CatalogMatch(compact_text(row.get("BLD NO.")), 92, "BLD NO. 精准命中", row)

        return None

    def _match_unique_oe_prefix(self, key: str, inquiry_oe: object) -> CatalogMatch | None:
        if len(key) < 5:
            return None

        rows: list[dict] = []
        fields: set[str] = set()
        for code_key, code_rows in self.by_oe.items():
            if not code_key.startswith(key) or code_key == key:
                continue
            rows.extend(code_rows)
            fields.update(self.by_oe_fields.get(code_key, set()))

        unique_rows = self._unique_rows(rows)
        if len(unique_rows) != 1:
            return None

        row = unique_rows[0]
        reason = "品牌号码组合前缀命中" if fields == {"OE NO.2"} else "OE 组合前缀命中"
        return CatalogMatch(
            compact_text(row.get("BLD NO.")),
            96,
            reason,
            row,
            matched_codes=((compact_text(inquiry_oe),) if compact_text(inquiry_oe) else ()),
        )

    def _match_oe_suffix_variant(self, key: str, inquiry_oe: object) -> CatalogMatch | None:
        suffix_match = re.fullmatch(r"(.+\d)[A-Z]+", key)
        if not suffix_match:
            return None

        base_key = suffix_match.group(1)
        if len(base_key) < 5:
            return None

        rows: list[dict] = []
        fields: set[str] = set()
        for code_key, code_rows in self.by_oe.items():
            if code_key.startswith(base_key):
                rows.extend(code_rows)
                fields.update(self.by_oe_fields.get(code_key, set()))

        unique_rows = self._unique_rows(rows)
        if len(unique_rows) != 1:
            return None

        row = unique_rows[0]
        reason = "品牌号码尾字母容错命中" if fields == {"OE NO.2"} else "OE 尾字母容错命中"
        return CatalogMatch(
            compact_text(row.get("BLD NO.")),
            89,
            reason,
            row,
            matched_codes=((compact_text(inquiry_oe),) if compact_text(inquiry_oe) else ()),
        )

    def _match_psa_352x(self, inquiry_oe: object) -> CatalogMatch | None:
        probe = psa_352x_key(inquiry_oe)
        if not probe:
            return None

        key, has_separator = probe
        psa_rows = self._unique_rows(self.by_psa_352x.get(key, []))
        exact_rows = self._unique_rows(self.by_oe.get(key, []))
        non_psa_exact_rows = [row for row in exact_rows if not self._row_is_psa_352x(row)]
        matched_codes = ((compact_text(inquiry_oe),) if compact_text(inquiry_oe) else ())

        if has_separator:
            if len(psa_rows) == 1:
                row = psa_rows[0]
                return CatalogMatch(compact_text(row.get("BLD NO.")), 97, "PSA 号码点号容错命中", row, matched_codes=matched_codes)
            if len(psa_rows) > 1:
                return self._ambiguous_match(psa_rows, 80, "PSA 号码命中多个 BLD，请人工确认", matched_codes)
            return None

        if psa_rows and non_psa_exact_rows:
            combined_rows = self._unique_rows(psa_rows + non_psa_exact_rows)
            if len(combined_rows) == 1:
                row = combined_rows[0]
                return CatalogMatch(compact_text(row.get("BLD NO.")), 97, "PSA 号码点号容错命中", row, matched_codes=matched_codes)
            return self._ambiguous_match(combined_rows, 80, "3520/3521 号码同时命中 PSA 与其他品牌，请人工确认", matched_codes)

        if len(psa_rows) == 1:
            row = psa_rows[0]
            return CatalogMatch(compact_text(row.get("BLD NO.")), 97, "PSA 号码点号容错命中", row, matched_codes=matched_codes)
        if len(psa_rows) > 1:
            return self._ambiguous_match(psa_rows, 80, "PSA 号码命中多个 BLD，请人工确认", matched_codes)
        return None

    def _match_split_oe_parts(self, oe_parts: list[str]) -> CatalogMatch | None:
        if len(oe_parts) <= 1:
            return None

        matches: list[tuple[str, CatalogMatch]] = []
        for part in oe_parts:
            part_key = normalize_code(part)
            manual_bld = self.manual_map.get(part_key)
            if manual_bld:
                row = self.by_bld.get(normalize_code(manual_bld))
                if row:
                    matches.append((part, CatalogMatch(compact_text(row.get("BLD NO.")), 100, "多号码人工确认映射", row)))
                    continue

            if part_key in self.by_bld:
                row = self.by_bld[part_key]
                matches.append((part, CatalogMatch(compact_text(row.get("BLD NO.")), 92, "BLD NO. 多号码精准命中", row)))
                continue

            psa_probe = psa_352x_key(part)
            psa_match = self._match_psa_352x(part)
            if psa_match:
                matches.append((part, psa_match))
                continue
            if psa_probe and psa_probe[1]:
                continue

            if part_key in self.by_oe:
                row = self.by_oe[part_key][0]
                matches.append((part, CatalogMatch(compact_text(row.get("BLD NO.")), 95, self._multi_exact_reason(part_key), row)))
                continue

            part_zero_o_key = zero_o_key(part)
            if part_zero_o_key and part_zero_o_key != part_key and part_zero_o_key in self.by_oe_zero_o:
                rows = self._unique_rows(self.by_oe_zero_o[part_zero_o_key])
                if len(rows) == 1:
                    row = rows[0]
                    matches.append((part, CatalogMatch(compact_text(row.get("BLD NO.")), 88, self._multi_tolerant_reason(part_zero_o_key), row)))

        unique: dict[str, CatalogMatch] = {}
        matched_parts = tuple(part for part, _ in matches)
        matched_keys = {normalize_code(part) for part in matched_parts}
        unmatched_parts = tuple(part for part in oe_parts if normalize_code(part) not in matched_keys)
        for _, match in matches:
            unique.setdefault(match.bld_no, match)

        if not unique:
            return None
        if len(unique) == 1:
            match = next(iter(unique.values()))
            return CatalogMatch(match.bld_no, match.score, match.reason, match.row, matched_codes=matched_parts, unmatched_codes=unmatched_parts)

        bld_list = " / ".join(unique)
        first = next(iter(unique.values()))
        return CatalogMatch(bld_list, 80, "多个号码命中不同 BLD，请人工确认", first.row, matched_codes=matched_parts, unmatched_codes=unmatched_parts)

    def _ambiguous_match(self, rows: list[dict], score: int, reason: str, matched_codes: tuple[str, ...]) -> CatalogMatch | None:
        unique_rows = self._unique_rows(rows)
        if not unique_rows:
            return None
        bld_list = " / ".join(compact_text(row.get("BLD NO.")) for row in unique_rows)
        return CatalogMatch(bld_list, score, reason, unique_rows[0], matched_codes=matched_codes)

    @staticmethod
    def _unique_rows(rows: list[dict]) -> list[dict]:
        seen = set()
        unique = []
        for row in rows:
            key = compact_text(row.get("BLD NO."))
            if key and key not in seen:
                seen.add(key)
                unique.append(row)
        return unique

    @staticmethod
    def _row_is_psa_352x(row: dict) -> bool:
        text = f"{compact_text(row.get('SERIES'))} {compact_text(row.get('Models'))}".upper()
        return any(brand in text for brand in PSA_352X_BRANDS)

    def _is_brand_code_key(self, key: str, *, tolerant: bool = False) -> bool:
        fields = self.by_oe_zero_o_fields if tolerant else self.by_oe_fields
        return fields.get(key) == {"OE NO.2"}

    def _exact_reason(self, key: str) -> str:
        return "品牌号码精准命中" if self._is_brand_code_key(key) else "OE 精准命中"

    def _tolerant_reason(self, key: str) -> str:
        return "品牌号码字符容错命中" if self._is_brand_code_key(key, tolerant=True) else "OE 字符容错命中"

    def _multi_exact_reason(self, key: str) -> str:
        return "品牌号码多号码精准命中" if self._is_brand_code_key(key) else "OE 多号码精准命中"

    def _multi_tolerant_reason(self, key: str) -> str:
        return "品牌号码多号码字符容错命中" if self._is_brand_code_key(key, tolerant=True) else "OE 多号码字符容错命中"


def load_manual_map(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    return {normalize_code(key): compact_text(value) for key, value in data.items() if compact_text(value)}


def save_manual_map(path: Path, mapping: dict[str, str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    clean = {normalize_code(key): compact_text(value) for key, value in mapping.items() if normalize_code(key) and compact_text(value)}
    with path.open("w", encoding="utf-8") as handle:
        json.dump(clean, handle, ensure_ascii=False, indent=2, sort_keys=True)


def catalog_summary(catalog: ProductCatalog) -> dict[str, int]:
    return {
        "products": len(catalog.rows),
        "oe_keys": len(catalog.by_oe),
        "bld_keys": len(catalog.by_bld),
        "manual_mappings": len(catalog.manual_map),
    }

from __future__ import annotations

import sqlite3
import threading
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.database import connect
from app.matcher import compact_text
from app.platform.audit_store import log_event
from app.platform.clock import now_text
from app.platform.sync_identity import material_key, stable_sync_id

from .specification import _material_values_from_data


_BOOTSTRAP_LOCK = threading.Lock()
_BOOTSTRAPPED_SOURCES: set[tuple[Path, Path, int | None]] = set()


def import_materials_from_excel(
    connection: sqlite3.Connection,
    material_path: Path,
    replace: bool = True,
    actor: str = "",
    *,
    commit: bool = True,
) -> int:
    workbook = load_workbook(material_path, read_only=True, data_only=True)
    try:
        if "材料数据" not in workbook.sheetnames:
            raise ValueError("材料数据文件里找不到工作表：材料数据")
        sheet = workbook["材料数据"]
        timestamp = now_text()
        rows: list[dict[str, Any]] = []
        ordinals: dict[str, int] = {}
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, max_col=11, values_only=True),
            start=2,
        ):
            data = {
                "model": values[0],
                "code": values[1],
                "category": values[2],
                "car": values[3],
                "part": values[4],
                "spec_text": "",
                "pieces": values[6],
                "thickness": values[8],
                "width": values[9],
                "length": values[10],
                "active": 1,
            }
            if not compact_text(data["model"]):
                continue
            if any(data[field] in (None, "") for field in ("pieces", "thickness", "width", "length")):
                continue
            row = _material_values_from_data(
                data,
                source=material_path.name,
                source_row=row_number,
                require_detail_fields=False,
            )
            key = material_key(row)
            ordinal = ordinals.get(key, 0) + 1
            ordinals[key] = ordinal
            row.update({"sync_id": stable_sync_id("material", key, ordinal), "created_at": timestamp, "updated_at": timestamp})
            rows.append(row)
    finally:
        workbook.close()
    if not rows:
        raise ValueError("材料数据里没有可导入的明细。")
    if replace:
        connection.execute("DELETE FROM material_items")
    connection.executemany(
        """
        INSERT INTO material_items
          (model, code, category, car, part, spec_text, pieces, thickness, width, length,
           active, source, source_row, sync_id, created_at, updated_at)
        VALUES
          (:model, :code, :category, :car, :part, :spec_text, :pieces, :thickness, :width,
           :length, :active, :source, :source_row, :sync_id, :created_at, :updated_at)
        """,
        rows,
    )
    log_event(
        connection,
        "导入材料数据",
        "material_data",
        material_path.name,
        f"导入 {len(rows)} 行材料明细",
        actor=actor,
    )
    if commit:
        connection.commit()
    return len(rows)


def bootstrap_materials_from_excel(database_path: Path, material_path: Path) -> None:
    source_mtime = material_path.stat().st_mtime_ns if material_path.exists() else None
    key = (database_path.resolve(), material_path.resolve(), source_mtime)
    with _BOOTSTRAP_LOCK:
        if key in _BOOTSTRAPPED_SOURCES:
            return
        with connect(database_path) as connection:
            row = connection.execute("SELECT COUNT(*) FROM material_items").fetchone()
            existing = int(row[0]) if row is not None else 0
            if existing == 0 and material_path.exists():
                import_materials_from_excel(
                    connection,
                    material_path,
                    replace=True,
                    actor="system",
                )
        _BOOTSTRAPPED_SOURCES.add(key)

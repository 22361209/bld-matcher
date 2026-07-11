from __future__ import annotations

import re
import sqlite3
import threading
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from app.database import connect
from app.matcher import compact_text
from app.platform.audit_store import log_event
from app.platform.clock import now_text


_BOOTSTRAP_LOCK = threading.Lock()
_BOOTSTRAPPED_SOURCES: set[tuple[Path, Path, int | None]] = set()


def _parse_required_float(value: object, label: str) -> float:
    text = compact_text(value)
    if not text:
        raise ValueError(f"{label}不能为空。")
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"{label}必须是数字：{text}") from exc


def _format_material_int(value: float) -> int | float:
    return int(value) if float(value).is_integer() else value


def _format_material_thickness(value: float) -> str:
    text = f"{value:.2f}".rstrip("0")
    return text + "0" if text.endswith(".") else text


def _format_material_spec(thickness: float, width: float, length: float) -> str:
    return f"{_format_material_thickness(thickness)}×{_format_material_int(width)}×{_format_material_int(length)}"


def _parse_material_spec_text(value: object) -> tuple[float, float, float]:
    text = compact_text(value)
    if not text:
        raise ValueError("规格尺寸不能为空。")
    normalized = re.sub(r"[×xX*＊/／\\\-－—]+", " ", text)
    parts = re.sub(r"\s+", " ", normalized).strip().split(" ")
    if len(parts) != 3:
        raise ValueError("规格尺寸请按“厚度 宽度 长度”填写，例如 2.5 357 1260。")
    try:
        return float(parts[0]), float(parts[1]), float(parts[2])
    except ValueError as exc:
        raise ValueError(f"规格尺寸必须包含 3 个数字：{text}") from exc


def _parse_material_spec_query(value: object) -> list[float]:
    text = compact_text(value)
    if not text:
        return []
    normalized = re.sub(r"[×xX*＊/／\\\-－—]+", " ", text)
    parts = re.sub(r"\s+", " ", normalized).strip().split(" ")
    if not 1 <= len(parts) <= 3:
        return []
    try:
        return [float(part) for part in parts]
    except ValueError:
        return []


def _material_spec_search(tokens: list[float]) -> tuple[str, list[object]]:
    if not tokens:
        return "", []
    epsilon = 0.000001

    def equal(field: str) -> str:
        return f"ABS({field} - ?) < ?"

    if len(tokens) == 1:
        value = tokens[0]
        return (
            f"({equal('thickness')} OR {equal('width')} OR {equal('length')})",
            [value, epsilon, value, epsilon, value, epsilon],
        )
    if len(tokens) == 2:
        first, second = tokens
        return (
            f"(({equal('thickness')} AND {equal('width')}) OR "
            f"({equal('thickness')} AND {equal('length')}) OR "
            f"({equal('width')} AND {equal('length')}))",
            [first, epsilon, second, epsilon] * 3,
        )
    first, second, third = tokens
    return (
        f"({equal('thickness')} AND {equal('width')} AND {equal('length')})",
        [first, epsilon, second, epsilon, third, epsilon],
    )


def _material_values_from_data(
    data: dict,
    *,
    source: str = "web",
    source_row: int = 0,
    require_detail_fields: bool = True,
) -> dict:
    model = compact_text(data.get("model"))
    if not model:
        raise ValueError("母件编码不能为空。")
    code = compact_text(data.get("code"))
    if require_detail_fields and not code:
        raise ValueError("零件编码不能为空。")
    part = compact_text(data.get("part"))
    if require_detail_fields and not part:
        raise ValueError("零件名称不能为空。")
    pieces = _parse_required_float(data.get("pieces"), "下料只数")
    if compact_text(data.get("spec_text")):
        thickness, width, length = _parse_material_spec_text(data.get("spec_text"))
    else:
        thickness = _parse_required_float(data.get("thickness"), "规格1")
        width = _parse_required_float(data.get("width"), "规格2")
        length = _parse_required_float(data.get("length"), "规格3")
    if pieces <= 0:
        raise ValueError("下料只数必须大于 0。")
    return {
        "model": model,
        "code": code,
        "category": compact_text(data.get("category")),
        "car": compact_text(data.get("car")),
        "part": part,
        "spec_text": _format_material_spec(thickness, width, length),
        "pieces": pieces,
        "thickness": thickness,
        "width": width,
        "length": length,
        "active": 1 if str(data.get("active", "1")) != "0" else 0,
        "source": source,
        "source_row": int(source_row or 0),
    }


def _material_changes(before: sqlite3.Row | None, after: dict) -> list[str]:
    labels = {
        "model": "母件编码",
        "code": "零件编码",
        "category": "类别",
        "car": "车型",
        "part": "零件名称",
        "spec_text": "规格尺寸",
        "pieces": "下料只数",
        "thickness": "规格1",
        "width": "规格2",
        "length": "规格3",
        "active": "状态",
    }
    if before is None:
        return ["新增材料明细"]
    return [
        f"{label}: {str(before[field] or '') or '(空)'} -> {str(after[field] or '') or '(空)'}"
        for field, label in labels.items()
        if str(before[field] or "") != str(after[field] or "")
    ]


def get_material_item(connection: sqlite3.Connection, item_id: int) -> sqlite3.Row | None:
    return connection.execute("SELECT * FROM material_items WHERE id = ?", (item_id,)).fetchone()


def upsert_material_item(
    connection: sqlite3.Connection,
    data: dict,
    actor: str = "",
    source: str = "web",
    *,
    commit: bool = True,
) -> int:
    timestamp = now_text()
    item_id = data.get("id")
    before = get_material_item(connection, int(item_id)) if item_id else None
    values = _material_values_from_data(data, source=source, source_row=before["source_row"] if before else 0)
    values["updated_at"] = timestamp
    if before:
        values["id"] = int(item_id)
        connection.execute(
            """
            UPDATE material_items
            SET model=:model, code=:code, category=:category, car=:car, part=:part,
                spec_text=:spec_text, pieces=:pieces, thickness=:thickness, width=:width,
                length=:length, active=:active, source=:source, updated_at=:updated_at
            WHERE id=:id
            """,
            values,
        )
        saved_id = int(item_id)
        changes = _material_changes(before, values)
        if changes:
            log_event(
                connection,
                "编辑材料明细",
                "material_item",
                f"{values['model']}-{values['code'] or values['part'] or values['id']}",
                "\n".join(changes[:20]),
                actor=actor,
            )
    else:
        values["created_at"] = timestamp
        cursor = connection.execute(
            """
            INSERT INTO material_items
              (model, code, category, car, part, spec_text, pieces, thickness, width, length,
               active, source, source_row, created_at, updated_at)
            VALUES
              (:model, :code, :category, :car, :part, :spec_text, :pieces, :thickness, :width,
               :length, :active, :source, :source_row, :created_at, :updated_at)
            """,
            values,
        )
        saved_id = int(cursor.lastrowid)
        log_event(
            connection,
            "新增材料明细",
            "material_item",
            f"{values['model']}-{values['code'] or values['part'] or saved_id}",
            "新增材料明细",
            actor=actor,
        )
    if commit:
        connection.commit()
    return saved_id


def import_materials_from_excel(
    connection: sqlite3.Connection,
    material_path: Path,
    replace: bool = True,
    actor: str = "",
    *,
    commit: bool = True,
) -> int:
    workbook = load_workbook(material_path, read_only=True, data_only=True)
    try:
        if "材料数据" not in workbook.sheetnames:
            raise ValueError("材料数据文件里找不到工作表：材料数据")
        sheet = workbook["材料数据"]
        timestamp = now_text()
        rows: list[dict] = []
        for row_number, values in enumerate(sheet.iter_rows(min_row=2, max_col=11, values_only=True), start=2):
            data = {
                "model": values[0],
                "code": values[1],
                "category": values[2],
                "car": values[3],
                "part": values[4],
                "spec_text": "",
                "pieces": values[6],
                "thickness": values[8],
                "width": values[9],
                "length": values[10],
                "active": 1,
            }
            if not compact_text(data["model"]):
                continue
            if any(data[field] in (None, "") for field in ("pieces", "thickness", "width", "length")):
                continue
            row = _material_values_from_data(
                data,
                source=material_path.name,
                source_row=row_number,
                require_detail_fields=False,
            )
            row.update({"created_at": timestamp, "updated_at": timestamp})
            rows.append(row)
    finally:
        workbook.close()
    if not rows:
        raise ValueError("材料数据里没有可导入的明细。")
    if replace:
        connection.execute("DELETE FROM material_items")
    connection.executemany(
        """
        INSERT INTO material_items
          (model, code, category, car, part, spec_text, pieces, thickness, width, length,
           active, source, source_row, created_at, updated_at)
        VALUES
          (:model, :code, :category, :car, :part, :spec_text, :pieces, :thickness, :width,
           :length, :active, :source, :source_row, :created_at, :updated_at)
        """,
        rows,
    )
    log_event(
        connection,
        "导入材料数据",
        "material_data",
        material_path.name,
        f"导入 {len(rows)} 行材料明细",
        actor=actor,
    )
    if commit:
        connection.commit()
    return len(rows)


def bootstrap_materials_from_excel(database_path: Path, material_path: Path) -> None:
    source_mtime = material_path.stat().st_mtime_ns if material_path.exists() else None
    key = (database_path.resolve(), material_path.resolve(), source_mtime)
    with _BOOTSTRAP_LOCK:
        if key in _BOOTSTRAPPED_SOURCES:
            return
        with connect(database_path) as connection:
            existing = int(connection.execute("SELECT COUNT(*) FROM material_items").fetchone()[0])
            if existing == 0 and material_path.exists():
                import_materials_from_excel(connection, material_path, replace=True, actor="system")
        _BOOTSTRAPPED_SOURCES.add(key)


def _filter_clauses(
    query: str = "",
    include_inactive: bool = False,
    only_inactive: bool = False,
) -> tuple[list[str], list[object]]:
    clauses: list[str] = []
    params: list[object] = []
    if only_inactive:
        clauses.append("active = 0")
    elif not include_inactive:
        clauses.append("active = 1")
    if query.strip():
        key = f"%{query.strip()}%"
        search = ["(model LIKE ? OR code LIKE ? OR category LIKE ? OR car LIKE ? OR part LIKE ? OR spec_text LIKE ?)"]
        search_params: list[object] = [key] * 6
        spec_clause, spec_params = _material_spec_search(_parse_material_spec_query(query))
        if spec_clause:
            search.append(spec_clause)
            search_params.extend(spec_params)
        clauses.append("(" + " OR ".join(search) + ")")
        params.extend(search_params)
    return clauses, params


def count_material_items(
    connection: sqlite3.Connection,
    query: str = "",
    include_inactive: bool = False,
    only_inactive: bool = False,
) -> int:
    sql = "SELECT COUNT(*) FROM material_items"
    clauses, params = _filter_clauses(query, include_inactive, only_inactive)
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    return int(connection.execute(sql, params).fetchone()[0] or 0)


def list_material_items(
    connection: sqlite3.Connection,
    query: str = "",
    include_inactive: bool = False,
    only_inactive: bool = False,
    limit: int = 3000,
    offset: int = 0,
) -> list[sqlite3.Row]:
    sql = """
        SELECT *, (width * length * 7.85 * thickness / pieces / 1000000.0) AS unit_weight
        FROM material_items
    """
    clauses, params = _filter_clauses(query, include_inactive, only_inactive)
    if clauses:
        sql += " WHERE " + " AND ".join(clauses)
    sql += " ORDER BY model, code, id LIMIT ? OFFSET ?"
    params.extend((limit, max(0, offset)))
    return connection.execute(sql, params).fetchall()


def deactivate_material_item(
    connection: sqlite3.Connection,
    item_id: int,
    actor: str = "",
    *,
    commit: bool = True,
) -> None:
    row = get_material_item(connection, item_id)
    connection.execute("UPDATE material_items SET active = 0, updated_at = ? WHERE id = ?", (now_text(), item_id))
    if row:
        log_event(
            connection,
            "停用材料明细",
            "material_item",
            f"{row['model']}-{row['code'] or row['part'] or row['id']}",
            "状态: 启用 -> 停用",
            actor=actor,
        )
    if commit:
        connection.commit()


def material_item_stats(connection: sqlite3.Connection) -> dict[str, int]:
    row = connection.execute(
        "SELECT COUNT(*) AS total, SUM(active = 1) AS active, SUM(active = 0) AS inactive FROM material_items"
    ).fetchone()
    models = int(
        connection.execute("SELECT COUNT(DISTINCT model) FROM material_items WHERE active = 1").fetchone()[0] or 0
    )
    return {
        "items": int(row["total"] or 0),
        "active": int(row["active"] or 0),
        "inactive": int(row["inactive"] or 0),
        "models": models,
    }


def rows_for_material_sheet(connection: sqlite3.Connection) -> dict[str, list[dict]]:
    rows_by_model: dict[str, list[dict]] = defaultdict(list)
    for row in connection.execute("SELECT * FROM material_items WHERE active = 1 ORDER BY model, code, id"):
        rows_by_model[str(row["model"])].append(
            {
                "source_row": row["source_row"],
                "model": row["model"],
                "code": row["code"],
                "category": row["category"],
                "car": row["car"],
                "part": row["part"],
                "spec_text": row["spec_text"],
                "pieces": float(row["pieces"]),
                "thickness": float(row["thickness"]),
                "width": float(row["width"]),
                "length": float(row["length"]),
            }
        )
    return rows_by_model

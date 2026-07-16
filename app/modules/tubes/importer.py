from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .domain import TubeImportRow, row_from_2026


def load_tube_rows(workbook_path: Path) -> list[TubeImportRow]:
    # 源文件的“2026”工作表声明的只读维度为 A1，流式模式会漏掉实际数据。
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    try:
        if "2026" not in workbook.sheetnames:
            raise ValueError("未找到“2026”明细表。")
        sheet = workbook["2026"]
        rows: list[TubeImportRow] = []
        for row_number, cells in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
            row = row_from_2026(cells, row_number)
            if row is not None:
                rows.append(row)
        if not rows:
            raise ValueError("“2026”明细表没有可导入的管件行。")
        return rows
    finally:
        workbook.close()

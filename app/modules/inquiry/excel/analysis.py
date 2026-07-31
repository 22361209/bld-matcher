from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.matcher import ProductCatalog, normalize_code, split_codes

from ..adjustments import (
    InquiryAdjustment,
    apply_adjustment,
    ensure_adjustments_consumed,
    workbook_row_adjustment_key,
)
from .pricing import numeric_price
from .reader import (
    _cell_from_values,
    _combined_match_text,
    _customer_code_text,
    _find_xlsx_inquiry_columns_from_rows,
    _find_xlsx_selected_header_row_from_rows,
    _manual_match_columns,
    _row_match_values,
    _selected_match_row_is_header,
)


PRODUCT_IMAGE_FIELDS = ("image_path", "image_path_2", "image_path_3", "image_path_4", "image_path_5")


def annotate_row_summary_with_match_columns(
    row_summary: dict,
    match_values: list[tuple[str, object]],
    match,
) -> dict:
    if not match or not match_values or not match.matched_codes:
        return row_summary

    matched_keys = {normalize_code(code) for code in match.matched_codes if normalize_code(code)}
    if not matched_keys:
        return row_summary

    hits = []
    for label, value in match_values:
        for part in split_codes(value) or ([value] if normalize_code(value) else []):
            if normalize_code(part) in matched_keys:
                hits.append(f"{label}列：{part}")

    if not hits:
        return row_summary

    prefix = "命中列：" + "，".join(dict.fromkeys(hits))
    row_summary["match_note"] = f"{prefix}；{row_summary['match_note']}" if row_summary.get("match_note") else prefix
    return row_summary


def summary_row(
    row_number: int,
    inquiry_oe: object,
    inquiry_name: object,
    match,
    customer_product_code: str = "",
) -> dict:
    parts = split_codes(inquiry_oe)
    match_note = ""
    price_cny = None
    product_status = ""
    image_fields = {field: "" for field in PRODUCT_IMAGE_FIELDS}
    if match and " / " not in (match.bld_no or ""):
        price_cny = numeric_price(match.row.get("price_cny"))
        product_status = str(match.row.get("product_status") or "").strip()
        image_fields = {
            field: str(match.row.get(field) or "").strip()
            for field in PRODUCT_IMAGE_FIELDS
        }

    if len(parts) > 1:
        if match and match.matched_codes:
            notes = [f"命中号码：{', '.join(match.matched_codes)}"]
            if match.unmatched_codes:
                notes.append(f"未命中号码：{', '.join(match.unmatched_codes)}")
            if " / " in match.bld_no:
                notes.append(f"命中 BLD：{match.bld_no}")
            match_note = "；".join(notes)
        elif not match:
            match_note = f"多个号码均未命中：{', '.join(parts)}"

    row = {
        "row": row_number,
        "oe": inquiry_oe,
        "name": inquiry_name,
        "customer_product_code": customer_product_code,
        "bld_no": match.bld_no if match else "",
        "price_cny": price_cny,
        "product_status": product_status,
        **image_fields,
        "reason": match.reason if match else "未找到",
        "score": match.score if match else 0,
        "match_note": match_note,
        "matched_oe_codes": [],
        "unmatched_oe_codes": [],
        "adjustment_allowed": bool(match and " / " not in (match.bld_no or "")),
    }
    if match and len(parts) > 1 and match.matched_codes:
        row["matched_oe_codes"] = list(match.matched_codes)
        row["unmatched_oe_codes"] = list(match.unmatched_codes)
    elif not match and len(parts) > 1:
        row["unmatched_oe_codes"] = list(parts)
    return row


def analyze_xlsx_with_bld(
    inquiry_path: Path,
    catalog: ProductCatalog,
    match_column: object = None,
    price_mode: str = "none",
    exchange_rate: float | None = None,
    customer_code_column: int | None = None,
    adjustments: dict[str, InquiryAdjustment] | None = None,
) -> dict:
    workbook = load_workbook(inquiry_path, read_only=True, data_only=True)
    try:
        summary = {"total": 0, "matched": 0, "unmatched": 0, "rows": []}
        consumed_adjustment_keys: set[str] = set()

        for sheet_number, sheet in enumerate(workbook.worksheets, start=1):
            first_rows = list(sheet.iter_rows(min_row=1, max_row=20, values_only=True))
            if not first_rows:
                continue

            manual_match_columns = _manual_match_columns(match_column)
            if not manual_match_columns:
                header_row, columns = _find_xlsx_inquiry_columns_from_rows(first_rows)
                selected_match_columns = [columns["oe"]]
            else:
                selected_match_columns = [column + 1 for column in manual_match_columns]
                header_row = _find_xlsx_selected_header_row_from_rows(
                    first_rows,
                    selected_match_columns,
                )
                columns = {}

            rows = sheet.iter_rows(min_row=header_row + 1, values_only=True)
            for row_index, values in enumerate(rows, start=header_row + 1):
                inquiry_name = _cell_from_values(values, columns["name"]) if "name" in columns else ""
                customer_product_code = (
                    _customer_code_text(_cell_from_values(values, customer_code_column + 1))
                    if customer_code_column is not None
                    else ""
                )
                match_values = _row_match_values(values, selected_match_columns)
                if _selected_match_row_is_header(match_values):
                    continue
                inquiry_oe = _combined_match_text(match_values)
                inquiry_desc = _cell_from_values(values, columns["description"]) if "description" in columns else ""
                if not str(inquiry_name or "").strip() and not str(inquiry_oe or "").strip():
                    continue

                match = catalog.match(inquiry_name, inquiry_oe, inquiry_desc)
                adjustment_key = workbook_row_adjustment_key(sheet_number, row_index)
                adjustment = (adjustments or {}).get(adjustment_key)
                match, tax_price_override, adjustment_note = apply_adjustment(
                    catalog,
                    match,
                    adjustment,
                )
                if adjustment is not None:
                    consumed_adjustment_keys.add(adjustment_key)
                summary["total"] += 1
                if match:
                    summary["matched"] += 1
                    row_summary = annotate_row_summary_with_match_columns(
                        summary_row(row_index, inquiry_oe, inquiry_name, match, customer_product_code),
                        match_values,
                        match,
                    )
                    row_summary["adjustment_key"] = adjustment_key
                    if tax_price_override is not None:
                        row_summary["price_cny"] = float(tax_price_override)
                    if adjustment_note:
                        row_summary["match_note"] = (
                            f"{row_summary['match_note']}；{adjustment_note}"
                            if row_summary.get("match_note")
                            else adjustment_note
                        )
                    summary["rows"].append(row_summary)
                else:
                    summary["unmatched"] += 1
                    row_summary = summary_row(row_index, inquiry_oe, inquiry_name, None, customer_product_code)
                    row_summary["adjustment_key"] = adjustment_key
                    summary["rows"].append(row_summary)

        ensure_adjustments_consumed(adjustments, consumed_adjustment_keys)
        return summary
    finally:
        workbook.close()

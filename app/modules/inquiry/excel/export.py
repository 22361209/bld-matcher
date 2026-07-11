from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import xlrd
from xlutils.copy import copy as copy_xls

from app.matcher import ProductCatalog
from app.product_status import product_status_header_for_price_mode

from .analysis import (
    analyze_xlsx_with_bld,
    annotate_row_summary_with_match_columns,
    summary_row,
)
from .cleanup import trim_unused_row_dimensions
from .pricing import match_export_price, match_export_status, price_export_header
from .reader import (
    _combined_match_text,
    _find_inquiry_columns,
    _find_xls_selected_header_row,
    _find_xlsx_inquiry_columns,
    _find_xlsx_selected_header_row,
    _manual_match_columns,
    _selected_match_row_is_header,
    _xls_match_values,
    _xlsx_match_values,
)


def _write_cell(
    sheet: Any,
    row: int,
    column: int,
    value: object,
    *,
    number_format: str | None = None,
) -> None:
    cell = sheet.cell(row, column)
    cell.value = value
    if number_format is not None:
        cell.number_format = number_format


def generate_xls_with_bld(
    inquiry_path: Path,
    output_path: Path,
    catalog: ProductCatalog,
    match_column: object = None,
    write_output: bool = True,
    price_mode: str = "none",
    exchange_rate: float | None = None,
) -> dict:
    book = xlrd.open_workbook(
        str(inquiry_path),
        formatting_info=True,
        ignore_workbook_corruption=True,
    )
    writable = copy_xls(book) if write_output else None
    summary = {"total": 0, "matched": 0, "unmatched": 0, "rows": []}
    manual_match_columns = _manual_match_columns(match_column)

    for sheet_index in range(book.nsheets):
        source_sheet = book.sheet_by_index(sheet_index)
        target_sheet = writable.get_sheet(sheet_index) if writable else None
        if source_sheet.nrows == 0:
            continue

        if not manual_match_columns:
            header_row, columns = _find_inquiry_columns(source_sheet)
            selected_match_columns = [columns["oe"]]
        else:
            header_row = _find_xls_selected_header_row(source_sheet, manual_match_columns)
            columns = {}
            selected_match_columns = manual_match_columns
        output_col = source_sheet.ncols
        price_header = price_export_header(price_mode)
        status_header = product_status_header_for_price_mode(price_mode) if price_header else ""
        status_col = output_col + 2 if price_header else None
        note_col = output_col + 3 if status_header else (output_col + 2 if price_header else output_col + 1)
        if target_sheet:
            target_sheet.write(header_row, output_col, "BLD NO.")
            if price_header:
                target_sheet.write(header_row, output_col + 1, price_header)
            if status_header and status_col is not None:
                target_sheet.write(header_row, status_col, status_header)
            target_sheet.write(header_row, note_col, "匹配说明")

        for row_index in range(header_row + 1, source_sheet.nrows):
            inquiry_name = source_sheet.cell_value(row_index, columns["name"]) if "name" in columns else ""
            match_values = _xls_match_values(
                source_sheet,
                row_index,
                selected_match_columns,
            )
            if _selected_match_row_is_header(match_values):
                continue
            inquiry_oe = _combined_match_text(match_values)
            inquiry_desc = (
                source_sheet.cell_value(row_index, columns["description"]) if "description" in columns else ""
            )
            if not str(inquiry_name).strip() and not str(inquiry_oe).strip():
                continue

            match = catalog.match(inquiry_name, inquiry_oe, inquiry_desc)
            summary["total"] += 1
            if match:
                if target_sheet:
                    target_sheet.write(row_index, output_col, match.bld_no)
                    if price_header:
                        price = match_export_price(match, price_mode, exchange_rate)
                        target_sheet.write(
                            row_index,
                            output_col + 1,
                            "" if price is None else price,
                        )
                    if status_header and status_col is not None:
                        target_sheet.write(
                            row_index,
                            status_col,
                            match_export_status(match, price_mode),
                        )
                summary["matched"] += 1
                row_summary = annotate_row_summary_with_match_columns(
                    summary_row(row_index + 1, inquiry_oe, inquiry_name, match),
                    match_values,
                    match,
                )
                if target_sheet:
                    target_sheet.write(row_index, note_col, row_summary["match_note"])
                summary["rows"].append(row_summary)
            else:
                if target_sheet:
                    target_sheet.write(row_index, output_col, "")
                    if price_header:
                        target_sheet.write(row_index, output_col + 1, "")
                    if status_header and status_col is not None:
                        target_sheet.write(row_index, status_col, "")
                summary["unmatched"] += 1
                row_summary = summary_row(row_index + 1, inquiry_oe, inquiry_name, None)
                if target_sheet:
                    target_sheet.write(row_index, note_col, row_summary["match_note"])
                summary["rows"].append(row_summary)

    if writable:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        writable.save(str(output_path))
    return summary


def generate_xlsx_with_bld(
    inquiry_path: Path,
    output_path: Path,
    catalog: ProductCatalog,
    match_column: object = None,
    write_output: bool = True,
    price_mode: str = "none",
    exchange_rate: float | None = None,
) -> dict:
    if not write_output:
        return analyze_xlsx_with_bld(
            inquiry_path,
            catalog,
            match_column=match_column,
            price_mode=price_mode,
            exchange_rate=exchange_rate,
        )

    workbook = load_workbook(inquiry_path)
    try:
        summary = {"total": 0, "matched": 0, "unmatched": 0, "rows": []}

        for sheet in workbook.worksheets:
            if sheet.max_row == 0:
                continue
            data_end_row = sheet.max_row

            manual_match_columns = _manual_match_columns(match_column)
            if not manual_match_columns:
                header_row, columns = _find_xlsx_inquiry_columns(sheet)
                selected_match_columns = [columns["oe"]]
            else:
                selected_match_columns = [column + 1 for column in manual_match_columns]
                header_row = _find_xlsx_selected_header_row(
                    sheet,
                    selected_match_columns,
                )
                columns = {}
            output_col = sheet.max_column + 1
            price_header = price_export_header(price_mode)
            status_header = product_status_header_for_price_mode(price_mode) if price_header else ""
            status_col = output_col + 2 if price_header else None
            note_col = output_col + 3 if status_header else (output_col + 2 if price_header else output_col + 1)
            _write_cell(sheet, header_row, output_col, "BLD NO.")
            if price_header:
                _write_cell(sheet, header_row, output_col + 1, price_header)
            if status_header and status_col is not None:
                _write_cell(sheet, header_row, status_col, status_header)
            _write_cell(sheet, header_row, note_col, "匹配说明")

            for row_index in range(header_row + 1, sheet.max_row + 1):
                inquiry_name = sheet.cell(row_index, columns["name"]).value if "name" in columns else ""
                match_values = _xlsx_match_values(
                    sheet,
                    row_index,
                    selected_match_columns,
                )
                if _selected_match_row_is_header(match_values):
                    continue
                inquiry_oe = _combined_match_text(match_values)
                inquiry_desc = sheet.cell(row_index, columns["description"]).value if "description" in columns else ""
                if not str(inquiry_name or "").strip() and not str(inquiry_oe or "").strip():
                    continue

                match = catalog.match(inquiry_name, inquiry_oe, inquiry_desc)
                summary["total"] += 1
                if match:
                    _write_cell(sheet, row_index, output_col, match.bld_no)
                    if price_header:
                        price = match_export_price(match, price_mode, exchange_rate)
                        _write_cell(
                            sheet,
                            row_index,
                            output_col + 1,
                            price,
                            number_format="0" if price_mode == "net" else "0.00",
                        )
                    if status_header and status_col is not None:
                        _write_cell(
                            sheet,
                            row_index,
                            status_col,
                            match_export_status(match, price_mode),
                        )
                    summary["matched"] += 1
                    row_summary = annotate_row_summary_with_match_columns(
                        summary_row(row_index, inquiry_oe, inquiry_name, match),
                        match_values,
                        match,
                    )
                    _write_cell(sheet, row_index, note_col, row_summary["match_note"])
                    summary["rows"].append(row_summary)
                else:
                    _write_cell(sheet, row_index, output_col, "")
                    if price_header:
                        _write_cell(sheet, row_index, output_col + 1, None)
                    if status_header and status_col is not None:
                        _write_cell(sheet, row_index, status_col, "")
                    summary["unmatched"] += 1
                    row_summary = summary_row(row_index, inquiry_oe, inquiry_name, None)
                    _write_cell(sheet, row_index, note_col, row_summary["match_note"])
                    summary["rows"].append(row_summary)
            trim_unused_row_dimensions(sheet, data_end_row)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return summary
    finally:
        workbook.close()


def generate_excel_with_bld(
    inquiry_path: Path,
    output_path: Path,
    catalog: ProductCatalog,
    match_column: object = None,
    write_output: bool = True,
    price_mode: str = "none",
    exchange_rate: float | None = None,
) -> dict:
    suffix = inquiry_path.suffix.lower()
    if suffix == ".xls":
        return generate_xls_with_bld(
            inquiry_path,
            output_path.with_suffix(".xls"),
            catalog,
            match_column=match_column,
            write_output=write_output,
            price_mode=price_mode,
            exchange_rate=exchange_rate,
        )
    if suffix == ".xlsx":
        return generate_xlsx_with_bld(
            inquiry_path,
            output_path.with_suffix(".xlsx"),
            catalog,
            match_column=match_column,
            write_output=write_output,
            price_mode=price_mode,
            exchange_rate=exchange_rate,
        )
    raise ValueError("客户询价文件仅支持 .xls 或 .xlsx。")

from __future__ import annotations

from collections.abc import Iterable
from pathlib import Path
import re
from string import ascii_uppercase
from typing import Any, cast

from openpyxl import load_workbook
import xlrd

from app.matcher import normalize_code, split_codes


INQUIRY_HEADERS = {
    "name": {"物料名称", "产品名称", "名称", "ITEM", "PART"},
    "oe": {
        "OE号",
        "OE",
        "OE NO",
        "OE NO.",
        "OE号码",
        "OE REFERENCE",
        "OE REF",
        "号码",
        "查询号码",
        "客户号码",
        "BLD号",
        "BLD NO",
        "BLD NO.",
    },
    "description": {"物料描述", "描述", "DESCRIPTION", "DESC"},
}


def _norm_header(value: object) -> str:
    return str(value or "").strip().upper().replace(" ", "")


MANUAL_HEADER_ALIASES = {_norm_header(alias) for aliases in INQUIRY_HEADERS.values() for alias in aliases} | {
    _norm_header(alias)
    for alias in {
        "SN",
        "NO",
        "NO.",
        "序号",
        "数量",
        "客户号码",
        "客户编码",
        "号码",
        "料号",
        "ITEM NO",
        "ITEM NO.",
        "PART NO",
        "PART NO.",
        "QTY",
        "QUANTITY",
        "رقم",
        "كمية",
    }
}


def _looks_like_match_code(value: object) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    upper = text.upper()
    if re.search(r"[A-Z]{3,}\s+[A-Z]{3,}", upper):
        return False
    parts = split_codes(text)
    return any(
        len(normalize_code(part)) >= 5 and any(char.isdigit() for char in normalize_code(part)) for part in parts
    )


def _looks_like_manual_header(value: object) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    return _norm_header(text) in MANUAL_HEADER_ALIASES or not _looks_like_match_code(text)


def _find_inquiry_columns(sheet) -> tuple[int, dict[str, int]]:
    aliases = {_norm_header(alias): key for key, names in INQUIRY_HEADERS.items() for alias in names}
    for row_index in range(min(sheet.nrows, 20)):
        columns: dict[str, int] = {}
        for col_index in range(sheet.ncols):
            key = aliases.get(_norm_header(sheet.cell_value(row_index, col_index)))
            if key:
                columns[key] = col_index
        if "oe" in columns:
            return row_index, columns
    raise ValueError("询价表没有找到可识别表头，需要包含 OE号。")


def _column_label(index: int) -> str:
    label = ""
    index += 1
    while index:
        index, remainder = divmod(index - 1, 26)
        label = ascii_uppercase[remainder] + label
    return label


def preview_inquiry_columns(
    inquiry_path: Path,
    max_rows: int = 8,
    max_cols: int = 12,
) -> dict:
    if inquiry_path.suffix.lower() == ".xls":
        book = xlrd.open_workbook(
            str(inquiry_path),
            formatting_info=True,
            ignore_workbook_corruption=True,
        )
        sheet = book.sheet_by_index(0)
        cols = min(sheet.ncols, max_cols)
        rows = []
        for row_index in range(min(sheet.nrows, max_rows)):
            rows.append([sheet.cell_value(row_index, col_index) for col_index in range(cols)])
        return {
            "sheet": sheet.name,
            "columns": [{"index": i, "label": _column_label(i)} for i in range(cols)],
            "rows": rows,
        }

    workbook = load_workbook(inquiry_path, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = []
        for values in sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
            rows.append(list(values[:max_cols]))
        cols = min(max((len(row) for row in rows), default=0), max_cols)
        rows = [row + [None] * (cols - len(row)) for row in rows]
        return {
            "sheet": sheet.title,
            "columns": [{"index": i, "label": _column_label(i)} for i in range(cols)],
            "rows": rows,
        }
    finally:
        workbook.close()


def _norm_openpyxl_cell(value: object) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def _find_xlsx_inquiry_columns(sheet) -> tuple[int, dict[str, int]]:
    aliases = {_norm_header(alias): key for key, names in INQUIRY_HEADERS.items() for alias in names}
    for row_index in range(1, min(sheet.max_row, 20) + 1):
        columns: dict[str, int] = {}
        for col_index in range(1, sheet.max_column + 1):
            key = aliases.get(_norm_openpyxl_cell(sheet.cell(row_index, col_index).value))
            if key:
                columns[key] = col_index
        if "oe" in columns:
            return row_index, columns
    raise ValueError("询价表没有找到可识别表头，需要包含 OE号。")


def _manual_match_columns(match_column: object) -> list[int]:
    if match_column is None:
        return []
    if isinstance(match_column, int | str):
        values = [match_column]
    elif isinstance(match_column, Iterable):
        values = list(match_column)
    else:
        values = [match_column]

    columns: list[int] = []
    seen = set()
    for value in values:
        try:
            index = int(cast(Any, value))
        except (TypeError, ValueError):
            continue
        if index < 0 or index in seen:
            continue
        seen.add(index)
        columns.append(index)
    return columns


def _combined_match_text(match_values: list[tuple[str, object]]) -> str:
    return "\n".join(str(value).strip() for _, value in match_values if str(value or "").strip())


def _selected_match_row_is_header(match_values: list[tuple[str, object]]) -> bool:
    values = [value for _, value in match_values if str(value or "").strip()]
    if not values:
        return False
    return all(_norm_header(value) in MANUAL_HEADER_ALIASES for value in values)


def _xls_match_values(
    sheet,
    row_index: int,
    match_columns: list[int],
) -> list[tuple[str, object]]:
    values = []
    for column_index in match_columns:
        if 0 <= column_index < sheet.ncols:
            values.append((_column_label(column_index), sheet.cell_value(row_index, column_index)))
    return values


def _xlsx_match_values(
    sheet,
    row_index: int,
    match_columns: list[int],
) -> list[tuple[str, object]]:
    values = []
    for column_index in match_columns:
        if 0 < column_index <= sheet.max_column:
            values.append((_column_label(column_index - 1), sheet.cell(row_index, column_index).value))
    return values


def _cell_from_values(row: tuple, column_index: int) -> object:
    return row[column_index - 1] if 0 < column_index <= len(row) else ""


def _row_match_values(
    values: tuple,
    match_columns: list[int],
) -> list[tuple[str, object]]:
    return [
        (_column_label(column_index - 1), _cell_from_values(values, column_index)) for column_index in match_columns
    ]


def _find_xls_selected_header_row(sheet, match_columns: int | list[int]) -> int:
    selected_columns = _manual_match_columns(match_columns)
    max_scan = min(sheet.nrows, 20)
    for row_index in range(max_scan):
        if not any(
            column_index < sheet.ncols and _looks_like_manual_header(sheet.cell_value(row_index, column_index))
            for column_index in selected_columns
        ):
            continue
        for next_row in range(row_index + 1, min(sheet.nrows, row_index + 5)):
            if any(
                column_index < sheet.ncols and _looks_like_match_code(sheet.cell_value(next_row, column_index))
                for column_index in selected_columns
            ):
                return row_index
    return 0


def _find_xlsx_selected_header_row(sheet, match_columns: int | list[int]) -> int:
    selected_columns = _manual_match_columns(match_columns)
    max_scan = min(sheet.max_row, 20)
    for row_index in range(1, max_scan + 1):
        if not any(
            column_index <= sheet.max_column and _looks_like_manual_header(sheet.cell(row_index, column_index).value)
            for column_index in selected_columns
        ):
            continue
        for next_row in range(row_index + 1, min(sheet.max_row, row_index + 4) + 1):
            if any(
                column_index <= sheet.max_column and _looks_like_match_code(sheet.cell(next_row, column_index).value)
                for column_index in selected_columns
            ):
                return row_index
    return 1


def _find_xlsx_inquiry_columns_from_rows(rows: list[tuple]) -> tuple[int, dict[str, int]]:
    aliases = {_norm_header(alias): key for key, names in INQUIRY_HEADERS.items() for alias in names}
    for row_index, row in enumerate(rows[:20], start=1):
        columns: dict[str, int] = {}
        for col_index, value in enumerate(row, start=1):
            key = aliases.get(_norm_openpyxl_cell(value))
            if key:
                columns[key] = col_index
        if "oe" in columns:
            return row_index, columns
    raise ValueError("询价表没有找到可识别表头，需要包含 OE号。")


def _find_xlsx_selected_header_row_from_rows(
    rows: list[tuple],
    match_columns: int | list[int],
) -> int:
    selected_columns = _manual_match_columns(match_columns)
    max_scan = min(len(rows), 20)
    for row_index in range(1, max_scan + 1):
        if not any(
            _looks_like_manual_header(_cell_from_values(rows[row_index - 1], column_index))
            for column_index in selected_columns
        ):
            continue
        for next_row in range(row_index + 1, min(max_scan, row_index + 4) + 1):
            if any(
                _looks_like_match_code(_cell_from_values(rows[next_row - 1], column_index))
                for column_index in selected_columns
            ):
                return row_index
    return 1

from __future__ import annotations

import csv
import io
import json
import os
import re
from copy import copy
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from app.helpers import safe_upload_name, unique_prefixed_path


ALLOWED_TEMPLATE_SUFFIXES = frozenset({".xlsx", ".xlsm"})
ALLOWED_SHIPMENT_SUFFIXES = frozenset({".xlsx", ".csv"})
CODE_ALIASES = {
    "商品编码",
    "产品编码",
    "货号",
    "型号",
    "物料编码",
    "客户编码",
    "BLD NO.",
    "BLD号",
    "ITEM CODE",
    "PART NO",
    "PART NO.",
    "SKU",
}
QTY_ALIASES = {"数量", "发货数量", "出货数量", "QTY", "QUANTITY"}


def _norm(value: object) -> str:
    return re.sub(r"[^A-Z0-9\u4E00-\u9FFF]+", "", str(value or "").strip().upper())


CODE_KEYS = {_norm(alias) for alias in CODE_ALIASES}
QTY_KEYS = {_norm(alias) for alias in QTY_ALIASES}


class ShippingTemplateStore:
    def __init__(self, root: Path) -> None:
        self.root = root
        self.file_dir = root / "files"
        self.manifest_path = root / "manifest.json"

    def load(self) -> list[dict]:
        if not self.manifest_path.is_file():
            return []
        try:
            data = json.loads(self.manifest_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return []
        return data if isinstance(data, list) else []

    def choices(self) -> tuple[list[str], list[dict]]:
        templates = sorted(
            [item for item in self.load() if item.get("status") == "ready"],
            key=lambda item: (str(item.get("customer") or ""), str(item.get("name") or "")),
        )
        customers = sorted(
            {
                str(item.get("customer") or "").strip()
                for item in templates
                if str(item.get("customer") or "").strip()
            }
        )
        return customers, templates

    def find(self, template_id: str) -> dict | None:
        return next((item for item in self.load() if item.get("id") == template_id), None)

    def add_uploaded(self, file, *, customer: str, name: str, actor: str) -> dict:
        stream = getattr(file, "stream", file)
        try:
            stream.seek(0)
        except (AttributeError, OSError):
            pass
        return self.add_bytes(
            str(getattr(file, "filename", "") or ""),
            stream.read(),
            customer=customer,
            name=name,
            actor=actor,
        )

    def add_bytes(self, original_name: str, payload: bytes, *, customer: str, name: str, actor: str) -> dict:
        suffix = Path(original_name).suffix.lower()
        if suffix not in ALLOWED_TEMPLATE_SUFFIXES:
            raise ValueError("模板文件仅支持 .xlsx 或 .xlsm。")
        template_id = datetime.now().strftime("%Y%m%d%H%M%S%f")
        safe_base = safe_upload_name(f"{customer}-{name}{suffix}")
        file_name = f"{template_id}-{safe_base}"
        self.file_dir.mkdir(parents=True, exist_ok=True)
        destination = self.file_dir / file_name
        temporary = destination.with_name(f".{destination.name}.{uuid4().hex}.upload{suffix}")
        try:
            temporary.write_bytes(payload)
            mapping = self.detect_mapping(temporary)
            preview = self.preview_workbook(temporary, max_rows=5, max_cols=6)
            os.replace(temporary, destination)
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            template = {
                "id": template_id,
                "customer": customer,
                "name": name,
                "status": "ready",
                "file_name": file_name,
                "original_name": original_name,
                "mapping": mapping,
                "preview_sheet": preview["sheet"],
                "created_at": now,
                "updated_at": now,
                "created_by": actor,
                "note": "",
            }
            templates = self.load()
            templates.append(template)
            self._save_manifest(templates)
            return template
        except Exception:
            destination.unlink(missing_ok=True)
            raise
        finally:
            temporary.unlink(missing_ok=True)

    def remove(self, template_id: str) -> None:
        templates = self.load()
        target = next((item for item in templates if item.get("id") == template_id), None)
        remaining = [item for item in templates if item.get("id") != template_id]
        self._save_manifest(remaining)
        path = self.template_path(target) if target else None
        if path:
            path.unlink(missing_ok=True)

    def template_path(self, template: dict | None) -> Path | None:
        if not template:
            return None
        file_name = str(template.get("file_name") or "").strip()
        if not file_name:
            return None
        path = (self.file_dir / file_name).resolve()
        return path if self.file_dir.resolve() in path.parents else None

    def _save_manifest(self, templates: list[dict]) -> None:
        self.root.mkdir(parents=True, exist_ok=True)
        temporary = self.manifest_path.with_name(f".{self.manifest_path.name}.{uuid4().hex}.tmp")
        try:
            temporary.write_text(json.dumps(templates, ensure_ascii=False, indent=2), encoding="utf-8")
            os.replace(temporary, self.manifest_path)
        finally:
            temporary.unlink(missing_ok=True)

    @staticmethod
    def preview_workbook(path: Path, *, max_rows: int = 8, max_cols: int = 8) -> dict:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = [list(values[:max_cols]) for values in sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True)]
            cols = min(max((len(row) for row in rows), default=0), max_cols)
            rows = [row + [None] * (cols - len(row)) for row in rows]
            return {"sheet": sheet.title, "columns": [chr(65 + index) for index in range(cols)], "rows": rows}
        finally:
            workbook.close()

    @staticmethod
    def detect_mapping(path: Path) -> dict:
        workbook = load_workbook(path)
        try:
            sheet = workbook.active
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 40)):
                code_col = qty_col = None
                for cell in row:
                    text = str(cell.value or "").strip()
                    if text in {"{{商品编码}}", "{商品编码}", "{{产品编码}}"}:
                        code_col = cell.column
                    elif text in {"{{数量}}", "{数量}", "{{发货数量}}"}:
                        qty_col = cell.column
                if code_col and qty_col:
                    return {
                        "mode": "placeholder",
                        "sheet": sheet.title,
                        "row": row[0].row,
                        "code_col": code_col,
                        "qty_col": qty_col,
                    }
            for row in sheet.iter_rows(max_row=min(sheet.max_row, 40), values_only=False):
                code_col = qty_col = None
                for cell in row:
                    key = _norm(cell.value)
                    if key in CODE_KEYS:
                        code_col = cell.column
                    if key in QTY_KEYS:
                        qty_col = cell.column
                if code_col and qty_col:
                    return {
                        "mode": "header",
                        "sheet": sheet.title,
                        "header_row": row[0].row,
                        "row": row[0].row + 1,
                        "code_col": code_col,
                        "qty_col": qty_col,
                    }
        finally:
            workbook.close()
        raise ValueError("模板里没有找到商品编码和数量列。")


class ShippingWorkbookAdapter:
    def parse_rows(self, path: Path) -> tuple[list[dict], dict]:
        if path.suffix.lower() == ".xlsx":
            return self._rows_from_xlsx(path)
        if path.suffix.lower() == ".csv":
            return self._rows_from_csv(path)
        raise ValueError("发货数据仅支持 .xlsx 或 .csv。")

    def generate(self, template_store: ShippingTemplateStore, template: dict, rows: list[dict], output_dir: Path) -> Path:
        template_path = template_store.template_path(template)
        if not template_path or not template_path.is_file():
            raise ValueError("当前模板还没有上传 Excel 文件。")
        mapping = template.get("mapping") or template_store.detect_mapping(template_path)
        workbook = load_workbook(template_path)
        output_path: Path | None = None
        temporary: Path | None = None
        try:
            sheet = workbook[mapping.get("sheet")] if mapping.get("sheet") in workbook.sheetnames else workbook.active
            start_row = int(mapping["row"])
            code_col = int(mapping["code_col"])
            qty_col = int(mapping["qty_col"])
            for offset, item in enumerate(rows):
                target_row = start_row + offset
                self._copy_row_style(sheet, start_row, target_row)
                sheet.cell(target_row, code_col).value = item["code"]
                sheet.cell(target_row, qty_col).value = item["quantity"]
            safe_customer = safe_upload_name(template.get("customer") or "客户")
            output_path = unique_prefixed_path(
                output_dir,
                f"shipping-notice-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{safe_customer}.xlsx",
            )
            temporary = output_path.with_name(f".{output_path.stem}.{uuid4().hex}.tmp.xlsx")
            workbook.save(temporary)
            os.replace(temporary, output_path)
        finally:
            workbook.close()
            if temporary is not None:
                temporary.unlink(missing_ok=True)
        if output_path is None:
            raise RuntimeError("Shipping notice output was not created.")
        return output_path

    @staticmethod
    def parse_filename_defaults(filename: str) -> tuple[str, str]:
        stem = Path(filename).stem.strip()
        parts = [part.strip() for part in re.split(r"[-_－—]+", stem, maxsplit=1) if part.strip()]
        if len(parts) == 2:
            return parts[0], parts[1]
        return "未分组客户", stem or "未命名模板"

    @staticmethod
    def _find_columns(rows) -> tuple[int, int, int]:
        header_index = code_index = qty_index = None
        for row_index, row in enumerate(rows):
            for col_index, value in enumerate(row):
                key = _norm(value)
                if key in CODE_KEYS:
                    code_index = col_index
                    header_index = row_index
                if key in QTY_KEYS:
                    qty_index = col_index
                    header_index = row_index if header_index is None else header_index
            if code_index is not None and qty_index is not None and header_index is not None:
                return header_index, code_index, qty_index
        raise ValueError("发货数据里没有找到商品编码和数量列。")

    def _rows_from_xlsx(self, path: Path) -> tuple[list[dict], dict]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            cached_rows = list(sheet.iter_rows(max_row=40, values_only=True))
            header_index, code_index, qty_index = self._find_columns(cached_rows)
            rows = []
            for row_number, row in enumerate(
                sheet.iter_rows(min_row=header_index + 2, values_only=True),
                start=header_index + 2,
            ):
                code = str(row[code_index] if code_index < len(row) else "" or "").strip()
                qty = row[qty_index] if qty_index < len(row) else ""
                if not code and not str(qty or "").strip():
                    continue
                rows.append({"row": row_number, "code": code, "quantity": qty})
            return rows, {"sheet": sheet.title, "header_row": header_index + 1}
        finally:
            workbook.close()

    def _rows_from_csv(self, path: Path) -> tuple[list[dict], dict]:
        text = path.read_bytes().decode("utf-8-sig", errors="ignore")
        all_rows = list(csv.reader(io.StringIO(text)))
        header_index, code_index, qty_index = self._find_columns(all_rows[:40])
        rows = []
        for row_number, row in enumerate(all_rows[header_index + 1 :], start=header_index + 2):
            code = str(row[code_index] if code_index < len(row) else "").strip()
            qty = row[qty_index] if qty_index < len(row) else ""
            if not code and not str(qty or "").strip():
                continue
            rows.append({"row": row_number, "code": code, "quantity": qty})
        return rows, {"sheet": "CSV", "header_row": header_index + 1}

    @staticmethod
    def _copy_row_style(sheet, source_row: int, target_row: int) -> None:
        if source_row == target_row:
            return
        for col_index in range(1, sheet.max_column + 1):
            source = sheet.cell(source_row, col_index)
            target = sheet.cell(target_row, col_index)
            if source.has_style:
                target._style = copy(source._style)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy(source.alignment)
            if source.fill:
                target.fill = copy(source.fill)
            if source.border:
                target.border = copy(source.border)
            if source.font:
                target.font = copy(source.font)

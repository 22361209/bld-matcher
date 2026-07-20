from __future__ import annotations

import hashlib
import json
import os
import shutil
import uuid
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage

from app.matcher import compact_text, find_catalog_header_row
from app.product_media import product_image_storage_name, validate_product_image_file

from .domain import ProductRecord


REQUIRED_FIELDS = ("BLD NO.", "SERIES", "ITEM", "OE NO.1", "Models", "产品状态", "导入单价")
SERIES_SELECTION_HEADERS = ("SERIES", "SERIES 2", "SERIES 3", "SERIES 4", "SERIES 5", "SERIES 6")
FIELD_LABELS = {
    "series": "SERIES",
    "item": "ITEM",
    "oe_no_1": "OE NO.1",
    "oe_no_2": "OE NO.2",
    "models": "Models",
    "product_status": "产品状态",
    "price_cny": "导入单价",
}


class CatalogImportPreviewChangedError(RuntimeError):
    def __init__(self) -> None:
        super().__init__("产品目录预览已变化，请重新上传并预览后再确认导入。")


@dataclass(frozen=True, slots=True)
class CatalogImage:
    row_number: int
    data: bytes
    suffix: str


@dataclass(frozen=True, slots=True)
class CatalogImportChoices:
    series: tuple[str, ...]
    items: tuple[str, ...]

    @property
    def series_keys(self) -> frozenset[str]:
        return frozenset(value.casefold() for value in self.series)

    @property
    def item_keys(self) -> frozenset[str]:
        return frozenset(value.casefold() for value in self.items)


@dataclass(frozen=True, slots=True)
class CatalogImportRow:
    row_number: int
    bld_no: str
    series: str
    item: str
    oe_no_1: str
    oe_no_2: str
    models: str
    product_status: str
    price_cny: float
    image: CatalogImage | None

    def values(self) -> dict[str, object]:
        return {
            "bld_no": self.bld_no,
            "series": self.series,
            "item": self.item,
            "oe_no_1": self.oe_no_1,
            "oe_no_2": self.oe_no_2,
            "models": self.models,
            "product_status": self.product_status,
            "price_cny": self.price_cny,
            "active": "1",
        }


@dataclass(frozen=True, slots=True)
class CatalogImportConflict:
    row: CatalogImportRow
    product: ProductRecord
    fields: tuple[dict[str, object], ...]
    all_fields: tuple[dict[str, object], ...]

    def web_payload(self) -> dict[str, object]:
        return {
            "row_number": self.row.row_number,
            "bld_no": self.row.bld_no,
            "fields": list(self.fields),
            "all_fields": list(self.all_fields),
        }


@dataclass(frozen=True, slots=True)
class CatalogImportPreview:
    rows: tuple[CatalogImportRow, ...]
    new_rows: tuple[CatalogImportRow, ...]
    conflicts: tuple[CatalogImportConflict, ...]
    unchanged_rows: tuple[CatalogImportRow, ...]
    digest: str

    def web_payload(self) -> dict[str, object]:
        return {
            "digest": self.digest,
            "new_count": len(self.new_rows),
            "conflict_count": len(self.conflicts),
            "unchanged_count": len(self.unchanged_rows),
            "conflicts": [conflict.web_payload() for conflict in self.conflicts],
            "new_rows": [{"row_number": row.row_number, "bld_no": row.bld_no} for row in self.new_rows],
        }


@dataclass(frozen=True, slots=True)
class CatalogImportResult:
    created_count: int
    updated_count: int
    kept_count: int


@dataclass(frozen=True, slots=True)
class CatalogImportStorage:
    catalog_path: Path
    image_dir: Path
    thumb_dir: Path


def _price(value: object, *, row_number: int) -> float:
    text = compact_text(value).replace(",", "").removeprefix("¥")
    try:
        parsed = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"第 {row_number} 行“导入单价”必须是数字。") from exc
    if not parsed.is_finite() or parsed < 0:
        raise ValueError(f"第 {row_number} 行“导入单价”必须大于或等于 0。")
    return float(parsed)


def _nonempty_row(row: tuple[object, ...]) -> bool:
    return any(compact_text(value) for value in row)


def _image_map(sheet, headers: list[str], header_index: int) -> dict[int, CatalogImage]:
    if "图片" not in headers:
        if sheet._images:
            raise ValueError("工作簿含有图片，但没有“图片”列。")
        return {}
    image_column = headers.index("图片")
    images: dict[int, CatalogImage] = {}
    errors: list[str] = []
    for image in sheet._images:  # openpyxl currently exposes drawings through this collection.
        marker = getattr(getattr(image, "anchor", None), "_from", None)
        if marker is None:
            errors.append("有一张图片无法识别所在单元格")
            continue
        row_number = marker.row + 1
        if marker.col != image_column:
            errors.append(f"第 {row_number} 行图片必须放在“图片”列")
            continue
        if marker.row <= header_index:
            errors.append("图片不能放在表头行")
            continue
        if row_number in images:
            errors.append(f"第 {row_number} 行“图片”列最多插入一张图片")
            continue
        image_format = str(getattr(image, "format", "png") or "png").lower()
        suffix = ".jpg" if image_format in {"jpg", "jpeg"} else f".{image_format}"
        data = image._data()
        file = FileStorage(stream=BytesIO(data), filename=f"catalog-image{suffix}")
        try:
            validate_product_image_file(file)
        except ValueError as exc:
            errors.append(f"第 {row_number} 行图片无效：{exc}")
            continue
        images[row_number] = CatalogImage(row_number=row_number, data=data, suffix=suffix)
    if errors:
        raise ValueError("目录图片校验失败：" + "；".join(errors[:10]) + "。")
    return images


def _validate_controlled_values(
    *,
    row_number: int,
    series_values: list[str],
    item: str,
    choices: CatalogImportChoices,
) -> list[str]:
    errors: list[str] = []
    if choices.series_keys:
        unknown_series = [value for value in series_values if value.casefold() not in choices.series_keys]
        if unknown_series:
            errors.append(f"第 {row_number} 行 SERIES 只能选择模板下拉选项：{'、'.join(unknown_series)}")
    if choices.item_keys and item.casefold() not in choices.item_keys:
        errors.append(f"第 {row_number} 行 ITEM 只能选择模板下拉选项：{item}")
    return errors


def read_catalog_import(path: Path, *, choices: CatalogImportChoices) -> tuple[CatalogImportRow, ...]:
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValueError("产品目录没有可读取的工作表。")
        raw_rows = list(sheet.iter_rows(values_only=True))
        header_index, headers = find_catalog_header_row(raw_rows)
        missing_headers = [field for field in REQUIRED_FIELDS if field not in headers]
        if missing_headers:
            raise ValueError(f"目录缺少必填列：{'、'.join(missing_headers)}。请下载并使用标准模板。")
        images = _image_map(sheet, headers, header_index)
        rows: list[CatalogImportRow] = []
        errors: list[str] = []
        seen: dict[str, int] = {}
        for row_number, raw in enumerate(raw_rows[header_index + 1 :], start=header_index + 2):
            if not _nonempty_row(raw) and row_number not in images:
                continue
            source = {header: value for header, value in zip(headers, raw) if header}
            missing = [field for field in REQUIRED_FIELDS if not compact_text(source.get(field))]
            if missing:
                errors.append(f"第 {row_number} 行缺少：{'、'.join(missing)}")
                continue
            series_values = [
                compact_text(source.get(header))
                for header in SERIES_SELECTION_HEADERS
                if compact_text(source.get(header))
            ]
            item = compact_text(source["ITEM"])
            row_errors = _validate_controlled_values(
                row_number=row_number,
                series_values=series_values,
                item=item,
                choices=choices,
            )
            if row_errors:
                errors.extend(row_errors)
                continue
            bld_no = compact_text(source["BLD NO."])
            key = bld_no.casefold()
            if key in seen:
                errors.append(f"第 {row_number} 行与第 {seen[key]} 行 BLD NO. 重复：{bld_no}")
                continue
            seen[key] = row_number
            try:
                price_cny = _price(source["导入单价"], row_number=row_number)
            except ValueError as exc:
                errors.append(str(exc))
                continue
            rows.append(
                CatalogImportRow(
                    row_number=row_number,
                    bld_no=bld_no,
                    series="\n".join(dict.fromkeys(series_values)),
                    item=item,
                    oe_no_1=compact_text(source["OE NO.1"]),
                    oe_no_2=compact_text(source.get("OE NO.2")),
                    models=compact_text(source["Models"]),
                    product_status=compact_text(source["产品状态"]),
                    price_cny=price_cny,
                    image=images.get(row_number),
                )
            )
        if errors:
            suffix = f"；另有 {len(errors) - 10} 处错误" if len(errors) > 10 else ""
            raise ValueError("目录校验失败：" + "；".join(errors[:10]) + suffix + "。")
        if not rows:
            raise ValueError("目录中没有可导入的产品数据。")
        return tuple(rows)
    finally:
        workbook.close()


def _field_differences(row: CatalogImportRow, product: ProductRecord) -> tuple[tuple[dict[str, object], ...], tuple[dict[str, object], ...]]:
    fields: list[dict[str, object]] = []
    all_fields: list[dict[str, object]] = []
    for field, label in FIELD_LABELS.items():
        before = getattr(product, field)
        after = getattr(row, field)
        all_fields.append({"label": label, "before": before if before is not None else "", "after": after, "changed": before != after})
        if before != after:
            fields.append({"label": label, "before": before if before is not None else "", "after": after})
    if row.image:
        before_image = "已有图片" if product.image_path else "无图片"
        after_image = "替换为 Excel 图片"
        all_fields.append({"label": "图片", "before": before_image, "after": after_image, "changed": True})
        fields.append({"label": "图片", "before": before_image, "after": after_image})
    return tuple(fields), tuple(all_fields)


def build_catalog_import_preview(rows: Iterable[CatalogImportRow], products: dict[str, ProductRecord]) -> CatalogImportPreview:
    stable_rows = tuple(rows)
    new_rows: list[CatalogImportRow] = []
    conflicts: list[CatalogImportConflict] = []
    unchanged: list[CatalogImportRow] = []
    digest_payload: list[dict[str, object]] = []
    for row in stable_rows:
        product = products.get(row.bld_no)
        fields, all_fields = _field_differences(row, product) if product else ((), ())
        if product is None:
            new_rows.append(row)
        elif fields:
            conflicts.append(CatalogImportConflict(row=row, product=product, fields=fields, all_fields=all_fields))
        else:
            unchanged.append(row)
        digest_payload.append(
            {
                "incoming": {**row.values(), "image": hashlib.sha256(row.image.data).hexdigest() if row.image else ""},
                "existing": product.web_payload() if product else None,
            }
        )
    digest = hashlib.sha256(
        json.dumps(digest_payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    ).hexdigest()
    return CatalogImportPreview(stable_rows, tuple(new_rows), tuple(conflicts), tuple(unchanged), digest)


class CatalogImportFileTransaction:
    def __init__(self, *, catalog_path: Path, image_dir: Path, thumb_dir: Path) -> None:
        self.catalog_path = catalog_path
        self.image_dir = image_dir
        self.thumb_dir = thumb_dir
        self.backup_dir = catalog_path.parent / "local-backups" / f"catalog-import-{uuid.uuid4().hex}"
        self.changes: list[tuple[Path, Path | None]] = []

    @staticmethod
    def _atomic_copy(source: Path, target: Path) -> None:
        target.parent.mkdir(parents=True, exist_ok=True)
        temporary = target.with_name(f".{target.name}.{uuid.uuid4().hex}.tmp")
        try:
            shutil.copy2(source, temporary)
            os.replace(temporary, target)
        finally:
            temporary.unlink(missing_ok=True)

    def _replace(self, source: Path, target: Path) -> None:
        if target == self.catalog_path:
            backup = self.backup_dir / "catalog" / target.name
        elif target.parent == self.image_dir:
            backup = self.backup_dir / "images" / target.name
        else:
            backup = self.backup_dir / "thumbs" / target.name
        backup = backup if target.exists() else None
        if backup:
            self._atomic_copy(target, backup)
        self.changes.append((target, backup))
        self._atomic_copy(source, target)

    def apply_images(self, rows: Iterable[CatalogImportRow]) -> dict[str, str]:
        image_paths: dict[str, str] = {}
        for row in rows:
            if not row.image:
                continue
            filename = product_image_storage_name(row.bld_no, row.image.suffix)
            staging = self.backup_dir / "staged" / filename
            staging.parent.mkdir(parents=True, exist_ok=True)
            staging.write_bytes(row.image.data)
            target = self.image_dir / filename
            self._replace(staging, target)
            thumb = self.thumb_dir / target.name
            if thumb.exists():
                thumb_backup = self.backup_dir / "thumbs" / thumb.name
                self._atomic_copy(thumb, thumb_backup)
                self.changes.append((thumb, thumb_backup))
                thumb.unlink()
            image_paths[row.bld_no] = f"data_product_images/{target.name}"
        return image_paths

    def apply_catalog(self, source_catalog: Path) -> None:
        self._replace(source_catalog, self.catalog_path)

    def rollback(self) -> None:
        errors: list[str] = []
        for target, backup in reversed(self.changes):
            try:
                if backup and backup.exists():
                    self._atomic_copy(backup, target)
                else:
                    target.unlink(missing_ok=True)
            except OSError:
                errors.append(target.name)
        if errors:
            raise RuntimeError("目录导入文件回滚失败，请检查备份目录。")

    def finalize(self) -> None:
        shutil.rmtree(self.backup_dir, ignore_errors=True)

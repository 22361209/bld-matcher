from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from .catalog_import import CatalogImportChoices, SERIES_SELECTION_HEADERS


CATALOG_HEADERS = (
    "BLD NO.",
    *SERIES_SELECTION_HEADERS,
    "ITEM",
    "OE NO.1",
    "Models",
    "产品状态",
    "导入单价",
    "OE NO.2",
    "图片",
)
TEMPLATE_DATA_ROWS = 500


def _option_range(name: str, sheet_name: str, column: str, count: int) -> DefinedName:
    return DefinedName(name, attr_text=f"'{sheet_name}'!${column}$2:${column}${count + 1}")


def _list_validation(name: str) -> DataValidation:
    return DataValidation(
        type="list",
        formula1=f"={name}",
        allow_blank=True,
        showErrorMessage=True,
        errorStyle="stop",
        errorTitle="仅可选择目录选项",
        error="请从下拉列表中选择，不要手动填写。",
        showInputMessage=True,
        promptTitle="受控目录字段",
        prompt="请选择现有产品库中的选项。",
    )


def build_catalog_import_template(choices: CatalogImportChoices) -> BytesIO:
    if not choices.series or not choices.items:
        raise ValueError("产品库暂无可选 SERIES 或 ITEM，暂时无法生成受控导入模板。")

    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError("Product import template has no active worksheet.")
    sheet.title = "产品目录"
    options = workbook.create_sheet("可选项")
    instructions = workbook.create_sheet("填写说明")

    sheet.append(CATALOG_HEADERS)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{chr(64 + len(CATALOG_HEADERS))}{TEMPLATE_DATA_ROWS + 1}"
    sheet.sheet_view.showGridLines = False
    header = sheet[1]
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_border = Border(bottom=Side(style="medium", color="163A5C"))
    for cell in header:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
    sheet.row_dimensions[1].height = 24
    widths = (22, 16, 16, 16, 16, 16, 16, 24, 25, 34, 18, 14, 22, 22, 18)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    options.append(["SERIES 选项", "ITEM 选项"])
    option_count = max(len(choices.series), len(choices.items))
    for index in range(option_count):
        options.append(
            [
                choices.series[index] if index < len(choices.series) else "",
                choices.items[index] if index < len(choices.items) else "",
            ]
        )
    options.sheet_view.showGridLines = False
    options.freeze_panes = "A2"
    options.column_dimensions["A"].width = 22
    options.column_dimensions["B"].width = 32
    for cell in options[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    workbook.defined_names.add(_option_range("CatalogSeriesOptions", options.title, "A", len(choices.series)))
    workbook.defined_names.add(_option_range("CatalogItemOptions", options.title, "B", len(choices.items)))
    for column in range(2, 2 + len(SERIES_SELECTION_HEADERS)):
        validation = _list_validation("CatalogSeriesOptions")
        sheet.add_data_validation(validation)
        validation.add(f"{chr(64 + column)}2:{chr(64 + column)}{TEMPLATE_DATA_ROWS + 1}")
    item_column = 2 + len(SERIES_SELECTION_HEADERS)
    item_validation = _list_validation("CatalogItemOptions")
    sheet.add_data_validation(item_validation)
    item_validation.add(f"{chr(64 + item_column)}2:{chr(64 + item_column)}{TEMPLATE_DATA_ROWS + 1}")
    options.sheet_state = "hidden"

    instructions.sheet_view.showGridLines = False
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 106
    instructions.append(["项目", "说明"])
    instructions.append(["必填字段", "BLD NO.、SERIES、ITEM、OE NO.1、Models、产品状态、导入单价。"])
    instructions.append(["SERIES 多选", "从 SERIES、SERIES 2 至 SERIES 6 依次选择；系统会合并为同一产品的多品牌 SERIES。每个格都只能从下拉选项选择。"])
    instructions.append(["ITEM", "只能从下拉选项选择。"])
    instructions.append(["可选字段", "OE NO.2 和图片。图片请插入到“图片”列对应行。"])
    instructions.append(["图片上限", "JPG、PNG、WebP；单张不超过 5 MB，任一边不超过 6000 像素，建议长边不超过 2000 像素。"])
    instructions.append(["BLD NO. 冲突", "上传后逐条预览。与产品库相同的 BLD NO. 默认保留，只有勾选后才会更新。"])
    for cell in instructions[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in instructions.iter_rows(min_row=2, max_col=2):
        row[0].font = Font(bold=True, color="1F4E78")
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    for row_number in range(2, instructions.max_row + 1):
        instructions.row_dimensions[row_number].height = 34

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer

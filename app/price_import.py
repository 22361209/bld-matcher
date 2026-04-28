from __future__ import annotations

import json
import re
import sqlite3
from pathlib import Path

import xlrd
from openpyxl import load_workbook

from .matcher import compact_text


BLD_HEADERS = {"BLD NO.", "BLD NO", "BLD号", "BLD 号", "BLD"}
PRICE_HEADERS = {"UNIT PRICE", "Unit Price", "含税单价", "单价", "价格", "PRICE"}


def _norm_header(value: object) -> str:
    return re.sub(r"[^A-Z0-9\u4e00-\u9fff]+", "", str(value or "").upper())


def _parse_price(value: object) -> float | None:
    text = compact_text(value).replace("¥", "").replace("￥", "").replace(",", "")
    if not text:
        return None
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def _find_columns(rows: list[list[object]]) -> tuple[int, int, int]:
    bld_keys = {_norm_header(item) for item in BLD_HEADERS}
    price_keys = {_norm_header(item) for item in PRICE_HEADERS}
    for row_index, row in enumerate(rows[:20]):
        bld_col = price_col = None
        for col_index, value in enumerate(row):
            key = _norm_header(value)
            if key in bld_keys:
                bld_col = col_index
            if key in price_keys:
                price_col = col_index
        if bld_col is not None and price_col is not None:
            return row_index, bld_col, price_col
    raise ValueError("没有找到 BLD NO. 和 Unit Price/含税单价 表头。")


def _read_rows(path: Path) -> list[list[object]]:
    if path.suffix.lower() == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.worksheets[0]
        return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]
    if path.suffix.lower() == ".xls":
        book = xlrd.open_workbook(path, ignore_workbook_corruption=True)
        sheet = book.sheet_by_index(0)
        return [[sheet.cell_value(r, c) for c in range(sheet.ncols)] for r in range(sheet.nrows)]
    raise ValueError("单价导入仅支持 .xls 或 .xlsx。")


def parse_price_file(path: Path, conn: sqlite3.Connection) -> dict:
    rows = _read_rows(path)
    header_row, bld_col, price_col = _find_columns(rows)
    preview = []
    counts = {"total": 0, "matched": 0, "missing": 0, "invalid_price": 0}

    for row_number, row in enumerate(rows[header_row + 1 :], start=header_row + 2):
        bld_no = compact_text(row[bld_col] if bld_col < len(row) else "")
        if not bld_no:
            continue
        price = _parse_price(row[price_col] if price_col < len(row) else "")
        product = conn.execute("SELECT bld_no, price_cny FROM products WHERE bld_no = ?", (bld_no,)).fetchone()
        status = "matched"
        if price is None:
            counts["invalid_price"] += 1
            status = "invalid_price"
        elif not product:
            counts["missing"] += 1
            status = "missing"
        else:
            counts["matched"] += 1
        counts["total"] += 1
        preview.append(
            {
                "row": row_number,
                "bld_no": bld_no,
                "price": price,
                "old_price": None if not product else product["price_cny"],
                "status": status,
            }
        )
    return {"counts": counts, "rows": preview}


def encode_rows(rows: list[dict]) -> str:
    return json.dumps(rows, ensure_ascii=False)


def decode_rows(payload: str) -> list[dict]:
    data = json.loads(payload)
    if not isinstance(data, list):
        raise ValueError("导入数据无效。")
    return data

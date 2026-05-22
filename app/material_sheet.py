from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PLAN_HEADERS = ["型号", "车型", "计划数/只", "L", "R", "冲压完成时间", "备注"]

DETAIL_HEADERS = [
    "型号",
    "车型",
    "计划数/只",
    "L",
    "R",
    "编码",
    "零件名称",
    "规格尺寸",
    "下料只数",
    "单位规格材料条数",
    "单件重量kg",
    "需求重量kg",
    "图片备注",
    "源表行",
]

SUMMARY_HEADERS = [
    "规格尺寸",
    "下料总数/只",
    "需求总重量kg",
    "单位规格材料总条数",
    "涉及型号",
    "明细拆分",
]


def normalize_model(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_number(value: Any, field_name: str, row_number: int) -> float:
    if value in (None, ""):
        raise ValueError(f"生产计划第 {row_number} 行缺少 {field_name}")
    try:
        return float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"生产计划第 {row_number} 行的 {field_name} 不是数字：{value}") from exc


def clean_sheet_title(value: Any) -> str:
    text = normalize_model(value)
    if not text:
        return datetime.now().strftime("%Y%m%d%H%M")
    text = re.sub(r"[\\/:*?\"<>|]", "", text)
    return text[:40] or datetime.now().strftime("%Y%m%d%H%M")


def safe_filename_part(value: str) -> str:
    text = re.sub(r"[\\/:*?\"<>|]", "", value).strip()
    text = text.replace("/", "").replace("\\", "")
    return text[:80] or datetime.now().strftime("%Y%m%d%H%M")


def format_int(value: float) -> int | float:
    return int(value) if float(value).is_integer() else value


def format_thickness(value: float) -> str:
    text = f"{value:.2f}".rstrip("0")
    if text.endswith("."):
        return text + "0"
    return text


def format_spec(thickness: float, width: float, length: float) -> str:
    return f"{format_thickness(thickness)}×{format_int(width)}×{format_int(length)}"


def sync_material_specs_from_dimensions(path: Path) -> int:
    values_wb = load_workbook(path, read_only=True, data_only=True)
    edit_wb = load_workbook(path)
    try:
        if "材料数据" not in values_wb.sheetnames:
            raise ValueError("材料数据文件里找不到工作表：材料数据")
        values_ws = values_wb["材料数据"]
        edit_ws = edit_wb["材料数据"]
        changed = 0
        for row_number, values in enumerate(values_ws.iter_rows(min_row=2, max_col=11, values_only=True), start=2):
            model = normalize_model(values[0])
            if not model:
                continue
            pieces = values[6]
            thickness = values[8]
            width = values[9]
            length = values[10]
            if any(value in (None, "") for value in [pieces, thickness, width, length]):
                continue
            try:
                spec_text = format_spec(float(thickness), float(width), float(length))
            except (TypeError, ValueError):
                continue
            cell = edit_ws.cell(row_number, 6)
            if str(cell.value or "") != spec_text:
                cell.value = spec_text
                changed += 1
        if changed:
            edit_wb.save(path)
        return changed
    finally:
        values_wb.close()
        edit_wb.close()


def create_plan_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "生产计划"
    ws.merge_cells("A1:G1")
    ws["A1"] = "26年4月冲压生产计划260423"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.append(PLAN_HEADERS)

    default_rows = [
        ["8018", "飞度", 1000, 500, 500, "2026/5/15", "其中300 对 做 279"],
        ["8060", "高7", 1400, 900, 500, "2026/5/15", ""],
        ["8081", "新捷达", 1600, 1000, 600, "2026/5/15", ""],
        ["8092", "赛欧3", 1000, "", 1000, "2026/5/15", "做 279"],
        ["8130", "索9", 1000, 500, 500, "2026/5/15", ""],
        ["8145", "8代凯美瑞", 1000, 500, 500, "2026/5/15", ""],
        ["8234", "雅尊", 1000, 500, 500, "2026/5/15", ""],
    ]
    for row in default_rows:
        ws.append(row)

    style_plan_sheet(ws)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def read_plan(path: Path) -> tuple[str, list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=True)
    try:
        ws = wb.active
        title = ws["A1"].value if ws["A1"].value else path.stem

        header_row = None
        header_map: dict[str, int] = {}
        for row in range(1, min(ws.max_row, 20) + 1):
            values = [normalize_model(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)]
            if "型号" in values and "计划数/只" in values:
                header_row = row
                header_map = {name: values.index(name) + 1 for name in PLAN_HEADERS if name in values}
                break
        if header_row is None:
            raise ValueError("生产计划里找不到表头，请包含：型号、车型、计划数/只、L、R、冲压完成时间、备注")
        for required in ["型号", "计划数/只"]:
            if required not in header_map:
                raise ValueError(f"生产计划缺少必需列：{required}")

        plans: list[dict[str, Any]] = []
        for row in range(header_row + 1, ws.max_row + 1):
            model = normalize_model(ws.cell(row, header_map["型号"]).value)
            if not model:
                continue
            qty = to_number(ws.cell(row, header_map["计划数/只"]).value, "计划数/只", row)
            plans.append(
                {
                    "model": model,
                    "car": ws.cell(row, header_map.get("车型", 0)).value if "车型" in header_map else "",
                    "qty": qty,
                    "left": ws.cell(row, header_map.get("L", 0)).value if "L" in header_map else "",
                    "right": ws.cell(row, header_map.get("R", 0)).value if "R" in header_map else "",
                    "finish_date": ws.cell(row, header_map.get("冲压完成时间", 0)).value
                    if "冲压完成时间" in header_map
                    else "",
                    "remark": ws.cell(row, header_map.get("备注", 0)).value if "备注" in header_map else "",
                }
            )
        if not plans:
            raise ValueError("生产计划里没有可计算的型号")
        return clean_sheet_title(title), plans
    finally:
        wb.close()


def read_materials(path: Path) -> dict[str, list[dict[str, Any]]]:
    wb = load_workbook(path, data_only=False)
    try:
        if "材料数据" not in wb.sheetnames:
            raise ValueError("材料数据文件里找不到工作表：材料数据")
        ws = wb["材料数据"]
        rows_by_model: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row_number, values in enumerate(ws.iter_rows(min_row=2, max_col=11, values_only=True), start=2):
            model = normalize_model(values[0])
            if not model:
                continue
            pieces = values[6]
            thickness = values[8]
            width = values[9]
            length = values[10]
            if any(value in (None, "") for value in [pieces, thickness, width, length]):
                continue
            rows_by_model[model].append(
                {
                    "source_row": row_number,
                    "model": model,
                    "code": values[1],
                    "category": values[2],
                    "car": values[3],
                    "part": values[4],
                    "spec_text": format_spec(float(thickness), float(width), float(length)),
                    "pieces": float(pieces),
                    "thickness": float(thickness),
                    "width": float(width),
                    "length": float(length),
                }
            )
        return rows_by_model
    finally:
        wb.close()


def material_data_stats(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"exists": False, "model_count": 0, "detail_count": 0, "sheet_count": 0, "updated_at": ""}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return {
            "exists": True,
            "invalid": True,
            "error": str(exc),
            "model_count": 0,
            "detail_count": 0,
            "sheet_count": 0,
            "updated_at": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
            "filename": path.name,
        }
    try:
        if "材料数据" not in wb.sheetnames:
            return {"exists": True, "invalid": True, "model_count": 0, "detail_count": 0, "sheet_count": len(wb.sheetnames)}
        ws = wb["材料数据"]
        models: set[str] = set()
        detail_count = 0
        for values in ws.iter_rows(min_row=2, max_col=11, values_only=True):
            model = normalize_model(values[0])
            if not model:
                continue
            models.add(model)
            required_values = [values[6], values[8], values[9], values[10]]
            if all(value not in (None, "") for value in required_values):
                detail_count += 1
        return {
            "exists": True,
            "invalid": False,
            "model_count": len(models),
            "detail_count": detail_count,
            "sheet_count": len(wb.sheetnames),
            "updated_at": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M"),
            "filename": path.name,
        }
    finally:
        wb.close()


def calculate(plans: list[dict[str, Any]], materials: dict[str, list[dict[str, Any]]]) -> tuple[list[dict[str, Any]], list[str]]:
    detail: list[dict[str, Any]] = []
    missing: list[str] = []
    for plan in plans:
        model = plan["model"]
        rows = materials.get(model, [])
        if not rows:
            missing.append(model)
            continue
        for material in rows:
            unit_weight = (
                material["width"]
                * material["length"]
                * 7.85
                * material["thickness"]
                / material["pieces"]
                / 1_000_000
            )
            total_weight = unit_weight * plan["qty"]
            bar_qty = plan["qty"] / material["pieces"]
            detail.append(
                {
                    **material,
                    **plan,
                    "calc_spec": format_spec(material["thickness"], material["width"], material["length"]),
                    "unit_weight": unit_weight,
                    "total_weight": total_weight,
                    "bar_qty": bar_qty,
                }
            )
    return detail, missing


def build_summary(detail: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    summary: dict[str, dict[str, Any]] = defaultdict(
        lambda: {"plan_qty": 0.0, "weight": 0.0, "bar_qty": 0.0, "models": set(), "details": []}
    )
    for item in detail:
        key = item["calc_spec"]
        summary[key]["plan_qty"] += item["qty"]
        summary[key]["weight"] += item["total_weight"]
        summary[key]["bar_qty"] += item["bar_qty"]
        summary[key]["models"].add(item["model"])
        summary[key]["details"].append(
            f"{item['model']}-{item['code']} {item['qty']:g}只/{item['pieces']:g}只="
            f"{item['bar_qty']:.2f}条，{item['total_weight']:.2f}kg"
        )
    return summary


def spec_sort_key(spec: str) -> tuple[float, float, float, str]:
    parts = spec.split("×")
    try:
        return float(parts[0]), float(parts[1]), float(parts[2]), spec
    except (IndexError, ValueError):
        return 0, 0, 0, spec


def style_plan_sheet(ws: Any) -> None:
    blue = "1F4E78"
    thin = Side(style="thin", color="B7B7B7")
    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"
    widths = [14, 16, 14, 10, 10, 18, 28]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_output_sheet(ws: Any, header_rows: tuple[int, ...] = (), title_merge: str | None = None) -> None:
    blue = "1F4E78"
    light = "D9EAF7"
    total_fill = "FFF2CC"
    thin = Side(style="thin", color="B7B7B7")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row_number in header_rows:
        for cell in ws[row_number]:
            cell.fill = PatternFill("solid", fgColor=blue)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        first_value = row[0].value
        if first_value in ("合计", "汇总结果"):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=total_fill if first_value == "合计" else light)
                cell.font = Font(bold=True)
        if row[0].row == 1 and first_value and "汇总" in str(first_value):
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=light)
                cell.font = Font(bold=True, size=14)
    if title_merge:
        ws.merge_cells(title_merge)
        ws[title_merge.split(":")[0]].alignment = Alignment(horizontal="center", vertical="center")
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), 56)
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = max(width, 10)
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.00"


def write_output(
    output_path: Path,
    title: str,
    plans: list[dict[str, Any]],
    detail: list[dict[str, Any]],
    missing: list[str],
) -> dict[str, Any]:
    summary = build_summary(detail)
    total_qty = sum(item["qty"] for item in detail)
    total_weight = sum(item["total_weight"] for item in detail)
    total_bars = sum(item["bar_qty"] for item in detail)

    wb = Workbook()
    ws_total = wb.active
    ws_total.title = "总汇总"
    ws_total.append([f"{title}料单汇总"])
    ws_total.append(["下料总数/只", format_int(total_qty)])
    ws_total.append(["需求总重量kg", total_weight])
    ws_total.append(["单位规格材料总条数", total_bars])
    if missing:
        ws_total.append(["未匹配型号", "、".join(missing)])
    else:
        ws_total.append([])
    ws_total.append(["规格尺寸", "下料总数/只", "需求总重量kg", "单位规格材料总条数"])
    for spec, data in sorted(summary.items(), key=lambda item: spec_sort_key(item[0])):
        ws_total.append([spec, format_int(data["plan_qty"]), data["weight"], data["bar_qty"]])
    ws_total.append(["合计", format_int(total_qty), total_weight, total_bars])

    ws_detail = wb.create_sheet("料单明细")
    ws_detail.append(DETAIL_HEADERS)
    for item in detail:
        ws_detail.append(
            [
                item["model"],
                item.get("car"),
                format_int(item["qty"]),
                item.get("left"),
                item.get("right"),
                item["code"],
                item["part"],
                item["calc_spec"],
                format_int(item["pieces"]),
                item["bar_qty"],
                item["unit_weight"],
                item["total_weight"],
                item.get("remark"),
                item["source_row"],
            ]
        )
    ws_detail.append(["合计", "", format_int(total_qty), "", "", "", "", "", "", total_bars, "", total_weight, "", ""])

    ws_summary = wb.create_sheet("按规格汇总")
    ws_summary.append(["汇总结果", "", "", "", "", ""])
    ws_summary.append(["下料总数/只", format_int(total_qty), "需求总重量kg", total_weight, "单位规格材料总条数", total_bars])
    ws_summary.append([])
    ws_summary.append(SUMMARY_HEADERS)
    for spec, data in sorted(summary.items(), key=lambda item: spec_sort_key(item[0])):
        ws_summary.append(
            [
                spec,
                format_int(data["plan_qty"]),
                data["weight"],
                data["bar_qty"],
                "、".join(sorted(data["models"])),
                "；".join(data["details"]),
            ]
        )
    ws_summary.append(["合计", format_int(total_qty), total_weight, total_bars, "", ""])

    ws_plan = wb.create_sheet("生产需求")
    ws_plan.append(PLAN_HEADERS)
    for plan in plans:
        ws_plan.append(
            [
                plan["model"],
                plan.get("car"),
                format_int(plan["qty"]),
                plan.get("left"),
                plan.get("right"),
                plan.get("finish_date"),
                plan.get("remark"),
            ]
        )

    if missing:
        ws_check = wb.create_sheet("检查提醒")
        ws_check.append(["类型", "内容"])
        for model in missing:
            ws_check.append(["材料数据未匹配型号", model])
        style_output_sheet(ws_check, header_rows=(1,))

    style_output_sheet(ws_total, header_rows=(6,), title_merge="A1:D1")
    style_output_sheet(ws_detail, header_rows=(1,))
    style_output_sheet(ws_summary, header_rows=(4,), title_merge="A1:F1")
    style_output_sheet(ws_plan, header_rows=(1,))

    ws_total.freeze_panes = "A7"
    ws_detail.freeze_panes = "A2"
    ws_summary.freeze_panes = "A5"
    ws_plan.freeze_panes = "A2"
    ws_total.auto_filter.ref = f"A6:D{max(6, ws_total.max_row - 1)}"
    ws_detail.auto_filter.ref = f"A1:{get_column_letter(ws_detail.max_column)}{max(1, ws_detail.max_row - 1)}"
    ws_summary.auto_filter.ref = f"A4:{get_column_letter(ws_summary.max_column)}{max(4, ws_summary.max_row - 1)}"
    ws_plan.auto_filter.ref = ws_plan.dimensions
    for row in ws_detail.iter_rows(min_row=2, max_row=max(2, ws_detail.max_row - 1), min_col=11, max_col=11):
        for cell in row:
            cell.number_format = "0.0000"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        "title": title,
        "plan_count": len(plans),
        "detail_count": len(detail),
        "spec_count": len(summary),
        "missing": missing,
        "total_qty": format_int(total_qty),
        "total_weight": total_weight,
        "total_bars": total_bars,
    }


def unique_output_path(output_dir: Path, title: str, filename_prefix: str = "") -> Path:
    filename = f"{filename_prefix}{safe_filename_part(title)}料单.xlsx"
    candidate = output_dir / filename
    if not candidate.exists():
        return candidate
    stem = Path(filename).stem
    counter = 2
    while True:
        candidate = output_dir / f"{stem}_{counter}.xlsx"
        if not candidate.exists():
            return candidate
        counter += 1


def generate_material_sheet(materials_path: Path, plan_path: Path, output_dir: Path) -> tuple[Path, dict[str, Any]]:
    if not materials_path.exists():
        raise ValueError("还没有材料数据文件，请先上传冲压料单.xlsx。")
    title, plans = read_plan(plan_path)
    materials = read_materials(materials_path)
    return generate_material_sheet_from_rows(materials, title, plans, output_dir)


def generate_material_sheet_from_materials(
    materials: dict[str, list[dict[str, Any]]],
    plan_path: Path,
    output_dir: Path,
    filename_prefix: str = "",
) -> tuple[Path, dict[str, Any]]:
    title, plans = read_plan(plan_path)
    return generate_material_sheet_from_rows(materials, title, plans, output_dir, filename_prefix=filename_prefix)


def generate_material_sheet_from_rows(
    materials: dict[str, list[dict[str, Any]]],
    title: str,
    plans: list[dict[str, Any]],
    output_dir: Path,
    filename_prefix: str = "",
) -> tuple[Path, dict[str, Any]]:
    detail, missing = calculate(plans, materials)
    if not detail:
        raise ValueError("没有生成任何料单明细，请检查型号是否和材料数据 A 列一致。")
    output_path = unique_output_path(output_dir, title, filename_prefix=filename_prefix)
    summary = write_output(output_path, title, plans, detail, missing)
    return output_path, summary

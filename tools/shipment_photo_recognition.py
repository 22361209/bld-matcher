#!/usr/bin/env python3
from __future__ import annotations

import argparse
import base64
import json
import os
import re
import subprocess
import sys
import tempfile
import time
import urllib.error
import urllib.request
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageOps
from pillow_heif import register_heif_opener


register_heif_opener()


IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff", ".heic", ".heif"}
DEFAULT_BASE_URL = "https://api.openai.com/v1"
DEFAULT_ENDPOINT_PATH = "/chat/completions"
OUTPUT_ROOT = Path("outputs") / "shipment_photo_recognition"

SYSTEM_PROMPT = """你是发货核对照片识别助手。任务是从货物/纸箱/托盘照片中识别每一张实际贴在货物上的白色标签，然后输出给程序处理的结构化 JSON。

标签通常为 60mm*90mm 白色标签，可能包含 PART NO、OE NO、BLD NO、C/NO、客户编码、产品名称、适配车型、QTY/PCS/数量、NW、GW、MEAS/纸箱尺寸、MADE IN CHINA、条形码等信息。
请只根据可见标签内容输出，不要猜测被遮挡或看不清的信息。不要自己生成表格，不要解释过程，不要使用 HTML，不要输出 <br>。

输出必须是纯 JSON，不要输出 Markdown。JSON 结构：
{
  "photo_summary": "简短说明照片内容和识别质量",
  "labels": [
    {
      "label_index": 1,
      "visible": true,
      "label_type": "part|customer|mixed|unknown",
      "numbers": ["标签上出现的所有主要号码，含 PART NO/OE/BLD/CNO/客户编码"],
      "part_no": "PART NO 或产品号码，无法确认则空字符串",
      "bld_no": "如果明确是 BLD 号则填写，否则空字符串",
      "oe_no": "如果明确是 OE 号则填写，否则空字符串",
      "customer_code": "如果明确是客户编码、C/NO、客户箱号则填写，否则空字符串",
      "product_name": "产品名称，无法确认则空字符串",
      "models": "适配车型，无法确认则空字符串",
      "quantity": 0,
      "carton_size": "纸箱尺寸，无法确认则空字符串",
      "barcode": "条形码数字/文字，无法确认则空字符串",
      "confidence": 0.0,
      "notes": "看不清、遮挡、反光、重复标签等备注"
    }
  ]
}

规则：
- 一张可见的物理标签对应 labels 里一条记录；不要把多张相同标签合并。
- 如果照片里有 3 张完全相同的 PART NO 标签，也必须输出 3 条 labels，后续程序会按 3 箱汇总。
- label_type 用于区分标签：包含 PART NO/QTY 的为 part；只包含 C/NO/客户名/MADE IN CHINA 的为 customer；两类都有为 mixed。
- quantity 必须来自标签上明确的 QTY/PCS/数量字段；看不清或没有数量时填 0，并在 notes 说明。
- PART NO 是货物号码时，必须放入 part_no，也要放入 numbers。比如 54501-8Y50B。
- C/NO、CNO、客户箱号等放入 customer_code。比如 C/NO: 186 则 customer_code 为 186。
- MEAS、NW、GW 不是产品号码，不要放入 numbers；MEAS 放入 carton_size，NW/GW 可写入 notes。
- 如果同一张标签上有多个号码，numbers 全部列出；不要把条形码误当唯一产品号码。
- product_name 只填真实产品名称；客户名、品牌名或 COLORMARKET 这类客户/市场文字不要当产品名称。
- confidence 使用 0 到 1，低于 0.75 的项后续会人工复核。
"""


@dataclass(frozen=True)
class PhotoJob:
    path: Path
    relative_name: str


def _load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key:
            os.environ.setdefault(key, value)


def _compact_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"<br\s*/?>", " ", text, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", text).strip()


def _safe_int(value: Any) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return max(0, int(value))
    match = re.search(r"\d+", str(value))
    return max(0, int(match.group(0))) if match else 0


def _safe_float(value: Any) -> float:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return 0.0
    if number < 0:
        return 0.0
    if number > 1:
        return 1.0
    return number


def _image_mime(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix in {".jpg", ".jpeg"}:
        return "image/jpeg"
    if suffix == ".png":
        return "image/png"
    if suffix == ".webp":
        return "image/webp"
    return "image/jpeg"


def _prepare_image(path: Path, max_side: int) -> tuple[bytes, str]:
    suffix = path.suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".webp"}:
        with Image.open(path) as image:
            image = ImageOps.exif_transpose(image)
            if max(image.size) > max_side:
                image.thumbnail((max_side, max_side))
            if image.mode not in {"RGB", "L"}:
                image = image.convert("RGB")
            with tempfile.NamedTemporaryFile(suffix=".jpg") as handle:
                image.save(handle.name, format="JPEG", quality=88, optimize=True)
                return Path(handle.name).read_bytes(), "image/jpeg"

    with Image.open(path) as image:
        image = ImageOps.exif_transpose(image).convert("RGB")
        if max(image.size) > max_side:
            image.thumbnail((max_side, max_side))
        with tempfile.NamedTemporaryFile(suffix=".jpg") as handle:
            image.save(handle.name, format="JPEG", quality=88, optimize=True)
            return Path(handle.name).read_bytes(), "image/jpeg"


def _data_url(path: Path, max_side: int) -> str:
    raw, mime = _prepare_image(path, max_side)
    encoded = base64.b64encode(raw).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def _extract_json_object(text: str) -> dict[str, Any]:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?", "", cleaned, flags=re.IGNORECASE).strip()
        cleaned = re.sub(r"```$", "", cleaned).strip()
    try:
        value = json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start < 0 or end <= start:
            raise
        value = json.loads(cleaned[start : end + 1])
    if not isinstance(value, dict):
        raise ValueError("模型返回的 JSON 顶层不是对象。")
    return value


def _chat_completion_request(
    *,
    api_key: str,
    base_url: str,
    endpoint_path: str,
    model: str,
    image_path: Path,
    timeout: int,
    max_side: int,
) -> dict[str, Any]:
    url = base_url.rstrip("/") + "/" + endpoint_path.strip("/")
    payload = {
        "model": model,
        "temperature": 0,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": (
                            "请识别这张发货照片中的每一张货物标签，按指定 JSON 结构输出。"
                            "重点读取标签上的号码、产品名称、数量、车型和纸箱尺寸。"
                        ),
                    },
                    {"type": "image_url", "image_url": {"url": _data_url(image_path, max_side), "detail": "high"}},
                ],
            },
        ],
    }
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=body,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            response_payload = json.loads(response.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"视觉接口返回 HTTP {exc.code}: {detail[:800]}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"视觉接口请求失败: {exc}") from exc

    try:
        content = response_payload["choices"][0]["message"]["content"]
    except (KeyError, IndexError, TypeError) as exc:
        raise RuntimeError(f"视觉接口返回格式无法识别: {response_payload}") from exc
    parsed = _extract_json_object(content)
    usage = response_payload.get("usage")
    if isinstance(usage, dict):
        parsed["_usage"] = usage
    response_model = response_payload.get("model")
    if response_model:
        parsed["_model"] = response_model
    return parsed


def _available_tesseract_languages(timeout: int) -> set[str]:
    result = subprocess.run(["tesseract", "--list-langs"], check=False, capture_output=True, text=True, timeout=timeout)
    if result.returncode != 0:
        return set()
    return {line.strip() for line in result.stdout.splitlines() if line.strip() and not line.startswith("List of")}


def _tesseract_language_arg(timeout: int) -> str:
    available = _available_tesseract_languages(timeout)
    preferred = [lang for lang in ("eng", "snum", "chi_sim") if lang in available]
    return "+".join(preferred or ["eng"])


def _tesseract_text(image_path: Path, timeout: int) -> str:
    if not shutil_which("tesseract"):
        raise RuntimeError("没有找到 tesseract 命令。")
    command = ["tesseract", str(image_path), "stdout", "-l", _tesseract_language_arg(timeout), "--psm", "6"]
    result = subprocess.run(command, check=False, capture_output=True, text=True, timeout=timeout)
    if result.returncode != 0:
        raise RuntimeError(result.stderr.strip() or "tesseract 识别失败。")
    return result.stdout.strip()


def shutil_which(command: str) -> str | None:
    for directory in os.environ.get("PATH", "").split(os.pathsep):
        candidate = Path(directory) / command
        if candidate.exists() and os.access(candidate, os.X_OK):
            return str(candidate)
    return None


def _numbers_from_text(text: str) -> list[str]:
    candidates = re.findall(r"[A-Z0-9][A-Z0-9\-./]{3,}[A-Z0-9]", text.upper())
    blocked = {"QTY", "PCS", "CTN", "MADE", "CHINA", "MODEL", "SIZE"}
    numbers: list[str] = []
    seen = set()
    for item in candidates:
        normalized = re.sub(r"[^A-Z0-9]", "", item)
        if len(normalized) < 4 or normalized in blocked or normalized in seen:
            continue
        seen.add(normalized)
        numbers.append(item)
    return numbers


def _quantity_from_text(text: str) -> int:
    patterns = [
        r"(?:QTY|QUANTITY|数量|PCS|件数)\s*[:：]?\s*(\d+)",
        r"(\d+)\s*(?:PCS|PC|件|只)\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return int(match.group(1))
    return 0


def _tesseract_result(image_path: Path, timeout: int) -> dict[str, Any]:
    text = _tesseract_text(image_path, timeout)
    numbers = _numbers_from_text(text)
    quantity = _quantity_from_text(text)
    label = {
        "label_index": 1,
        "visible": bool(text),
        "label_type": "unknown",
        "numbers": numbers,
        "part_no": "",
        "bld_no": "",
        "oe_no": "",
        "customer_code": "",
        "product_name": "",
        "models": "",
        "quantity": quantity,
        "carton_size": "",
        "barcode": "",
        "confidence": 0.35 if text else 0.0,
        "notes": "本地 OCR 草稿；无法可靠区分多张标签，请人工复核。",
    }
    return {"photo_summary": "本地 OCR 草稿", "labels": [label] if text else [], "ocr_text": text}


def _normalize_label(raw: Any, default_index: int) -> dict[str, Any]:
    item = raw if isinstance(raw, dict) else {}
    numbers = item.get("numbers")
    if isinstance(numbers, str):
        numbers = [part.strip() for part in re.split(r"[,，;；、\n]+", numbers) if part.strip()]
    if not isinstance(numbers, list):
        numbers = []
    numbers = [_compact_text(value) for value in numbers if _compact_text(value)]
    return {
        "label_index": _safe_int(item.get("label_index")) or default_index,
        "visible": bool(item.get("visible", True)),
        "label_type": _compact_text(item.get("label_type")),
        "numbers": numbers,
        "part_no": _compact_text(item.get("part_no")),
        "bld_no": _compact_text(item.get("bld_no")),
        "oe_no": _compact_text(item.get("oe_no")),
        "customer_code": _compact_text(item.get("customer_code")),
        "product_name": _compact_text(item.get("product_name")),
        "models": _compact_text(item.get("models")),
        "quantity": _safe_int(item.get("quantity")),
        "carton_size": _compact_text(item.get("carton_size")),
        "barcode": _compact_text(item.get("barcode")),
        "confidence": _safe_float(item.get("confidence")),
        "notes": _compact_text(item.get("notes")),
    }


def _normalize_result(raw: dict[str, Any]) -> dict[str, Any]:
    labels = raw.get("labels", [])
    if not isinstance(labels, list):
        labels = []
    return {
        "photo_summary": _compact_text(raw.get("photo_summary")),
        "ocr_text": _compact_text(raw.get("ocr_text")),
        "labels": [_normalize_label(label, index) for index, label in enumerate(labels, start=1)],
    }


def _normalize_usage(raw: Any) -> dict[str, int]:
    usage = raw if isinstance(raw, dict) else {}
    return {
        "prompt_tokens": _safe_int(usage.get("prompt_tokens")),
        "completion_tokens": _safe_int(usage.get("completion_tokens")),
        "total_tokens": _safe_int(usage.get("total_tokens")),
    }


def _label_key(label: dict[str, Any]) -> tuple[str, str, str, str]:
    numbers = [label.get("part_no"), label.get("bld_no"), label.get("oe_no"), label.get("customer_code"), *(label.get("numbers") or [])]
    number_key = next((_compact_text(value) for value in numbers if _compact_text(value)), "未识别号码")
    return (
        number_key,
        _compact_text(label.get("bld_no")),
        _compact_text(label.get("product_name")),
        _compact_text(label.get("models")),
    )


def _summary_rows(photo_results: list[dict[str, Any]], run_date: str) -> list[dict[str, Any]]:
    groups: dict[tuple[str, str, str, str], dict[str, Any]] = {}
    for photo in photo_results:
        for label in photo["result"].get("labels", []):
            key = _label_key(label)
            row = groups.setdefault(
                key,
                {
                    "date": run_date,
                    "label_numbers": key[0],
                    "bld_no": key[1],
                    "product_name": key[2],
                    "quantity": 0,
                    "models": key[3],
                    "cartons": 0,
                    "photos": set(),
                    "low_confidence": 0,
                    "notes": [],
                },
            )
            row["quantity"] += _safe_int(label.get("quantity"))
            row["cartons"] += 1
            row["photos"].add(photo["relative_name"])
            if _safe_float(label.get("confidence")) < 0.75:
                row["low_confidence"] += 1
            if label.get("notes"):
                row["notes"].append(label["notes"])

    rows = []
    for row in groups.values():
        rows.append(
            {
                **row,
                "photos": "；".join(sorted(row["photos"])),
                "notes": "；".join(dict.fromkeys(row["notes"])),
            }
        )
    return sorted(rows, key=lambda item: (item["label_numbers"], item["product_name"], item["models"]))


def _auto_width(sheet) -> None:
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(60, len(value)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.column_dimensions[column_letter].width = max(10, max_length + 2)


def _write_sheet(sheet, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for row in rows:
        sheet.append(row)
    _auto_width(sheet)
    sheet.freeze_panes = "A2"


def write_workbook(photo_results: list[dict[str, Any]], output_path: Path, run_date: str) -> None:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "汇总"
    summary = _summary_rows(photo_results, run_date)
    _write_sheet(
        summary_sheet,
        ["日期", "标签上的号码", "BLD号", "产品名称", "数量", "车型", "箱数", "来源照片", "低置信标签数", "备注"],
        [
            [
                row["date"],
                row["label_numbers"],
                row["bld_no"],
                row["product_name"],
                row["quantity"],
                row["models"],
                row["cartons"],
                row["photos"],
                row["low_confidence"],
                row["notes"],
            ]
            for row in summary
        ],
    )

    detail_rows = []
    for photo in photo_results:
        for label in photo["result"].get("labels", []):
            detail_rows.append(
                [
                    run_date,
                    photo["relative_name"],
                    label["label_index"],
                    label.get("label_type", ""),
                    "；".join(label.get("numbers") or []),
                    label.get("part_no", ""),
                    label.get("bld_no", ""),
                    label.get("oe_no", ""),
                    label.get("customer_code", ""),
                    label.get("product_name", ""),
                    label.get("quantity", 0),
                    label.get("models", ""),
                    label.get("carton_size", ""),
                    label.get("barcode", ""),
                    label.get("confidence", 0),
                    label.get("notes", ""),
                ]
            )
    detail_sheet = workbook.create_sheet("标签明细")
    _write_sheet(
        detail_sheet,
        ["日期", "照片", "标签序号", "标签类型", "标签上的号码", "PART NO", "BLD号", "OE号", "客户编码", "产品名称", "数量", "车型", "纸箱尺寸", "条形码", "置信度", "备注"],
        detail_rows,
    )

    photo_sheet = workbook.create_sheet("照片清单")
    _write_sheet(
        photo_sheet,
        ["照片", "识别状态", "标签数", "耗时秒", "输入Token", "输出Token", "总Token", "模型", "照片说明", "错误"],
        [
            [
                photo["relative_name"],
                photo["status"],
                len(photo["result"].get("labels", [])),
                photo.get("seconds", 0),
                photo.get("usage", {}).get("prompt_tokens", 0),
                photo.get("usage", {}).get("completion_tokens", 0),
                photo.get("usage", {}).get("total_tokens", 0),
                photo.get("model", ""),
                photo["result"].get("photo_summary", ""),
                photo.get("error", ""),
            ]
            for photo in photo_results
        ],
    )

    raw_sheet = workbook.create_sheet("原始结果")
    _write_sheet(
        raw_sheet,
        ["照片", "JSON"],
        [[photo["relative_name"], json.dumps(photo, ensure_ascii=False, indent=2)] for photo in photo_results],
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def find_photos(input_dir: Path) -> list[PhotoJob]:
    root = input_dir.resolve()
    paths = [path for path in root.rglob("*") if path.is_file() and path.suffix.lower() in IMAGE_SUFFIXES]
    return [PhotoJob(path, path.relative_to(root).as_posix()) for path in sorted(paths)]


def recognize_photo(job: PhotoJob, args: argparse.Namespace) -> dict[str, Any]:
    started = time.time()
    provider = args.provider
    try:
        if provider == "tesseract":
            raw = _tesseract_result(job.path, args.timeout)
            usage = {}
            response_model = "tesseract"
        else:
            api_key = (
                os.environ.get("SHIPMENT_VISION_API_KEY")
                or os.environ.get("DASHSCOPE_API_KEY")
                or os.environ.get("QWEN_API_KEY")
                or os.environ.get("OPENAI_API_KEY")
            )
            model = args.model or os.environ.get("SHIPMENT_VISION_MODEL")
            base_url = args.base_url or os.environ.get("SHIPMENT_VISION_BASE_URL") or DEFAULT_BASE_URL
            endpoint_path = args.endpoint_path or os.environ.get("SHIPMENT_VISION_ENDPOINT_PATH") or DEFAULT_ENDPOINT_PATH
            if not api_key:
                raise RuntimeError("缺少 SHIPMENT_VISION_API_KEY、DASHSCOPE_API_KEY、QWEN_API_KEY 或 OPENAI_API_KEY。")
            if not model:
                raise RuntimeError("缺少 --model 或 SHIPMENT_VISION_MODEL。")
            raw = _chat_completion_request(
                api_key=api_key,
                base_url=base_url,
                endpoint_path=endpoint_path,
                model=model,
                image_path=job.path,
                timeout=args.timeout,
                max_side=args.max_side,
            )
            usage = raw.pop("_usage", {})
            response_model = _compact_text(raw.pop("_model", "")) or model
        result = _normalize_result(raw)
        return {
            "relative_name": job.relative_name,
            "path": str(job.path),
            "status": "ok",
            "seconds": round(time.time() - started, 2),
            "model": response_model,
            "usage": _normalize_usage(usage),
            "result": result,
            "error": "",
        }
    except Exception as exc:
        return {
            "relative_name": job.relative_name,
            "path": str(job.path),
            "status": "error",
            "seconds": round(time.time() - started, 2),
            "model": "",
            "usage": _normalize_usage({}),
            "result": {"photo_summary": "", "labels": []},
            "error": str(exc),
        }


def default_output_path(input_dir: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    safe_name = re.sub(r"[^A-Za-z0-9\u4E00-\u9FFF_-]+", "-", input_dir.name).strip("-") or "shipment"
    return OUTPUT_ROOT / f"shipment-photo-{timestamp}-{safe_name}.xlsx"


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="识别发货照片标签并汇总生成 Excel。")
    parser.add_argument("input_dir", type=Path, help="照片文件夹，可指向 NAS 同步到本机/容器内的目录。")
    parser.add_argument("--output", type=Path, help="输出 Excel 路径，默认写入 outputs/shipment_photo_recognition/。")
    parser.add_argument("--json-output", type=Path, help="可选：保存结构化识别 JSON。")
    parser.add_argument("--provider", choices=["openai-compatible", "tesseract"], default="openai-compatible")
    parser.add_argument("--model", help="视觉模型名。也可用 SHIPMENT_VISION_MODEL 环境变量。")
    parser.add_argument("--base-url", help="OpenAI-compatible API base URL，默认 https://api.openai.com/v1。")
    parser.add_argument("--endpoint-path", default="", help="接口路径，默认 /chat/completions。")
    parser.add_argument("--date", default=date.today().isoformat(), help="写入 Excel 的发货日期，默认今天。")
    parser.add_argument("--timeout", type=int, default=180, help="单张照片识别超时时间秒数。")
    parser.add_argument("--max-side", type=int, default=2200, help="发送给视觉模型前压缩图片的最长边。")
    parser.add_argument("--limit", type=int, default=0, help="只处理前 N 张，用于试跑。")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    _load_env_file(Path(".env"))
    args = parse_args(argv)
    input_dir = args.input_dir.expanduser().resolve()
    if not input_dir.exists() or not input_dir.is_dir():
        print(f"照片文件夹不存在：{input_dir}", file=sys.stderr)
        return 2

    photos = find_photos(input_dir)
    if args.limit > 0:
        photos = photos[: args.limit]
    if not photos:
        print(f"没有找到图片文件：{input_dir}", file=sys.stderr)
        return 1

    print(f"找到 {len(photos)} 张照片，开始识别。")
    results = []
    for index, job in enumerate(photos, start=1):
        print(f"[{index}/{len(photos)}] {job.relative_name}")
        result = recognize_photo(job, args)
        label_count = len(result["result"].get("labels", []))
        status = result["status"]
        print(f"  {status}，标签 {label_count}，耗时 {result['seconds']}s")
        if result["error"]:
            print(f"  错误：{result['error']}")
        results.append(result)

    output_path = (args.output or default_output_path(input_dir)).expanduser().resolve()
    write_workbook(results, output_path, args.date)
    json_output = (args.json_output.expanduser().resolve() if args.json_output else output_path.with_suffix(".json"))
    json_output.parent.mkdir(parents=True, exist_ok=True)
    json_output.write_text(json.dumps(results, ensure_ascii=False, indent=2), encoding="utf-8")

    total_labels = sum(len(item["result"].get("labels", [])) for item in results)
    failed = sum(1 for item in results if item["status"] != "ok")
    print(f"完成：照片 {len(results)} 张，标签 {total_labels} 张，失败 {failed} 张。")
    print(f"Excel：{output_path}")
    print(f"JSON：{json_output}")
    return 0 if failed == 0 else 3


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
